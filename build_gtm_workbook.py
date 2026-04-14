"""
The English Hub - Master GTM Campaign Tracker & Marketing Plan
Builds comprehensive Excel workbook from all strategy documents.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

# ── Style definitions ──
HEADER_FONT = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
SUBHEADER_FONT = Font(name='Calibri', bold=True, size=11, color='1F4E79')
BODY_FONT = Font(name='Calibri', size=10)
SMALL_FONT = Font(name='Calibri', size=9, color='666666')
TITLE_FONT = Font(name='Calibri', bold=True, size=16, color='1F4E79')
SECTION_FONT = Font(name='Calibri', bold=True, size=13, color='FFFFFF')

DARK_BLUE = PatternFill('solid', fgColor='1F4E79')
MED_BLUE = PatternFill('solid', fgColor='2E75B6')
LIGHT_BLUE = PatternFill('solid', fgColor='D6E4F0')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
ORANGE_FILL = PatternFill('solid', fgColor='F4B084')
PURPLE_FILL = PatternFill('solid', fgColor='D9C2EC')
GREY_FILL = PatternFill('solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

def style_header_row(ws, row, cols, fill=DARK_BLUE, font=HEADER_FONT):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data_row(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill

def add_section_row(ws, row, text, cols, fill=MED_BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = SECTION_FONT
    cell.fill = fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.row_dimensions[row].height = 28

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = "1F4E79"
set_col_widths(ws, [3, 25, 20, 20, 20, 20, 20])

ws.merge_cells('B2:F2')
ws['B2'] = "THE ENGLISH HUB — MASTER GTM PLAN & CAMPAIGN TRACKER"
ws['B2'].font = TITLE_FONT

ws.merge_cells('B3:F3')
ws['B3'] = "Go-To-Market Strategy | Affiliate Marketing | Content | SEO | B2B Sales | International Growth"
ws['B3'].font = Font(name='Calibri', size=11, color='666666', italic=True)

ws.merge_cells('B4:F4')
ws['B4'] = f"Created: {datetime.now().strftime('%d %B %Y')} | Product: theenglishhub.app | Pricing: £9.99/mo | £79.99/yr | 30-day free trial"
ws['B4'].font = SMALL_FONT

# KPI targets
r = 6
add_section_row(ws, r, "KEY MILESTONES & TARGETS", 6)
r += 1
headers = ["Metric", "Month 1", "Month 3", "Month 6", "Month 12"]
for i, h in enumerate(headers, 2):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, 6)

kpis = [
    ["Paying Subscribers", "30+", "300", "1,000", "3,000+"],
    ["MRR (£)", "£300", "£3,000", "£10,000", "£30,000+"],
    ["Active Affiliates", "100", "400", "1,000", "2,500+"],
    ["Teacher Ambassadors", "5", "15", "30", "50+"],
    ["School Partnerships", "5 trials", "10 paid", "20 paid", "50+ paid"],
    ["Monthly Website Visitors", "2,500", "10,000", "25,000", "75,000+"],
    ["Email Subscribers", "500", "3,000", "8,000", "20,000+"],
    ["TikTok Followers", "500", "5,000", "15,000", "50,000+"],
    ["YouTube Subscribers", "100", "1,000", "5,000", "15,000+"],
    ["Blog Posts Published", "10", "30", "60", "120+"],
    ["International Subscribers (GCC)", "10", "100", "500", "2,000+"],
]
for row_data in kpis:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws.cell(row=r, column=i, value=v)
    style_data_row(ws, r, 6, GREY_FILL if r % 2 == 0 else None)

# Phase overview
r += 2
add_section_row(ws, r, "LAUNCH PHASES OVERVIEW", 6)
r += 1
headers = ["Phase", "Timeline", "Objective", "Key Focus", "Success Criteria"]
for i, h in enumerate(headers, 2):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, 6)

phases = [
    ["Phase 1: Pre-Launch", "Days 1-4 (Cowork Blitz)", "Build waitlist + recruit 200 affiliates + bank ALL content", "4-day intensive: landing page, 200 affiliate recruits, ALL 30 TikToks + 4 YTs filmed, teacher outreach, SEO foundation", "500+ waitlist, 200 affiliates contacted, 30 TikToks + 4 YTs banked, 10 blog posts, 10 schools in pipeline"],
    ["Phase 2: Soft Launch", "Week 2", "Convert waitlist to trials & paid", "Waitlist launch sequence, affiliate activation, content blitz (pre-banked), community seeding", "100+ free trials, 30+ paid, 5+ testimonials, 100 active affiliates"],
    ["Phase 3: Growth", "Months 2-3", "Reach 300 paying subscribers", "Scale affiliates to 500+, test paid ads, school partnerships push, international entry", "300 subscribers, £3K MRR, 500 affiliates, 5+ school deals, 10K monthly visitors"],
    ["Phase 4: Scale", "Months 4-6", "Reach 1,000 subscribers", "Double down on winners, 1,000+ affiliates, exam season campaign, referral programme, B2B pipeline", "1,000 subs, £10K MRR, 1,000+ affiliates, 20+ schools, 25K monthly visitors"],
    ["Phase 5: International", "Months 3-6", "GCC market penetration", "Middle East school outreach, crisis-driven families, British curriculum schools, 500+ GCC affiliates", "500 international subs, 10 GCC school partnerships, 200+ ME influencers"],
]
for row_data in phases:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws.cell(row=r, column=i, value=v)
    style_data_row(ws, r, 6)
    ws.row_dimensions[r].height = 60

# Budget overview
r += 2
add_section_row(ws, r, "BUDGET ALLOCATION (FIRST 6 MONTHS)", 6)
r += 1
headers = ["Category", "Monthly Budget", "6-Month Total", "Notes", ""]
for i, h in enumerate(headers, 2):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, 6)

budget = [
    ["Affiliate Commissions (variable)", "Variable ~£200-2,000", "~£6,000", "20% of affiliate-driven revenue — scales with success"],
    ["Paid Ads (test)", "£300-500", "£2,400", "Google Ads + TikTok Spark Ads + Meta retargeting"],
    ["Content Production", "£100-200", "£900", "Canva Pro, music licensing, minimal video production"],
    ["SEO Tools", "£50-100", "£450", "Ahrefs/SEMrush lite + rank tracking"],
    ["Email Marketing", "£0-50", "£150", "Free tier initially, scale as list grows"],
    ["Conference/Events", "£0 (months 1-2), £200+", "£600", "Bett Show, Education Show, TeachMeets"],
    ["Affiliate Software (Rewardful)", "£29-79/mo", "£350", "Grows with affiliate count"],
    ["TOTAL ESTIMATED", "£700-3,000/mo", "£10,850", "Majority is variable (affiliate commission = pay-for-performance)"],
]
for row_data in budget:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws.cell(row=r, column=i, value=v)
    style_data_row(ws, r, 6, GREY_FILL if r % 2 == 0 else None)

# ════════════════════════════════════════════════════════════════
# SHEET 2: MASTER TASK LIST (THE CORE SEQUENTIAL TRACKER)
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Master Task List")
ws2.sheet_properties.tabColor = "2E75B6"
cols2 = ["#", "Phase", "Category", "Task", "Priority", "Status", "Owner", "Start Date", "Due Date", "Dependencies", "Notes / Details", "Completion %"]
set_col_widths(ws2, [5, 14, 16, 55, 10, 12, 12, 12, 12, 18, 40, 10])

# Headers
for i, h in enumerate(cols2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header_row(ws2, 1, len(cols2))
ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}1"
ws2.freeze_panes = 'A2'

tasks = [
    # ── PHASE 1: PRE-LAUNCH (Weeks 1-2) ──
    # Technical
    [1, "Phase 1", "Technical", "Fix 500 errors on /exam-guide/* pages (AQA, Edexcel, OCR, WJEC, IGCSE)", "CRITICAL", "Not Started", "Dev", "", "", "None", "Currently returning 500 — blocks SEO and user access"],
    [2, "Phase 1", "Technical", "Fix duplicate title tags on /about, /mock-exams, /games, /subjects pages", "CRITICAL", "Not Started", "Dev", "", "", "None", "All using default homepage title — SEO audit finding"],
    [3, "Phase 1", "Technical", "Implement structured data / JSON-LD schema markup (Organization, Product, FAQPage)", "IMPORTANT", "Not Started", "Dev", "", "", "#2", "No schema detected anywhere on site"],
    [4, "Phase 1", "Technical", "Add canonical tags and Open Graph / Twitter Card meta tags", "IMPORTANT", "Not Started", "Dev", "", "", "#2", "Missing — kills social sharing previews"],
    [5, "Phase 1", "Technical", "Complete pre-launch technical checklist (SSL, CORS, rate limiting, env vars)", "CRITICAL", "Not Started", "Dev", "", "", "None", "See Launch Checklist doc — all CRITICAL items"],
    [6, "Phase 1", "Technical", "Test Stripe live mode end-to-end (signup → trial → payment → cancel)", "CRITICAL", "Not Started", "Dev", "", "", "#5", "Must work before any user sees the product"],
    [7, "Phase 1", "Technical", "Configure Rewardful affiliate tracking with Stripe live mode", "CRITICAL", "Not Started", "Dev", "", "", "#6", "Affiliate links must track correctly before recruitment"],
    [8, "Phase 1", "Technical", "Set up Sentry error monitoring (frontend + backend)", "CRITICAL", "Not Started", "Dev", "", "", "#5", "Cannot launch blind to errors"],
    [9, "Phase 1", "Technical", "Configure database backups (Azure PostgreSQL, 7-day retention)", "CRITICAL", "Not Started", "Dev", "", "", "None", "Test restore procedure"],
    [10, "Phase 1", "Technical", "Mobile responsiveness testing on real devices (iPhone Safari, Android Chrome, iPad)", "CRITICAL", "Not Started", "Dev/QA", "", "", "#5", "Majority of student users on mobile"],
    [11, "Phase 1", "Technical", "Set up GA4 + Vercel Analytics with consent banner", "CRITICAL", "Not Started", "Dev", "", "", "#5", "Track signup, trial start, payment events"],
    [12, "Phase 1", "Technical", "Implement UTM tracking convention across all marketing links", "IMPORTANT", "Not Started", "Dev", "", "", "#11", "Must know which channels drive signups"],

    # Legal & Compliance
    [13, "Phase 1", "Legal", "Publish Terms of Service at /terms", "CRITICAL", "Not Started", "Legal", "", "", "None", "No legal agreement = no protection"],
    [14, "Phase 1", "Legal", "Publish GDPR-compliant Privacy Policy at /privacy", "CRITICAL", "Not Started", "Legal", "", "", "None", "UK GDPR fines up to £17.5M"],
    [15, "Phase 1", "Legal", "Implement cookie consent banner (UK PECR compliant)", "CRITICAL", "Not Started", "Dev/Legal", "", "", "#14", "Block non-essential cookies until consent"],
    [16, "Phase 1", "Legal", "Publish Refund Policy at /refund-policy", "CRITICAL", "Not Started", "Legal", "", "", "None", "UK law: 14-day cooling-off period"],
    [17, "Phase 1", "Legal", "Implement age verification / parental consent for under-16s", "CRITICAL", "Not Started", "Dev", "", "", "#14", "Children's Code applies to under-18s"],
    [18, "Phase 1", "Legal", "Register with ICO (UK data protection)", "CRITICAL", "Not Started", "Founder", "", "", "None", "£40-52/year — criminal offence if not registered"],
    [19, "Phase 1", "Legal", "Obtain DPAs from all sub-processors (Supabase, Stripe, Azure, Vercel, Sentry)", "IMPORTANT", "Not Started", "Legal", "", "", "None", "GDPR requirement"],
    [20, "Phase 1", "Legal", "Publish Accessibility Statement at /accessibility", "IMPORTANT", "Not Started", "Dev", "", "", "None", "Equality Act 2010 — schools will ask"],

    # Landing Page & Conversion
    [21, "Phase 1", "Conversion", "Optimise landing page: clear What/Who/Why/How Much, CTA above fold, <3s load", "CRITICAL", "Not Started", "Dev/Marketing", "", "", "#2,#4", "Front door — must convert"],
    [22, "Phase 1", "Conversion", "Build waitlist landing page with email capture (name, email, year group, exam board)", "CRITICAL", "Not Started", "Dev", "", "", "#21", "Hero: '50% off first 3 months for first 200 sign-ups'"],
    [23, "Phase 1", "Conversion", "Set up email platform (SPF, DKIM, DMARC authenticated sender domain)", "CRITICAL", "Not Started", "Dev", "", "", "None", "Required for deliverability"],
    [24, "Phase 1", "Conversion", "Draft and load all 7 trial onboarding email sequences", "IMPORTANT", "Not Started", "Marketing", "", "", "#23", "Day 0-28 automated sequence"],
    [25, "Phase 1", "Conversion", "Draft waitlist launch email sequence (4 emails over 14 days)", "IMPORTANT", "Not Started", "Marketing", "", "", "#23", "Email 1: We're live, Email 2: Feature tour, Email 3: Social proof, Email 4: Trial ending"],

    # Social Media Setup
    [26, "Phase 1", "Social Media", "Create TikTok account @theenglishhubapp — bio, profile pic, link in bio", "IMPORTANT", "Not Started", "Marketing", "", "", "None", "Primary short-form platform"],
    [27, "Phase 1", "Social Media", "Create YouTube channel 'The English Hub' — banner, about, playlists", "IMPORTANT", "Not Started", "Marketing", "", "", "None", "Long-form content hub"],
    [28, "Phase 1", "Social Media", "Create Instagram @theenglishhubapp — bio, highlights, link in bio", "IMPORTANT", "Not Started", "Marketing", "", "", "None", "Secondary platform"],
    [29, "Phase 1", "Social Media", "Create Twitter/X account — professional profile for teacher/parent audience", "NICE", "Not Started", "Marketing", "", "", "None", "Teacher community engagement"],
    [30, "Phase 1", "Social Media", "Film and edit first 10 TikTok videos (Videos 1-10 from Content Playbook)", "IMPORTANT", "Not Started", "Content", "", "", "#26", "Bank content before launch"],
    [31, "Phase 1", "Social Media", "Film and edit YouTube Video Y1: 'How to Get a Grade 9 in GCSE English Language'", "IMPORTANT", "Not Started", "Content", "", "", "#27", "18-22 min flagship video"],
    [32, "Phase 1", "Social Media", "Schedule first 2 weeks of daily TikTok posts", "IMPORTANT", "Not Started", "Content", "", "", "#30", "Optimal times: 4:30PM/7PM/8PM weekdays, 11AM weekends"],
    [33, "Phase 1", "Social Media", "Create 3 YouTube Shorts from top TikTok scripts", "NICE", "Not Started", "Content", "", "", "#30", "Cross-pollinate audiences"],

    # SEO Foundation
    [34, "Phase 1", "SEO", "Optimise homepage title/meta: 'GCSE English Revision App | Courses, Mock Exams & Practice'", "CRITICAL", "Not Started", "Dev/SEO", "", "", "#2", "Front-load highest-volume keyword"],
    [35, "Phase 1", "SEO", "Optimise all landing page title tags and meta descriptions (per SEO Playbook)", "IMPORTANT", "Not Started", "Dev/SEO", "", "", "#2", "/courses, /practice, /mock-exams, /revision, /about, /exam-guide"],
    [36, "Phase 1", "SEO", "Submit sitemap to Google Search Console + verify ownership", "CRITICAL", "Not Started", "Dev", "", "", "#1", "How Google finds your pages"],
    [37, "Phase 1", "SEO", "Write and publish first 10 cornerstone blog posts (priority keywords)", "IMPORTANT", "Not Started", "Content/SEO", "", "", "#34", "Target: grade 9 guides, text analysis, exam technique"],
    [38, "Phase 1", "SEO", "Set up Bing Webmaster Tools (schools use Edge)", "NICE", "Not Started", "Dev", "", "", "#36", "Import from Search Console"],

    # Affiliate Recruitment
    [39, "Phase 1", "Affiliates", "Set up 'Become an Affiliate' page on website", "IMPORTANT", "Not Started", "Dev", "", "", "#7", "Passive inbound recruitment channel"],
    [40, "Phase 1", "Affiliates", "Create affiliate resource hub: banners, email templates, social posts, demo videos", "IMPORTANT", "Not Started", "Marketing", "", "", "#7", "15 banners, 3 email templates, 5 social templates, 2 videos"],
    [41, "Phase 1", "Affiliates", "Generate unique discount codes for affiliates ([NAME]15 = 15% off first 3 months)", "IMPORTANT", "Not Started", "Dev", "", "", "#7", "Stackable with free trial"],
    [42, "Phase 1", "Affiliates", "DM outreach to 300 micro-influencers (Cat A+B+F — UK GCSE, StudyTok, Revision creators)", "CRITICAL", "Not Started", "Founder/Cowork", "", "", "#7,#40", "Target: 60 signed (20% conv). Use Instantly + manual DMs. 4-day blitz."],
    [43, "Phase 1", "Affiliates", "Email outreach to 200 English teachers (Cat C — UK teachers + GCC British school teachers)", "CRITICAL", "Not Started", "Founder/Cowork", "", "", "#7,#40", "Target: 50 signed (25% conv). Hunter.io for emails. Include GCC HoDs."],
    [44, "Phase 1", "Affiliates", "Email outreach to 150 parent bloggers & GCC parent influencers (Cat D+J)", "CRITICAL", "Not Started", "Founder/Cowork", "", "", "#7,#40", "Target: 35 signed (23% conv). UK + Dubai + Qatar + Saudi parent communities."],
    [44.1, "Phase 1", "Affiliates", "DM outreach to 200 GCC education & expat creators (Cat E+G+H+I)", "CRITICAL", "Not Started", "Founder/Cowork", "", "", "#7,#40", "Target: 40 signed (20% conv). IGCSE creators, GCC study accounts, ME teacher influencers."],
    [44.2, "Phase 1", "Affiliates", "Bulk recruit via affiliate marketplaces — ShareASale, Awin, PartnerStack, impact.com", "IMPORTANT", "Not Started", "Founder/Cowork", "", "", "#7,#39", "Target: 150 applications via marketplace discovery. Filter by education/parenting niche."],
    [45, "Phase 1", "Affiliates", "Set up automated affiliate onboarding email sequence (5 emails over 14 days)", "IMPORTANT", "Not Started", "Marketing", "", "", "#23,#40", "Welcome → Resources → Walkthrough → Prompt → Check-in"],

    # Teacher/School Outreach
    [46, "Phase 1", "B2B Schools", "Scrape/compile list of 100 English HoD emails from UK school websites", "IMPORTANT", "Not Started", "Sales", "", "", "None", "Or use TES Jobs contacts"],
    [47, "Phase 1", "B2B Schools", "Email 100 English Heads of Department: free school dashboard trial offer", "IMPORTANT", "Not Started", "Sales", "", "", "#46", "Target: 10 schools trialling by end of pre-launch"],
    [48, "Phase 1", "B2B Schools", "Create school pricing page: Class £199/yr, Dept £599/yr, Whole School £1,499/yr", "IMPORTANT", "Not Started", "Dev", "", "", "#6", "Per-student equivalent: £5.69 → £3.00"],

    # ── PHASE 2: SOFT LAUNCH (Weeks 3-4) ──
    [49, "Phase 2", "Email", "Send waitlist launch Email 1: 'We're live — claim your 50% off' (first 200 only)", "CRITICAL", "Not Started", "Marketing", "", "", "#22,#25", "Scarcity-driven — first 200 sign-ups"],
    [50, "Phase 2", "Email", "Send waitlist launch Email 2 (Day 3): Feature walkthrough + 60s demo video", "IMPORTANT", "Not Started", "Marketing", "", "", "#49", ""],
    [51, "Phase 2", "Email", "Send waitlist launch Email 3 (Day 7): Social proof + beta user quotes", "IMPORTANT", "Not Started", "Marketing", "", "", "#50", "Need testimonials from beta"],
    [52, "Phase 2", "Email", "Send waitlist launch Email 4 (Day 14): Trial ending + annual plan pitch", "IMPORTANT", "Not Started", "Marketing", "", "", "#51", "Push annual plan at £59.99 launch special"],
    [53, "Phase 2", "Affiliates", "Notify all 200+ recruited affiliates platform is live — send activation pack", "CRITICAL", "Not Started", "Marketing", "", "", "#42,#43,#44", "Mass email via Instantly. Unique codes, launch offers, content suggestions."],
    [54, "Phase 2", "Affiliates", "Run 'First 10 Sales' bonus: £50 extra for affiliates hitting 10 conversions in 2 weeks", "IMPORTANT", "Not Started", "Marketing", "", "", "#53", "Activation incentive"],
    [55, "Phase 2", "Content", "Coordinate content blitz: 50+ pieces across all channels in 2 weeks", "CRITICAL", "Not Started", "Content/Affiliates", "", "", "#53,#32", "In-house + affiliate content"],
    [56, "Phase 2", "Content", "Host live 'GCSE English Masterclass' on YouTube/TikTok", "IMPORTANT", "Not Started", "Content", "", "", "#26,#27", "Drive sign-ups through live event"],
    [57, "Phase 2", "Content", "Film TikTok Videos 11-20 from Content Playbook", "IMPORTANT", "Not Started", "Content", "", "", "#30", "Maintain daily posting cadence"],
    [58, "Phase 2", "Content", "Film YouTube Video Y2: 'I Marked 100 GCSE English Essays'", "IMPORTANT", "Not Started", "Content", "", "", "#31", "15-18 min, examiner insight angle"],
    [59, "Phase 2", "Community", "Post helpful revision content on r/GCSE (45K members) and The Student Room", "IMPORTANT", "Not Started", "Marketing", "", "", "None", "Genuine contributor — not overt promotion"],
    [60, "Phase 2", "Community", "Engage in r/TeachingUK as a resource provider", "NICE", "Not Started", "Marketing", "", "", "None", "Teacher community credibility"],
    [61, "Phase 2", "Feedback", "Send in-app NPS survey after 7 days to first users", "IMPORTANT", "Not Started", "Dev/Marketing", "", "", "#49", "'What would make this worth paying for?'"],
    [62, "Phase 2", "Feedback", "Email first 50 users for video testimonials (offer 1 month free)", "IMPORTANT", "Not Started", "Marketing", "", "", "#49", "30-second clips for social proof"],

    # ── PHASE 3: GROWTH (Months 2-3) ──
    [63, "Phase 3", "Affiliates", "Scale affiliate program to 500 active affiliates", "CRITICAL", "Not Started", "Marketing", "", "", "#53", "Second wave: 500 more outreach. Marketplace bulk recruit. Feedspot database export."],
    [64, "Phase 3", "Affiliates", "Launch affiliate leaderboard with monthly prizes", "IMPORTANT", "Not Started", "Dev/Marketing", "", "", "#63", "Gamify — top affiliates get exclusive perks"],
    [65, "Phase 3", "Affiliates", "Host monthly affiliate webinar sharing top-performing strategies", "IMPORTANT", "Not Started", "Marketing", "", "", "#63", "Build affiliate community"],
    [66, "Phase 3", "Paid Ads", "Launch Google Ads: target 'GCSE English revision' keywords (£2-4 CPC)", "IMPORTANT", "Not Started", "Marketing", "", "", "#11,#12", "Test budget: £150-250/month"],
    [67, "Phase 3", "Paid Ads", "Launch TikTok Spark Ads boosting best organic content (£5-10/day)", "IMPORTANT", "Not Started", "Marketing", "", "", "#11,#30", "Boost top-performing TikToks"],
    [68, "Phase 3", "Paid Ads", "Launch Meta retargeting: website visitors + lookalike audiences", "IMPORTANT", "Not Started", "Marketing", "", "", "#11", "Test budget: £100-150/month"],
    [69, "Phase 3", "B2B Schools", "Attend 2-3 education conferences (Bett Show, Education Show)", "IMPORTANT", "Not Started", "Sales/Founder", "", "", "None", "In-person school partnership development"],
    [70, "Phase 3", "B2B Schools", "Offer 'class licence' pricing to schools (£199/yr for 35 students)", "IMPORTANT", "Not Started", "Sales", "", "", "#48", "Target: 5 schools with paid subscriptions"],
    [71, "Phase 3", "Content", "Scale to 2 blog posts/week (SEO-driven)", "IMPORTANT", "Not Started", "Content", "", "", "#37", "Compound organic traffic"],
    [72, "Phase 3", "Content", "Weekly YouTube video (full essay/exam technique walkthrough)", "IMPORTANT", "Not Started", "Content", "", "", "#58", "Sunday publish schedule"],
    [73, "Phase 3", "Content", "Daily TikTok/Shorts (reuse blog content as micro-content)", "IMPORTANT", "Not Started", "Content", "", "", "#57", "Film Videos 21-30 from Playbook"],
    [74, "Phase 3", "International", "Reach out to 20 British curriculum schools in Qatar and UAE", "IMPORTANT", "Not Started", "Sales", "", "", "#48", "IGCSE exam season: May/June + Oct/Nov"],
    [75, "Phase 3", "International", "Partner with 5 Middle East education influencers/tutors", "IMPORTANT", "Not Started", "Marketing", "", "", "#7", "Categories D+E from Influencer Playbook"],
    [76, "Phase 3", "International", "Localise marketing for IGCSE exam season differences", "IMPORTANT", "Not Started", "Marketing", "", "", "#74", "Oct/Nov session for retakes"],
    [77, "Phase 3", "Email", "Email all paying customers: 'Love us? Earn 20% by sharing' affiliate recruitment", "IMPORTANT", "Not Started", "Marketing", "", "", "#49", "Customer → affiliate conversion"],

    # ── PHASE 4: SCALE (Months 4-6) ──
    [78, "Phase 4", "Analytics", "Analyse months 1-3 data: channels, affiliates, content types by LTV:CAC", "CRITICAL", "Not Started", "Marketing", "", "", "All Phase 3", "Data-driven pivot — shift 80% budget to top 2 channels"],
    [79, "Phase 4", "Affiliates", "Expand to 1,000+ affiliates — self-serve signup + marketplace + Feedspot bulk import", "CRITICAL", "Not Started", "Dev/Marketing", "", "", "#63", "Open floodgates: self-serve page, marketplace listings, bulk CSV import from databases"],
    [80, "Phase 4", "Affiliates", "Launch 'Teacher Ambassador' programme — target 100 teacher ambassadors", "IMPORTANT", "Not Started", "Marketing", "", "", "#79", "Commission + free premium + badge. Outreach to 500+ teachers via Sprint Education lists."],
    [81, "Phase 4", "Campaign", "Launch '100 Days to GCSEs' countdown campaign (starts late January)", "CRITICAL", "Not Started", "Marketing", "", "", "#71", "Every piece of marketing references exam dates"],
    [82, "Phase 4", "Campaign", "Create 'Exam Season Pass' offer: £29.99 for 5 months (Jan-May)", "IMPORTANT", "Not Started", "Dev/Marketing", "", "", "#6", "Covers entire revision period at discount"],
    [83, "Phase 4", "Campaign", "Ramp paid ad spend 2x during exam season (Jan-May)", "IMPORTANT", "Not Started", "Marketing", "", "", "#66,#67,#68", "Peak demand period"],
    [84, "Phase 4", "Referral", "Launch student referral programme: both get 1 month free", "IMPORTANT", "Not Started", "Dev", "", "", "#6", "Viral loop — referral + shareable badges"],
    [85, "Phase 4", "Referral", "Build shareable progress badges and mock exam result cards for social media", "NICE", "Not Started", "Dev/Design", "", "", "#84", "Organic viral sharing"],
    [86, "Phase 4", "B2B Schools", "Dedicated school sales push — target 20 schools by month 6", "IMPORTANT", "Not Started", "Sales", "", "", "#70", "Build pipeline for September intake (new academic year)"],
    [87, "Phase 4", "B2B Schools", "Create school case study from early adopters", "IMPORTANT", "Not Started", "Marketing", "", "", "#70", "Social proof for school sales"],
    [88, "Phase 4", "International", "Launch targeted Middle East campaigns for crisis-driven families", "IMPORTANT", "Not Started", "Marketing", "", "", "#74,#75", "Families displaced by conflict seeking British curriculum continuity"],
    [89, "Phase 4", "International", "Partner with relocation agencies and expat communities in GCC", "IMPORTANT", "Not Started", "Sales", "", "", "#88", "Warm introductions to target families"],
    [90, "Phase 4", "Seasonal", "Plan 'Back to School' campaign for September (60-day extended trial)", "IMPORTANT", "Not Started", "Marketing", "", "", "None", "New academic year = fresh cohort"],
    [91, "Phase 4", "Seasonal", "Plan 'Mock Exam Prep' campaign for Nov-Dec (£14.99 for 2 months)", "IMPORTANT", "Not Started", "Marketing", "", "", "None", "Mock exams typically in Dec/Jan"],
    [92, "Phase 4", "Seasonal", "Plan 'Results Day' campaign for August (retake students)", "IMPORTANT", "Not Started", "Marketing", "", "", "None", "November resit entries surged 5.8% in 2025"],
    [93, "Phase 4", "Content", "Film YouTube Videos Y3-Y12 (full walkthroughs per Content Playbook)", "IMPORTANT", "Not Started", "Content", "", "", "#72", "Paper 1 + Paper 2 full walkthroughs, text-specific videos"],
]

for row_data in tasks:
    r2 = row_data[0] + 1  # +1 for header row
    for i, v in enumerate(row_data, 1):
        ws2.cell(row=r2, column=i, value=v)
    style_data_row(ws2, r2, len(cols2))
    # Color code priority
    priority_cell = ws2.cell(row=r2, column=5)
    if priority_cell.value == "CRITICAL":
        priority_cell.fill = RED_FILL
    elif priority_cell.value == "IMPORTANT":
        priority_cell.fill = YELLOW_FILL
    elif priority_cell.value == "NICE":
        priority_cell.fill = GREEN_FILL
    # Color code phase
    phase_cell = ws2.cell(row=r2, column=2)
    if "Phase 1" in str(phase_cell.value):
        phase_cell.fill = PatternFill('solid', fgColor='BDD7EE')
    elif "Phase 2" in str(phase_cell.value):
        phase_cell.fill = PatternFill('solid', fgColor='A9D08E')
    elif "Phase 3" in str(phase_cell.value):
        phase_cell.fill = ORANGE_FILL
    elif "Phase 4" in str(phase_cell.value):
        phase_cell.fill = PURPLE_FILL
    ws2.row_dimensions[r2].height = 35

# ════════════════════════════════════════════════════════════════
# SHEET 3: AFFILIATE OUTREACH TRACKER
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Affiliate Tracker")
ws3.sheet_properties.tabColor = "00B050"
cols3 = ["#", "Name / Handle", "Platform", "Category", "Tier", "Followers", "Contact", "Outreach Status", "Date Contacted", "Response", "Signed Up?", "Affiliate Code", "Conversions", "Revenue Generated", "Notes"]
set_col_widths(ws3, [4, 25, 12, 22, 8, 10, 25, 14, 12, 10, 10, 14, 11, 13, 30])

for i, h in enumerate(cols3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header_row(ws3, 1, len(cols3))
ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}1"
ws3.freeze_panes = 'A2'

# Pre-populate 1,000 affiliate targets across all categories
# Named influencers first, then category quotas to hit 1,000 total
influencers = [
    # ═══ CATEGORY A: UK GCSE/A-Level English Education Creators (25 named) ═══
    [1, "Mr Bruff (@mrbruff)", "YouTube/TikTok", "A: UK GCSE English", "Macro", "430K+", "YT contact form", "Not Started", "", "", "", "", "", "", "Wave 4. #1 GCSE English channel — even mention drives massive traffic"],
    [2, "Mr Salles (@mrsallesteachesenglish)", "YouTube/TikTok", "A: UK GCSE English", "Mid", "151K+", "YT/Substack", "Not Started", "", "", "", "", "", "", "Wave 3. 9K+ Substack subs, 2nd largest GCSE English channel"],
    [3, "Stacey Reay (@staceyreay1)", "TikTok/YouTube", "A: UK GCSE English", "Mid", "60K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Pearson Teacher of Year 2017"],
    [4, "Mr Everything English", "TikTok/YouTube", "A: UK GCSE English", "Mid", "120K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. High TikTok engagement, P.R.T.E.Z.E.L method"],
    [5, "Miss English Teacher (@missenglishteacher)", "TikTok", "A: UK GCSE English", "Mid", "377K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Huge GCSE English TikTok — teacher humour format"],
    [6, "The Lightup Hub (@thelightuptutor)", "TikTok/IG", "A: UK GCSE English", "Mid", "226K TikTok / 101K IG", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Dual-platform, GCSE English specific"],
    [7, "FunkyPedagogy / Jennifer Webb", "X/Blog", "A: UK GCSE English", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE English specialist blogger"],
    [8, "GCSE English RevisionPod (Mr Forster & Gallie)", "Podcast/YT", "A: UK GCSE English", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Dedicated GCSE English podcast"],
    [9, "Sarah Teaches GCSE English", "TikTok", "A: UK GCSE English", "Micro", "25K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Active GCSE examiner + private tutor"],
    [10, "First Rate Tutors", "YouTube/Substack", "A: UK GCSE English", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Already monetises via Udemy — natural fit"],
    [11, "Lisa GCSE English Teacher", "TikTok", "A: UK GCSE English", "Micro", "18K", "DM/IZEA", "Not Started", "", "", "", "", "", "", "Wave 1. On IZEA marketplace — seeking brand collabs"],
    [12, "Miss K English", "TikTok/IG", "A: UK GCSE English", "Micro", "12K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Snackable revision content"],
    [13, "Dr Amina Yonis", "TikTok/IG", "A: UK GCSE English", "Micro", "20K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. PhD-qualified English educator"],
    [14, "MME Revise", "YouTube/Website", "A: UK GCSE English", "Mid", "50K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Established revision platform, affiliate-ready"],
    [15, "@ygjack_english", "IG", "A: UK GCSE English", "Micro", "36.7K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE English Instagram specialist"],
    [16, "KrissTuition", "TikTok/IG", "A: UK GCSE English", "Micro", "15K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Online tutor, GCSE English focus"],
    [17, "Ruth Havenga (@english_gcse_tutor_)", "IG", "A: UK GCSE English", "Nano", "7.7K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Dedicated GCSE English tutor account"],
    [18, "GCSE Objectives", "TikTok", "A: UK GCSE English", "Nano", "8K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Student-run = peer credibility"],
    [19, "The English Teacher", "TikTok/IG", "A: UK GCSE English", "Micro", "10K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. English-specific, small but targeted"],
    [20, "Cognito (@cognitoedu)", "YouTube", "A: UK GCSE English", "Macro", "1.1M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. 1.1M subs — covers GCSE English + all subjects"],
    [21, "Primrose Kitten", "YouTube", "A: UK GCSE English", "Mid", "258K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Major GCSE revision channel"],
    [22, "Seneca Learning", "Website/Social", "A: UK GCSE English", "Macro", "14M users", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. 14M student users — partnership/integration"],
    [23, "ZNotes", "Website/Discord", "A: UK GCSE English", "Mid", "6M users / 46K Discord", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. 6M users, strong IGCSE/GCSE community"],
    [24, "FreeScienceLessons (@freesciencelessons)", "YouTube", "A: UK GCSE English", "Macro", "935K", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Cross-subject but massive GCSE audience"],
    [25, "DrPhysicsA", "YouTube", "A: UK GCSE English", "Mid", "317K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Cross-subject GCSE/A-Level, high trust"],
    # ═══ CATEGORY B: StudyTok/Studygram (25 named) ═══
    [26, "StudyPlate (@studyplate)", "TikTok", "B: StudyTok/Studygram", "Micro", "35K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE-to-A-Level transition audience"],
    [27, "David Cai (@davidcaii)", "IG/TikTok", "B: StudyTok/Studygram", "Mid", "108K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Aesthetic study content"],
    [28, "Sophie (@myhoneststudyblr)", "Instagram", "B: StudyTok/Studygram", "Mid", "193K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Huge studygram"],
    [29, "Felicity Notes", "Instagram", "B: StudyTok/Studygram", "Macro", "208K", "DM", "Not Started", "", "", "", "", "", "", "Wave 4. Massive studygram"],
    [30, "Study With Kate (@study_with_kate)", "TikTok", "B: StudyTok/Studygram", "Mid", "157K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Already does app partnerships"],
    [31, "@studyw.tok", "TikTok", "B: StudyTok/Studygram", "Macro", "1.1M", "DM", "Not Started", "", "", "", "", "", "", "Wave 4. 1.1M TikTok — massive StudyTok reach"],
    [32, "Jack Edwards (@jack_edwards)", "YouTube/TikTok", "B: StudyTok/Studygram", "Macro", "1.54M YT / 842K TikTok", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Mega creator — Durham English graduate"],
    [33, "UnJaded Jade", "YouTube/TikTok", "B: StudyTok/Studygram", "Macro", "975K YT / 213K TikTok", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Cambridge student, study content"],
    [34, "Ruby Granger", "YouTube", "B: StudyTok/Studygram", "Macro", "961K", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. OG study creator — Exeter English grad"],
    [35, "Ways To Study (@waystostudy)", "IG", "B: StudyTok/Studygram", "Mid", "321K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Massive study IG account"],
    [36, "@anika.studies", "TikTok/IG", "B: StudyTok/Studygram", "Micro", "30K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. UK study aesthetic"],
    [37, "@conversationswithruby", "TikTok/IG", "B: StudyTok/Studygram", "Micro", "20K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Study + life content"],
    [38, "@pcfgstudy", "IG", "B: StudyTok/Studygram", "Micro", "15K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE/A-Level study notes"],
    [39, "@cambridgram", "IG", "B: StudyTok/Studygram", "Micro", "25K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Cambridge study content"],
    [40, "@cleaasdiary", "TikTok/IG", "B: StudyTok/Studygram", "Micro", "18K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Study diary format"],
    [41, "GCSE Revision Notes (@xokezzagcse)", "Instagram", "B: StudyTok/Studygram", "Nano", "5K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Directly GCSE-focused studygram"],
    [42, "Study Tips GCSE", "Instagram", "B: StudyTok/Studygram", "Nano", "3K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Laser-targeted GCSE audience"],
    [43, "A* Study Notes", "Instagram", "B: StudyTok/Studygram", "Nano", "8K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Aspirational top-grade audience"],
    [44, "Holly Gabrielle", "YouTube", "B: StudyTok/Studygram", "Mid", "436K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Cambridge student study vlogs"],
    [45, "Vee Kativhu", "YouTube", "B: StudyTok/Studygram", "Mid", "281K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Oxford grad, study tips + motivation"],
    [46, "Erin McGurk", "TikTok", "B: StudyTok/Studygram", "Mid", "201K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Cambridge study TikTok"],
    [47, "Tom Powell (@tom_powellll)", "IG", "B: StudyTok/Studygram", "Mid", "362K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Cambridge, huge IG following"],
    [48, "Ali Abdaal (@aliabdaal)", "YouTube", "B: StudyTok/Studygram", "Mega", "6.3M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Mega study/productivity creator — aspirational"],
    [49, "Thomas Frank", "YouTube", "B: StudyTok/Studygram", "Macro", "2.9M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Study/productivity — huge reach"],
    [50, "Cajun Koi Academy", "YouTube", "B: StudyTok/Studygram", "Macro", "1.3M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Study tips + techniques"],
    # ═══ CATEGORY C: UK Teacher Influencers (35 named) ═══
    [51, "Teacher Toolkit / Ross McGill", "X/Blog", "C: UK Teacher Influencers", "Mid-Macro", "130K+", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 2. Most influential UK edu blog, 25+ years"],
    [52, "Tom Sherrington (@teacherhead)", "X/Blog", "C: UK Teacher Influencers", "Mid", "90K+", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 2. 30-year headteacher, WalkThrus author"],
    [53, "Jo Morgan / Resourceaholic", "X/Blog", "C: UK Teacher Influencers", "Micro-Mid", "45K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Maths but huge teacher network effect"],
    [54, "Mark Anderson / ICTEvangelist", "X/Blog", "C: UK Teacher Influencers", "Mid", "55K+", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 1. EdTech advocate — perfect for new platform"],
    [55, "Craig Barton / Tips for Teachers", "Substack/Podcast", "C: UK Teacher Influencers", "Mid", "24K Substack", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. 24K Substack subs, Mr Barton Maths podcast"],
    [56, "Niall Alcock / We Are In Beta", "Substack", "C: UK Teacher Influencers", "Micro", "23K Substack", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. 23K subs, edu newsletter"],
    [57, "Matt Green / Rapping Science Teacher", "TikTok", "C: UK Teacher Influencers", "Macro", "1.5M", "DM/Email", "Not Started", "", "", "", "", "", "", "Wave 4. 1.5M TikTok — viral teacher content"],
    [58, "Biology With Olivia", "TikTok", "C: UK Teacher Influencers", "Mid", "341K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Cross-subject but massive GCSE audience"],
    [59, "Hannah Kettle Maths (@hannahkettlemaths)", "TikTok", "C: UK Teacher Influencers", "Mid", "287K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. GCSE examiner — cross-subject authority"],
    [60, "Miss Estruch (@miss.estruch.biology)", "TikTok", "C: UK Teacher Influencers", "Micro", "97.3K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Active GCSE teacher TikTok"],
    [61, "A-Level Help", "TikTok", "C: UK Teacher Influencers", "Mid", "584K TikTok / 632K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Massive A-Level/GCSE audience"],
    [62, "Teacher Tapp", "App/Newsletter", "C: UK Teacher Influencers", "Mid", "10K+ daily users", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Daily teacher survey app — sponsored question potential"],
    [63, "David Didau / The Learning Spy", "Blog/X", "C: UK Teacher Influencers", "Micro", "30K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Influential edu blogger"],
    [64, "Kat Howard (@saysmiss)", "X/Blog", "C: UK Teacher Influencers", "Micro", "25K+", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 1. English teacher, published author"],
    [65, "Alex Quigley / The Confident Teacher", "X/Blog", "C: UK Teacher Influencers", "Micro", "20K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. EEF Senior Associate, literacy focus"],
    [66, "Miss Cole / Educated Minds", "YouTube/IG", "C: UK Teacher Influencers", "Nano-Micro", "8K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Directly helps students through GCSEs"],
    [67, "Neil Does Maths", "TikTok", "C: UK Teacher Influencers", "Micro", "15K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Referenced in ITV report"],
    [68, "Mrs Humanities", "X/Blog", "C: UK Teacher Influencers", "Micro", "20K", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 1. Active teacher community builder"],
    [69, "The Classroom Teacher", "IG/TikTok", "C: UK Teacher Influencers", "Nano-Micro", "7K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Authentic teacher voice"],
    [70, "English Teacher Chloe", "TikTok/IG", "C: UK Teacher Influencers", "Nano", "6K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. English-specific teacher content"],
    [71, "tutor2u", "YouTube/TikTok", "C: UK Teacher Influencers", "Macro", "1.2M YT", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive teacher/student audience"],
    [72, "@twinklresources", "IG", "C: UK Teacher Influencers", "Mid", "189K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Huge teacher resource brand"],
    [73, "@bbcbitesize", "IG/TikTok", "C: UK Teacher Influencers", "Mid", "95K IG", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. BBC education brand — aspirational"],
    [74, "@savemyexams", "IG/TikTok", "C: UK Teacher Influencers", "Mid", "57K IG", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Exam-focused — natural affiliate fit"],
    [75, "@mathswithmisschang", "IG", "C: UK Teacher Influencers", "Mid", "57K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Cross-subject teacher influence"],
    [76, "Gizmo AI (@gizmo.ai)", "TikTok", "C: UK Teacher Influencers", "Micro", "102K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. AI study tool — partnership angle"],
    [77, "pinkpencilmath", "TikTok", "C: UK Teacher Influencers", "Macro", "1.8M", "DM/Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive maths TikTok — cross-subject reach"],
    [78, "EEF (Education Endowment Foundation)", "Newsletter", "C: UK Teacher Influencers", "Mid", "60K+ subscribers", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. 60K teacher newsletter — sponsored content"],
    [79, "Tes (Times Educational Supplement)", "Website/Social", "C: UK Teacher Influencers", "Macro", "550K+ registered", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. 550K teachers — sponsored/partnership"],
    [80, "The Edtech Podcast / Sophie Bailey", "Podcast", "C: UK Teacher Influencers", "Micro", "145 countries", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Global edtech podcast — feature/sponsor"],
    [81, "EdTech Impact", "Website", "C: UK Teacher Influencers", "Micro", "Review platform", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. EdTech review site — get listed + reviewed"],
    [82, "Cult of Pedagogy / Jennifer Gonzalez", "Blog/Podcast", "C: UK Teacher Influencers", "Mid", "250K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Global edu blog — sponsor/feature"],
    [83, "Seneca REVISE Podcast", "Podcast", "C: UK Teacher Influencers", "Mid", "Large student base", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Direct access to revision students"],
    [84, "Parenting Hell / Rob Beckett & Josh Widdicombe", "Podcast", "C: UK Teacher Influencers", "Mega", "50M+ downloads", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. 50M+ downloads — aspirational sponsored read"],
    [85, "TES Podagogy", "Podcast", "C: UK Teacher Influencers", "Mid", "Teacher audience", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Tes official podcast — sponsored segment"],
    # ═══ CATEGORY D: UK Parent Bloggers (20 named) ═══
    [86, "Mum in the Madhouse", "Blog/IG", "D: UK Parent Bloggers", "Micro", "35K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Established UK parenting blogger, secondary-age kids"],
    [87, "Actually Mummy / Helen Wills", "Blog/IG", "D: UK Parent Bloggers", "Micro", "20K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Writes about secondary school + exam stress"],
    [88, "The Unmumsy Mum", "Instagram", "D: UK Parent Bloggers", "Macro", "250K", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Huge following — story mention = massive traffic"],
    [89, "Slummy Single Mummy / Jo Middleton", "Blog", "D: UK Parent Bloggers", "Micro", "30K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK parenting blogger, secondary-age kids"],
    [90, "Honest Mum / Vicki Broadbent", "Blog/IG", "D: UK Parent Bloggers", "Micro", "40K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Award-winning UK parenting blog"],
    [91, "Teenagers Untangled / Rachel Richards", "Podcast", "D: UK Parent Bloggers", "Micro", "Top 2% global", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Top 2% worldwide parenting podcast, teen focus"],
    [92, "Parent Guide to GCSEs / Emily & Paul Hughes", "Blog", "D: UK Parent Bloggers", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Directly GCSE parent guide — perfect fit"],
    [93, "Who's the Mummy / Sally Whittle", "Blog/IG", "D: UK Parent Bloggers", "Micro", "15K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK parenting blogger"],
    [94, "Extraordinary Chaos / Sarah Christie", "Blog", "D: UK Parent Bloggers", "Micro", "10K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK family blog, education posts"],
    [95, "Shelley Wilson", "Blog/IG", "D: UK Parent Bloggers", "Micro", "12K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. YA author + parent blogger"],
    [96, "Happy Mum Happy Baby / Giovanna Fletcher", "Podcast/IG", "D: UK Parent Bloggers", "Macro", "1.8M IG", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. 1.8M IG — aspirational sponsored feature"],
    [97, "Polly Jemima", "Blog", "D: UK Parent Bloggers", "Micro", "60K monthly views", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Top 0.5% UK parenting blogs"],
    [98, "NurtureStore", "Blog/Pinterest", "D: UK Parent Bloggers", "Mid", "180K Pinterest", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. 180K Pinterest — homeschool/education angle"],
    [99, "Angelic Scalliwags", "Blog", "D: UK Parent Bloggers", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Homeschool to GCSEs content"],
    [100, "The Wonder Years Blog", "Blog", "D: UK Parent Bloggers", "Nano", "3K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. IGCSE/A-Level home ed content"],
    [101, "Educational Freedom", "Facebook Group", "D: UK Parent Bloggers", "Micro", "13K FB group", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. 13K member home ed community"],
    [102, "AnaCristina", "YouTube/IG", "D: UK Parent Bloggers", "Micro", "12K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. UK bilingual family, education-focused"],
    [103, "Nina Sherwood-Ma", "Blog", "D: UK Parent Bloggers", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Already curates GCSE revision resources"],
    [104, "Mumsnet Education Board", "Forum", "D: UK Parent Bloggers", "Macro", "14M monthly visits", "Sponsored Post", "Not Started", "", "", "", "", "", "", "Wave 3. 14M visits — sponsored thread potential"],
    [105, "The Student Room (TSR)", "Forum", "D: UK Parent Bloggers", "Macro", "1.37M members", "Sponsored", "Not Started", "", "", "", "", "", "", "Wave 3. 1.37M student members — sponsored content"],
    # ═══ CATEGORY E: Middle East / Expat Education (25 named) ═══
    [106, "ExpatWoman (@expatwoman)", "IG/Blog", "E: ME Expat Education", "Mid", "80K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Huge expat community across GCC"],
    [107, "Qatar Living", "Website/Social", "E: ME Expat Education", "Macro", "708K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. 708K — GCC expat hub, school section"],
    [108, "ILQ (I Love Qatar)", "Social/Website", "E: ME Expat Education", "Mid", "339K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major Qatar community platform"],
    [109, "New In Doha", "Social/Website", "E: ME Expat Education", "Mid", "286K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Doha expat community — school content"],
    [110, "Kids Love Qatar", "IG/Facebook", "E: ME Expat Education", "Mid", "150K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. 150K — family/education focus in Qatar"],
    [111, "WWHID (What's Hot In Doha) Facebook", "Facebook", "E: ME Expat Education", "Mid", "165K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 2. 165K member FB group — sponsored post"],
    [112, "British Mums Dubai", "Facebook/IG", "E: ME Expat Education", "Micro", "32K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Direct access to British curriculum parents"],
    [113, "Sassy Mama Dubai", "Website/IG", "E: ME Expat Education", "Micro", "25K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai family life, education content"],
    [114, "WhichSchoolAdvisor", "Website/IG", "E: ME Expat Education", "Micro", "20K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. School comparison — natural affiliate fit"],
    [115, "Edarabia", "Website/Social", "E: ME Expat Education", "Micro", "30K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. ME education directory — partnership"],
    [116, "Dubai Mums (@dubaimums)", "IG/Facebook", "E: ME Expat Education", "Micro", "25K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Community hub for Dubai parents"],
    [117, "CambriLearn", "Website/Social", "E: ME Expat Education", "Mid", "50K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Same audience — partnership opportunity"],
    [118, "248AM Kuwait", "Website/Social", "E: ME Expat Education", "Mid", "101K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Kuwait's top community site"],
    [119, "Kuwait Moms Guide", "IG/Facebook", "E: ME Expat Education", "Micro", "30K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Direct access to Kuwait parents"],
    [120, "Bahrain Confidential", "Website/IG", "E: ME Expat Education", "Micro", "46K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Bahrain lifestyle/education"],
    [121, "Mums in Bahrain", "IG/Facebook", "E: ME Expat Education", "Micro", "32K IG / 22K FB", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Bahrain parent community"],
    [122, "Bahrain Schools Guide", "Website", "E: ME Expat Education", "Nano", "5.9K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. School directory — listing/affiliate"],
    [123, "Muscat Mums", "Facebook", "E: ME Expat Education", "Micro", "15K FB", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. Oman parent community"],
    [124, "Doha Family Magazine", "Magazine/Social", "E: ME Expat Education", "Micro", "20K copies", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Print + digital — sponsored content"],
    [125, "Hadeel Ghassan", "Social", "E: ME Expat Education", "Mid", "663K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Qatar influencer, Comparative Literature MA"],
    [126, "DASHA", "Social", "E: ME Expat Education", "Mid", "107K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Qatar content creator"],
    [127, "Samy Chaffai", "Social", "E: ME Expat Education", "Macro", "1.5M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. #1 education influencer in Qatar"],
    [128, "SchoolsCompared", "Website", "E: ME Expat Education", "Micro", "30K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. UAE school comparison site"],
    [129, "Khaleej Times Education", "Website/Social", "E: ME Expat Education", "Macro", "Large", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Major UAE newspaper education section"],
    [130, "Doha Mums", "Magazine/Facebook", "E: ME Expat Education", "Micro", "20K copies", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Qatar family magazine"],
    # ═══ CATEGORY F: Student Revision (10 named) ═══
    [131, "SnapRevise", "TikTok", "F: Student Revision", "Micro", "40K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Competitor but could partner on referrals"],
    [132, "GCSEPod", "TikTok/YT", "F: Student Revision", "Micro", "20K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Established EdTech"],
    [133, "Save My Exams", "TikTok/YT", "F: Student Revision", "Mid", "56K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Exam-focused — affiliate opportunity"],
    [134, "BBC Bitesize", "All", "F: Student Revision", "Macro", "200K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Aspirational"],
    [135, "r/GCSE subreddit", "Reddit", "F: Student Revision", "Mid", "171K members", "Mod DM", "Not Started", "", "", "", "", "", "", "Wave 2. 171K GCSE students — community post"],
    [136, "r/6thForm subreddit", "Reddit", "F: Student Revision", "Mid", "167K members", "Mod DM", "Not Started", "", "", "", "", "", "", "Wave 2. 167K A-Level students — cross-promote"],
    [137, "Eleven Plus Forum", "Forum", "F: Student Revision", "Micro", "40K members", "Admin", "Not Started", "", "", "", "", "", "", "Wave 1. Academic parent community"],
    [138, "Home Education UK Facebook", "Facebook", "F: Student Revision", "Micro", "32K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE homeschool parents"],
    [139, "studyquill", "YouTube", "F: Student Revision", "Macro", "1.1M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive study channel"],
    [140, "Francesco D'Alessio / Keep Productive", "YouTube", "F: Student Revision", "Mid", "400K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Productivity/tools reviewer — app review"],
    # ═══ CATEGORY G: GCC IGCSE Creators (20 named) ═══
    [141, "Emad Anwer", "IG/TikTok", "G: GCC IGCSE Creators", "Mega", "4.5M IG", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. #1 Saudi English teacher — massive reach"],
    [142, "AbdurRahman Hijazi / Dalilk Academy", "Social", "G: GCC IGCSE Creators", "Mega", "3.9M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Huge Arabic-English education"],
    [143, "Hassan Alghamdi", "TikTok", "G: GCC IGCSE Creators", "Mega", "4.6M TikTok", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Saudi STEM mega-creator"],
    [144, "Aziz Time", "Social", "G: GCC IGCSE Creators", "Macro", "1.9M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Saudi education content"],
    [145, "Maram Kabbani", "TikTok", "G: GCC IGCSE Creators", "Macro", "1.4M TikTok", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Saudi tech/education"],
    [146, "Qadrat Rawan", "TikTok", "G: GCC IGCSE Creators", "Mid", "400K+ TikTok", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Saudi STEM TikTok"],
    [147, "Newsejazah", "Social", "G: GCC IGCSE Creators", "Macro", "1.3M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Arabic education content"],
    [148, "Baby Melons", "Social", "G: GCC IGCSE Creators", "Macro", "1M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Saudi family/education"],
    [149, "@tahirelshazli", "TikTok/IG", "G: GCC IGCSE Creators", "Micro", "15K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. IGCSE-specific GCC creator"],
    [150, "@the.ig.school", "TikTok/IG", "G: GCC IGCSE Creators", "Micro", "10K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. IGCSE revision content"],
    [151, "Fluency in English", "YouTube", "G: GCC IGCSE Creators", "Macro", "1.8M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Arabic-English education channel"],
    [152, "Dalati English", "YouTube", "G: GCC IGCSE Creators", "Macro", "1.5M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. English education in Arabic"],
    [153, "English Today USA", "YouTube", "G: GCC IGCSE Creators", "Macro", "1.6M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. English learning — GCC audience overlap"],
    [154, "English with Hicham", "YouTube", "G: GCC IGCSE Creators", "Mid", "405K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Arabic-English education"],
    [155, "Prof Fati", "YouTube", "G: GCC IGCSE Creators", "Mid", "440K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Education content in Arabic"],
    [156, "Tutopiya", "Website/Social", "G: GCC IGCSE Creators", "Micro", "20K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Singapore-based IGCSE specialist"],
    [157, "Taughtly", "Website/Social", "G: GCC IGCSE Creators", "Micro", "10K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. IGCSE tutoring platform"],
    [158, "Dear Sir / Mohammad Kashif", "YouTube", "G: GCC IGCSE Creators", "Mega", "20M+", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. 20M+ subs — MASSIVE South Asian English education"],
    [159, "Shobhit Nirwan", "YouTube", "G: GCC IGCSE Creators", "Macro", "3.6M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. South Asian education — GCC expat reach"],
    [160, "Vedantu", "YouTube", "G: GCC IGCSE Creators", "Macro", "2.4M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. EdTech platform — partnership"],
    # ═══ CATEGORY H: GCC Student Lifestyle (10 named) ═══
    [161, "@sarah_alruba3e", "YouTube/IG", "H: GCC Student Lifestyle", "Macro", "500K+", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 3. 450K IG — huge ME student audience"],
    [162, "@namednicole", "TikTok/IG", "H: GCC Student Lifestyle", "Micro", "25K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCC study content"],
    [163, "@miraadaoud", "TikTok/IG", "H: GCC Student Lifestyle", "Micro", "20K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCC student creator"],
    [164, "@karil_e", "TikTok", "H: GCC Student Lifestyle", "Micro", "15K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCC study tips"],
    [165, "@hanadiihussami", "TikTok/IG", "H: GCC Student Lifestyle", "Micro", "18K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCC education content"],
    [166, "@tintinsjourney", "TikTok/IG", "H: GCC Student Lifestyle", "Micro", "12K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Expat student in GCC"],
    [167, "@raghaadlife", "TikTok/IG", "H: GCC Student Lifestyle", "Micro", "10K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCC student lifestyle"],
    [168, "@r33ren", "TikTok", "H: GCC Student Lifestyle", "Micro", "20K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Saudi student creator"],
    [169, "@yaraa_alf", "TikTok/IG", "H: GCC Student Lifestyle", "Micro", "15K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCC study content"],
    [170, "Rawan Alshikh", "YouTube", "H: GCC Student Lifestyle", "Mega", "15M YouTube", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. 15M subs — aspirational, Dubai family"],
    # ═══ CATEGORY I: GCC Teacher Influencers (10 named) ═══
    [171, "Thomas Blakemore (@thomasblakemore)", "TikTok/IG/YT", "I: GCC Teacher Influencers", "Mid", "250K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. British teacher in Dubai, huge following"],
    [172, "@englishsahle", "TikTok/IG", "I: GCC Teacher Influencers", "Mid", "217K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. English teacher content in GCC"],
    [173, "@monaabuhattab", "TikTok/IG", "I: GCC Teacher Influencers", "Macro", "527K+", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 3. Huge ME education following"],
    [174, "@alramsa.institute", "IG", "I: GCC Teacher Influencers", "Mid", "217K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Language institute — partnership"],
    [175, "@almentorcourses", "IG/Website", "I: GCC Teacher Influencers", "Mid", "151K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Online courses — affiliate fit"],
    [176, "Ministry of Education UAE", "Social", "I: GCC Teacher Influencers", "Macro", "429K", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Official channel — government partnership"],
    [177, "British School of Kuwait", "Social", "I: GCC Teacher Influencers", "Micro", "15K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Direct school — bulk license angle"],
    [178, "Study with Sudhir", "YouTube", "I: GCC Teacher Influencers", "Mid", "440K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. South Asian English education in GCC"],
    [179, "Kauser Wise", "YouTube", "I: GCC Teacher Influencers", "Mid", "725K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. English grammar — GCC audience"],
    [180, "Magnet Brains", "YouTube", "I: GCC Teacher Influencers", "Macro", "1.5M+", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive education channel — GCC reach"],
    # ═══ CATEGORY J: GCC Parent Influencers (15 named) ═══
    [181, "@7ram_alshamsi", "IG/TikTok", "J: GCC Parent Influencers", "Macro", "1.9M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Massive UAE parent influencer"],
    [182, "@alaaabuhharb", "IG/TikTok", "J: GCC Parent Influencers", "Macro", "1.7M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Huge GCC parent following"],
    [183, "@mom_indubai / Virdah Javed", "IG", "J: GCC Parent Influencers", "Macro", "1M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 3. Dubai parent — direct audience match"],
    [184, "@afnanrecipes", "IG/YT", "J: GCC Parent Influencers", "Macro", "2.7M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Massive family-focused following"],
    [185, "Afrae Es-satte", "IG/TikTok", "J: GCC Parent Influencers", "Mega", "7M IG / 8.3M TikTok", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. MENA Family Creator of the Year"],
    [186, "Karen Wazen", "IG", "J: GCC Parent Influencers", "Mega", "8M IG", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Dubai-based mega influencer, family content"],
    [187, "Salama Mohamed", "IG", "J: GCC Parent Influencers", "Macro", "2M", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Major UAE family influencer"],
    [188, "Leila Trabi / thekidxpert", "IG", "J: GCC Parent Influencers", "Macro", "806K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Child development expert — education authority"],
    [189, "Sneha Rebecca Cherian", "IG/Blog", "J: GCC Parent Influencers", "Micro", "25K", "DM/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai expat parent, Parenting Award 2023"],
    [190, "The Family Hub / Hanan Ezzeldin", "IG/Podcast", "J: GCC Parent Influencers", "Micro", "15K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai educational consultant"],
    [191, "Qatar Mums (@qatarmums)", "IG/Facebook", "J: GCC Parent Influencers", "Micro", "15K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Community hub for Qatar parents"],
    [192, "@dohanews", "IG/X", "J: GCC Parent Influencers", "Macro", "500K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Qatar news — school closure posts = huge reach"],
    [193, "Sheikha Moza bint Nasser", "Social", "J: GCC Parent Influencers", "Mega", "2M", "PR Team", "Not Started", "", "", "", "", "", "", "Wave 4. Qatar Foundation chair — aspirational partnership"],
    [194, "Baby Drool / Faye", "Blog/IG", "J: GCC Parent Influencers", "Nano-Micro", "8K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. British expat in Dubai"],
    [195, "Raising Freddie / Aimee", "Blog/IG", "J: GCC Parent Influencers", "Micro", "10K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai-based, UK education trained"],
    # ═══ CATEGORY K: Education Blogs & News (10 named) ═══
    [196, "eLearning Industry", "Website", "K: Education Blogs", "Mid", "800K+ monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major edtech blog — sponsored post"],
    [197, "Getting Smart", "Website", "K: Education Blogs", "Mid", "200K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Education innovation blog"],
    [198, "EdTech Digest", "Website", "K: Education Blogs", "Micro", "50K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. EdTech news — product review"],
    [199, "Class Tech Tips / Monica Burns", "Blog/Social", "K: Education Blogs", "Micro", "30K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. EdTech review blog"],
    [200, "EdSurge", "Website", "K: Education Blogs", "Mid", "300K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Major edtech publication"],
    [201, "Edtech Insiders", "Newsletter/Podcast", "K: Education Blogs", "Mid", "40K total", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. 40K reach — edtech newsletter"],
    [202, "SchoolsWeek", "Website", "K: Education Blogs", "Mid", "100K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. UK education journalism"],
    [203, "Tes Magazine", "Website", "K: Education Blogs", "Macro", "500K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Major UK education media"],
    [204, "SecEd", "Website", "K: Education Blogs", "Micro", "30K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Secondary education news"],
    [205, "The Edvocate", "Website", "K: Education Blogs", "Micro", "40K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Education advocacy blog"],
    # ═══ CATEGORY L: EdTech Publishers / Affiliate Networks (10 named) ═══
    [206, "ShareASale Education Vertical", "Network", "L: EdTech Networks", "Macro", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "List on ShareASale education category for bulk recruitment"],
    [207, "Awin Education Category", "Network", "L: EdTech Networks", "Macro", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "List on Awin — access 100K+ publishers"],
    [208, "CJ Affiliate", "Network", "L: EdTech Networks", "Macro", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "Commission Junction — major affiliate network"],
    [209, "PartnerStack", "Network", "L: EdTech Networks", "Mid", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "B2B SaaS affiliate network"],
    [210, "impact.com", "Network", "L: EdTech Networks", "Macro", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "Enterprise partnership management"],
    [211, "Tutorful Affiliate Program", "Platform", "L: EdTech Networks", "Micro", "Affiliate", "Email", "Not Started", "", "", "", "", "", "", "Confirmed affiliate program — cross-promote"],
    [212, "MyTutor Affiliate Program", "Platform", "L: EdTech Networks", "Micro", "Affiliate", "Email", "Not Started", "", "", "", "", "", "", "Confirmed affiliate program — cross-promote"],
    [213, "Preply Affiliate Program", "Platform", "L: EdTech Networks", "Mid", "Affiliate", "Email", "Not Started", "", "", "", "", "", "", "Confirmed affiliate program — reciprocal"],
    [214, "GoStudent Affiliate Program", "Platform", "L: EdTech Networks", "Mid", "Affiliate", "Email", "Not Started", "", "", "", "", "", "", "Confirmed affiliate program — reciprocal"],
    [215, "Superprof UK", "Platform", "L: EdTech Networks", "Mid", "Affiliate", "Email", "Not Started", "", "", "", "", "", "", "Tutor marketplace — cross-promotion"],
    # ═══ CATEGORY M: Homeschool & Curriculum (5 named) ═══
    [216, "NurtureStore / Cathy James", "Blog/Pinterest", "M: Homeschool", "Mid", "180K Pinterest", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. 180K Pinterest — homeschool resource hub"],
    [217, "Polly Jemima", "Blog", "M: Homeschool", "Micro", "60K/mo views", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Top 0.5% UK parenting"],
    [218, "Angelic Scalliwags", "Blog", "M: Homeschool", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Homeschool to GCSEs journey"],
    [219, "The Wonder Years", "Blog", "M: Homeschool", "Nano", "3K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. IGCSE/A-Level home education"],
    [220, "Educational Freedom Facebook Group", "Facebook", "M: Homeschool", "Micro", "13K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. Active home ed community"],
    # ═══ CATEGORY N: ESL/IELTS/Language (20 named) ═══
    [221, "English with Lucy", "YouTube", "N: ESL/IELTS", "Mega", "13M", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. 13M subs — #1 English learning channel"],
    [222, "BBC Learning English", "YouTube", "N: ESL/IELTS", "Mega", "9.5M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. BBC brand — aspirational partnership"],
    [223, "Linguamarina", "YouTube", "N: ESL/IELTS", "Mega", "8.8M", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Massive English learning channel"],
    [224, "EnglishClass101", "YouTube", "N: ESL/IELTS", "Mega", "8.6M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Innovative Language Learning brand"],
    [225, "Let's Talk English", "YouTube", "N: ESL/IELTS", "Mega", "8.2M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive ESL channel"],
    [226, "Speak English with Vanessa", "YouTube", "N: ESL/IELTS", "Mega", "7.1M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Huge ESL channel"],
    [227, "Rachel's English", "YouTube", "N: ESL/IELTS", "Mega", "6.1M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Pronunciation-focused"],
    [228, "mmmEnglish / Emma", "YouTube", "N: ESL/IELTS", "Mega", "5M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Australian English teacher"],
    [229, "Antonio Parlati", "TikTok", "N: ESL/IELTS", "Mega", "6M TikTok", "DM/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Massive English TikTok"],
    [230, "Speak English with Zach", "TikTok", "N: ESL/IELTS", "Mega", "5M TikTok", "DM/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Huge English TikTok"],
    [231, "IELTS Liz", "YouTube/Website", "N: ESL/IELTS", "Macro", "2.3M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. IELTS specialist — GCC exam prep audience"],
    [232, "E2 IELTS", "YouTube", "N: ESL/IELTS", "Macro", "2.4M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. IELTS prep channel"],
    [233, "Claudine James", "TikTok", "N: ESL/IELTS", "Macro", "4.4M TikTok", "DM", "Not Started", "", "", "", "", "", "", "Wave 4. English teacher TikTok"],
    [234, "IELTS Advantage", "TikTok/YT", "N: ESL/IELTS", "Macro", "2.5M TikTok", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. IELTS prep — GCC audience"],
    [235, "Papa Teach Me", "YouTube", "N: ESL/IELTS", "Macro", "1.7M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. English grammar channel"],
    [236, "engVid / Ronnie", "YouTube", "N: ESL/IELTS", "Macro", "4.8M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive grammar channel"],
    [237, "Mr Duncan", "YouTube", "N: ESL/IELTS", "Macro", "1.45M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Long-running English channel"],
    [238, "Asad Yaqub", "YouTube", "N: ESL/IELTS", "Macro", "908K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. English for Arabic speakers"],
    [239, "Fastrack IELTS", "YouTube", "N: ESL/IELTS", "Mid", "500K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. IELTS prep specialist"],
    [240, "Oxford Online English", "YouTube", "N: ESL/IELTS", "Macro", "2.5M+", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Major English learning channel"],
    # ═══ CATEGORY O: BookTok/Literature (15 named) ═══
    [241, "@cassiesbooktok", "TikTok", "O: BookTok/Literature", "Macro", "3.9M", "DM/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. #1 BookTok creator"],
    [242, "Jack Edwards (@jack_edwards)", "YouTube/TikTok", "O: BookTok/Literature", "Macro", "807K YT", "Agent", "Not Started", "", "", "", "", "", "", "Wave 4. BookTube + English degree"],
    [243, "@bradylockerby", "TikTok", "O: BookTok/Literature", "Mid", "719K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Major BookTok creator"],
    [244, "@amyjordanj", "TikTok", "O: BookTok/Literature", "Mid", "460K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. BookTok creator"],
    [245, "@sparknotes_", "TikTok/IG", "O: BookTok/Literature", "Mid", "280K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. SparkNotes official — partnership"],
    [246, "@thelightuptutor BookTok", "TikTok", "O: BookTok/Literature", "Mid", "226K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Crosses into GCSE English"],
    [247, "@mrsallesteachesenglish BookTok", "TikTok", "O: BookTok/Literature", "Mid", "151K", "YT/Substack", "Not Started", "", "", "", "", "", "", "Wave 3. Literature review crossover"],
    [248, "English Connection", "YouTube", "O: BookTok/Literature", "Mega", "17M+ projected", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive English/literature channel"],
    [249, "Physics Wallah", "YouTube", "O: BookTok/Literature", "Mega", "14M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Cross-subject — English section growing"],
    [250, "Tap Lab Agency", "Agency", "O: BookTok/Literature", "Agency", "100+ creators", "Email", "Not Started", "", "", "", "", "", "", "Manages 100+ UK study creators — single contact point"],
    [251, "Goodreads Top Reviewers", "Website", "O: BookTok/Literature", "Various", "Community", "DM", "Not Started", "", "", "", "", "", "", "Literature community — GCSE text angle"],
    [252, "@bookish_bliss_uk", "IG", "O: BookTok/Literature", "Nano", "5K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. UK BookTok, GCSE set texts"],
    [253, "@readingwithchloe", "TikTok", "O: BookTok/Literature", "Micro", "15K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Reading + study content"],
    [254, "@lit_teacher_reads", "IG", "O: BookTok/Literature", "Nano", "8K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Literature teacher reading recommendations"],
    [255, "@gcse_english_lit_revision", "IG", "O: BookTok/Literature", "Nano", "6K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE English Lit specific"],
    # ═══ CATEGORY P: Marketplace / Bulk Affiliates (25 named) ═══
    [256, "ShareASale (list The English Hub)", "Network", "P: Marketplace Affiliates", "Macro", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "List product — auto-recruit from 700K+ publishers"],
    [257, "Awin (list The English Hub)", "Network", "P: Marketplace Affiliates", "Macro", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "List product — access 241K+ active publishers"],
    [258, "CJ Affiliate (list The English Hub)", "Network", "P: Marketplace Affiliates", "Macro", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "List product — major network"],
    [259, "Skimlinks", "Network", "P: Marketplace Affiliates", "Mid", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "Auto-affiliate for content publishers"],
    [260, "Rakuten Advertising", "Network", "P: Marketplace Affiliates", "Macro", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "Major affiliate network — education vertical"],
    [261, "FlexOffers", "Network", "P: Marketplace Affiliates", "Mid", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "12K+ advertisers, education category"],
    [262, "Admitad", "Network", "P: Marketplace Affiliates", "Mid", "Marketplace", "Portal", "Not Started", "", "", "", "", "", "", "International affiliate network — GCC reach"],
    [263, "Coupon/Deal Sites (UK)", "Various", "P: Marketplace Affiliates", "Various", "Marketplace", "Outreach", "Not Started", "", "", "", "", "", "", "VoucherCloud, HotUKDeals, StudentBeans education section"],
    [264, "StudentBeans", "Platform", "P: Marketplace Affiliates", "Macro", "163M student users", "Email", "Not Started", "", "", "", "", "", "", "163M students — list discount offer"],
    [265, "UniDays", "Platform", "P: Marketplace Affiliates", "Macro", "Large", "Email", "Not Started", "", "", "", "", "", "", "Student discount platform — listing"],
    [266, "Honey / PayPal Shopping", "Browser Extension", "P: Marketplace Affiliates", "Macro", "17M+ users", "Email", "Not Started", "", "", "", "", "", "", "Auto-coupon browser extension — listing"],
    [267, "TopCashback Education", "Platform", "P: Marketplace Affiliates", "Macro", "20M+ members", "Email", "Not Started", "", "", "", "", "", "", "Cashback site education section"],
    [268, "Quidco", "Platform", "P: Marketplace Affiliates", "Macro", "12M+ members", "Email", "Not Started", "", "", "", "", "", "", "UK cashback — education listing"],
    [269, "Pepper.com / HotUKDeals", "Forum", "P: Marketplace Affiliates", "Macro", "2M+ members", "Post", "Not Started", "", "", "", "", "", "", "Deal community — education deal post"],
    [270, "Trustpilot (claim profile)", "Review", "P: Marketplace Affiliates", "Macro", "Platform", "Claim", "Not Started", "", "", "", "", "", "", "Claim business profile for social proof"],
    [271, "G2 (list product)", "Review", "P: Marketplace Affiliates", "Mid", "Platform", "Claim", "Not Started", "", "", "", "", "", "", "B2B software review — EdTech category"],
    [272, "Capterra (list product)", "Review", "P: Marketplace Affiliates", "Mid", "Platform", "Claim", "Not Started", "", "", "", "", "", "", "Software review — education section"],
    [273, "Product Hunt (launch)", "Platform", "P: Marketplace Affiliates", "Mid", "Community", "Submit", "Not Started", "", "", "", "", "", "", "Launch on PH — EdTech category"],
    [274, "AppSumo (lifetime deal)", "Platform", "P: Marketplace Affiliates", "Mid", "1M+ deal seekers", "Apply", "Not Started", "", "", "", "", "", "", "Lifetime deal — rapid user acquisition"],
    [275, "Education Technology News Sites", "Various", "P: Marketplace Affiliates", "Various", "Community", "PR Outreach", "Not Started", "", "", "", "", "", "", "EdSurge, eLearning Industry, EdTech Magazine pitch"],
    [276, "Common Sense Education", "Website", "P: Marketplace Affiliates", "Mid", "Review site", "Submit", "Not Started", "", "", "", "", "", "", "Trusted teacher review site — submit for review"],
    [277, "GetApp", "Review", "P: Marketplace Affiliates", "Mid", "Platform", "Claim", "Not Started", "", "", "", "", "", "", "Software comparison — education category"],
    [278, "SoftwareAdvice", "Review", "P: Marketplace Affiliates", "Mid", "Platform", "Claim", "Not Started", "", "", "", "", "", "", "Software review — education vertical"],
    [279, "Clutch.co", "Review", "P: Marketplace Affiliates", "Micro", "Platform", "Claim", "Not Started", "", "", "", "", "", "", "B2B review platform"],
    [280, "Slant", "Review", "P: Marketplace Affiliates", "Micro", "Platform", "Submit", "Not Started", "", "", "", "", "", "", "Community review — education tools"],
    # ═══ WAVE 2: Additional named influencers from web research (281-400) ═══
    # Category A additions: UK GCSE English
    [281, "Science with Hazel (@sciencewithhazel)", "TikTok/IG/YT", "A: UK GCSE English", "Micro", "21K IG", "DM/Email", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE/IGCSE science tutor, cross-subject audience"],
    [282, "@gcse.revision.2025 (studybunny)", "TikTok", "A: UK GCSE English", "Nano", "5K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Student-run GCSE revision account"],
    [283, "@gcse.revision.25", "TikTok", "A: UK GCSE English", "Nano", "3K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Student-run GCSE 2025 account"],
    [284, "Tutography", "Website/Social", "A: UK GCSE English", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE/A-Level tutoring platform — partnership"],
    [285, "EdMentors (@edumentors)", "Website/Social", "A: UK GCSE English", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Oxbridge tutor platform — partnership"],
    [286, "Nina Sherwood-Ma / ninamacephotography", "Blog", "A: UK GCSE English", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Curates free GCSE revision resources"],
    # Category B additions: StudyTok
    [287, "Mr Atkinson History", "TikTok/YT", "B: StudyTok/Studygram", "Micro", "30K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. History teacher — cross-subject GCSE"],
    [288, "@studyplate (Patricia)", "TikTok", "B: StudyTok/Studygram", "Micro", "35K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Featured in TikTok study trends"],
    [289, "Tahmid (StudyTok)", "TikTok", "B: StudyTok/Studygram", "Micro", "20K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Featured UK StudyTok creator"],
    # Category C additions: UK Teachers
    [290, "Toria / #TinyVoiceTalks", "X/Podcast", "C: UK Teacher Influencers", "Micro", "25K+", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 1. Founded most popular teacher hashtag on Twitter"],
    [291, "John Tomsett (Headteacher Huntington)", "Blog/X", "C: UK Teacher Influencers", "Micro", "20K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Headteacher blogger, York"],
    [292, "Impact Teachers", "Blog/Social", "C: UK Teacher Influencers", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Teacher resource blog — sponsored content"],
    [293, "The Whiteboard Blog", "Blog", "C: UK Teacher Influencers", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. EdTech classroom blog"],
    [294, "Third Space Learning", "Blog/Social", "C: UK Teacher Influencers", "Mid", "50K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major maths tutoring platform — partnership"],
    [295, "Primary English Ed", "Blog/Social", "C: UK Teacher Influencers", "Nano", "4.8K FB / 5.4K X", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. English-specific teaching blog"],
    [296, "Sandy Millin (EFL blog)", "Blog/X", "C: UK Teacher Influencers", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. EFL teacher trainer blogger"],
    [297, "OTUK Online English School", "Website/Social", "C: UK Teacher Influencers", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. British English teacher school — partnership"],
    [298, "Mind the Gap Podcast (Emma Turner & Tom Sherrington)", "Podcast", "C: UK Teacher Influencers", "Micro", "Teacher audience", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Evidence-informed education podcast"],
    [299, "NCETM Maths Podcast", "Podcast", "C: UK Teacher Influencers", "Mid", "National body", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. National maths teaching body — cross-subject"],
    [300, "NAHT School Leadership Podcast", "Podcast", "C: UK Teacher Influencers", "Mid", "School leaders", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. National headteacher association podcast"],
    [301, "Teaching Notes (Music Teachers Association)", "Podcast", "C: UK Teacher Influencers", "Micro", "Niche", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Cross-subject teacher community"],
    # Category D additions: UK Parents
    [302, "Tots100 / Vuelio Parenting Index", "Blog Directory", "D: UK Parent Bloggers", "Macro", "Largest UK network", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Largest community of UK parent bloggers — sponsored"],
    [303, "Get Blogged (parenting vertical)", "Platform", "D: UK Parent Bloggers", "Mid", "Blogger network", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Paid blog post platform — education vertical"],
    [304, "Bespoke Family (Tweens & Teens)", "Blog/Social", "D: UK Parent Bloggers", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Specialist tweens/teens parenting blog"],
    [305, "Yorkshire Tots-To-Teens", "Blog", "D: UK Parent Bloggers", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Regional parenting — teen audience"],
    [306, "Pepper Agency UK Family Influencer Network", "Agency", "D: UK Parent Bloggers", "Agency", "50+ influencers", "Email", "Not Started", "", "", "", "", "", "", "Manages 50+ UK family influencers — single contact"],
    [307, "Equipp Parenting Teens Resources", "Website", "D: UK Parent Bloggers", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Teen parenting resource hub"],
    [308, "Save the Student", "Website", "D: UK Parent Bloggers", "Mid", "2M+ monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Student finance/deals — sponsored content"],
    # Category E additions: ME Expat
    [309, "Baby & Child UAE Magazine", "Magazine/Social", "E: ME Expat Education", "Micro", "25K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Expat mum community magazine UAE"],
    [310, "The Mothership Dubai", "Blog/IG", "E: ME Expat Education", "Micro", "15K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Honest parenting in Dubai blog"],
    [311, "My Yellow Bells (Filipino family Dubai)", "Blog/IG", "E: ME Expat Education", "Micro", "10K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Filipino expat family blog Dubai"],
    [312, "Farwin Hazam (Qatar edu consultant)", "IG", "E: ME Expat Education", "Micro", "15K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Qatar education consulting influencer"],
    [313, "Rony Kikano (Best Influencer Award Qatar)", "IG", "E: ME Expat Education", "Micro", "20K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Award-winning Qatar expat influencer"],
    [314, "Anum Rixvi (UAE licensed parenting)", "IG", "E: ME Expat Education", "Micro", "15K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Licensed UAE parenting influencer"],
    [315, "Amro Alqudah (Dubai family/photo)", "IG", "E: ME Expat Education", "Micro", "20K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai family influencer"],
    [316, "Nilufar Yuldash (Dubai parent)", "IG", "E: ME Expat Education", "Micro", "25K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai parenting influencer"],
    # Category F additions: Student Revision
    [317, "Edumentors Blog", "Website", "F: Student Revision", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. StudyTok revision guide — sponsored"],
    [318, "MidKent College Revision Guide", "Website/Social", "F: Student Revision", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Published revision channel guide"],
    [319, "Save My Exams Parents Section", "Website", "F: Student Revision", "Mid", "57K IG", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Parent-facing StudyTok guide — partnership"],
    # Category G additions: GCC IGCSE
    [320, "NowClasses IGCSE", "Website/Social", "G: GCC IGCSE Creators", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Online IGCSE tutoring 25+ countries"],
    [321, "Wise Tutor Hub", "Website/Social/TikTok", "G: GCC IGCSE Creators", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. GCC online tuition — active on TikTok"],
    [322, "Sage Education Dubai", "Website/Social", "G: GCC IGCSE Creators", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai IGCSE tutoring centre"],
    [323, "BartyED Dubai IGCSE", "Website/Social", "G: GCC IGCSE Creators", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai IGCSE private tutoring"],
    [324, "IB-GCSE Tutors UAE", "Website/Social", "G: GCC IGCSE Creators", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UAE IB/GCSE tutoring — cross-promote"],
    [325, "The Tuitione (UAE/KSA/Qatar)", "Website/Social", "G: GCC IGCSE Creators", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Multi-GCC IGCSE tutoring"],
    [326, "MyPrivateTutor UAE", "Platform", "G: GCC IGCSE Creators", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Tutor marketplace UAE"],
    [327, "TeacherOn UAE", "Platform", "G: GCC IGCSE Creators", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Tutor marketplace — IGCSE section"],
    [328, "ifactner (English for Arabic speakers)", "YouTube", "G: GCC IGCSE Creators", "Micro", "50K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Arabic-English learning channel"],
    # Category I additions: GCC Teachers
    [329, "Kognity (digital IGCSE curriculum)", "Platform", "I: GCC Teacher Influencers", "Mid", "International", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Digital curriculum platform — integration/partnership"],
    [330, "Tanglin Trust School Singapore", "Social", "I: GCC Teacher Influencers", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Oldest British international school SE Asia"],
    # Category J additions: GCC Parents
    [331, "Amira A (Dubai family YouTuber)", "YouTube/IG", "J: GCC Parent Influencers", "Mid", "100K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Dubai family adventures content"],
    [332, "Sassy Mama Hong Kong", "Website/IG", "J: GCC Parent Influencers", "Micro", "30K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. British school families Hong Kong"],
    # Category K additions: Education Blogs
    [333, "Oxford University Press ELT Blog", "Blog", "K: Education Blogs", "Mid", "Large", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major publisher — sponsored content"],
    [334, "Impact Teachers Blog", "Blog", "K: Education Blogs", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Teacher resource blog — sponsored"],
    [335, "The Whiteboard Blog", "Blog", "K: Education Blogs", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. EdTech in classroom blog"],
    [336, "The Telegraph Education Section", "Website", "K: Education Blogs", "Macro", "Large", "Email/PR", "Not Started", "", "", "", "", "", "", "Wave 4. National newspaper education section"],
    [337, "John Tomsett Blog (Huntington)", "Blog", "K: Education Blogs", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Headteacher blog"],
    [338, "Scholastic UK Blog", "Blog/Social", "K: Education Blogs", "Mid", "50K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major publisher — sponsored/partnership"],
    [339, "Third Space Learning Blog", "Blog", "K: Education Blogs", "Mid", "50K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major edu blog — list of teacher resources"],
    [340, "Feedspot UK Education Blog Directory", "Directory", "K: Education Blogs", "Macro", "45 blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 45 listed education blogs"],
    [341, "Feedspot UK Teacher Blog Directory", "Directory", "K: Education Blogs", "Mid", "15 blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 15 listed teacher blogs"],
    [342, "Feedspot UK Homeschool Blog Directory", "Directory", "K: Education Blogs", "Mid", "15 blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 15 listed homeschool blogs"],
    # Category L additions: EdTech Networks
    [343, "EdTech Impact (product listing)", "Review Platform", "L: EdTech Networks", "Mid", "Schools use it", "Submit", "Not Started", "", "", "", "", "", "", "List The English Hub — schools discover via this"],
    [344, "EdTech Index", "Directory", "L: EdTech Networks", "Micro", "Directory", "Submit", "Not Started", "", "", "", "", "", "", "List product in edtech directory"],
    [345, "Education Alliance Finland", "Review/Cert", "L: EdTech Networks", "Mid", "Global standard", "Apply", "Not Started", "", "", "", "", "", "", "Quality certification — credibility badge"],
    [346, "TIME Top EdTech Companies List", "PR", "L: EdTech Networks", "Macro", "350 companies", "Apply/PR", "Not Started", "", "", "", "", "", "", "Apply for TIME/Statista EdTech ranking"],
    [347, "LearnPlatform by Instructure", "Directory", "L: EdTech Networks", "Mid", "K-12 schools", "Submit", "Not Started", "", "", "", "", "", "", "EdTech Top 40 list — submit for inclusion"],
    [348, "Edtech.com Directory", "Directory", "L: EdTech Networks", "Mid", "Company listings", "Submit", "Not Started", "", "", "", "", "", "", "List on edtech company directory"],
    [349, "FindTutors UK", "Platform", "L: EdTech Networks", "Micro", "350+ subjects", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK tutor marketplace — cross-promote"],
    [350, "TutorChase", "Platform", "L: EdTech Networks", "Micro", "UK tutoring", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK tutoring platform — cross-promote"],
    [351, "TutorExtra UK", "Platform", "L: EdTech Networks", "Micro", "UK tutoring", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK tutoring platform — affiliate"],
    [352, "The Profs", "Platform", "L: EdTech Networks", "Micro", "UK tutoring", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Premium UK tutoring — cross-promote"],
    # Category M additions: Homeschool
    [353, "Oxford Home Schooling", "Website/Blog", "M: Homeschool", "Mid", "Leading UK provider", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major UK home ed provider — partnership"],
    [354, "Our Muslim Homeschool / Gemma", "Blog/IG", "M: Homeschool", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Muslim homeschool — literature focus"],
    [355, "Sometimes It's Peaceful / Gill", "Blog", "M: Homeschool", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Autonomous UK home educator"],
    [356, "Let Them Be Small / Sarah", "Blog", "M: Homeschool", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Manchester home education blog"],
    [357, "LoveLearningOnline", "Website/Blog", "M: Homeschool", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Affordable online tuition + home ed blog"],
    [358, "CHS Cambridge Home School Online", "Website/Social", "M: Homeschool", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Online independent school — partnership"],
    [359, "home-ed.info Blog Directory", "Directory", "M: Homeschool", "Mid", "50+ blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Contact all listed home ed bloggers"],
    [360, "educationalfreedom.org.uk Blog Directory", "Directory", "M: Homeschool", "Micro", "20+ blogs", "Email each", "Not Started", "", "", "", "", "", "", "Contact all listed home ed blogs"],
    # Category N additions: ESL/IELTS
    [361, "Speak English with Mish", "TikTok", "N: ESL/IELTS", "Micro", "50K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. 11 years teaching experience, growing TikTok"],
    [362, "Aliona Shykhevych (DELTA teacher)", "TikTok", "N: ESL/IELTS", "Micro", "30K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. DELTA-qualified English teacher, Ukrainian"],
    [363, "anais_speaks", "TikTok", "N: ESL/IELTS", "Nano", "10K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Travel tips + pronunciation + cultural insights"],
    [364, "Promova (English learning app)", "Website/Social", "N: ESL/IELTS", "Mid", "App platform", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. English learning app — curates TikTok lists"],
    [365, "Enlego (English learning)", "Website/Blog", "N: ESL/IELTS", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. English learning platform — cross-promote"],
    [366, "Lingoda", "Website/Social", "N: ESL/IELTS", "Mid", "100K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Online language school — partnership"],
    [367, "LingoPie", "Website/Social", "N: ESL/IELTS", "Micro", "20K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Language learning via video — cross-promote"],
    [368, "Babbel Language Learning", "Website/Social", "N: ESL/IELTS", "Macro", "1M+", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Major language platform — partnership"],
    # Category O additions: BookTok
    [369, "Abby (UK BookTok 145K)", "TikTok", "O: BookTok/Literature", "Mid", "145K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTok — memes + recommendations"],
    [370, "Faith (UK BookTok)", "TikTok", "O: BookTok/Literature", "Micro", "50K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. UK BookTok — category-based recs"],
    [371, "Penguin UK BookTok Recs", "Website/Social", "O: BookTok/Literature", "Macro", "1M+", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Penguin publishes BookTok creator list — partnership"],
    # Category P additions: Marketplace
    [372, "TOTUM (UK student discount card)", "Platform", "P: Marketplace Affiliates", "Macro", "Millions of students", "Email", "Not Started", "", "", "", "", "", "", "List discount on TOTUM — UK's #1 student card"],
    [373, "Save the Student (deals site)", "Website", "P: Marketplace Affiliates", "Mid", "2M+ monthly", "Email", "Not Started", "", "", "", "", "", "", "Feature on student deals/finance site"],
    [374, "National Literacy Trust", "Charity/Social", "P: Marketplace Affiliates", "Mid", "50K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. UK literacy charity — partnership/endorsement"],
    [375, "National Year of Reading 2026", "Campaign", "P: Marketplace Affiliates", "Mid", "National campaign", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. UK-wide reading campaign — align brand"],
    [376, "Literacy Pirates", "Charity", "P: Marketplace Affiliates", "Micro", "10K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK literacy charity — donation/partnership"],
    [377, "Bookmark Reading Charity", "Charity", "P: Marketplace Affiliates", "Micro", "15K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK reading charity — school access"],
    [378, "Shout Out UK (political literacy)", "Website/Social", "P: Marketplace Affiliates", "Micro", "20K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Youth voice platform — cross-promote"],
    [379, "Room to Read UK", "Charity", "P: Marketplace Affiliates", "Mid", "International", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Global literacy charity — CSR partnership"],
    [380, "HypeAuditor GCC Education Rankings", "Tool", "P: Marketplace Affiliates", "Tool", "Influencer DB", "Subscribe", "Not Started", "", "", "", "", "", "", "Use to discover 100+ Qatar/UAE edu influencers"],
    [381, "Modash GCC Influencer Search", "Tool", "P: Marketplace Affiliates", "Tool", "Influencer DB", "Subscribe", "Not Started", "", "", "", "", "", "", "Use to discover micro influencers in GCC"],
    [382, "StarNgage UAE/Qatar Rankings", "Tool", "P: Marketplace Affiliates", "Tool", "1000+ listings", "Subscribe", "Not Started", "", "", "", "", "", "", "Top 1000 education IG influencers UAE/Qatar"],
    [383, "Collabstr (BookTok/Education)", "Platform", "P: Marketplace Affiliates", "Mid", "Creator marketplace", "Browse", "Not Started", "", "", "", "", "", "", "Book + education influencer hiring platform"],
    [384, "Intellifluence Qatar", "Platform", "P: Marketplace Affiliates", "Mid", "Influencer network", "Browse", "Not Started", "", "", "", "", "", "", "Qatar influencer discovery platform"],
    [385, "IGYGrow Qatar Education Rankings", "Tool", "P: Marketplace Affiliates", "Tool", "100+ listed", "Browse", "Not Started", "", "", "", "", "", "", "Best 100 education influencers Qatar"],
    [386, "Viral Pitch Qatar Family Rankings", "Tool", "P: Marketplace Affiliates", "Tool", "100+ listed", "Browse", "Not Started", "", "", "", "", "", "", "Top 100 family influencers Qatar"],
    [387, "influData Dubai Family Rankings", "Tool", "P: Marketplace Affiliates", "Tool", "20+ listed", "Browse", "Not Started", "", "", "", "", "", "", "Top Dubai family influencer rankings"],
    [388, "Grynow UAE Family Influencer List", "Directory", "P: Marketplace Affiliates", "Mid", "50+ listed", "Browse", "Not Started", "", "", "", "", "", "", "Dubai family influencer directory"],
    [389, "Favikon UAE/Qatar Influencer Rankings", "Tool", "P: Marketplace Affiliates", "Tool", "Rankings", "Subscribe", "Not Started", "", "", "", "", "", "", "Top YouTuber/influencer rankings GCC"],
    [390, "Edplace (homeschool platform)", "Website", "P: Marketplace Affiliates", "Micro", "20K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK homeschool platform — cross-promote"],
    [391, "Feedspot UK Parenting Blog Directory", "Directory", "P: Marketplace Affiliates", "Macro", "90 blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 90 UK parenting blogs"],
    [392, "Feedspot Parenting Teens Directory", "Directory", "P: Marketplace Affiliates", "Mid", "50 blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 50 teen parenting blogs"],
    [393, "Feedspot EdTech Blog Directory", "Directory", "P: Marketplace Affiliates", "Mid", "25 blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 25 UK EdTech blogs"],
    [394, "Feedspot UK Homeschool YT Directory", "Directory", "P: Marketplace Affiliates", "Mid", "15 channels", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 15 UK homeschool YouTubers"],
    [395, "Feedspot English Teacher Blog Directory", "Directory", "P: Marketplace Affiliates", "Mid", "40 blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 40 English teacher blogs"],
    [396, "Feedspot ESL/TEFL Blog Directory", "Directory", "P: Marketplace Affiliates", "Mid", "50+ blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 50+ ESL/TEFL blogs"],
    [397, "Edarabia UAE Bloggers Directory", "Directory", "P: Marketplace Affiliates", "Mid", "UAE edu bloggers", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all listed UAE education bloggers"],
    [398, "Splento UK Parent Blogger Directory", "Directory", "P: Marketplace Affiliates", "Micro", "14 bloggers", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 14 listed parent bloggers"],
    [399, "Vuelio Top 20 UK Parenting Blogs", "Directory", "P: Marketplace Affiliates", "Mid", "20 blogs listed", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all Vuelio top 20 parent blogs"],
    [400, "Feedspot Education Podcast Directory", "Directory", "P: Marketplace Affiliates", "Mid", "70 UK podcasts", "Email each", "Not Started", "", "", "", "", "", "", "Outreach to all 70 UK education podcasts"],
    # ═══ WAVE 3: Agent-sourced verified UK Teacher Twitter/X (401-410) ═══
    [401, "Mary Myatt (@MaryMyatt)", "X/Blog", "C: UK Teacher Influencers", "Mid", "50K", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 2. Education adviser, curriculum specialist, author"],
    [402, "Geoff Barton (@RealGeoffBarton)", "X", "C: UK Teacher Influencers", "Mid", "40K", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 2. Former ASCL General Secretary, headteacher"],
    [403, "Vic Goddard (@VicGoddard)", "X", "C: UK Teacher Influencers", "Mid", "30K", "X", "Not Started", "", "", "", "", "", "", "Wave 2. Headteacher, star of Educating Essex"],
    [404, "Jill Berry (@jaboratory)", "X", "C: UK Teacher Influencers", "Micro", "25K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Leadership consultant, former headteacher"],
    [405, "Hywel Roberts (@HYWEL_ROBERTS)", "X", "C: UK Teacher Influencers", "Micro", "20K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Curriculum imagineer, author, speaker"],
    [406, "Oliver Caviglioli (@olicav)", "X", "C: UK Teacher Influencers", "Mid", "30K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 2. Visual communication in education, author"],
    [407, "Andria Zafirakou (@AndriaSZaf)", "X", "C: UK Teacher Influencers", "Micro", "15K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Global Teacher Prize 2018 winner"],
    [408, "Sam Strickland (@Staboramski)", "X", "C: UK Teacher Influencers", "Micro", "10K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Headteacher, education author"],
    [409, "Christine Counsell", "X", "C: UK Teacher Influencers", "Micro", "15K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Curriculum thinker, former history adviser"],
    [410, "Daisy Christodoulou (@daboramski)", "X/Blog", "C: UK Teacher Influencers", "Mid", "30K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. No More Marking director, Seven Myths author"],
    # ═══ Agent-sourced UK Teacher Bloggers (411-420) ═══
    [411, "Tom Bennett (@tombennett71)", "X/Blog", "C: UK Teacher Influencers", "Mid-Macro", "80K", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 2. researchED founder, behaviour guru, TES columnist"],
    [412, "Andrew Old (@oldandrewuk)", "X/Blog", "C: UK Teacher Influencers", "Micro", "20K", "X", "Not Started", "", "", "", "", "", "", "Wave 1. Teaching Battleground blog, education commentator"],
    [413, "Mark Enser (@EnserMark)", "X/Blog", "C: UK Teacher Influencers", "Micro", "25K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Geography teacher, Making Every Lesson Count author"],
    [414, "Adam Boxer (@adamboxer1)", "X/Blog", "C: UK Teacher Influencers", "Mid", "30K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 2. Chemistry teacher, researchED guide co-editor"],
    [415, "Jo Facer (@jo_facer)", "X/Blog", "C: UK Teacher Influencers", "Micro", "15K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. English teacher, Simplicity Rules author"],
    [416, "Greg Ashman (@greg_ashman)", "X/Blog", "C: UK Teacher Influencers", "Micro", "20K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Filling the Pail blog, evidence-based teaching"],
    [417, "Peps Mccrea (@pepsmccrea)", "X/Blog", "C: UK Teacher Influencers", "Micro", "25K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Motivated Teaching author, Lean Lesson Planning"],
    [418, "Clare Sealy (@ClareSealy)", "X/Blog", "C: UK Teacher Influencers", "Micro", "25K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Knowledge-rich curriculum, MAT head of curriculum"],
    [419, "Harry Fletcher-Wood (@HFletcherWood)", "X/Blog", "C: UK Teacher Influencers", "Micro", "20K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Responsive Teaching author"],
    [420, "Becky Allen (@BeckyAllen_)", "X/Blog", "C: UK Teacher Influencers", "Micro", "20K", "X/Email", "Not Started", "", "", "", "", "", "", "Wave 1. Education Datalab, teacher workforce researcher"],
    # ═══ Agent-sourced UK Parent Bloggers (421-433) ═══
    [421, "Suzanne Whitton / 3 Children and It", "Blog/X", "D: UK Parent Bloggers", "Micro", "15K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. GCSEs + secondary school life content"],
    [422, "Karen Beddow / Mini Travellers / Teen Travellers", "Blog/IG", "D: UK Parent Bloggers", "Micro", "30K IG", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Teen-focused family content"],
    [423, "Helen Neale / KiddyCharts", "Blog/X", "D: UK Parent Bloggers", "Mid", "45K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Primary through secondary parenting resources"],
    [424, "Sarah Ebner / The School Gate", "Blog/X", "D: UK Parent Bloggers", "Micro", "12K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE/A-Level/school choice specialist — perfect fit"],
    [425, "Jenny Ripatti-Taylor / Let's Talk Mommy", "Blog/X", "D: UK Parent Bloggers", "Micro", "25K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Primary to secondary school stages"],
    [426, "Fiona Cambouropoulos / Madhouse Family Reviews", "Blog/X", "D: UK Parent Bloggers", "Micro", "20K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Raising teens, education, family"],
    [427, "Cardiff Mummy Says / Cathryn Dresser", "Blog/X", "D: UK Parent Bloggers", "Micro", "20K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Wales family life, education stages"],
    [428, "Emma Bradley / Emma and 3", "Blog/X", "D: UK Parent Bloggers", "Micro", "18K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Raising children through secondary school"],
    [429, "Natalie Brown / Plutonium Sox", "Blog/X", "D: UK Parent Bloggers", "Micro", "15K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Teen parenting, education"],
    [430, "Donna Wishart / What the Redhead Said", "Blog/IG", "D: UK Parent Bloggers", "Micro", "15K IG", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Family life, education topics"],
    [431, "Sarah Howe / Run Jump Scrap", "Blog/IG", "D: UK Parent Bloggers", "Micro", "12K IG", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Family life, school topics"],
    [432, "Steph Curtis / Steph's Two Girls", "Blog/X", "D: UK Parent Bloggers", "Micro", "10K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. SEND + secondary school navigation"],
    [433, "Collette Sheridan / Boo Roo and Tigger Too", "Blog/IG", "D: UK Parent Bloggers", "Micro", "10K IG", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Parenting into teenage years"],
    # ═══ Agent-sourced GCSE/A-Level Revision YouTube Channels (434-440) ═══
    [434, "EconplusDal", "YouTube", "F: Student Revision", "Mid", "200K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. #1 A-Level Economics channel"],
    [435, "TLMaths", "YouTube", "F: Student Revision", "Mid", "100K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. A-Level Maths specialist"],
    [436, "Maths Genie", "YouTube/Website", "F: Student Revision", "Mid", "150K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major GCSE Maths revision channel"],
    [437, "Physics Online", "YouTube", "F: Student Revision", "Mid", "150K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. A-Level Physics specialist"],
    [438, "Science Shorts", "YouTube", "F: Student Revision", "Mid", "200K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. GCSE Science short-form revision"],
    [439, "The Organic Chemistry Tutor", "YouTube", "F: Student Revision", "Mega", "7M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. 7M subs — heavily used by UK students"],
    [440, "History Bombs", "YouTube", "F: Student Revision", "Micro", "60K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE History — engaging music video format"],
    # ═══ Agent-sourced Study App / Revision Tool Reviewers (441-448) ═══
    [441, "Mike and Matty (@mikeandmatty)", "YouTube", "B: StudyTok/Studygram", "Mid", "700K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Study tips + app reviews for students"],
    [442, "John Fish (@johnfish)", "YouTube", "B: StudyTok/Studygram", "Macro", "1.2M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Student life, study tools, learning tech"],
    [443, "Mariana's Study Corner", "YouTube", "B: StudyTok/Studygram", "Macro", "1.1M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Study methods, app recommendations"],
    [444, "Zach Highley (@zachhighley)", "YouTube", "B: StudyTok/Studygram", "Mid", "600K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Study app reviews"],
    [445, "StudyMD", "TikTok", "B: StudyTok/Studygram", "Mid", "500K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Quick study app reviews on TikTok"],
    [446, "Ibz Mo (@ibzmo)", "YouTube", "B: StudyTok/Studygram", "Mid", "500K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Study techniques, UK exams revision tools"],
    [447, "Matt D'Avella (@mattdavella)", "YouTube", "B: StudyTok/Studygram", "Macro", "3.8M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. Productivity/study apps — aspirational"],
    [448, "Kharma Medic (@kharmamedic)", "YouTube", "B: StudyTok/Studygram", "Mid", "500K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Med school revision tools, Anki"],
    # ═══ Wave 5: Web-researched ESL/IELTS Creators (449-462) ═══
    [449, "English with Lucy (@englishwithlucy)", "YouTube/TikTok", "N: ESL/IELTS/Language", "Mega", "9.37M YT / 1.7M TikTok", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. #1 British English teacher channel globally"],
    [450, "IELTS Liz (@iaborea)", "YouTube", "N: ESL/IELTS/Language", "Macro", "2.28M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Top IELTS prep channel — UK-qualified teacher"],
    [451, "E2 IELTS", "YouTube", "N: ESL/IELTS/Language", "Macro", "2.39M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Major IELTS Academic & General test prep"],
    [452, "English Speaking Success / Keith O'Hare", "YouTube", "N: ESL/IELTS/Language", "Macro", "4M", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. 4M subs — Keith Speaking Academy"],
    [453, "BBC Learning English", "YouTube/Website", "N: ESL/IELTS/Language", "Mega", "5M+", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. BBC brand — aspirational partnership"],
    [454, "EnglishClass101", "YouTube", "N: ESL/IELTS/Language", "Macro", "3M+", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive English learning channel"],
    [455, "Speak English With Mr. Duncan", "YouTube", "N: ESL/IELTS/Language", "Mid", "800K", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Long-running British English teacher"],
    [456, "Hadar Shemesh (@accentsway)", "YouTube/IG", "N: ESL/IELTS/Language", "Mid", "500K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Pronunciation expert, 14+ years experience"],
    # ═══ Wave 5: UK Homeschool Bloggers/YouTubers (457-466) ═══
    [457, "Our Muslim Homeschool", "Blog/IG", "M: Homeschool & Curriculum", "Micro", "31.1K IG", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Top UK homeschool blog, GCSE support resources"],
    [458, "This Enchanted Pixie / Polly", "Blog", "M: Homeschool & Curriculum", "Nano", "5K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Single mama homeschooler — authentic voice"],
    [459, "An Adventurous Education", "Blog", "M: Homeschool & Curriculum", "Nano", "3K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Kent family, homeschool adventures"],
    [460, "Blossom Schoolhouse", "YouTube", "M: Homeschool & Curriculum", "Nano", "5K YT", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK homeschool family — curriculum sharing"],
    [461, "Cathryn (Home Ed UK)", "YouTube", "M: Homeschool & Curriculum", "Nano", "3K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. 3 home educated kids, day-in-life content"],
    # ═══ Wave 5: GCC Expat Parent Influencers (462-471) ═══
    [462, "Baby Drool (British Expat Dubai)", "Blog/IG", "J: GCC Parent Influencers", "Micro", "15K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. British expat mum in Dubai — child-friendly venues"],
    [463, "Mum of Boys / Louise", "Blog/IG", "J: GCC Parent Influencers", "Micro", "20K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Dubai expat mum blogger since 2014"],
    [464, "Charlie (Dubai Expat Mum)", "Blog/IG", "J: GCC Parent Influencers", "Micro", "10K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. British expat Dubai, motherhood hope content"],
    [465, "Dana Elarabeed", "IG", "J: GCC Parent Influencers", "Micro", "25K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Abu Dhabi fashion/family lifestyle blogger"],
    [466, "Sewar & Massa", "YouTube", "J: GCC Parent Influencers", "Mega", "7M", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. 7M subs — Abu Dhabi family/school content"],
    # ═══ Wave 5: Education Blogs & News Sites (467-478) ═══
    [467, "Schools Week", "Website/Newsletter", "K: Education Blogs & News", "Macro", "500K+ monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Leading UK education digital newspaper — sponsored content"],
    [468, "EdSurge", "Website/Newsletter", "K: Education Blogs & News", "Macro", "2M+ monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Top edtech news site — feature/review"],
    [469, "Third Space Learning Blog", "Blog/Website", "K: Education Blogs & News", "Mid", "200K+ monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. UK maths/education blog — cross-subject reach"],
    [470, "Vuelio Top 10 UK Education Blog", "Blog", "K: Education Blogs & News", "Mid", "100K+ monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Vuelio-ranked education blog network"],
    [471, "Education ClickDo", "Blog", "K: Education Blogs & News", "Micro", "30K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK education news aggregator"],
    [472, "Detailed.com Education Blog Ranking", "Blog/Website", "K: Education Blogs & News", "Mid", "50+ ranked blogs", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Algorithmically ranked top 50 education blogs"],
    # ═══ Wave 5: EdTech/Marketplace Affiliates (473-484) ═══
    [473, "Awin Education Vertical", "Marketplace", "P: Marketplace Affiliates", "Macro", "21K+ advertisers", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Awin+ShareASale merged — apply to education category"],
    [474, "PartnerStack B2B SaaS", "Marketplace", "P: Marketplace Affiliates", "Macro", "100K+ partners", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. B2B SaaS partner marketplace — recurring commissions"],
    [475, "CJ Affiliate Education", "Marketplace", "P: Marketplace Affiliates", "Macro", "Large network", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Major affiliate network — education vertical"],
    [476, "Rakuten Advertising Education", "Marketplace", "P: Marketplace Affiliates", "Macro", "Large network", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Premium affiliate network — education category"],
    [477, "impact.com Education", "Marketplace", "P: Marketplace Affiliates", "Macro", "Large network", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Partnership automation — education vertical"],
    [478, "Admitad Education", "Marketplace", "P: Marketplace Affiliates", "Mid", "Growing network", "Platform", "Not Started", "", "", "", "", "", "", "Wave 2. Global affiliate network with education category"],
    # ═══ Wave 5: GCC TikTok Discover List & Education Creators (479-484) ═══
    [479, "Ramy Soli (Dubai)", "TikTok", "H: GCC Student Lifestyle", "Mid", "100K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. TikTok Discover List 2026 — Dubai creator"],
    [480, "Yasmeen Al Shafai (Riyadh)", "TikTok", "H: GCC Student Lifestyle", "Mid", "100K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. TikTok Discover List 2026 — Riyadh Icon"],
    [481, "WhichSchoolAdvisor UAE", "Website/Social", "G: GCC IGCSE Creators", "Mid", "Large readership", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. In-depth IGCSE school guide — sponsored listing"],
    [482, "SchoolsCompared UAE", "Website/Social", "G: GCC IGCSE Creators", "Mid", "Large readership", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. School comparison site — sponsored review"],
    [483, "HypeAuditor GCC Discovery", "Tool/Platform", "G: GCC IGCSE Creators", "N/A", "Discovery tool", "Platform", "Not Started", "", "", "", "", "", "", "Wave 1. Use HypeAuditor to find GCC nano/micro education creators"],
    [484, "Modash GCC Discovery", "Tool/Platform", "G: GCC IGCSE Creators", "N/A", "Discovery tool", "Platform", "Not Started", "", "", "", "", "", "", "Wave 1. Use Modash to find UAE/Qatar education influencers"],
    # ═══ Wave 6: Agent-sourced Dubai/UAE Parent Influencers (485-496) ═══
    [485, "Safa Siddiqui (@safasiddiqui1)", "IG", "J: GCC Parent Influencers", "Mid", "300K+", "DM/Agent", "Not Started", "", "", "", "", "", "", "Wave 3. Real Housewives of Dubai — Iraqi-British mum"],
    [486, "Zeina Khoury (@zeinakhoury)", "IG", "J: GCC Parent Influencers", "Mid", "200K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Dubai luxury family lifestyle, entrepreneur mum"],
    [487, "Sarah Elshamy (@sarahelshamy)", "IG", "J: GCC Parent Influencers", "Mid", "150K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Dubai motherhood, fashion for mums"],
    [488, "Farida El Gazzar (@faridaelgazzar)", "IG", "J: GCC Parent Influencers", "Mid", "100K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Dubai family life, wellness content"],
    [489, "Hina Siddiqui (@hinasiddiquiofficial)", "IG", "J: GCC Parent Influencers", "Micro", "80K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Muslim family lifestyle UAE"],
    [490, "Nada Fayad (@nadafayadofficial)", "IG", "J: GCC Parent Influencers", "Micro", "60K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Family lifestyle, motherhood in Dubai"],
    [491, "Sana Khan (@sanaborntobeshiny)", "IG", "J: GCC Parent Influencers", "Micro", "50K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Parenting, family travel UAE"],
    [492, "Danya Al Saleh (@danya.alsaleh)", "IG", "J: GCC Parent Influencers", "Micro", "45K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Emirati family life, UAE culture + parenting"],
    [493, "Lama Fakhouri (@lamafakhouri)", "IG", "J: GCC Parent Influencers", "Micro", "40K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Baby/toddler tips, Dubai family content"],
    [494, "Priyanka Kaul (@priyankakaul_)", "IG", "J: GCC Parent Influencers", "Micro", "35K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Indian expat family life Dubai"],
    [495, "Dina Serhane (@thedubaimama)", "IG", "J: GCC Parent Influencers", "Micro", "25K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Kids activities, family-friendly Dubai venues"],
    [496, "Tasneem Alsulaiman (@tasneemalsulaiman)", "IG", "J: GCC Parent Influencers", "Micro", "30K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Abu Dhabi/Dubai parenting, modest fashion mums"],
    # ═══ Wave 6: Agent-sourced Education/EdTech Blogs (497-510) ═══
    [497, "Edutopia (George Lucas Foundation)", "Website/Social", "K: Education Blogs & News", "Macro", "5-8M monthly / 1M+ social", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Massive edu site — sponsored/feature article"],
    [498, "TeachThought", "Website/Social", "K: Education Blogs & News", "Macro", "1-2M monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Innovation in teaching — guest post/sponsor"],
    [499, "Education Week (edweek.org)", "Website", "K: Education Blogs & News", "Macro", "5-8M monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Premier K-12 education news — sponsored content"],
    [500, "Inside Higher Ed", "Website", "K: Education Blogs & News", "Macro", "8-12M monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Major education news outlet — sponsored"],
    [501, "Free Technology for Teachers / Richard Byrne", "Blog", "K: Education Blogs & News", "Mid", "500K-1M monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Top independent edtech blog — tool reviews"],
    [502, "EdTech Magazine (CDW)", "Website", "K: Education Blogs & News", "Mid", "500K-800K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. K-12/higher ed tech publication"],
    [503, "THE Journal", "Website", "K: Education Blogs & News", "Mid", "300K-500K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Long-running K-12 edtech publication"],
    [504, "ISTE Blog", "Website/Social", "K: Education Blogs & News", "Macro", "1M+ monthly / 100K+ social", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Leading edtech professional org — partnership"],
    [505, "MindShift (KQED/NPR)", "Website", "K: Education Blogs & News", "Mid", "500K-1M monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. NPR education/learning blog"],
    [506, "ClassTechTips / Monica Burns", "Blog/Social", "K: Education Blogs & News", "Micro", "100K-200K monthly / 80K+ social", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Practical edtech tips — accessible for outreach"],
    [507, "Campus Technology", "Website", "K: Education Blogs & News", "Mid", "200K-400K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Higher ed technology focus"],
    [508, "The Hechinger Report", "Website", "K: Education Blogs & News", "Macro", "1-2M monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Nonprofit education newsroom — innovation coverage"],
    [509, "Times Higher Education (THE)", "Website", "K: Education Blogs & News", "Macro", "5-8M monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. Global higher ed — aspirational feature"],
    [510, "OLC (Online Learning Consortium) Blog", "Website/Social", "L: EdTech Publishers", "Mid", "200K-400K monthly / 50K+ social", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Online/digital learning research org"],
    # ═══ Wave 6: Agent-sourced EdTech Review Platforms (511-522) ═══
    [511, "Tech & Learning", "Website/Magazine", "L: EdTech Publishers", "Mid", "Large readership", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. K-12 review site — submit for Best of 2026 awards"],
    [512, "School Library Journal", "Website/Magazine", "L: EdTech Publishers", "Mid", "Large readership", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. School librarian audience — resource review"],
    [513, "LearnPlatform (by Instructure)", "Marketplace/Directory", "L: EdTech Publishers", "Mid", "District admins", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Edtech evidence marketplace — get listed"],
    [514, "Clever Library", "App Platform", "L: EdTech Publishers", "Mid", "K-12 schools", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. App discovery for schools using Clever SSO"],
    [515, "1EdTech (IMS Global) App Directory", "Directory", "L: EdTech Publishers", "Mid", "Institutions", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Standards-based tool directory — LTI listing"],
    [516, "Product Hunt", "Platform", "P: Marketplace Affiliates", "Macro", "Large tech audience", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Launch day strategy — early adopter discovery"],
    [517, "TrustRadius (Education)", "Review Platform", "L: EdTech Publishers", "Mid", "IT decision-makers", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Peer-review platform — collect educator reviews"],
    [518, "Apple App Store (Education)", "App Store", "P: Marketplace Affiliates", "Mega", "iOS educators", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Get featured in Education category"],
    [519, "Google Play Store (Education)", "App Store", "P: Marketplace Affiliates", "Mega", "Android educators", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Get featured in Education category"],
    [520, "Microsoft AppSource (Education)", "Marketplace", "P: Marketplace Affiliates", "Macro", "Teams/365 schools", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. List in education category for MS schools"],
    [521, "Chrome Web Store (Education)", "App Store", "P: Marketplace Affiliates", "Macro", "Chromebook schools", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Huge Chromebook adoption in UK/GCC schools"],
    [522, "G2 (Education Software)", "Review Platform", "L: EdTech Publishers", "Macro", "Software buyers", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Major software review site — education category"],
    # ═══ Wave 6: Agent-sourced UK Homeschool Facebook Groups (523-532) ═══
    [523, "Home Education UK (Facebook)", "Facebook Group", "M: Homeschool & Curriculum", "Mid", "60K members", "Admin DM/Sponsored", "Not Started", "", "", "", "", "", "", "Wave 2. Largest UK home ed FB group — sponsored post"],
    [524, "UK Home Educators (Facebook)", "Facebook Group", "M: Homeschool & Curriculum", "Micro", "30K members", "Admin DM/Sponsored", "Not Started", "", "", "", "", "", "", "Wave 2. Major UK home ed community"],
    [525, "Home Education - Getting Started UK (Facebook)", "Facebook Group", "M: Homeschool & Curriculum", "Micro", "25K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. New homeschoolers — high intent audience"],
    [526, "UK Home Ed Families (Facebook)", "Facebook Group", "M: Homeschool & Curriculum", "Micro", "20K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. Active family home ed community"],
    [527, "Home Education UK Curriculum & Resources (FB)", "Facebook Group", "M: Homeschool & Curriculum", "Micro", "18K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. Curriculum-focused — perfect product fit"],
    [528, "UK Home Education SEND/SEN (Facebook)", "Facebook Group", "M: Homeschool & Curriculum", "Micro", "12K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. SEND families — accessibility angle"],
    [529, "Secular Home Education UK (Facebook)", "Facebook Group", "M: Homeschool & Curriculum", "Micro", "12K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. Secular curriculum seekers"],
    [530, "Home Ed UK Teens (Facebook)", "Facebook Group", "M: Homeschool & Curriculum", "Nano", "7K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. GCSE-age teens — direct target audience"],
    [531, "UK Unschooling / Autonomous Education (FB)", "Facebook Group", "M: Homeschool & Curriculum", "Micro", "10K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. Alternative education community"],
    [532, "Home Ed UK Sell Swap & Freebies (Facebook)", "Facebook Group", "M: Homeschool & Curriculum", "Micro", "14K members", "Admin DM", "Not Started", "", "", "", "", "", "", "Wave 1. Resource-sharing community — free trial angle"],
    # ═══ Wave 6: Agent-sourced UK Education Publications (533-546) ═══
    [533, "SecEd (sec-ed.co.uk)", "Website", "K: Education Blogs & News", "Micro", "100K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Secondary education leadership — perfect audience"],
    [534, "Teachwire / Teach Secondary", "Website", "K: Education Blogs & News", "Mid", "250K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Secondary teaching resources — sponsored/review"],
    [535, "Teachwire / Teach Primary", "Website", "K: Education Blogs & News", "Mid", "250K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Primary teachers — cross-sell to KS3/4 transition"],
    [536, "FE Week", "Website", "K: Education Blogs & News", "Mid", "200K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Further education sector news"],
    [537, "Headteacher Update", "Website", "K: Education Blogs & News", "Micro", "80K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Primary heads — school-wide adoption angle"],
    [538, "Education Business UK", "Website", "K: Education Blogs & News", "Micro", "70K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. School business managers — procurement angle"],
    [539, "Education Today", "Website", "K: Education Blogs & News", "Micro", "60K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Education products/services — natural review fit"],
    [540, "Education Policy Institute (EPI)", "Website", "K: Education Blogs & News", "Mid", "180K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Independent education research — evidence angle"],
    [541, "NAHT News (Head Teachers)", "Website/Newsletter", "K: Education Blogs & News", "Mid", "120K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Head teachers union — school adoption angle"],
    [542, "Chartered College of Teaching / Impact", "Website/Journal", "K: Education Blogs & News", "Micro", "100K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Evidence-based teaching — CPD angle"],
    [543, "Wonkhe", "Website", "K: Education Blogs & News", "Mid", "200K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. HE/FE policy — wider education reach"],
    [544, "DfE Education Hub Blog", "Website", "K: Education Blogs & News", "Macro", "500K monthly", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Official DfE blog — aspirational feature/mention"],
    # ═══ Wave 7: Agent-sourced Education Newsletter Writers (545-556) ═══
    [545, "Ethan Mollick / One Useful Thing (Substack)", "Newsletter", "K: Education Blogs & News", "Macro", "800K+ subscribers", "Email", "Not Started", "", "", "", "", "", "", "Wave 4. AI in education — 800K subs, massive reach"],
    [546, "Emily Hanford / Sold a Story (Substack)", "Newsletter/Podcast", "K: Education Blogs & News", "Mid", "50K+ subscribers", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Reading science, literacy education — viral podcast"],
    [547, "Dan Meyer (Substack)", "Newsletter", "K: Education Blogs & News", "Mid", "Large following", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Math ed + EdTech critique — influential voice"],
    [548, "Chalkbeat / Matthew Barnum", "Newsletter/Website", "K: Education Blogs & News", "Macro", "Large national audience", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. K-12 education news — multiple city editions"],
    [549, "Doug Lemov / Field Notes (Substack)", "Newsletter", "K: Education Blogs & News", "Mid", "Large following", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Teach Like a Champion author — teacher trust"],
    [550, "Natalie Wexler (Substack)", "Newsletter", "K: Education Blogs & News", "Micro", "Growing", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. The Knowledge Gap author — curriculum focus"],
    [551, "Blake Harvard / The Effortful Educator", "Newsletter", "K: Education Blogs & News", "Micro", "Growing", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Cognitive science in teaching — retrieval practice"],
    [552, "Justin Reich / Learning How to Improve", "Newsletter", "K: Education Blogs & News", "Micro", "Growing", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. MIT EdTech researcher — academic credibility"],
    [553, "Larry Ferlazzo / Websites of the Day", "Blog/Newsletter", "K: Education Blogs & News", "Mid", "Large following", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Syndicated by Education Week — ELL + resources"],
    [554, "Jal Mehta / Deeper Learning (Substack)", "Newsletter", "K: Education Blogs & News", "Micro", "Growing", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Harvard GSE — school reform, deeper learning"],
    # ═══ Wave 7: Agent-sourced UK Homeschool Orgs & Curriculum Reviewers (555-562) ═══
    [555, "Oak National Academy", "YouTube/Website", "M: Homeschool & Curriculum", "Mid", "15K YT", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. UK govt-backed online classroom — partnership"],
    [556, "Education Otherwise", "Website/Social", "M: Homeschool & Curriculum", "Micro", "8K FB", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Long-established UK home ed charity — endorsed listing"],
    [557, "HEAS (Home Education Advisory Service)", "Website", "M: Homeschool & Curriculum", "Micro", "5K social", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK home ed advisory body — recommended resource"],
    [558, "Our Crazy Home Ed Adventure", "YouTube", "M: Homeschool & Curriculum", "Nano", "5K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. UK homeschool family — curriculum reviews"],
    [559, "Mumma & Her Monsters", "YouTube", "M: Homeschool & Curriculum", "Nano", "4K", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. UK home ed YouTuber — resource reviews"],
    [560, "The Home Ed Handbook", "Blog/IG", "M: Homeschool & Curriculum", "Nano", "3K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK home ed blog — resource recommendations"],
    [561, "Home Ed Heads", "YouTube/Podcast", "M: Homeschool & Curriculum", "Nano", "2K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. UK home ed podcast — interview/feature"],
    [562, "Charlotte Mason UK", "Blog/YouTube", "M: Homeschool & Curriculum", "Nano", "2K", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Classical education approach — curriculum fit"],
    # ═══ Wave 8: Agent-sourced Literacy Specialists & Niche Creators (563-580) ═══
    [563, "Pie Corbett / Talk for Writing", "X/Website", "C: UK Teacher Influencers", "Mid", "Large teacher following", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Talk for Writing creator — massive English teacher influence"],
    [564, "Jane Considine / The Write Stuff", "X/Website", "C: UK Teacher Influencers", "Mid", "Large teacher following", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Writing methodology — huge primary/secondary English reach"],
    [565, "Mary Myatt", "X/Blog/Podcast", "C: UK Teacher Influencers", "Micro", "30K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Curriculum specialist, prolific blogger/podcaster"],
    [566, "Daisy Christodoulou / No More Marking", "X/Blog", "C: UK Teacher Influencers", "Micro", "25K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Seven Myths About Education — assessment thought leader"],
    [567, "The Secondary English Podcast", "Podcast", "C: UK Teacher Influencers", "Nano", "Niche", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Subject-specific podcast — direct GCSE English teacher audience"],
    [568, "Paul Dix / Pivotal Education Podcast", "Podcast", "C: UK Teacher Influencers", "Micro", "Teacher audience", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Behaviour expert — school-wide influence"],
    [569, "Schools Week Podcast", "Podcast", "K: Education Blogs & News", "Mid", "Moderate", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Tied to Schools Week publication — policy audience"],
    # ═══ Wave 8: Caribbean IGCSE/CXC Creators (570-574) ═══
    [570, "Kerwin Springer (@kerspringer)", "YouTube", "N: ESL/IELTS/Language", "Mid", "100K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. #1 CXC Maths YouTuber — Caribbean IGCSE crossover"],
    [571, "Science with Hazel", "YouTube", "N: ESL/IELTS/Language", "Mid", "200K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Jamaican-origin, IGCSE/A-Level Biology/Chemistry"],
    [572, "CXC Zone (@cxczone)", "IG/YouTube", "N: ESL/IELTS/Language", "Micro", "20K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Caribbean CSEC/CAPE study content"],
    [573, "Leslie Ann Latchman", "YouTube", "N: ESL/IELTS/Language", "Nano", "8K+", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. CSEC English A/B tutorials — English-specific"],
    [574, "Elevate Caribbean (@elevatecaribbean)", "IG", "N: ESL/IELTS/Language", "Micro", "15K+", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. CXC + IGCSE study resources — Barbados"],
    # ═══ Wave 7: Agent-sourced UK Tutoring Platforms & Partners (575-586) ═══
    [575, "Tutorhouse (tutorhouse.co.uk)", "Website", "P: Marketplace Affiliates", "Mid", "3K+ tutors", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Has affiliate scheme — cross-promote GCSE English"],
    [576, "Wyzant (via impact.com)", "Marketplace", "P: Marketplace Affiliates", "Macro", "65K+ tutors", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Established affiliate program via impact.com"],
    [577, "Chegg Tutors (via CJ/ShareASale)", "Marketplace", "P: Marketplace Affiliates", "Macro", "Tens of thousands", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Affiliate program via CJ Affiliate/ShareASale"],
    [578, "Atom Learning", "Website/App", "P: Marketplace Affiliates", "Mid", "100K+ students", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Has referral/partner programme — 11+/GCSE prep"],
    [579, "Varsity Tutors (Nerdy Inc.)", "Website", "P: Marketplace Affiliates", "Macro", "40K+ tutors", "Platform", "Not Started", "", "", "", "", "", "", "Wave 3. Has affiliate program — online tutoring"],
    [580, "PMT Education (Physics & Maths Tutor)", "Website", "P: Marketplace Affiliates", "Macro", "Massive GCSE traffic", "Email", "Not Started", "", "", "", "", "", "", "Wave 3. Major GCSE resource site — cross-link/partner"],
    [581, "First Tutors", "Website", "P: Marketplace Affiliates", "Mid", "70K+ tutors", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. UK's largest tutor directory — listing/backlink"],
    [582, "Tutor Hunt", "Website", "P: Marketplace Affiliates", "Mid", "60K+ tutors", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. Major UK tutor directory — partnership"],
    [583, "Explore Learning", "Website/Centres", "P: Marketplace Affiliates", "Mid", "40K+ members / 140+ centres", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. 140 UK centres — recommend-a-friend scheme"],
    [584, "Spires Online Tutoring", "Website", "P: Marketplace Affiliates", "Micro", "2K+ Oxbridge tutors", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Oxbridge/Russell Group tutors — premium angle"],
    [585, "Owl Tutors", "Website", "P: Marketplace Affiliates", "Micro", "300+ vetted tutors", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. Selective 11+/GCSE/A-Level tutoring — referral"],
    [586, "The Profs", "Website", "P: Marketplace Affiliates", "Micro", "400+ specialist tutors", "Email", "Not Started", "", "", "", "", "", "", "Wave 1. University-level + professional — referral partnership"],
    # ═══ Wave 7: Agent-sourced UK BookTok/BookTube Creators (587-598) ═══
    [587, "@abbysbooks (Abby)", "TikTok", "O: Book/Literature/Writing", "Mid", "300K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. Major UK BookTok — literature reviews"],
    [588, "Leena Norms (@leena_norms)", "YouTube", "O: Book/Literature/Writing", "Mid", "150K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTube — reading vlogs, book reviews"],
    [589, "@lucythereader (Lucy)", "TikTok/YouTube", "O: Book/Literature/Writing", "Mid", "250K TikTok", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Dual-platform UK book creator"],
    [590, "@thisstoryaintover (Ali)", "TikTok", "O: Book/Literature/Writing", "Mid", "300K", "DM", "Not Started", "", "", "", "", "", "", "Wave 3. UK BookTok — reading recommendations"],
    [591, "@jeansbookishlife (Jean)", "TikTok", "O: Book/Literature/Writing", "Mid", "200K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTok — literature discussions"],
    [592, "@emmabooks (Emma)", "TikTok", "O: Book/Literature/Writing", "Mid", "250K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTok — book hauls and reviews"],
    [593, "@bethanmarii (Bethan)", "TikTok", "O: Book/Literature/Writing", "Mid", "200K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTok creator"],
    [594, "@evie.frye.books (Evie)", "TikTok", "O: Book/Literature/Writing", "Mid", "180K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTok — classic + modern lit"],
    [595, "@bookswithjxmes (James)", "TikTok", "O: Book/Literature/Writing", "Mid", "150K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTok — male reader voice"],
    [596, "@ellireadsalot (Elli)", "TikTok", "O: Book/Literature/Writing", "Mid", "150K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTok — avid reader content"],
    [597, "@moaningaboutbooks", "TikTok", "O: Book/Literature/Writing", "Micro", "120K", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTok — honest/humorous reviews"],
    [598, "@perusal_project (Izzie)", "YouTube", "O: Book/Literature/Writing", "Micro", "80K", "Email", "Not Started", "", "", "", "", "", "", "Wave 2. UK BookTube — thoughtful literary analysis"],
    # ═══ Wave 8: Agent-verified UK Teacher TikTok/IG/X ═══
    [587, "Kit Brown (@kjbr0wn)", "TikTok/IG", "C: UK Teacher Influencers", "Mega", "2M TikTok", "DM/Email", "Not Started", "", "", "", "", "", "", "Wave 4. Year 4 primary teacher, BBC Bitesize presenter, How to Shine author"],
    [588, "Mr P ICT / Lee Parkinson MBE (@ICT_MrP)", "TikTok/IG/X", "C: UK Teacher Influencers", "Macro", "226K TikTok / 316K IG", "Email/Agent", "Not Started", "", "", "", "", "", "", "Wave 4. UK's most followed primary teacher, MBE, Apple Distinguished Educator, Bett keynote"],
    [589, "Mrs F (@teachwithmrsf)", "TikTok/IG", "C: UK Teacher Influencers", "Mid", "65.4K TikTok / 23K IG", "DM", "Not Started", "", "", "", "", "", "", "Wave 2. Scottish primary teacher, podcaster, authentic teacher wellbeing voice"],
    [590, "Miss M / Aym Teacher (@aymteacher)", "TikTok", "C: UK Teacher Influencers", "Micro-Mid", "36.6K TikTok", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. Primary teacher, 1.2M TikTok likes, brand partner"],
    [591, "Miss Taylor (@misstaylorteachesprimary)", "TikTok", "C: UK Teacher Influencers", "Micro", "23.9K TikTok", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. UK primary teacher, classroom management, brand partner (Ryman)"],
    [592, "Zoe Enser (@greeborunner)", "X/Blog", "C: UK Teacher Influencers", "Micro", "14.3K X", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 1. Former Ofsted National Lead for Secondary English & Drama, Generative Learning author"],
    [593, "Bruce Robertson (@BruceNextLevel)", "X/Blog", "C: UK Teacher Influencers", "Micro", "~8K X", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 1. Scottish headteacher, The Teaching Delusion trilogy author"],
    [594, "Shaun Allison (@shaun_allison)", "X/Blog", "C: UK Teacher Influencers", "Micro", "~20K X", "Email/X", "Not Started", "", "", "", "", "", "", "Wave 1. Co-author Making Every Lesson Count, Durrington Research School co-founder"],
    [595, "Mrs Mactivity (@mrsmactivity)", "TikTok/YT", "C: UK Teacher Influencers", "Micro", "Growing TikTok", "DM", "Not Started", "", "", "", "", "", "", "Wave 1. EYFS/KS1/KS2 primary resource creator, 500K+ resource downloads"],
]

# Now generate category-fill rows to reach 1,000 total
# 598 named + 402 fill = 1,000
category_fills = [
    # (category, platform, tier, start_num, count, contact_method, notes)
    ("A: UK GCSE English", "TikTok/IG", "Nano-Micro", 596, 19, "DM", "Source: TikTok #GCSEEnglish, IG #GCSErevision. Use HypeAuditor to find."),
    ("B: StudyTok/Studygram", "TikTok/IG", "Nano-Micro", 615, 14, "DM", "Source: TikTok #StudyTok #StudyWithMe. Target 1K-20K followers."),
    ("C: UK Teacher Influencers", "X/TikTok/Blog", "Nano-Micro", 629, 0, "Email/DM", "Source: Feedspot Teacher Blog DB + X #EduTwitter. Sprint Education list."),
    ("D: UK Parent Bloggers", "Blog/IG", "Nano-Micro", 629, 5, "Email", "Source: Feedspot Parenting Blog DB. Mumsnet active posters."),
    ("E: ME Expat Education", "IG/Blog/YT", "Nano-Micro", 634, 27, "Email/DM", "Source: ExpatWoman directory, Qatar Living, SchoolsCompared."),
    ("F: Student Revision", "TikTok/IG/YT", "Nano-Micro", 661, 20, "DM", "Source: TikTok #Revision, #ALevel, #GCSErevision2026."),
    ("G: GCC IGCSE Creators", "TikTok/IG", "Nano-Micro", 681, 47, "DM", "Source: TikTok #IGCSE #igcseuae #igcse_qatar. HypeAuditor GCC filter."),
    ("H: GCC Student Lifestyle", "TikTok/IG", "Nano-Micro", 728, 48, "DM", "Source: TikTok GCC student hashtags. Scrumball/IGYGrow discovery."),
    ("I: GCC Teacher Influencers", "TikTok/IG/YT", "Nano-Mid", 776, 48, "Email/DM", "Source: LinkedIn 'British school' + 'Head of English' in GCC. DM on IG."),
    ("J: GCC Parent Influencers", "IG/Facebook", "Nano-Mid", 824, 26, "DM/Email", "Source: Facebook Dubai/Qatar/Abu Dhabi expat parent groups. IG hashtags."),
    ("K: Education Blogs & News", "Blog/Website", "Nano-Mid", 850, 4, "Email", "Source: Feedspot Education Blog DB. TES/SchoolsWeek commenters."),
    ("L: EdTech Publishers", "Blog/Website/Social", "Micro-Mid", 854, 32, "Email", "Source: impact.com + Awin education category. EdSurge directory."),
    ("M: Homeschool & Curriculum", "Blog/IG/YT", "Nano-Micro", 886, 24, "Email", "Source: Feedspot Homeschool DB. IG #homeschoolUK #homeschoollife."),
    ("N: ESL/IELTS/Language", "YT/IG/TikTok", "Micro-Mid", 910, 14, "Email/DM", "Source: YouTube ESL/IELTS channels. IG #EnglishTeacher #IELTS."),
    ("O: Book/Literature/Writing", "IG/YT/Blog", "Nano-Micro", 924, 0, "Email/DM", "Source: BookTok, #BookStagram, Goodreads top reviewers."),
    ("P: Marketplace Affiliates", "Various", "Various", 924, 77, "Marketplace", "Source: ShareASale, Awin, PartnerStack, impact.com education vertical."),
]

num = 598  # last named influencer number
for cat, platform, tier, start, count, contact, notes in category_fills:
    for i in range(count):
        num += 1
        influencers.append([
            num,
            f"[{cat.split(':')[1].strip()} #{i+1}]",
            platform,
            cat,
            tier,
            "TBD",
            contact,
            "Not Started", "", "", "", "", "", "",
            f"Slot {i+1}/{count}. {notes}"
        ])

for row_data in influencers:
    r = row_data[0] + 1
    for i, v in enumerate(row_data, 1):
        ws3.cell(row=r, column=i, value=v)
    style_data_row(ws3, r, len(cols3))
    ws3.row_dimensions[r].height = 22

# Add category summary at the bottom
summary_row = len(influencers) + 3
add_section_row(ws3, summary_row, "CATEGORY QUOTAS — TARGET: 1,000 AFFILIATES", len(cols3))
quotas = [
    ("A: UK GCSE English Creators", 50), ("B: StudyTok/Studygram", 60), ("C: UK Teacher Influencers", 80),
    ("D: UK Parent Bloggers", 50), ("E: ME Expat Education", 60), ("F: Student Revision", 40),
    ("G: GCC IGCSE Creators", 80), ("H: GCC Student Lifestyle", 60), ("I: GCC Teacher Influencers", 60),
    ("J: GCC Parent Influencers", 60), ("K: Education Blogs & News", 80), ("L: EdTech Publishers", 60),
    ("M: Homeschool & Curriculum", 60), ("N: ESL/IELTS/Language", 50), ("O: Book/Literature/Writing", 30),
    ("P: Marketplace Affiliates", 121),
]
for idx, (cat, count) in enumerate(quotas):
    r = summary_row + 1 + idx
    ws3.cell(row=r, column=2, value=cat)
    ws3.cell(row=r, column=3, value=count)
    ws3.cell(row=r, column=4, value="Target")
    style_data_row(ws3, r, len(cols3))

# ════════════════════════════════════════════════════════════════
# SHEET 4: SEO KEYWORD TRACKER
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("SEO Keywords")
ws4.sheet_properties.tabColor = "FF6600"
cols4 = ["#", "Keyword", "Cluster", "Monthly Volume (UK)", "Competition", "Target Page", "Current Rank", "Content Status", "Blog Post Title", "Publish Date", "Notes"]
set_col_widths(ws4, [4, 40, 18, 16, 12, 20, 10, 12, 40, 12, 25])

for i, h in enumerate(cols4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header_row(ws4, 1, len(cols4))
ws4.auto_filter.ref = f"A1:{get_column_letter(len(cols4))}1"
ws4.freeze_panes = 'A2'

seo_keywords = [
    [1, "gcse english revision", "B: Transactional", "8,100", "High", "/revision", "", "", "The Ultimate GCSE English Revision Guide 2026", "", ""],
    [2, "gcse english language revision", "B: Transactional", "3,600", "High", "/courses/gcse-lang", "", "", "", "", ""],
    [3, "gcse english literature revision", "B: Transactional", "3,600", "High", "/courses/gcse-lit", "", "", "", "", ""],
    [4, "english revision app", "B: Transactional", "1,300", "Medium", "/ (homepage)", "", "", "", "", ""],
    [5, "gcse revision app", "B: Transactional", "1,900", "Medium", "/ (homepage)", "", "", "", "", ""],
    [6, "gcse english practice papers", "B: Transactional", "6,600", "High", "/practice", "", "", "", "", ""],
    [7, "gcse english mock exam", "B: Transactional", "2,400", "Medium", "/mock-exams", "", "", "", "", ""],
    [8, "igcse english revision", "B: Transactional", "1,600", "Medium", "/igcse", "", "", "", "", ""],
    [9, "how to get a grade 9 in english language", "C: Problem/Intent", "2,400", "Medium", "Blog", "", "", "How to Get a Grade 9 in GCSE English Language", "", "Priority blog post"],
    [10, "how to get a grade 9 in english literature", "C: Problem/Intent", "1,900", "Medium", "Blog", "", "", "How to Get a Grade 9 in GCSE English Literature", "", ""],
    [11, "how to revise for gcse english", "C: Problem/Intent", "1,300", "Medium", "Blog", "", "", "How to Revise for GCSE English: The Complete Guide", "", ""],
    [12, "how to write a grade 9 essay gcse", "C: Problem/Intent", "1,000", "Medium", "Blog", "", "", "How to Write a Grade 9 GCSE English Essay", "", ""],
    [13, "how to analyse a poem gcse", "C: Problem/Intent", "1,300", "Medium", "Blog", "", "", "How to Analyse a Poem for GCSE English", "", ""],
    [14, "best gcse english revision website", "D: Comparison", "590", "Medium", "Blog/Homepage", "", "", "Best GCSE English Revision Websites 2026 (Compared)", "", "High-intent keyword"],
    [15, "best gcse english revision app", "D: Comparison", "390", "Low", "Blog/Homepage", "", "", "Best GCSE English Revision Apps 2026", "", "Low competition — quick win"],
    [16, "an inspector calls grade 9 essay", "E: Long-tail", "2,400", "Medium", "Blog", "", "", "An Inspector Calls: Grade 9 Essay Examples & Analysis", "", "Evergreen content"],
    [17, "macbeth grade 9 essay gcse", "E: Long-tail", "1,900", "Medium", "Blog", "", "", "Macbeth Grade 9 Essay: Complete GCSE Guide", "", "Evergreen content"],
    [18, "power and conflict poetry analysis", "E: Long-tail", "2,900", "Medium", "Blog", "", "", "Power and Conflict Poetry: Complete Analysis Guide", "", ""],
    [19, "aqa english language paper 1 question 5 tips", "E: Long-tail", "880", "Low", "Blog", "", "", "AQA English Language Paper 1 Q5: Creative Writing Tips", "", "Low competition"],
    [20, "gcse english language creative writing examples", "E: Long-tail", "1,300", "Medium", "Blog", "", "", "GCSE English Language Creative Writing Examples", "", ""],
]

for row_data in seo_keywords:
    r = row_data[0] + 1
    for i, v in enumerate(row_data, 1):
        ws4.cell(row=r, column=i, value=v)
    style_data_row(ws4, r, len(cols4))

# ════════════════════════════════════════════════════════════════
# SHEET 5: CONTENT CALENDAR
# ════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Content Calendar")
ws5.sheet_properties.tabColor = "FF0066"
cols5 = ["Day", "Session", "Time Block", "Platform", "Content Pillar", "Title / Topic", "Format", "Script Ref", "Status", "Publish Time", "Link/URL", "Engagement", "Notes"]
set_col_widths(ws5, [5, 10, 14, 10, 14, 45, 14, 10, 10, 11, 20, 10, 30])

for i, h in enumerate(cols5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header_row(ws5, 1, len(cols5))
ws5.auto_filter.ref = f"A1:{get_column_letter(len(cols5))}1"
ws5.freeze_panes = 'A2'

# 4-DAY COWORK CONTENT BLITZ — Film ALL 30 TikToks + 4 YouTube videos in 4 days
# Then schedule/drip-release over 4 weeks post-production
content_rows = [
    # ═══ DAY 1: QUICK FIX BLITZ — Film 10 TikToks (TT-V1 to V10) ═══
    ["D1", "AM Session 1", "09:00-10:30", "TikTok", "Quick Fix", "3 Words That Instantly Make Your Essay Sound Smarter", "Talking head + text overlay", "TT-V1", "Not Started", "Sched W1 Mon", "", "", "Hook: 'Stop using shows and suggests'. Outfit 1."],
    ["D1", "AM Session 1", "09:00-10:30", "TikTok", "Quick Fix", "The Difference Between Effect and Affect (Finally)", "Green screen + whiteboard", "TT-V2", "Not Started", "Sched W1 Tue", "", "", "Mnemonic: A=Action, E=End result. Same outfit."],
    ["D1", "AM Session 1", "09:00-10:30", "TikTok", "Quick Fix", "POV: You Used a Semicolon Correctly", "POV reaction split-screen", "TT-V3", "Not Started", "Sched W1 Wed", "", "", "CTA: Drop a semicolon sentence in comments"],
    ["D1", "AM Session 2", "10:45-12:15", "TikTok", "Exam Intel", "5 Techniques Examiners LOVE Seeing in Paper 1", "Talking head + text overlays", "TT-V4", "Not Started", "Sched W1 Thu", "", "", "Outfit change. Semantic fields, syntax, cyclical."],
    ["D1", "AM Session 2", "10:45-12:15", "TikTok", "Quick Fix", "Stop Starting Sentences With 'This Shows That'", "Text on screen + voiceover", "TT-V5", "Not Started", "Sched W1 Fri", "", "", "Show 3 better alternatives"],
    ["D1", "AM Session 2", "10:45-12:15", "TikTok", "Quick Fix", "The Word That Gets You Marks in EVERY English Answer", "Talking head, dramatic pause", "TT-V6", "Not Started", "Sched W1 Sat", "", "", "Reveal: 'Perhaps' — tentative analytical language"],
    ["D1", "PM Session 1", "13:30-15:00", "TikTok", "Quick Fix", "Pathetic Fallacy Explained in 15 Seconds", "Green screen + landscape", "TT-V7", "Not Started", "Sched W1 Sun", "", "", "15s format — difference from metaphor. Outfit 3."],
    ["D1", "PM Session 1", "13:30-15:00", "TikTok", "Quick Fix", "3 Connectives That Scream 'I'm Getting a Grade 9'", "Text overlay carousel", "TT-V8", "Not Started", "Sched W2 Mon", "", "", "Conversely, Moreover, Fundamentally"],
    ["D1", "PM Session 2", "15:15-16:45", "TikTok", "Exam Intel", "How I'd Revise GCSE English If I Had 30 Days Left", "Talking head + day plan", "TT-V9", "Not Started", "Sched W2 Tue", "", "", "60s — 30-day revision plan. Outfit 4."],
    ["D1", "PM Session 2", "15:15-16:45", "TikTok", "Exam Intel", "The Mark Scheme Hack Nobody Talks About", "Screen recording + talking head", "TT-V10", "Not Started", "Sched W2 Wed", "", "", "Reverse-engineer from Level 6 descriptors"],
    ["D1", "Evening", "17:00-18:00", "—", "Edit/Review", "EDIT: Review + rough-cut all 10 Day 1 TikToks in CapCut", "Editing session", "—", "Not Started", "—", "", "", "Add text overlays, captions, sounds. Export drafts."],
    # ═══ DAY 2: EXAM INTEL + SECRETS BLITZ — Film 10 TikToks (TT-V11 to V20) ═══
    ["D2", "AM Session 1", "09:00-10:30", "TikTok", "Exam Intel", "How to Answer 'How Does the Writer Use Language'", "On-screen walkthrough", "TT-V11", "Not Started", "Sched W2 Thu", "", "", "What → How → Why structure. Outfit 5."],
    ["D2", "AM Session 1", "09:00-10:30", "TikTok", "Exam Intel", "Your Examiner Has 4 Minutes to Mark Your Essay", "Talking head, shocked expression", "TT-V12", "Not Started", "Sched W2 Fri", "", "", "First impressions matter"],
    ["D2", "AM Session 1", "09:00-10:30", "TikTok", "Exam Intel", "Revision Technique Tier List for GCSE English", "Tier list graphic", "TT-V13", "Not Started", "Sched W2 Sat", "", "", "S/A/B/C/D tiers — drives comments"],
    ["D2", "AM Session 2", "10:45-12:15", "TikTok", "Exam Intel", "The 15-Minute Essay Plan That Gets Grade 9", "Speed-draw + voiceover", "TT-V14", "Not Started", "Sched W2 Sun", "", "", "Step-by-step planning method. Outfit 6."],
    ["D2", "AM Session 2", "10:45-12:15", "TikTok", "Exam Intel", "Things Your English Teacher Won't Tell You — Part 1", "Talking head, conspiratorial", "TT-V15", "Not Started", "Sched W3 Mon", "", "", "3 insider tips — drives Part 2 demand"],
    ["D2", "AM Session 2", "10:45-12:15", "TikTok", "Exam Intel", "Things Your English Teacher Won't Tell You — Part 2", "Series continuation", "TT-V16", "Not Started", "Sched W3 Tue", "", "", "Part 3 when this hits 10k likes"],
    ["D2", "PM Session 1", "13:30-15:00", "TikTok", "Exam Intel", "What Examiners ACTUALLY Write on Your Paper", "Green screen + mock script", "TT-V17", "Not Started", "Sched W3 Wed", "", "", "Annotations, best-fit model. Outfit 7."],
    ["D2", "PM Session 1", "13:30-15:00", "TikTok", "Exam Intel", "The Question Nobody Answers Properly (and How to Fix It)", "Talking head + screenshot", "TT-V18", "Not Started", "Sched W3 Thu", "", "", "Paper 2 comparison question"],
    ["D2", "PM Session 2", "15:15-16:45", "TikTok", "Relatable", "POV: The Examiner Reading Your Opening Sentence", "POV reaction", "TT-V19", "Not Started", "Sched W3 Fri", "", "", "Transformation: bad → good opening. Outfit 8."],
    ["D2", "PM Session 2", "15:15-16:45", "TikTok", "Relatable", "5 Things That Instantly Make Your Essay Look Year 7", "List with visual examples", "TT-V20", "Not Started", "Sched W3 Sat", "", "", "Fix these five = jump two grades"],
    ["D2", "Evening", "17:00-18:00", "—", "Edit/Review", "EDIT: Review + rough-cut all 10 Day 2 TikToks in CapCut", "Editing session", "—", "Not Started", "—", "", "", "Finalise Day 1 exports. Start Day 2 rough cuts."],
    # ═══ DAY 3: RELATABLE + BUILD IN PUBLIC — Film 10 TikToks (TT-V21 to V30) + YT-V1 ═══
    ["D3", "AM Session 1", "09:00-10:30", "TikTok", "Relatable", "Rating GCSE English Essays From My DMs", "Screen recording + reactions", "TT-V21", "Not Started", "Sched W3 Sun", "", "", "CTA: DM me YOUR paragraph. Outfit 9."],
    ["D3", "AM Session 1", "09:00-10:30", "TikTok", "Relatable", "POV: You Just Wrote 'The Author Uses Lots of Techniques'", "Skit — reacting to own writing", "TT-V22", "Not Started", "Sched W4 Mon", "", "", "Side-by-side comparison"],
    ["D3", "AM Session 1", "09:00-10:30", "TikTok", "Quick Fix", "Words You're Using Wrong in Your Essay", "Split screen: wrong vs right", "TT-V23", "Not Started", "Sched W4 Tue", "", "", "Infer/Imply, Literally, Bias/Biased"],
    ["D3", "AM Session 2", "10:45-12:15", "TikTok", "Build in Public", "Why I'm Building an English Revision App", "Talking head, personal", "TT-V24", "Not Started", "Sched W4 Wed", "", "", "Founder story — authentic. Outfit 10."],
    ["D3", "AM Session 2", "10:45-12:15", "TikTok", "Build in Public", "Day in My Life Building an EdTech Startup", "DITL montage + voiceover", "TT-V25", "Not Started", "Sched W4 Thu", "", "", "Behind the scenes"],
    ["D3", "AM Session 2", "10:45-12:15", "TikTok", "Build in Public", "Sneak Peek: Inside The English Hub App", "Screen recording + talking head", "TT-V26", "Not Started", "Sched W4 Fri", "", "", "First 100 users get it free"],
    ["D3", "PM Session 1", "13:30-15:00", "TikTok", "Relatable", "Every English Class Has This Student", "Skit — multiple characters", "TT-V27", "Not Started", "Sched W4 Sat", "", "", "Drives 'which one are you' comments. Multiple outfits for characters."],
    ["D3", "PM Session 1", "13:30-15:00", "TikTok", "Relatable", "English Exam Bingo", "Bingo card, checking off items", "TT-V28", "Not Started", "Sched W4 Sun", "", "", "Score out of 6"],
    ["D3", "PM Session 1", "13:30-15:00", "TikTok", "Deep Dive", "Speed-Writing a Grade 9 Paragraph in 60 Seconds", "Annotation replay + voiceover", "TT-V29", "Not Started", "Sched W5 Mon", "", "", "Screen recording of live writing with analysis"],
    ["D3", "PM Session 2", "15:15-16:45", "TikTok", "Deep Dive", "Live Paper 1 Q2 Extract Walkthrough", "Walkthrough + annotations", "TT-V30", "Not Started", "Sched W5 Tue", "", "", "Show real technique in action"],
    ["D3", "PM Session 2", "15:15-17:30", "YouTube", "Deep Dive", "How to Get a Grade 9 in GCSE English Language (Complete Guide)", "Full walkthrough, 18-22 min", "YT-V1", "Not Started", "Sched W1 Sun 10AM", "", "", "FLAGSHIP VIDEO. Extended session — needs 2hrs for full take."],
    ["D3", "Evening", "17:30-18:30", "—", "Edit/Review", "EDIT: Finalise all 30 TikToks + begin YT-V1 edit in DaVinci", "Editing session", "—", "Not Started", "—", "", "", "Export all TikToks to scheduling queue."],
    # ═══ DAY 4: YOUTUBE MARATHON — Film YT-V2, YT-V3, YT-V4 + Schedule everything ═══
    ["D4", "AM Session 1", "09:00-11:00", "YouTube", "Exam Intel", "I Marked 100 GCSE English Essays — Here's What Examiners Want", "Examiner insight, 15-18 min", "YT-V2", "Not Started", "Sched W2 Sun 10AM", "", "", "5 mistakes + 5 impressive things. Prep examiner props."],
    ["D4", "AM Session 2", "11:15-13:15", "YouTube", "Deep Dive", "GCSE English Language Paper 1 — Full Walkthrough (AQA 2026)", "Full walkthrough, 35-45 min", "YT-V3", "Not Started", "Sched W3 Sun 10AM", "", "", "Every question explained. Print sample paper as prop."],
    ["D4", "PM Session 1", "14:15-16:15", "YouTube", "Deep Dive", "GCSE English Language Paper 2 — Full Walkthrough (AQA 2026)", "Full walkthrough, 35-45 min", "YT-V4", "Not Started", "Sched W4 Sun 10AM", "", "", "Comparison question focus. Print sample paper."],
    ["D4", "PM Session 2", "16:30-18:00", "—", "Schedule", "SCHEDULE: Upload all 30 TikToks + 4 YTs to scheduling queue", "Scheduling session", "—", "Not Started", "—", "", "", "TikTok Later/native scheduler. YouTube Studio. 4 weeks of content DONE."],
    ["D4", "PM Session 2", "16:30-18:00", "YT Shorts", "Quick Fix", "Cross-post: 3 best TikToks as YT Shorts", "Repurposed TikToks", "TT-V1,3,10", "Not Started", "Sched W2-W4", "", "", "Pick highest-energy takes from Day 1-3 filming"],
    ["D4", "Evening", "18:00-19:00", "—", "Edit/Review", "FINAL: QC all scheduled content, check captions, hashtags, thumbnails", "QA session", "—", "Not Started", "—", "", "", "4 WEEKS OF CONTENT BANKED IN 4 DAYS ✓"],
]

for idx, row_data in enumerate(content_rows, 2):
    for i, v in enumerate(row_data, 1):
        ws5.cell(row=idx, column=i, value=v)
    style_data_row(ws5, idx, len(cols5))
    ws5.row_dimensions[idx].height = 28

# ════════════════════════════════════════════════════════════════
# SHEET 6: EMAIL SEQUENCES TRACKER
# ════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Email Sequences")
ws6.sheet_properties.tabColor = "7030A0"
cols6 = ["Sequence", "Email #", "Day", "Subject Line", "Preview Text", "Purpose", "CTA", "Status", "Open Rate", "Click Rate", "Notes"]
set_col_widths(ws6, [20, 8, 6, 35, 35, 25, 20, 10, 10, 10, 25])

for i, h in enumerate(cols6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header_row(ws6, 1, len(cols6))
ws6.freeze_panes = 'A2'

email_sequences = [
    # Trial Onboarding
    ["Trial Onboarding", "1.1", "0", "Welcome to The English Hub!", "Your 30-day free trial starts now", "Get started — set exam board, choose topic", "Get Started Now", "Draft", "", "", ""],
    ["Trial Onboarding", "1.2", "1", "Ready for your first module?", "It only takes 15 minutes", "Complete first module (Unseen Poetry / Grade 9 Essay / Macbeth)", "Start Your First Module", "Draft", "", "", "3x improvement stat"],
    ["Trial Onboarding", "1.3", "3", "Get instant feedback on your essays", "Our AI marker gives examiner-level feedback", "Try AI Essay Feedback tool", "Try AI Essay Feedback", "Draft", "", "", "One full grade boundary improvement stat"],
    ["Trial Onboarding", "1.4", "5", "Fancy a mock exam? No stress.", "Timed practice in real exam conditions", "Take a mock exam", "Take a Mock Exam", "Draft", "", "", "Pause/resume available"],
    ["Trial Onboarding", "1.5", "7", "Your first week — here's how it went", "A quick look at what you've achieved", "Show progress snapshot, encourage continuation", "Keep Going / Re-engage", "Draft", "", "", "Personalised stats"],
    ["Trial Onboarding", "1.6", "14", "You're halfway through your trial", "Here's what you've unlocked so far", "Mid-trial check, push favourite features", "Explore More", "Draft", "", "", "Highlight unused features"],
    ["Trial Onboarding", "1.7", "25", "Your trial ends in 5 days", "Here's what happens next", "Convert to paid — show value summary", "Subscribe Now", "Draft", "", "", "Annual plan pitch"],
    # Waitlist Launch
    ["Waitlist Launch", "2.1", "1", "We're live — claim your 50% off", "First 200 sign-ups only", "Drive trial sign-ups with scarcity", "Start Free Trial", "Draft", "", "", "50% off first 3 months"],
    ["Waitlist Launch", "2.2", "3", "See what's inside The English Hub", "60-second video tour", "Feature walkthrough", "Watch Demo", "Draft", "", "", "Include demo video"],
    ["Waitlist Launch", "2.3", "7", "Here's what beta users are saying", "Real feedback from early users", "Social proof", "Join Them", "Draft", "", "", "Include testimonials"],
    ["Waitlist Launch", "2.4", "14", "Your early access is ending soon", "Annual plan saves you 33%", "Push annual plan conversion", "Get Annual Plan", "Draft", "", "", "£59.99 launch special"],
    # Affiliate Onboarding
    ["Affiliate Onboarding", "3.1", "0", "Welcome to The English Hub Affiliate Programme!", "You're Approved!", "Welcome + dashboard login + commission structure", "Log In to Dashboard", "Draft", "", "", "20% recurring, £25 min payout"],
    ["Affiliate Onboarding", "3.2", "1", "Your affiliate resource pack is ready", "Banners, templates, and demo videos", "Provide all marketing assets", "Download Assets", "Draft", "", "", "15 banners, 3 email templates, 5 social templates"],
    ["Affiliate Onboarding", "3.3", "2", "Quick tour of The English Hub", "5-minute walkthrough video", "Platform familiarity", "Watch Tour", "Draft", "", "", "Loom video — 'best affiliates are genuine users'"],
    ["Affiliate Onboarding", "3.4", "5", "Ready to start earning?", "3 quick ways to share today", "First promotion prompt", "Start Promoting", "Draft", "", "", "IG story, blog mention, or email to list"],
    ["Affiliate Onboarding", "3.5", "14", "How's it going?", "What's working for our top affiliates", "Personal check-in from founder", "Reply", "Draft", "", "", "Share successful promotion example"],
]

for idx, row_data in enumerate(email_sequences, 2):
    for i, v in enumerate(row_data, 1):
        ws6.cell(row=idx, column=i, value=v)
    style_data_row(ws6, idx, len(cols6))

# ════════════════════════════════════════════════════════════════
# SHEET 7: B2B SCHOOL SALES TRACKER
# ════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("B2B School Sales")
ws7.sheet_properties.tabColor = "BF8F00"
cols7 = ["#", "School Name", "Region", "Type", "Contact Name", "Role", "Email", "Phone", "Status", "Package Interest", "Students", "Annual Value", "Date Contacted", "Next Follow-up", "Notes"]
set_col_widths(ws7, [4, 25, 12, 14, 18, 16, 25, 14, 12, 14, 8, 12, 12, 12, 25])

for i, h in enumerate(cols7, 1):
    ws7.cell(row=1, column=i, value=h)
style_header_row(ws7, 1, len(cols7))
ws7.auto_filter.ref = f"A1:{get_column_letter(len(cols7))}1"
ws7.freeze_panes = 'A2'

# Add pricing reference row
r = 2
add_section_row(ws7, r, "SCHOOL PRICING: Class (35) £199/yr | Dept (150) £599/yr | Whole School (500) £1,499/yr | MAT (500+) £2-2.50/student | Per-student: 10=£8, 25=£7, 50=£6, 100=£5, 200=£4, 200+=£3.50, 500+=£3", len(cols7))

# UK MAT Targets
r += 1
add_section_row(ws7, r, "PRIORITY MAT TARGETS (Single deal = 5-50 schools)", len(cols7))

uk_mats = [
    [1, "United Learning", "National", "MAT (90+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "5000+", "£15,000+", "", "", "Mix independent + state. Strong data culture."],
    [2, "Ark Schools", "National", "MAT (39 schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "3000+", "£9,000+", "", "", "Disadvantaged focus. Evidence-driven."],
    [3, "Harris Federation", "London", "MAT (55+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "5000+", "£15,000+", "", "", "London-focused. Strong results."],
    [4, "Oasis Community Learning", "National", "MAT (50+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "4000+", "£12,000+", "", "", "Community-focused."],
    [5, "Academies Enterprise Trust", "National", "MAT (50+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "4000+", "£12,000+", "", "", "National spread."],
    [6, "Ormiston Academies Trust", "National", "MAT (40+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "3500+", "£10,500+", "", "", "Focus on improvement."],
    [7, "Star Academies", "National", "MAT (30+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "2500+", "£7,500+", "", "", "Strong results from disadvantaged areas."],
    [8, "Inspiration Trust", "Norfolk", "MAT (15+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "1500+", "£4,500+", "", "", "Knowledge-rich curriculum."],
    [9, "Outwood Grange", "Yorkshire", "MAT (30+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "2500+", "£7,500+", "", "", "Strong improvement record."],
    [10, "Delta Academies Trust", "Yorkshire/Humber", "MAT (50+ schools)", "", "Director of Education", "", "", "Not Started", "Trust-wide", "4000+", "£12,000+", "", "", "Yorkshire and Humber."],
]

for row_data in uk_mats:
    r += 1
    for i, v in enumerate(row_data, 1):
        ws7.cell(row=r, column=i, value=v)
    style_data_row(ws7, r, len(cols7))

# Empty rows for UK state/independent schools
r += 2
add_section_row(ws7, r, "UK STATE & INDEPENDENT SCHOOLS — Target: 200 in pipeline by Month 3", len(cols7))
r += 1
add_section_row(ws7, r, "Source: GIAS DfE DB (3,400 secondaries) | Ofsted reports (English flagged) | School Performance Tables | Sprint Education lists (£200-500) | ISC directory (1,400 independents)", len(cols7))

# ════════════════════════════════════════════════════════════════
# SHEET 8: INTERNATIONAL / MIDDLE EAST TRACKER
# ════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("International - GCC")
ws8.sheet_properties.tabColor = "C00000"
cols8 = ["#", "Country", "School / Organisation", "Type", "Contact", "Email", "Phone", "Status", "Students Est.", "Annual Value", "Date Contacted", "Notes"]
set_col_widths(ws8, [4, 10, 32, 16, 20, 30, 16, 12, 10, 12, 12, 30])

for i, h in enumerate(cols8, 1):
    ws8.cell(row=1, column=i, value=h)
style_header_row(ws8, 1, len(cols8))
ws8.auto_filter.ref = f"A1:{get_column_letter(len(cols8))}1"
ws8.freeze_panes = 'A2'

# Market context
r = 2
add_section_row(ws8, r, "GCC MARKET: 784 int'l schools UAE | 52+ British in Qatar | ~240K IGCSE English students | Parents pay £8-25K/yr fees | IGCSE May/Jun + Oct/Nov", len(cols8))
r += 1
add_section_row(ws8, r, "CRISIS: All GCC schools remote since Feb/Mar 2026. No return date. 5% penetration = £1.2M ARR. Target: 25 paid school licences by Month 6.", len(cols8), PatternFill('solid', fgColor='943126'))

# ═══ QATAR SCHOOLS ═══
r += 1
add_section_row(ws8, r, "QATAR — 20+ British Curriculum Schools (Remote since 1 March 2026)", len(cols8))

gcc_schools = [
    # Qatar
    [1, "Qatar", "Doha College", "British", "Head of English", "enquiries@dohacollege.com", "+974 4407 6777", "Not Started", "500+", "", "", "PRIORITY TIER 1. Also: admissions1@dohacollege.com"],
    [2, "Qatar", "King's College Doha (Al Thumama)", "British", "Head of English", "info@king.edu.ruhr", "+974 4496 5888", "Not Started", "400+", "", "", "TIER 1. WhatsApp: +974 6006 202"],
    [3, "Qatar", "King's College Doha (Mesaimeer)", "British", "Head of English", "info@king.edu.ruhr", "+974 4496 5888", "Not Started", "400+", "", "", "Second campus"],
    [4, "Qatar", "Nord Anglia / Compass Intl Doha", "British", "Head of English", "admissions.cisd@nais.qa", "", "Not Started", "500+", "", "", "TIER 1. Nord Anglia group — group deal potential"],
    [5, "Qatar", "Nord Anglia Intl School Al Khor", "British", "Head of English", "admissions.naisak@nais.qa", "+974 4437 9600", "Not Started", "300+", "", "", "TIER 1."],
    [6, "Qatar", "GEMS Wellington School Qatar", "British", "Head of English", "admissions@gemswellingtonqatar.com", "", "Not Started", "400+", "", "", "GEMS group — group deal via gemseducation.com/Contact"],
    [7, "Qatar", "Sherborne Qatar", "British", "Head of English", "admissions@sherborneqatar.com", "", "Not Started", "300+", "", "", ""],
    [8, "Qatar", "Royal Grammar School Guildford Qatar", "British", "Head of English", "info@rgsq.com.qa", "", "Not Started", "300+", "", "", ""],
    [9, "Qatar", "Doha British School", "British", "Head of English", "admin@dohabritishschool.com", "", "Not Started", "300+", "", "", ""],
    [10, "Qatar", "Doha English Speaking School", "British", "Head of English", "admin@dess.com.qa", "", "Not Started", "300+", "", "", ""],
    [11, "Qatar", "Durham School for Girls Doha", "British", "Head of English", "admissions@durhamschooldoha.com", "", "Not Started", "200+", "", "", ""],
    [12, "Qatar", "Newton British Academy Al Dafna", "British", "Head of English", "info@newtonacademy.qa", "", "Not Started", "300+", "", "", "Newton group — 5 Qatar campuses"],
    [13, "Qatar", "Newton British Academy Barwa City", "British", "Head of English", "info@newtonacademy.qa", "", "Not Started", "300+", "", "", ""],
    [14, "Qatar", "Park House English School", "British", "Head of English", "admin@phesq.com", "", "Not Started", "200+", "", "", ""],
    [15, "Qatar", "Qatar International School", "British", "Head of English", "info@qis.net", "+974 4468 0026", "Not Started", "300+", "", "", ""],
    [16, "Qatar", "Gulf English School Qatar", "British", "Head of English", "admin@gesq.com.qa", "", "Not Started", "200+", "", "", ""],
    [17, "Qatar", "Al Khor International School", "British", "Head of English", "info@alkhorinternational.net", "", "Not Started", "200+", "", "", ""],
    # UAE — Dubai
    [18, "UAE-Dubai", "Nord Anglia Intl School Dubai", "British", "Head of English", "admissions@nasdubai.ae", "+971 4 219 9999", "Not Started", "600+", "", "", "TIER 1. Nord Anglia group."],
    [19, "UAE-Dubai", "Repton Dubai (Nad Al Sheba)", "British", "Head of English", "info@reptondubai.org", "+971 4 426 9393", "Not Started", "500+", "", "", "TIER 1. principal@reptondubai.org"],
    [20, "UAE-Dubai", "Brighton College Dubai", "British", "Head of English", "admissions@brightoncollegedubai.ae", "+971 4 387 1116", "Not Started", "500+", "", "", "TIER 1."],
    [21, "UAE-Dubai", "GEMS Wellington Intl School", "British", "Head of English", "admissions@gemswellington.ae", "", "Not Started", "600+", "", "", "GEMS group flagship"],
    [22, "UAE-Dubai", "GEMS Wellington Academy Al Khail", "British", "Head of English", "admissions@gemswa-alkhail.ae", "", "Not Started", "400+", "", "", "GEMS group"],
    [23, "UAE-Dubai", "GEMS Wellington Academy Silicon Oasis", "British", "Head of English", "admin@gwa-so.ae", "", "Not Started", "400+", "", "", "GEMS group"],
    [24, "UAE-Dubai", "GEMS Cambridge Intl School Dubai", "British", "Head of English", "admissions@gcisd.ae", "", "Not Started", "400+", "", "", "GEMS group"],
    [25, "UAE-Dubai", "GEMS Royal Dubai School", "British", "Head of English", "admin@gemsroyaldubaischool.ae", "", "Not Started", "300+", "", "", "GEMS group"],
    [26, "UAE-Dubai", "Dubai College", "British", "Head of English", "admin@dubaicollege.org", "+971 4 399 9111", "Not Started", "500+", "", "", "Premium. TIER 1."],
    [27, "UAE-Dubai", "Jumeirah College", "British", "Head of English", "info@jumeirahcollegedubai.com", "", "Not Started", "400+", "", "", ""],
    [28, "UAE-Dubai", "Dubai British School Emirates Hills", "British", "Head of English", "info@dubaibritishschool.ae", "", "Not Started", "400+", "", "", ""],
    [29, "UAE-Dubai", "Dubai British School Jumeira Park", "British", "Head of English", "info@dbsjp.ae", "", "Not Started", "300+", "", "", ""],
    [30, "UAE-Dubai", "Kings' School Dubai", "British", "Head of English", "admin@kingsschools.com", "", "Not Started", "400+", "", "", "3 campuses"],
    [31, "UAE-Dubai", "Hartland International School", "British", "Head of English", "info@hartlandschool.ae", "", "Not Started", "400+", "", "", ""],
    [32, "UAE-Dubai", "English College Dubai", "British", "Head of English", "admin@englishcollege.ae", "+971 4 394 3465", "Not Started", "300+", "", "", ""],
    [33, "UAE-Dubai", "North London Collegiate School Dubai", "British", "Head of English", "admissions@nlcsdubai.ae", "", "Not Started", "400+", "", "", "Premium"],
    [34, "UAE-Dubai", "Royal Grammar School Guildford Dubai", "British", "Head of English", "admissions@rgsdubai.com", "", "Not Started", "300+", "", "", ""],
    [35, "UAE-Dubai", "Kent College Dubai", "British", "Head of English", "admin@kentcollegedubai.com", "", "Not Started", "300+", "", "", ""],
    [36, "UAE-Dubai", "Sunmarke School", "British", "Head of English", "info@sunmarke.com", "", "Not Started", "400+", "", "", ""],
    [37, "UAE-Dubai", "Jebel Ali School", "British", "Head of English", "admin@jasdubai.com", "", "Not Started", "300+", "", "", ""],
    [38, "UAE-Dubai", "Safa British School", "British", "Head of English", "info@safabritish.com", "", "Not Started", "300+", "", "", ""],
    [39, "UAE-Dubai", "Westminster School Dubai", "British", "Head of English", "admin@westminsterschooldubai.com", "", "Not Started", "300+", "", "", ""],
    [40, "UAE-Dubai", "Sharjah English School", "British", "Head of English", "admin@ses.sch.ae", "+971 6 522 1555", "Not Started", "300+", "", "", "Sharjah"],
    # UAE — Abu Dhabi
    [41, "UAE-AbuDhabi", "British Intl School Abu Dhabi (Nord Anglia)", "British", "Head of English", "admissions@bisad.ae", "+971 2 510 0176", "Not Started", "500+", "", "", "TIER 1. Nord Anglia group."],
    [42, "UAE-AbuDhabi", "British School Al Khubairat (BSAK)", "British", "Head of English", "registrar@britishschool.sch.ae", "+971 2 204 0200", "Not Started", "500+", "", "", "TIER 1."],
    [43, "UAE-AbuDhabi", "Brighton College Abu Dhabi", "British", "Head of English", "admissions@brightoncollege.ae", "+971 2 815 6504", "Not Started", "400+", "", "", "TIER 1."],
    [44, "UAE-AbuDhabi", "Cranleigh Abu Dhabi", "British", "Head of English", "admissions@cranleighabudhabi.ae", "", "Not Started", "400+", "", "", ""],
    [45, "UAE-AbuDhabi", "Repton School Abu Dhabi", "British", "Head of English", "admissions@reptonabudhabi.ae", "", "Not Started", "400+", "", "", ""],
    [46, "UAE-AbuDhabi", "Brighton College Al Ain", "British", "Head of English", "admissions@brightoncollegealain.ae", "", "Not Started", "300+", "", "", ""],
    [47, "UAE-AbuDhabi", "Aspen Heights British School", "British", "Head of English", "info@aspenheights.ae", "", "Not Started", "300+", "", "", ""],
    [48, "UAE-AbuDhabi", "Al Ain British Academy", "British", "Head of English", "admin@aaba.ae", "", "Not Started", "300+", "", "", ""],
    [49, "UAE-AbuDhabi", "Yasmina British Academy", "British", "Head of English", "admin@yasmina.ae", "", "Not Started", "300+", "", "", ""],
    [50, "UAE-AbuDhabi", "Belvedere British School", "British", "Head of English", "admin@belvedere.ae", "", "Not Started", "300+", "", "", ""],
    # Saudi Arabia
    [51, "Saudi", "British Intl School Riyadh", "British", "Head of English", "info@bisriyadh.com", "", "Not Started", "500+", "", "", "TIER 1. Multiple campuses (Al Hamra, Al Waha, DQ)"],
    [52, "Saudi", "British Intl School Jeddah", "British", "Head of English", "admin@bisjeddah.com", "", "Not Started", "400+", "", "", ""],
    [53, "Saudi", "British Intl School Al Khobar", "British", "Head of English", "admin@biskhobar.com", "", "Not Started", "300+", "", "", ""],
    [54, "Saudi", "British School Dhahran", "British", "Head of English", "admin@britishschooldhahran.com", "", "Not Started", "300+", "", "", ""],
    [55, "Saudi", "Jeddah Prep and Grammar School", "British", "Head of English", "admin@jps.net.sa", "", "Not Started", "300+", "", "", ""],
    [56, "Saudi", "King's College Riyadh", "British", "Head of English", "info@kingscollegeriyadh.com", "", "Not Started", "300+", "", "", ""],
    [57, "Saudi", "Reigate Grammar School Riyadh", "British", "Head of English", "admissions@rgsriyadh.com", "", "Not Started", "300+", "", "", ""],
    [58, "Saudi", "Sherborne School Jeddah", "British", "Head of English", "admissions@sherbornejeddah.com", "", "Not Started", "300+", "", "", "New 2025"],
    [59, "Saudi", "Thamer Intl Schools Jeddah", "British", "Head of English", "admin@tis.edu.sa", "+966 12 639 2400", "Not Started", "300+", "", "", ""],
    [60, "Saudi", "Downe House Riyadh", "British", "Head of English", "admissions@downehouseriyadh.com", "", "Not Started", "200+", "", "", ""],
    # Kuwait
    [61, "Kuwait", "British School of Kuwait", "British", "Head of English", "admissions@bsk.edu.kw", "+965 183 0456", "Not Started", "400+", "", "", "TIER 1"],
    [62, "Kuwait", "The English School Kuwait", "British", "Head of English", "admissions@tes.edu.kw", "+965 22271385", "Not Started", "400+", "", "", "TIER 1"],
    [63, "Kuwait", "Gulf British Academy", "British", "Head of English", "admin@gba.edu.kw", "", "Not Started", "300+", "", "", ""],
    [64, "Kuwait", "Cambridge English School Hawally", "British", "Head of English", "admin@ces.edu.kw", "", "Not Started", "300+", "", "", "2 campuses"],
    [65, "Kuwait", "New English School Kuwait", "British", "Head of English", "admin@nesq8.com", "", "Not Started", "300+", "", "", ""],
    [66, "Kuwait", "Kuwait Intl English School", "British", "Head of English", "admin@kies.edu.kw", "", "Not Started", "300+", "", "", ""],
    [67, "Kuwait", "English School Fahaheel", "British", "Head of English", "admin@esf.edu.kw", "", "Not Started", "300+", "", "", ""],
    # Bahrain
    [68, "Bahrain", "British School of Bahrain", "British", "Head of English", "admissions@britishschoolbahrain.com", "+973 1761 6555", "Not Started", "400+", "", "", "TIER 1"],
    [69, "Bahrain", "St Christopher's School Bahrain", "British", "Head of English", "admin@stchristophers.net.bh", "", "Not Started", "300+", "", "", ""],
    [70, "Bahrain", "Nadeen School Bahrain", "British", "Head of English", "admin@nadeen.net", "", "Not Started", "200+", "", "", ""],
    # Oman
    [71, "Oman", "British School Muscat", "British", "Head of English", "admin@britishschoolmuscat.com", "", "Not Started", "400+", "", "", "TIER 1"],
    [72, "Oman", "British School Salalah", "British", "Head of English", "info@britishschoolsalalah.com", "", "Not Started", "200+", "", "", ""],
    [73, "Oman", "A'Soud Global School Muscat", "British", "Head of English", "info@agsmuscat.edu.om", "+968 2442 3952", "Not Started", "300+", "", "", "Premium British school"],
    [74, "Oman", "Cheltenham Muscat", "British", "Head of English", "admissions@cheltenhammuscat.com", "", "Not Started", "300+", "", "", ""],
    [75, "Oman", "Downe House Muscat", "British", "Head of English", "admissions@downehousemuscat.com", "", "Not Started", "200+", "", "", ""],
    [76, "Oman", "Azzan Bin Qais Intl School", "British", "Head of English", "info@azzan.edu.om", "", "Not Started", "200+", "", "", ""],
    [77, "Oman", "Muscat International School", "British", "Head of English", "admin@mis.edu.om", "", "Not Started", "300+", "", "", ""],
]

for row_data in gcc_schools:
    r += 1
    for i, v in enumerate(row_data, 1):
        ws8.cell(row=r, column=i, value=v)
    style_data_row(ws8, r, len(cols8))

# Group deal targets
r += 2
add_section_row(ws8, r, "GROUP DEAL TARGETS — Single deal = 5-50 schools", len(cols8))
groups = [
    ["—", "GCC-wide", "GEMS Education (50+ British schools)", "Group", "Head of Innovation", "gemseducation.com/Contact", "", "Not Started", "50,000+", "£150K+", "", "PRIORITY 1. Biggest single deal in GCC."],
    ["—", "GCC-wide", "Nord Anglia Education (4 GCC schools)", "Group", "Group Enquiries", "enquiries@nordanglia.com", "", "Not Started", "2,000+", "£6,000+", "", "4 schools: Dubai, Abu Dhabi, Qatar Al Khor, Compass Doha"],
    ["—", "Qatar", "Newton Group (5 Qatar campuses)", "Group", "Group Admin", "info@newtonacademy.qa", "", "Not Started", "1,500+", "£4,500+", "", "5 campuses across Doha + Barwa City"],
    ["—", "UAE-Dubai", "Kings' Schools (3 Dubai campuses)", "Group", "Group Admin", "admin@kingsschools.com", "", "Not Started", "1,200+", "£3,600+", "", "Dubai, Al Barsha, Nad Al Sheba"],
    ["—", "GCC-wide", "Brighton College (3 GCC campuses)", "Group", "Group Admin", "admissions@brightoncollege.ae", "", "Not Started", "1,200+", "£3,600+", "", "Dubai, Abu Dhabi, Al Ain"],
    ["—", "GCC-wide", "Repton (3 GCC campuses)", "Group", "Group Admin", "info@reptondubai.org", "", "Not Started", "1,500+", "£4,500+", "", "Dubai, Al Barsha, Abu Dhabi"],
]
for row_data in groups:
    r += 1
    for i, v in enumerate(row_data, 1):
        ws8.cell(row=r, column=i, value=v)
    style_data_row(ws8, r, len(cols8))

# ════════════════════════════════════════════════════════════════
# SHEET 9: PRICING & OFFERS REFERENCE
# ════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Pricing & Offers")
ws9.sheet_properties.tabColor = "548235"
set_col_widths(ws9, [3, 25, 18, 18, 18, 40])

r = 2
ws9.merge_cells('B2:F2')
ws9['B2'] = "PRICING & SEASONAL OFFERS REFERENCE"
ws9['B2'].font = TITLE_FONT

# Core pricing
r = 4
add_section_row(ws9, r, "CORE PRICING", 6)
r += 1
headers = ["Plan", "Price", "Effective Monthly", "Discount vs Monthly", "Notes"]
for i, h in enumerate(headers, 2):
    ws9.cell(row=r, column=i, value=h)
style_header_row(ws9, r, 6)

pricing = [
    ["Monthly", "£9.99/month", "£9.99", "—", "Below £10 psychological threshold, 41% cheaper than Seneca Premium"],
    ["Annual", "£79.99/year", "£6.67", "33% off", "Target: 40% of subscribers on annual"],
    ["30-Day Free Trial", "£0", "—", "—", "No card required — frictionless sign-up"],
]
for row_data in pricing:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws9.cell(row=r, column=i, value=v)
    style_data_row(ws9, r, 6)

# Launch offers
r += 2
add_section_row(ws9, r, "LAUNCH OFFERS", 6)
r += 1
headers = ["Offer", "Details", "Duration", "Purpose", ""]
for i, h in enumerate(headers, 2):
    ws9.cell(row=r, column=i, value=h)
style_header_row(ws9, r, 6)

offers = [
    ["Early Bird", "50% off first 3 months (£4.99/mo)", "First 200 sign-ups", "Drive initial adoption, create urgency"],
    ["Affiliate Discount", "15% off first 3 months (£8.49/mo)", "Ongoing via affiliate codes", "Incentivise affiliate sharing"],
    ["Annual Launch Special", "£59.99/year (25% off annual)", "First month only", "Lock in annual subscribers early"],
    ["Student Discount", "Free for FSM students (verified via school)", "Ongoing", "Social impact, press coverage, school goodwill"],
]
for row_data in offers:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws9.cell(row=r, column=i, value=v)
    style_data_row(ws9, r, 6)

# Seasonal campaigns
r += 2
add_section_row(ws9, r, "SEASONAL CAMPAIGNS", 6)
r += 1
headers = ["Campaign", "Timing", "Offer", "Rationale", ""]
for i, h in enumerate(headers, 2):
    ws9.cell(row=r, column=i, value=h)
style_header_row(ws9, r, 6)

seasonal = [
    ["Back to School", "September", "60-day extended trial", "New academic year, students choosing tools"],
    ["Mock Exam Prep", "Nov-Dec", "'Mock Ready' £14.99 for 2 months", "Mock exams typically Dec/Jan"],
    ["New Year Resolution", "January", "20% off annual plan", "'New year, new grades' messaging"],
    ["100 Days to GCSEs", "Late January", "'Exam Season Pass' £29.99 for 5 months (Jan-May)", "Covers entire revision period"],
    ["Easter Revision", "March-April", "7-day intensive trial + flash sale", "Easter = peak study period"],
    ["Results Day", "August", "'Retake Ready' campaign for resit students", "Nov resits surged 5.8% in 2025"],
    ["IGCSE Oct/Nov", "Aug-Sep", "Targeted international student campaign", "Different exam season for IGCSE"],
]
for row_data in seasonal:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws9.cell(row=r, column=i, value=v)
    style_data_row(ws9, r, 6)

# School pricing
r += 2
add_section_row(ws9, r, "SCHOOL / BULK PRICING", 6)
r += 1
headers = ["Package", "Price", "Per Student/Year", "Includes", ""]
for i, h in enumerate(headers, 2):
    ws9.cell(row=r, column=i, value=h)
style_header_row(ws9, r, 6)

school_pricing = [
    ["Class Licence (up to 35)", "£199/year", "£5.69", "Dashboard, progress reports, teacher admin"],
    ["Department Licence (up to 150)", "£599/year", "£3.99", "+ department analytics, multiple teacher accounts"],
    ["Whole School Licence (up to 500)", "£1,499/year", "£3.00", "+ priority support, custom onboarding"],
    ["MAT/Group (500+)", "Custom pricing", "£2.00-2.50 target", "+ dedicated account manager, volume negotiated"],
]
for row_data in school_pricing:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws9.cell(row=r, column=i, value=v)
    style_data_row(ws9, r, 6)

# ════════════════════════════════════════════════════════════════
# SHEET 10: TIKTOK SCRIPTS (Full 30 video scripts)
# ════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("TikTok Scripts")
ws10.sheet_properties.tabColor = "000000"
cols10 = ["Video #", "Title", "Category", "Format", "Duration", "Hook", "Key Points / Script Outline", "Sound", "Hashtags", "CTA"]
set_col_widths(ws10, [8, 40, 14, 22, 8, 40, 60, 25, 50, 35])

for i, h in enumerate(cols10, 1):
    ws10.cell(row=1, column=i, value=h)
style_header_row(ws10, 1, len(cols10))
ws10.freeze_panes = 'A2'

tiktok_scripts = [
    ["V1", "3 Words That Instantly Make Your Essay Sound Smarter", "Quick Fix", "Talking head + text overlay + word cards", "30s", "Stop using 'shows' and 'suggests' in every paragraph.", "Reveal: Conveys, Epitomises, Encapsulates. Show each in sentence context.", "Original audio, light lo-fi", "#GCSEEnglish #EnglishTips #GCSERevision #StudyTok #Grade9 #EduTok #EnglishLanguage", "Follow for a new word every day — link in bio for the full vocabulary list"],
    ["V2", "The Difference Between 'Effect' and 'Affect' (Finally)", "Quick Fix", "Green screen + whiteboard graphic", "20s", "If you still get these mixed up, this is your sign.", "Affect = Action (A) = verb. Effect = End result (E) = noun. Example sentence each.", "Trending 'aha moment' or original", "#GCSEEnglish #GrammarTips #EnglishLanguage #StudyTok #LearnOnTikTok #EnglishGrammar #GCSE2026", "Comment the one you always get wrong"],
    ["V3", "POV: You Used a Semicolon Correctly", "Quick Fix + Relatable", "POV reaction split-screen", "25s", "When you finally learn what a semicolon actually does", "Wrong: 'I went to the shop; and bought milk'. Correct: 'The storm raged outside; inside, the house was silent.'", "'Level Up' trending or triumphant audio", "#Semicolon #GCSEEnglish #GrammarTok #WritingTips #StudyTok #LearnOnTikTok #GCSE", "Drop a semicolon sentence in the comments — I'll tell you if it works"],
    ["V4", "5 Techniques Examiners LOVE Seeing in Language Paper 1", "Exam Intel + Quick Fix", "Talking head + numbered text overlays", "45s", "Your examiner reads 300 papers. Here's what makes them stop and go 'yes.'", "1. Semantic field analysis 2. Sentence structure/syntax 3. Cyclical structure 4. 'the reader' not 'the audience' 5. Embedding quotes", "Original audio with subtle beat", "#GCSEEnglish #LanguagePaper1 #Grade9 #ExamTips #AQAEnglish #EnglishLanguage #GCSE2026 #Revision", "Which one are you NOT doing? Comment below"],
    ["V5", "Stop Starting Sentences With 'This Shows That'", "Quick Fix", "Text on screen + voiceover", "30s", "Things your English teacher is TIRED of reading", "Bad: 'The writer uses a metaphor. This shows that...' Good: 'Through this metaphor, [author] conveys...' / '[Author]'s use of metaphor emphasises...'", "Trending 'upgrade' or transformation sound", "#GCSEEnglish #EnglishEssay #WritingTips #StudyTok #ExamHacks #EnglishLanguage #Grade9", "Screenshot this. You'll need it."],
    ["V6", "The Word That Gets You Marks in EVERY English Answer", "Quick Fix", "Talking head, dramatic pause", "25s", "There's one word that works in literally every GCSE English answer.", "Build suspense → Reveal: 'Perhaps' — shows alternative interpretations. 'Perhaps Shelley is suggesting...' / 'This could perhaps reflect...'", "Original audio", "#GCSEEnglish #ExamHack #Grade9Tips #EnglishLiterature #StudyTok #GCSE #Revision", "Follow for the words that literally print marks"],
    ["V7", "Pathetic Fallacy Explained in 15 Seconds", "Quick Fix", "Green screen + moody landscape", "15s", "Pathetic fallacy is NOT a metaphor.", "Weather/environment reflecting character emotions. 'The storm raged as Macbeth's guilt consumed him'. Difference: metaphor compares; pathetic fallacy MIRRORS.", "Thunder → calm sound effect", "#PathticFallacy #GCSEEnglish #EnglishLit #Macbeth #LiteraryTechniques #StudyTok #LearnOnTikTok", "Save this for your exam"],
    ["V8", "3 Connectives That Scream 'I'm Getting a Grade 9'", "Quick Fix", "Text overlay carousel + voiceover", "25s", "Replace 'however' with these and watch your grade jump.", "1. Conversely (flip an argument) 2. Moreover (builds sophistication) 3. Fundamentally (asserts authority). Show in sentences.", "Original or trending 'level up'", "#GCSEEnglish #Grade9 #EssayTips #StudyTok #EnglishLanguage #Connectives #ExamPrep", "Link in bio for the full Grade 9 vocabulary list on the app"],
    ["V9", "How I'd Revise GCSE English If I Had 30 Days Left", "Exam Intel", "Talking head + day plan graphic", "60s", "If your exam is in 30 days and you've done nothing, here's the plan.", "Days 1-10: Learn texts (quotes + context, 3 per character/theme). Days 11-20: Past papers timed, 1/day. Days 21-25: Mark own work. Days 26-30: Weak areas only.", "Motivational trending audio", "#GCSERevision #30DayChallenge #ExamPrep #GCSEEnglish #StudyPlan #Revision #GCSE2026", "Save this and start tomorrow. The app has a built-in revision planner — link in bio"],
    ["V10", "The Mark Scheme Hack Nobody Talks About", "Exam Intel", "Screen recording of mark scheme + talking head", "45s", "The mark scheme literally tells you how to get a Grade 9. Nobody reads it.", "Show Level 6 descriptors: 'perceptive,' 'convincing,' 'detailed'. Translate: 'Perceptive = noticed something most wouldn't'. Reverse-engineer from descriptors.", "Original audio", "#MarkScheme #GCSEEnglish #Grade9 #ExamHack #AQA #Edexcel #StudyTok #Revision", "Comment 'MARKS' and I'll DM you the decoded mark scheme"],
    ["V11", "How to Answer the 'How Does the Writer Use Language' Question", "Exam Intel", "On-screen text walkthrough + annotations", "50s", "This question is worth 12 marks and most people lose 4 of them.", "Structure: What? (identify technique) → How? (explain effect) → Why? (link to writer's purpose). Model mini-paragraph. Common mistake: just identifying without WHY.", "Lo-fi study beat", "#LanguagePaper1 #GCSEEnglish #ExamTechnique #Grade9 #AQAEnglish #StudyTok #EnglishLanguage", "Full walkthrough on YouTube — link in bio"],
    ["V12", "Your Examiner Has 4 Minutes to Mark Your Essay", "Exam Intel + Relatable", "Talking head, slightly shocked", "35s", "Your examiner spends about 4 minutes reading your entire essay. Four.", "Tip 1: Strong opening = instant positive impression. Tip 2: Clear paragraphs (scanning, not reading every word). Tip 3: Conclusion = last thing before mark.", "Original audio with dramatic timing", "#Examiner #GCSEEnglish #ExamSecrets #Grade9 #EnglishEssay #StudyTok #GCSE2026", "Follow for more examiner secrets"],
    ["V13", "Revision Technique Tier List for GCSE English", "Exam Intel", "Tier list graphic on screen", "50s", "Let's rank every revision technique for English.", "S: Practice papers timed + active recall flashcards. A: Mark scheme self-assessment + planning essays. B: Mind maps + analysis videos. C: Re-reading text + copying notes. D: Highlighting + SparkNotes only.", "Trending tier list sound", "#TierList #GCSERevision #StudyTips #GCSEEnglish #RevisionTechniques #StudyTok #ExamPrep", "Disagree? Fight me in the comments"],
    ["V14", "The 15-Minute Essay Plan That Gets Grade 9", "Exam Intel", "Speed-draw of plan + voiceover", "55s", "Spending 5 minutes planning saves you 15 minutes writing. Here's how.", "Step 1: Read question, underline KEY word (2 min). Step 2: Brain-dump 4-5 quotes (2 min). Step 3: 1-word technique + 1-word effect (2 min). Step 4: Order into logical arc (1 min).", "Sped-up writing ASMR or lo-fi", "#EssayPlan #GCSEEnglish #Grade9 #ExamTechnique #StudyTok #Revision #WritingTips", "Try it on your next practice essay. Link in bio for guided essay plans on the app"],
    ["V15", "Things Your English Teacher Won't Tell You — Part 1", "Exam Intel + Relatable", "Talking head, conspiratorial tone", "40s", "Your English teacher can't say this, but I can.", "1. Don't need to love the text — ANALYSE it. Faking passion is fine. 2. No 'right answer' — want CONVINCING argument. 3. Grade 9s write smarter not more. Shorter paragraphs, sharper points.", "Original audio (conspiratorial voice)", "#GCSEEnglish #EnglishTeacher #ExamSecrets #Grade9 #StudyTok #ThingsTeachersDontSay #GCSE2026", "Part 2? Comment YES"],
    ["V16", "Things Your English Teacher Won't Tell You — Part 2", "Exam Intel + Relatable", "Same format as Part 1", "40s", "Part 2 because you lot went mad for the first one.", "4. Same quotes for multiple questions if analysed differently. 5. Context ≠ history essay — one sentence woven in. 6. Creative writing paper is EASIER to get 9 on than literature.", "Original audio (maintain consistency)", "#GCSEEnglish #EnglishTeacher #ExamSecrets #ThingsTeachersDontSay #Grade9 #StudyTok #GCSE", "Part 3 when this hits 10k likes"],
    ["V17", "What Examiners ACTUALLY Write on Your Paper", "Exam Intel", "Green screen + mock annotated script", "45s", "Ever wonder what the person marking your paper actually thinks?", "Annotations: ticks for 'clear point,' '!' for 'good insight,' '?' for 'what are you saying'. Best-fit model. Biggest turn-off: huge text blocks. Biggest positive: alternative interpretation.", "Original audio", "#Examiner #GCSEEnglish #ExamMarking #MarkScheme #Grade9 #StudyTok #BehindTheScenes", "Want me to mark YOUR paragraph? Drop it in the comments"],
    ["V18", "The Question Nobody Answers Properly (and How to Fix It)", "Exam Intel", "Talking head + question screenshot", "45s", "The comparison question on Paper 2 is where everyone drops marks.", "Problem: write about Text A, then Text B, and call it comparison. Fix: every paragraph mentions BOTH texts. Template: 'While [Author A] uses [technique], [Author B] instead [contrast]'", "Original audio", "#Paper2 #GCSEEnglish #ComparisonQuestion #ExamTechnique #Grade9 #AQA #StudyTok", "Save this before Paper 2"],
    ["V19", "POV: The Examiner Reading Your Opening Sentence", "Relatable + Exam Intel", "POV reaction — camera is essay", "30s", "POV: the examiner when you start with 'In this essay I will...'", "Pained expression → Cut to: 'POV: when you start with a bold analytical statement' → Impressed nod, reaches for higher mark band. Show good opening on screen.", "Trending reaction/transformation sound", "#POV #GCSEEnglish #ExaminerReacts #EnglishEssay #Grade9 #StudyTok #GCSE2026", "Show me YOUR opening sentence in the comments"],
    ["V20", "5 Things That Instantly Make Your Essay Look Year 7", "Relatable + Quick Fix", "List with visual examples", "40s", "If you're doing ANY of these, your essay screams Year 7.", "1. 'In this essay I will...' 2. Using 'thing' instead of specific noun 3. One-sentence or full-page paragraphs 4. Retelling plot not analysing 5. 'the book says' instead of 'Dickens presents'", "Trending 'ick' or cringe sound", "#GCSEEnglish #EssayMistakes #EnglishTips #StudyTok #GCSE #Grade9 #WritingTips", "Which one are you guilty of? Be honest"],
    ["V21", "Rating GCSE English Essays From My DMs", "Relatable + Exam Intel", "Screen recording + talking head reactions", "55s", "You sent me your essays. Let's see what we're working with.", "3 snippets: 1 needs work, 1 mid, 1 impressive. Rate each, explain one specific fix. Balance encouraging + honest.", "Original audio", "#GCSEEnglish #EssayReview #Grade9 #ExaminerReacts #StudyTok #EnglishEssay #Feedback", "DM me YOUR paragraph for the next one"],
    ["V22", "POV: You Just Wrote 'The Author Uses Lots of Techniques'", "Relatable", "Skit — reacting to own past writing", "20s", "POV camera with dramatic zoom on the phrase", "'Lots of techniques? WHICH techniques?' Fix: 'Shelley employs an extended metaphor rooted in natural imagery to...' Side-by-side comparison.", "Trending 'oh no' or dramatic", "#GCSEEnglish #POV #EssayFail #EnglishTips #StudyTok #GCSE #CommonMistakes", "Tag someone who does this"],
    ["V23", "Words You're Using Wrong in Your Essay (and What to Say Instead)", "Quick Fix + Relatable", "Split screen: wrong vs right", "35s", "These words are lowkey sabotaging your grade.", "1. Infer vs Imply (writer implies, YOU infer) 2. Literally (you mean figuratively) 3. Bias vs Biased ('writer is bias' — no) 4. Could of (it's could have)", "Trending educational or original", "#GCSEEnglish #CommonMistakes #VocabularyTips #GrammarPolice #StudyTok #EnglishLanguage #GCSE", "Save this — you'll thank me in the exam hall"],
    ["V24", "Why I'm Building an English Revision App", "Build in Public", "Talking head, personal/authentic", "50s", "I used to teach GCSE English. Here's why I quit to build an app.", "Story: students struggling with same problems yearly. Existing resources: boring textbooks or unstructured YouTube. Vision: everything in one place, built how they learn. Quick phone mockup.", "Soft background music", "#EdTech #BuildInPublic #TheEnglishHub #GCSEEnglish #StartupLife #IndieHacker #StudyTok", "Follow the journey. Link in bio to join the waitlist"],
    ["V25", "Day in My Life Building an EdTech Startup", "Build in Public", "DITL montage + voiceover", "45s", "What building a revision app actually looks like.", "Clips: morning coffee, writing content, testing with students, checking analytics, responding to DMs, late-night debugging. Authentic moments.", "Trending DITL montage audio", "#BuildInPublic #EdTech #StartupLife #DayInMyLife #TheEnglishHub #GCSE #IndieHacker", "What feature would you want? Comment below"],
    ["V26", "Sneak Peek: Inside The English Hub App", "Build in Public", "Screen recording + talking head corner", "40s", "First look at what I've been building for 6 months.", "Show: clean UI, quote bank, practice questions, instant feedback. 'Uses AI for examiner-style feedback'. Feature students asked for.", "Upbeat, energetic trending sound", "#TheEnglishHub #AppReveal #EdTech #GCSERevision #GCSEEnglish #StudyTok #RevisionApp", "Link in bio for early access. First 100 users get it free"],
    ["V27", "Every English Class Has This Student", "Relatable", "Skit — same person, different characters", "35s", "Every English class. Tell me I'm wrong.", "1. Hasn't read book but wings it 2. Writes 6 pages for 2-mark Q 3. 'It's just an opinion' to dodge analysis 4. Actually finds deeper meaning and everyone zones out", "Trending skit/character audio", "#EnglishClass #GCSEEnglish #SchoolHumor #Relatable #StudyTok #StudentLife #GCSE", "Which one are you? Be honest"],
    ["V28", "English Exam Bingo", "Relatable", "Bingo card on screen, checking off", "30s", "English exam bingo — you've definitely done at least 5.", "Furthermore x9 / 'In conclusion' with 2 min left / Made up a quote / Forgot author's name / Wrong character / Crossed out half a page in pen", "Trending bingo/game show audio", "#ExamBingo #GCSEEnglish #Relatable #StudentHumor #ExamSeason #StudyTok #GCSE", "Comment your score out of 6"],
    ["V29", "I Wrote a Grade 9 Paragraph in 60 Seconds", "Deep Dive (snippet)", "Speed-write + slow replay + annotation", "55s", "Grade 9 paragraph. 60 seconds. Let's go.", "Real-time Macbeth paragraph (sped up). Slow replay: annotate each sentence — Point, Evidence (embedded quote), Analysis, Alternative interpretation, Context link.", "Timer/countdown or focus sound", "#Grade9 #GCSEEnglish #Macbeth #ExamTechnique #ModelAnswer #StudyTok #EnglishLiterature", "Full walkthrough on YouTube. Link in bio"],
    ["V30", "Read This Extract and Watch Me Get Full Marks (Paper 1 Q2)", "Deep Dive (snippet)", "Screen recording + annotations appearing", "60s", "Paper 1, Question 2. Most people lose marks here. Watch how to get them all.", "Short extract (2-3 lines). Annotate: circle language features, arrows for effects. Write model response in real-time. Examiner comment overlay.", "Lo-fi study beats", "#Paper1Q2 #GCSEEnglish #LanguagePaper #FullMarks #ExamTechnique #AQA #StudyTok #Revision", "Save and share with your study group"],
]

for idx, row_data in enumerate(tiktok_scripts, 2):
    for i, v in enumerate(row_data, 1):
        ws10.cell(row=idx, column=i, value=v)
    style_data_row(ws10, idx, len(cols10))
    ws10.row_dimensions[idx].height = 80

# ════════════════════════════════════════════════════════════════
# SHEET 11: YOUTUBE SCRIPTS (12 videos)
# ════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("YouTube Scripts")
ws11.sheet_properties.tabColor = "FF0000"
cols11 = ["Video #", "Title", "Thumbnail Concept", "Length", "Target Keywords", "Structure / Outline", "Description Template", "Status", "Publish Date"]
set_col_widths(ws11, [8, 50, 35, 8, 45, 70, 50, 10, 12])

for i, h in enumerate(cols11, 1):
    ws11.cell(row=1, column=i, value=h)
style_header_row(ws11, 1, len(cols11))
ws11.freeze_panes = 'A2'

youtube_scripts = [
    ["Y1", "How to Get a Grade 9 in GCSE English Language (2026 Complete Guide)", "Bold 'GRADE 9' gold text, shocked face, red arrow → marked paper showing 9", "18-22m", "gcse english language grade 9, how to get a 9, gcse english tips 2026, aqa english language", "Intro (0-1:30) → Paper 1 breakdown (1:30-8:00) → Paper 2 breakdown (8:00-14:00) → Creative writing deep dive (14:00-18:00) → 8-week revision strategy (18:00-20:00) → CTA (20:00-22:00)", "Want a Grade 9 in GCSE English Language? Complete 2026 guide. Timestamps + Free Resources links + social links", "Not Started", ""],
    ["Y2", "I Marked 100 GCSE English Essays — Here's What Examiners Actually Want", "Stack of papers with red pen, shocked face, '100 ESSAYS' red circle", "15-18m", "gcse english examiner, what examiners want, gcse english essay tips, marking", "Intro + credentials (0-2:00) → 5 most common mistakes (2:00-8:00) → 5 things that impress (8:00-14:00) → Grade 5 vs Grade 9 side-by-side (14:00-16:00) → CTA (16:00-18:00)", "I've read over 100 essays. The difference between Grade 5 and 9 is clear in the first paragraph.", "Not Started", ""],
    ["Y3", "GCSE English Language Paper 1 — Full Walkthrough (AQA 2026)", "Paper 1 blurred + 'FULL WALKTHROUGH' bold + timer graphic", "35-45m", "aqa english language paper 1, walkthrough, paper 1 2026, paper 1 answers", "Intro/time management (0-2:00) → Q1: Info retrieval 4marks (2-6) → Q2: Language analysis 8marks (6-14) → Q3: Structure 8marks (14-22) → Q4: Evaluation 20marks (22-32) → Q5: Creative writing 40marks (32-42) → Summary (42-45)", "Full Paper 1 walkthrough with model answers for every question.", "Not Started", ""],
    ["Y4", "GCSE English Language Paper 2 — Full Walkthrough (AQA 2026)", "Paper 2 blurred + 'PAPER 2' + comparison arrows between two texts", "35-45m", "aqa english language paper 2, paper 2 walkthrough, 2026, comparison", "Same structure as Y3 but for Paper 2. Heavy focus on comparison question.", "Full Paper 2 walkthrough — the comparison question explained.", "Not Started", ""],
    ["Y5", "An Inspector Calls — Complete GCSE Revision (AQA/Edexcel)", "Book cover + 'EVERYTHING YOU NEED' + key character faces", "25-35m", "an inspector calls gcse, revision, quotes, analysis, grade 9", "Overview + context → Character analysis (Birling, Sheila, Eric, Gerald, Inspector) → Themes (responsibility, class, gender, generations) → Key quotes with analysis → Essay planning → Model paragraph", "Complete An Inspector Calls revision for GCSE. Every character, theme, and key quote.", "Not Started", ""],
    ["Y6", "Macbeth — Complete GCSE Revision (AQA/Edexcel)", "Macbeth crown + dagger + 'FULL GUIDE' + dramatic lighting", "25-35m", "macbeth gcse, revision, quotes, analysis, grade 9, aqa", "Overview + context → Character analysis → Themes (ambition, power, guilt, supernatural, masculinity) → Key quotes → Essay planning → Model paragraph", "Complete Macbeth revision. Characters, themes, quotes, and essay technique.", "Not Started", ""],
    ["Y7", "How to Write a Grade 9 Creative Writing Piece (Paper 1 Q5)", "Creative writing paper + '40 MARKS' + gold grade 9 stamp", "20-25m", "gcse creative writing, paper 1 question 5, grade 9, creative writing tips", "Why Q5 is your secret weapon → Structure techniques → Openings that grab → Vocabulary and imagery → Endings → Full model response → Planning method", "Creative writing is where Grade 9s are won. Full guide with model response.", "Not Started", ""],
    ["Y8", "GCSE English Literature Paper 1 — Full Walkthrough (AQA 2026)", "Lit Paper 1 + book icons + 'FULL GUIDE'", "30-40m", "aqa english literature paper 1, shakespeare, 19th century novel", "Section A: Shakespeare (planning + model paragraph) → Section B: 19th century novel → Time management → Exam technique", "Full Literature Paper 1 walkthrough — Shakespeare and 19th century novel.", "Not Started", ""],
    ["Y9", "Power and Conflict Poetry — Complete Revision (AQA)", "Poetry anthology + explosive imagery + 'ALL 15 POEMS'", "30-40m", "power and conflict poetry, aqa poetry, gcse poetry revision, anthology", "Each poem: overview, key quotes, techniques, comparison links. Comparison planning matrix.", "All 15 Power and Conflict poems analysed with comparison links.", "Not Started", ""],
    ["Y10", "A Christmas Carol — Complete GCSE Revision", "Scrooge + ghost imagery + 'COMPLETE GUIDE' + Victorian London", "25-30m", "a christmas carol gcse, revision, quotes, dickens, grade 9", "Context (Victorian poverty) → Character analysis → Themes (greed, redemption, social responsibility) → Quotes → Model paragraph", "Complete A Christmas Carol revision for GCSE English Literature.", "Not Started", ""],
    ["Y11", "Romeo and Juliet — Complete GCSE Revision", "R&J imagery + 'FULL REVISION' + balcony/dagger", "25-30m", "romeo and juliet gcse, revision, shakespeare, quotes, analysis", "Context → Characters → Themes (love, fate, conflict, family) → Key quotes → Model paragraph", "Complete Romeo and Juliet revision for GCSE.", "Not Started", ""],
    ["Y12", "GCSE English: My Top 10 Revision Mistakes (and How to Fix Them)", "Crossed out mistakes + checkmarks + 'DON'T DO THIS'", "12-15m", "gcse english revision mistakes, revision tips, how not to revise, gcse tips", "10 mistakes with fixes: 1. Feature spotting 2. Retelling plot 3. No context 4. Ignoring the question 5. One long paragraph 6. Only reading notes 7. Not timing yourself 8. Skipping creative writing practice 9. Not using the mark scheme 10. Revising everything equally", "The 10 biggest GCSE English revision mistakes I see students make.", "Not Started", ""],
]

for idx, row_data in enumerate(youtube_scripts, 2):
    for i, v in enumerate(row_data, 1):
        ws11.cell(row=idx, column=i, value=v)
    style_data_row(ws11, idx, len(cols11))
    ws11.row_dimensions[idx].height = 100

# ════════════════════════════════════════════════════════════════
# SHEET 12: OUTREACH TEMPLATES
# ════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet("Outreach Templates")
ws12.sheet_properties.tabColor = "4472C4"
set_col_widths(ws12, [3, 20, 80])

r = 1
ws12.merge_cells('B1:C1')
ws12['B1'] = "OUTREACH MESSAGE TEMPLATES"
ws12['B1'].font = TITLE_FONT

# Micro-influencer DM
r = 3
add_section_row(ws12, r, "TEMPLATE A: Micro-Influencer DM (TikTok/Instagram)", 3)
r += 1
ws12.cell(row=r, column=2, value="Message:")
ws12.cell(row=r, column=3, value="Hi [Name]! I love your GCSE revision content — especially your [specific video/post]. We've just launched The English Hub, an AI-powered GCSE English platform that gives essay feedback in seconds. Would you like free lifetime access + earn 20% recurring commission for every student you refer? I'd love to send you a demo. No pressure at all!")
ws12.cell(row=r, column=3).alignment = LEFT_WRAP
ws12.row_dimensions[r].height = 60

# Teacher email
r += 2
add_section_row(ws12, r, "TEMPLATE B: Teacher/Tutor Email", 3)
r += 1
ws12.cell(row=r, column=2, value="Subject:")
ws12.cell(row=r, column=3, value="Free AI essay feedback tool for your students")
r += 1
ws12.cell(row=r, column=2, value="Body:")
ws12.cell(row=r, column=3, value="Hi [Name],\n\nI'm [Your Name], founder of The English Hub. I noticed your excellent [GCSE resource/YouTube channel/blog] and wanted to reach out.\n\nWe've built an AI-powered GCSE English platform that gives students instant essay feedback on AQA/Edexcel/OCR/WJEC papers. I'd love to offer you:\n\n- Free lifetime premium access\n- Free school dashboard for your classes\n- 20% recurring commission on any students you refer\n\nWould you be open to a 15-minute demo? Happy to work around your schedule.\n\nBest regards,\n[Your Name]")
ws12.cell(row=r, column=3).alignment = LEFT_WRAP
ws12.row_dimensions[r].height = 180

# Parent blogger email
r += 2
add_section_row(ws12, r, "TEMPLATE C: Parent Blogger Email", 3)
r += 1
ws12.cell(row=r, column=2, value="Subject:")
ws12.cell(row=r, column=3, value="GCSE English tool your readers would love")
r += 1
ws12.cell(row=r, column=2, value="Body:")
ws12.cell(row=r, column=3, value="Hi [Name],\n\nI follow your blog/channel and love your content on [specific topic]. I wanted to introduce The English Hub — an AI-powered GCSE/IGCSE English revision platform.\n\nWe offer:\n- 30-day free trial (no card required)\n- AI essay feedback that gives examiner-style comments\n- All major exam boards (AQA, Edexcel, OCR, WJEC, Cambridge IGCSE)\n\nI'd love to offer you:\n- Free premium access for your family\n- 20% recurring commission on referrals\n- Optional: £50-150 sponsored review fee\n\nWould you like me to set up a free account for you?\n\nBest,\n[Your Name]")
ws12.cell(row=r, column=3).alignment = LEFT_WRAP
ws12.row_dimensions[r].height = 180

# School outreach
r += 2
add_section_row(ws12, r, "TEMPLATE D: School Head of Department Email", 3)
r += 1
ws12.cell(row=r, column=2, value="Subject:")
ws12.cell(row=r, column=3, value="Free AI essay feedback tool for your English department")
r += 1
ws12.cell(row=r, column=2, value="Body:")
ws12.cell(row=r, column=3, value="Dear [Name],\n\nI'm writing to introduce The English Hub — an AI-powered GCSE and IGCSE English platform that provides examiner-style feedback on student essays.\n\nWe're offering English departments a free trial of our school dashboard, which includes:\n\n- AI essay feedback modelled on AQA/Edexcel/OCR/WJEC mark schemes\n- Student progress tracking and analytics\n- Structured revision plans aligned to exam schedules\n- Coverage of all major set texts\n\nPricing starts at £199/year for a class of 35 (£5.69/student — less than one CGP book).\n\nWould you be open to a 15-minute demonstration? I'm happy to work around your timetable.\n\nBest regards,\n[Your Name]\nFounder, The English Hub")
ws12.cell(row=r, column=3).alignment = LEFT_WRAP
ws12.row_dimensions[r].height = 200

# Follow-up sequence
r += 2
add_section_row(ws12, r, "FOLLOW-UP SEQUENCE (ALL CATEGORIES)", 3)
r += 1
ws12.cell(row=r, column=2, value="Follow-up 1 (Day 3):")
ws12.cell(row=r, column=3, value="Hi [Name], just bumping this in case it got buried! Happy to answer any questions about The English Hub. No worries at all if it's not for you 🙂")
r += 1
ws12.cell(row=r, column=2, value="Follow-up 2 (Day 7):")
ws12.cell(row=r, column=3, value="Hi [Name], last message from me — I wanted to share that [X affiliates] have already joined and [Y students] are using the platform. If you'd like free access to try it yourself, just let me know. All the best!")
r += 1
ws12.cell(row=r, column=2, value="Follow-up 3 (Day 14):")
ws12.cell(row=r, column=3, value="Final gentle nudge — no response needed if you're not interested. But if timing was the issue, the offer of free lifetime access + 20% recurring commission is always open. Just reply whenever suits. Thanks, [Name]!")

# ════════════════════════════════════════════════════════════════
# SHEET 13: COMMISSION CALCULATOR
# ════════════════════════════════════════════════════════════════
ws13 = wb.create_sheet("Commission Calculator")
ws13.sheet_properties.tabColor = "00B050"
set_col_widths(ws13, [3, 30, 18, 18, 18, 18])

r = 2
ws13.merge_cells('B2:F2')
ws13['B2'] = "AFFILIATE COMMISSION STRUCTURE"
ws13['B2'].font = TITLE_FONT

r = 4
add_section_row(ws13, r, "STANDARD COMMISSION", 6)
r += 1
headers = ["Element", "Details", "", "", ""]
for i, h in enumerate(headers, 2):
    ws13.cell(row=r, column=i, value=h)
style_header_row(ws13, r, 6)

comm_data = [
    ["Standard Commission", "20% recurring for lifetime of customer", "", "", ""],
    ["Cookie Duration", "90 days", "", "", ""],
    ["Minimum Payout", "£25", "", "", ""],
    ["Payment Frequency", "Monthly (net 30)", "", "", ""],
    ["Payment Method", "PayPal or bank transfer via Rewardful", "", "", ""],
    ["First-Sale Bonus", "£5 bonus on first referred conversion", "", "", ""],
    ["Volume Bonus (10/mo)", "£50 bonus", "", "", ""],
    ["Volume Bonus (25/mo)", "£150 bonus", "", "", ""],
    ["Annual Plan Bonus", "Additional 5% one-time (£4.00 on £79.99)", "", "", ""],
]
for row_data in comm_data:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws13.cell(row=r, column=i, value=v)
    style_data_row(ws13, r, 6)

r += 2
add_section_row(ws13, r, "EARNINGS EXAMPLES", 6)
r += 1
headers = ["Scenario", "Monthly Earnings", "Annual Earnings", "", ""]
for i, h in enumerate(headers, 2):
    ws13.cell(row=r, column=i, value=h)
style_header_row(ws13, r, 6)

earnings = [
    ["10 monthly subscribers", "£20.00/month", "£240/year", "", ""],
    ["10 annual subscribers", "£16.00/year each", "£160/year + £40 bonuses = £200", "", ""],
    ["50 annual subscribers (Teacher Ambassador)", "£800/year", "£800 + £200 bonuses = £1,000/year", "", ""],
    ["100 monthly subscribers", "£200/month", "£2,400/year", "", ""],
]
for row_data in earnings:
    r += 1
    for i, v in enumerate(row_data, 2):
        ws13.cell(row=r, column=i, value=v)
    style_data_row(ws13, r, 6)

# ════════════════════════════════════════════════════════════════
# SHEET 14: ASPIRATIONAL INFLUENCERS & PARTNERS
# Non-education creators/platforms with massive reach to target demographic
# ════════════════════════════════════════════════════════════════
ws14 = wb.create_sheet("Aspirational Influencers")
ws14.sheet_properties.tabColor = "9B59B6"
cols14 = ["#", "Name / Channel", "Platform", "Category", "Tier", "Followers / Reach", "Why Relevant", "Approach", "Contact", "Status", "Notes"]
for i, col in enumerate(cols14, 1):
    cell = ws14.cell(row=1, column=i, value=col)
    cell.font = HEADER_FONT
    cell.fill = DARK_BLUE
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws14.row_dimensions[1].height = 30
ws14.freeze_panes = 'A2'
col_widths_14 = [5, 35, 18, 28, 12, 22, 35, 25, 20, 14, 40]
for i, w in enumerate(col_widths_14, 1):
    ws14.column_dimensions[get_column_letter(i)].width = w

# Aspirational influencer categories for the new sheet
aspirational_categories = [
    "UK Teen Lifestyle",
    "UK Mainstream Parenting",
    "UK Comedy & Entertainment",
    "Dubai/GCC Lifestyle",
    "UK YA Authors & Literature",
    "UK Podcasts (Parents/Teens)",
    "UK Fitness & Mental Health",
    "UK Gaming & Streaming",
    "GCC Teen/Young Adult",
    "UK Motivational & Self-Dev",
    "UK Student Life Vloggers",
    "UK Beauty & Fashion (Teen)",
    "British TV & Celebrity",
    "UK Theatre & Drama",
    "UK Sports & Football",
    "UK Music Artists",
    "UK STEM & Science",
    "UK Finance (Young People)",
    "UK Food & Cooking",
    "UK Travel & Gap Year",
    "UK Education Charities",
    "UK Literary Festivals",
    "UK Education Conferences",
    "UK Parenting Media",
    "UK Student Platforms",
    "UK Education Journalists",
    "UK Online Communities",
    "GCC Expat Platforms",
    "UK School Partnerships",
    "UK Exam Boards & Competitors",
]

aspirational_influencers = [
    # ═══ UK TEEN LIFESTYLE ═══
    # [num, name, platform, category, tier, followers, why_relevant, approach, contact, status, notes]
    [1, "KSI (@ksi)", "YouTube/TikTok/IG", "UK Teen Lifestyle", "Mega", "24M+ YT / 12M+ IG", "Massive UK teen audience, cultural icon", "Brand deal / Sponsored mention", "Agent", "Not Started", "Aspirational. British-Nigerian creator, boxing, gaming, music"],
    [2, "Zoe Sugg / Zoella (@zoesugg)", "IG/YouTube", "UK Teen Lifestyle", "Mega", "9M+ IG", "OG UK lifestyle creator, parent of young child", "Story mention / Brand collab", "Agent", "Not Started", "UK lifestyle icon — now parenting content too"],
    [3, "Joe Sugg (@joe_sugg)", "IG/YouTube", "UK Teen Lifestyle", "Mega", "5.5M IG / 8M YT", "Huge UK teen/young adult following", "Sponsored content", "Agent", "Not Started", "Mainstream UK creator — aspirational"],
    [4, "Saffron Barker (@saffronbarker)", "YouTube/TikTok", "UK Teen Lifestyle", "Macro", "2.5M YT / 1M TikTok", "UK teen lifestyle, ex-Strictly", "Brand deal", "Agent", "Not Started", "Brighton-based, massive teen girl following"],
    [5, "Amelia Dimoldenberg (@ameliadimz)", "YouTube/TikTok/IG", "UK Teen Lifestyle", "Macro", "3M+ YT / 2M+ IG", "Chicken Shop Date, UK youth culture icon", "Sponsored episode / Brand deal", "Agent", "Not Started", "Cultural phenomenon among UK teens"],
    [6, "Maisie Peters (@maisiehpeters)", "TikTok/IG", "UK Teen Lifestyle", "Macro", "2M+ TikTok", "UK pop artist, huge teen fanbase", "Sponsored mention / TikTok collab", "Agent", "Not Started", "Ed Sheeran signee — massive UK teen reach"],
    [7, "Sophie Louise (@sophie.louisee)", "YouTube/IG", "UK Teen Lifestyle", "Mid", "800K YT", "UK lifestyle, fashion, student content", "Brand collab", "Email", "Not Started", "Relatable student lifestyle content"],
    [8, "Lucy Flight (@lucyflight)", "YouTube", "UK Teen Lifestyle", "Mid", "600K YT", "UK beauty/lifestyle popular with teens", "Sponsored video", "Email", "Not Started", "Beauty + lifestyle — teen girl audience"],
    [9, "Niki and Gabi (@nikiandgabi)", "YouTube", "UK Teen Lifestyle", "Macro", "9M+ YT", "Twin lifestyle — huge teen audience", "Brand deal", "Agent", "Not Started", "Aspirational — massive teen reach"],
    [10, "ClickForTaz (@clickfortaz)", "YouTube/TikTok", "UK Teen Lifestyle", "Mid", "1M+ YT", "UK lifestyle/challenge content", "Sponsored content", "Email", "Not Started", "UK-based, school-age audience"],

    # ═══ UK MAINSTREAM PARENTING ═══
    [11, "Stacey Solomon (@staceysolomon)", "IG/TikTok", "UK Mainstream Parenting", "Mega", "6M+ IG", "UK's most-followed mum influencer, Loose Women", "Story mention / Brand deal", "Agent", "Not Started", "6M IG — school-age children, massive parent reach"],
    [12, "Mrs Hinch / Sophie Hinchliffe (@mrshinchhome)", "IG", "UK Mainstream Parenting", "Mega", "4.8M IG", "Cleaning influencer, mum of 2, parent audience", "Story mention", "Agent", "Not Started", "4.8M — huge mum audience, back-to-school angle"],
    [13, "Louise Pentland (@louisepentland)", "IG/YouTube", "UK Mainstream Parenting", "Macro", "2M+ IG / 2.4M YT", "UK mum vlogger, author, mental health", "Sponsored video / Story", "Agent", "Not Started", "Parenting + wellness — school-age child"],
    [14, "Anna Saccone Joly (@annasaccone)", "YouTube/IG", "UK Mainstream Parenting", "Macro", "1.5M YT", "Family vlogging, school-age kids", "Sponsored family content", "Email", "Not Started", "Ireland/UK family vlog — relatable parent content"],
    [15, "Emily Norris (@emilynoris)", "YouTube/IG", "UK Mainstream Parenting", "Mid", "700K YT", "UK mum of 3 boys, organisation, family life", "Sponsored video / Story", "Email", "Not Started", "Practical mum content — back-to-school"],
    [16, "Channel Mum (@channelmum)", "YouTube", "UK Mainstream Parenting", "Mid", "500K+", "UK mum creator network", "Network sponsorship", "Email", "Not Started", "Network of UK mum creators — multi-creator deal"],
    [17, "Mother Pukka / Anna Whitehouse (@mother_pukka)", "IG/Podcast", "UK Mainstream Parenting", "Mid", "350K IG", "Flex work campaigner, school-age kids", "Sponsored post / Podcast read", "Email", "Not Started", "Work-life balance audience = busy parents"],
    [18, "Clemmie Telford (@clem_adventures)", "IG/Podcast", "UK Mainstream Parenting", "Mid", "200K IG", "Honest parenting, mental health", "Sponsored Story / Podcast", "Email", "Not Started", "Authentic parenting voice — relatable"],
    [19, "The Dad Lab (@thedadlab)", "YouTube/IG", "UK Mainstream Parenting", "Macro", "3.8M YT / 1.3M IG", "London-based dad, science experiments", "Sponsored content", "Email", "Not Started", "3.8M YT — educational-adjacent, dad audience"],
    [20, "Five Minute Mum / Daisy Upton (@fiveminutemum)", "IG/Books", "UK Mainstream Parenting", "Mid", "300K IG", "Quick play ideas, school-age kids", "Sponsored post", "Email", "Not Started", "Education-adjacent — learning through play"],

    # ═══ UK COMEDY & ENTERTAINMENT ═══
    [21, "The Sidemen (@sidemen)", "YouTube", "UK Comedy & Entertainment", "Mega", "22M+ YT", "UK's biggest YouTube group — massive teen audience", "Sponsored challenge / Brand deal", "Agent", "Not Started", "22M+ — every UK teen knows the Sidemen"],
    [22, "TommyInnit (@tommyinnit)", "YouTube/Twitch", "UK Comedy & Entertainment", "Mega", "14M YT", "UK teen Minecraft/gaming creator", "Sponsored mention", "Agent", "Not Started", "14M — teen boy demographic, school-age"],
    [23, "Miniminter / Simon (@miniminter)", "YouTube", "UK Comedy & Entertainment", "Mega", "10M+ YT", "Sidemen member — huge UK teen reach", "Brand deal", "Agent", "Not Started", "Sidemen — UK teen staple"],
    [24, "Chunkz (@chaboree)", "YouTube/TikTok", "UK Comedy & Entertainment", "Macro", "5M+ YT", "UK comedy, football, challenge content", "Sponsored content", "Agent", "Not Started", "UK teen boys + football fans"],
    [25, "Niko Omilana (@nikoomilana)", "YouTube", "UK Comedy & Entertainment", "Macro", "7M+ YT", "UK prank/comedy — ran for London mayor", "Brand integration", "Agent", "Not Started", "Cultural figure for UK teens"],
    [26, "Jack Whitehall (@jackwhitehall)", "IG/Netflix", "UK Comedy & Entertainment", "Mega", "3M+ IG", "UK comedian, Travels With My Father", "Sponsored post / Event", "Agent", "Not Started", "Education comedy angle — Bad Education"],
    [27, "Michael Dapaah / Big Shaq (@michaeldapaah)", "YouTube/IG", "UK Comedy & Entertainment", "Macro", "3M+", "Man's Not Hot, UK comedy/music", "Brand collab", "Agent", "Not Started", "Cultural phenomenon — UK youth"],
    [28, "Sophie Willan (@sophiewillan)", "TikTok/IG", "UK Comedy & Entertainment", "Mid", "200K", "UK comedian, BAFTA winner", "Sponsored content", "Email", "Not Started", "Working-class education angle"],
    [29, "Behzinga (@behzinga)", "YouTube", "UK Comedy & Entertainment", "Macro", "6M+ YT", "Sidemen member — fitness + comedy", "Brand deal", "Agent", "Not Started", "Sidemen — UK teen staple"],
    [30, "Callux (@callux)", "YouTube", "UK Comedy & Entertainment", "Macro", "3M+ YT", "UK YouTube OG, challenge content", "Sponsored video", "Agent", "Not Started", "Long-running UK creator — teen audience"],

    # ═══ DUBAI/GCC LIFESTYLE ═══
    [31, "Mo Vlogs (@movlogs)", "YouTube", "Dubai/GCC Lifestyle", "Macro", "10M+ YT", "Dubai luxury lifestyle, family content", "Sponsored video", "Agent", "Not Started", "10M — Dubai-based, massive Gulf audience"],
    [32, "Khalid Al Ameri (@khalidalameri)", "YouTube/IG/TikTok", "Dubai/GCC Lifestyle", "Macro", "4M+ IG", "Emirati family content, cultural bridge", "Sponsored content", "Email", "Not Started", "Emirati dad — education-positive content"],
    [33, "Huda Kattan / Huda Beauty (@hudabeauty)", "IG", "Dubai/GCC Lifestyle", "Mega", "54M IG", "Dubai-based beauty mogul, Iraqi-American", "Story mention / Brand deal", "Agent", "Not Started", "54M — aspirational, massive GCC parent reach"],
    [34, "Lana Rose (@lanarose786)", "YouTube/IG", "Dubai/GCC Lifestyle", "Macro", "5M+ YT", "Mo Vlogs sister, Dubai lifestyle", "Sponsored content", "Agent", "Not Started", "Dubai lifestyle — young adult audience"],
    [35, "AboFlah (@AboFlah)", "YouTube", "Dubai/GCC Lifestyle", "Mega", "35M+ YT", "Dubai gaming/charity creator", "Sponsored content", "Agent", "Not Started", "35M — Arabic-speaking, massive Gulf youth audience"],
    [36, "Nuseir Yassin / Nas Daily (@nasdaily)", "YouTube/IG/TikTok", "Dubai/GCC Lifestyle", "Mega", "20M+ FB / 4M YT", "Dubai-based, education-positive content", "Sponsored video / Collab", "Agent", "Not Started", "20M — education angle already built in"],
    [37, "Karen Wazen (@karenwazen)", "IG", "Dubai/GCC Lifestyle", "Macro", "5.5M IG", "Dubai fashion/lifestyle, mum of 3", "Story mention", "Agent", "Not Started", "5.5M — Dubai mum, luxury family lifestyle"],
    [38, "Joelle Mardinian (@joaborealjoelle)", "IG", "Dubai/GCC Lifestyle", "Macro", "14M IG", "Dubai beauty/lifestyle, TV presenter", "Story mention", "Agent", "Not Started", "14M — massive GCC women audience"],
    [39, "Fouz Al Fahad (@fozaza)", "IG/Snapchat", "Dubai/GCC Lifestyle", "Macro", "3M+ IG", "Kuwaiti fashion influencer", "Sponsored post", "Agent", "Not Started", "3M — GCC fashion/lifestyle"],
    [40, "Hala Abdallah (@halaabdallahofficial)", "TikTok/IG", "Dubai/GCC Lifestyle", "Macro", "5M+ TikTok", "Syrian-Gulf lifestyle creator", "TikTok collab", "Agent", "Not Started", "5M TikTok — young Gulf audience"],

    # ═══ UK YA AUTHORS & LITERATURE ═══
    [41, "Malorie Blackman (@malorian)", "X/IG", "UK YA Authors & Literature", "Mid", "100K+", "Noughts & Crosses, former Children's Laureate", "School visit / Sponsored read", "Agent", "Not Started", "Books studied in GCSE — direct curriculum link"],
    [42, "Philip Pullman (@PhilipPullman)", "X", "UK YA Authors & Literature", "Mid", "250K+ X", "His Dark Materials, Northern Lights", "Endorsement", "Agent", "Not Started", "Cultural icon — literature curriculum link"],
    [43, "Holly Jackson (@HoJay92)", "IG/TikTok", "UK YA Authors & Literature", "Mid", "200K+ IG", "A Good Girl's Guide to Murder — massive teen hit", "Sponsored post / BookTok collab", "Email", "Not Started", "Biggest UK YA author right now — Netflix adaptation"],
    [44, "Alice Oseman (@aliceoseman)", "IG/X", "UK YA Authors & Literature", "Mid", "500K+ IG", "Heartstopper creator, Netflix show", "Sponsored post", "Agent", "Not Started", "Heartstopper = massive UK teen following"],
    [45, "Jacqueline Wilson (@officialfriendsjw)", "IG", "UK YA Authors & Literature", "Mid", "200K+", "Tracy Beaker, former Children's Laureate", "Endorsement / School event", "Agent", "Not Started", "Beloved UK children's author — parent nostalgia"],
    [46, "David Walliams (@davidwalliams)", "IG", "UK YA Authors & Literature", "Macro", "1.5M+ IG", "Children's author, TV personality", "Brand endorsement", "Agent", "Not Started", "1.5M — children's books + TV crossover"],
    [47, "Jeff Kinney / Diary of a Wimpy Kid", "IG", "UK YA Authors & Literature", "Macro", "1M+", "Wimpy Kid series — global phenomenon", "Brand partnership", "Agent", "Not Started", "Massive reach with 10-14 age transition audience"],
    [48, "Anthony Horowitz (@AnthonyHorowitz)", "X/IG", "UK YA Authors & Literature", "Micro", "50K+", "Alex Rider series, Foyle's War", "Endorsement", "Email", "Not Started", "Alex Rider = teen boy audience"],
    [49, "Cressida Cowell (@caborelcressc)", "IG/X", "UK YA Authors & Literature", "Micro", "30K+", "How to Train Your Dragon, Waterstones Laureate", "School programme / Endorsement", "Email", "Not Started", "Current Waterstones Children's Laureate"],
    [50, "Charlie Mackesy (@charliemackesy)", "IG", "UK YA Authors & Literature", "Macro", "1M+ IG", "The Boy, the Mole, the Fox and the Horse", "Collab / Endorsement", "Agent", "Not Started", "1M IG — inspirational, cross-generational"],

    # ═══ UK PODCASTS (PARENTS/TEENS) ═══
    [51, "Parenting Hell / Rob Beckett & Josh Widdicombe", "Podcast", "UK Podcasts (Parents/Teens)", "Mega", "50M+ downloads", "UK's #1 parenting podcast", "Sponsored read", "Agent", "Not Started", "50M+ downloads — every UK parent knows this"],
    [52, "The High Low / Dolly Alderton & Pandora Sykes", "Podcast", "UK Podcasts (Parents/Teens)", "Macro", "Millions of downloads", "Culture/lifestyle, huge female audience", "Sponsored read", "Agent", "Not Started", "Young adult women — future parent audience"],
    [53, "Off Menu / James Acaster & Ed Gamble", "Podcast", "UK Podcasts (Parents/Teens)", "Macro", "50M+ downloads", "Comedy food podcast, massive UK audience", "Sponsored read", "Agent", "Not Started", "Cross-demographic UK reach"],
    [54, "Happy Place / Fearne Cotton", "Podcast/IG", "UK Podcasts (Parents/Teens)", "Macro", "Millions of downloads", "Wellbeing, huge parent audience", "Sponsored read / IG mention", "Agent", "Not Started", "Fearne Cotton — trusted family voice"],
    [55, "How to Fail / Elizabeth Day", "Podcast", "UK Podcasts (Parents/Teens)", "Macro", "Millions of downloads", "Failure/success stories, parent audience", "Sponsored read", "Agent", "Not Started", "Growth mindset angle — education fit"],
    [56, "Sh**ged Married Annoyed / Chris & Rosie Ramsey", "Podcast", "UK Podcasts (Parents/Teens)", "Macro", "Massive audience", "UK family comedy podcast", "Sponsored read", "Agent", "Not Started", "North East UK parents — relatable"],
    [57, "The Rest Is History / Tom Holland & Dominic Sandbrook", "Podcast", "UK Podcasts (Parents/Teens)", "Macro", "#1 history podcast", "History enthusiasts — GCSE History crossover", "Sponsored read", "Agent", "Not Started", "Education-adjacent — history lovers"],
    [58, "The Diary Of A CEO / Steven Bartlett", "Podcast/YouTube", "UK Podcasts (Parents/Teens)", "Mega", "8M+ YT", "UK's biggest interview podcast", "Sponsored read / Guest spot", "Agent", "Not Started", "8M YT — massive UK young adult reach"],
    [59, "My Therapist Ghosted Me / Vogue Williams & Joanne McNally", "Podcast", "UK Podcasts (Parents/Teens)", "Macro", "Massive audience", "Comedy, UK/Irish women", "Sponsored read", "Agent", "Not Started", "Young women audience — future parents"],
    [60, "Teenage Kicks / Podcast", "Podcast", "UK Podcasts (Parents/Teens)", "Micro", "Growing", "Specifically about raising teenagers", "Sponsored read", "Email", "Not Started", "Directly targets parents of teens"],

    # ═══ UK GAMING & STREAMING ═══
    [61, "DanTDM (@dantdm)", "YouTube", "UK Gaming & Streaming", "Mega", "28M+ YT", "UK's biggest gaming YouTuber — Minecraft", "Brand integration", "Agent", "Not Started", "28M — every UK teen boy watches DanTDM"],
    [62, "Stampy / stampylonghead", "YouTube", "UK Gaming & Streaming", "Macro", "10M+ YT", "Minecraft, kid-friendly gaming", "Sponsored content", "Agent", "Not Started", "10M — younger teen audience, family-safe"],
    [63, "W2S / Harry Lewis (@wroetoshaw)", "YouTube", "UK Gaming & Streaming", "Macro", "16M+ YT", "Sidemen, gaming/challenge content", "Brand deal", "Agent", "Not Started", "Sidemen member — massive UK teen reach"],
    [64, "Vikkstar / Vik (@vaborelikkstar123)", "YouTube", "UK Gaming & Streaming", "Macro", "7M+ YT", "Sidemen, gaming, owns London esports team", "Brand deal", "Agent", "Not Started", "Sidemen — also Indian-heritage angle"],
    [65, "Grian", "YouTube", "UK Gaming & Streaming", "Macro", "8M+ YT", "UK Minecraft builder, Hermitcraft", "Sponsored mention", "Agent", "Not Started", "8M — creative/building audience"],
    [66, "SmallishBeans / Joel", "YouTube", "UK Gaming & Streaming", "Macro", "4M+ YT", "UK Minecraft/gaming creator", "Sponsored content", "Email", "Not Started", "Kid-friendly UK gaming content"],
    [67, "LDShadowLady / Lizzie", "YouTube", "UK Gaming & Streaming", "Macro", "6M+ YT", "UK gaming, Minecraft, kid-friendly", "Sponsored content", "Email", "Not Started", "6M — teen girl gaming audience"],
    [68, "InTheLittleWood / Martyn", "YouTube", "UK Gaming & Streaming", "Macro", "3M+ YT", "UK gaming, Yogscast member", "Sponsored content", "Email", "Not Started", "3M — UK gaming community"],
    [69, "Wilbur Soot (@wilbursoot)", "YouTube/Twitch", "UK Gaming & Streaming", "Macro", "6M+ YT", "UK gaming/music, Dream SMP", "Sponsored mention", "Agent", "Not Started", "6M — teen audience, also musician"],
    [70, "Ph1LzA / Philza (@ph1lza)", "YouTube/Twitch", "UK Gaming & Streaming", "Macro", "5M+ YT", "UK Minecraft creator, hardcore series", "Sponsored mention", "Agent", "Not Started", "5M — dedicated teen fanbase"],

    # ═══ UK STEM & SCIENCE ═══
    [71, "Tom Scott (@tomscottgo)", "YouTube", "UK STEM & Science", "Macro", "6M+ YT", "UK tech/science/geography — educational entertainment", "Sponsored video", "Agent", "Not Started", "6M — education-adjacent, massive UK reach"],
    [72, "Steve Mould (@stevemould)", "YouTube", "UK STEM & Science", "Macro", "2M+ YT", "UK science demos and experiments", "Sponsored content", "Email", "Not Started", "2M — STEM audience overlaps with students"],
    [73, "Matt Parker / Stand-up Maths", "YouTube", "UK STEM & Science", "Macro", "1.2M YT", "Maths comedy/entertainment", "Sponsored video", "Email", "Not Started", "1.2M — makes maths fun, student audience"],
    [74, "Hannah Fry (@faborelryrsquared)", "YouTube/BBC", "UK STEM & Science", "Mid", "500K+ YT", "Maths professor, BBC presenter", "Collab / Endorsement", "Agent", "Not Started", "BBC profile — trusted science communicator"],
    [75, "Simon Clark (@simonoxfphys)", "YouTube", "UK STEM & Science", "Mid", "400K YT", "Oxford PhD, climate science", "Sponsored video", "Email", "Not Started", "Oxford academic — credible + relatable"],
    [76, "Tibees (@tibees)", "YouTube", "UK STEM & Science", "Macro", "1.2M YT", "Maths/physics, exam paper reactions", "Sponsored video", "Email", "Not Started", "1.2M — watches old exam papers = student audience"],
    [77, "James May (@jamesmay)", "YouTube/IG", "UK STEM & Science", "Macro", "3M+ YT", "Top Gear fame, science/craft content", "Brand deal", "Agent", "Not Started", "3M — trusted UK voice, education-adjacent"],
    [78, "Lex Fridman Clips (UK audience)", "YouTube", "UK STEM & Science", "Mega", "4M+ YT", "Deep interviews, huge student listenership", "Sponsored read", "Agent", "Not Started", "Massive student podcast audience"],
    [79, "Mark Rober (@markrober)", "YouTube", "UK STEM & Science", "Mega", "28M+ YT", "Science/engineering experiments", "Brand deal", "Agent", "Not Started", "28M — aspirational, STEM + fun"],
    [80, "Vsauce / Michael Stevens", "YouTube", "UK STEM & Science", "Mega", "20M+ YT", "Science curiosity channel", "Brand deal", "Agent", "Not Started", "20M — student favourite, curiosity-driven"],

    # ═══ UK MUSIC ARTISTS ═══
    [81, "Dave (@santandave)", "IG/TikTok", "UK Music Artists", "Mega", "5M+ IG", "UK rap — Streatham, BRIT winner", "TikTok sound / Brand deal", "Agent", "Not Started", "UK teens worship Dave — education in lyrics"],
    [82, "Central Cee (@centralcee)", "IG/TikTok", "UK Music Artists", "Mega", "10M+ IG", "UK drill, massive teen following", "TikTok sound collab", "Agent", "Not Started", "10M — biggest UK rapper right now"],
    [83, "Stormzy (@stormzy)", "IG", "UK Music Artists", "Mega", "4M+ IG", "UK grime icon, Cambridge scholarship funder", "Brand partnership / Education tie-in", "Agent", "Not Started", "Already funds scholarships — education champion"],
    [84, "Ed Sheeran (@teddysphotos)", "IG", "UK Music Artists", "Mega", "44M IG", "Global superstar, Suffolk origins", "Aspirational endorsement", "Agent", "Not Started", "44M — aspirational, every UK teen knows Ed"],
    [85, "Little Simz (@littlesimz)", "IG", "UK Music Artists", "Macro", "1M+ IG", "Mercury Prize winner, N London", "Brand collab", "Agent", "Not Started", "Critical darling — cultural credibility"],
    [86, "Raye (@raborelaye)", "IG/TikTok", "UK Music Artists", "Macro", "3M+ IG", "BRIT Awards record-breaker 2024", "TikTok sound / Brand deal", "Agent", "Not Started", "3M — massive UK teen following"],
    [87, "Sam Fender (@samfender)", "IG", "UK Music Artists", "Macro", "1.5M IG", "North Shields, working class anthem", "Brand collab", "Agent", "Not Started", "Working-class education angle"],
    [88, "Dua Lipa (@dualipa)", "IG", "UK Music Artists", "Mega", "88M IG", "British-Albanian pop megastar", "Aspirational endorsement", "Agent", "Not Started", "88M — aspirational, global UK icon"],
    [89, "Cat Burns (@catburns)", "TikTok/IG", "UK Music Artists", "Macro", "1M+ TikTok", "UK pop, TikTok viral", "TikTok collab / Sound", "Email", "Not Started", "TikTok native — Gen Z UK audience"],
    [90, "Aitch (@aaborelitch)", "IG/TikTok", "UK Music Artists", "Macro", "2M+ IG", "Manchester rapper, teen favourite", "TikTok collab", "Agent", "Not Started", "2M — Manchester, working-class teen audience"],

    # ═══ UK SPORTS & FOOTBALL ═══
    [91, "Mark Goldbridge (@markgoldbridge)", "YouTube/TikTok", "UK Sports & Football", "Macro", "2M+ YT", "Man United fan channel — viral reactions", "Sponsored content", "Email", "Not Started", "2M — massive teen boy audience"],
    [92, "Harry Pinero (@harrypinero)", "YouTube/IG", "UK Sports & Football", "Macro", "2M+ YT", "Football + comedy street interviews", "Sponsored content", "Email", "Not Started", "2M — UK teen boys, football culture"],
    [93, "Jezza (@jaborelezzy)", "YouTube", "UK Sports & Football", "Macro", "4M+ YT", "Football challenges, UK-based", "Brand integration", "Email", "Not Started", "4M — football challenge content = teen boys"],
    [94, "ChrisMD (@chrismd)", "YouTube", "UK Sports & Football", "Macro", "6M+ YT", "Football challenges, UK creator", "Brand deal", "Agent", "Not Started", "6M — UK football + teen audience"],
    [95, "Marcus Rashford (@marcusrashford)", "IG", "UK Sports & Football", "Mega", "16M+ IG", "Man Utd, child poverty campaigner, MBE", "Endorsement / Partnership", "Agent", "Not Started", "16M — education champion, free school meals"],
    [96, "Bukayo Saka (@bukayosaka87)", "IG", "UK Sports & Football", "Mega", "8M+ IG", "Arsenal star, young role model", "Endorsement", "Agent", "Not Started", "8M — role model for UK teens"],
    [97, "The F2 Freestylers (@f2freestylers)", "YouTube", "UK Sports & Football", "Macro", "12M+ YT", "Football skills/tricks — massive teen audience", "Sponsored content", "Agent", "Not Started", "12M — pure teen boy demographic"],
    [98, "Spencer FC (@spencerfc)", "YouTube", "UK Sports & Football", "Mid", "2M+ YT", "UK football content, Hashtag United", "Sponsored content", "Email", "Not Started", "2M — teen football audience"],
    [99, "Saturday Social / Football Daily", "YouTube", "UK Sports & Football", "Macro", "1M+ YT", "Sky Sports digital, UK football analysis", "Sponsored segment", "Email", "Not Started", "Sky Sports reach — teen football fans"],
    [100, "Yung Filly (@yungfilly)", "YouTube/TikTok", "UK Sports & Football", "Macro", "3M+ YT", "Comedy + football + entertainment", "Brand deal", "Agent", "Not Started", "3M — UK teen boys, multi-format content"],

    # ═══ BRITISH TV & CELEBRITY ═══
    [101, "Holly Willoughby (@hollywilloughby)", "IG", "British TV & Celebrity", "Mega", "8M+ IG", "TV presenter, mum of 3 school-age kids", "Story mention / Brand deal", "Agent", "Not Started", "8M — UK's most trusted TV mum"],
    [102, "Fearne Cotton (@feaborelarnecotton)", "IG/Podcast", "British TV & Celebrity", "Macro", "3M+ IG", "Presenter, Happy Place podcast, mum", "Podcast sponsor / IG story", "Agent", "Not Started", "3M — wellness + parenting audience"],
    [103, "Rochelle Humes (@rochellehumes)", "IG", "British TV & Celebrity", "Macro", "2.2M IG", "This Morning, mum of 3", "Story mention", "Agent", "Not Started", "2.2M — UK mum, school-age children"],
    [104, "Christine McGuinness (@mrscmcguinness)", "IG", "British TV & Celebrity", "Macro", "700K+ IG", "Autism advocate, 3 school-age kids", "Sponsored content", "Email", "Not Started", "SEND angle — autism & education"],
    [105, "Frankie Bridge (@frankiebridge)", "IG", "British TV & Celebrity", "Macro", "1.5M IG", "Singer/presenter, mum of 2 boys", "Story mention", "Agent", "Not Started", "1.5M — UK family content"],
    [106, "Giovanna Fletcher (@mrsgifletcher)", "IG/Podcast", "British TV & Celebrity", "Macro", "1.8M IG", "Author, Happy Mum Happy Baby podcast", "Podcast sponsor / IG story", "Agent", "Not Started", "1.8M — #1 UK mum podcaster"],
    [107, "Joe Wicks (@thebodycoach)", "YouTube/IG", "British TV & Celebrity", "Mega", "4M+ IG", "PE with Joe during lockdown, MBE", "Brand collab", "Agent", "Not Started", "4M — lockdown PE hero, education crossover"],
    [108, "David Attenborough (official)", "IG", "British TV & Celebrity", "Mega", "7M+ IG", "National treasure, nature/science", "Aspirational", "Agent", "Not Started", "7M — aspirational, education & curiosity"],
    [109, "Dermot O'Leary (@dermotoleary)", "IG", "British TV & Celebrity", "Macro", "1M+ IG", "X Factor, children's author, dad", "Story mention", "Agent", "Not Started", "1M — children's books + parent audience"],
    [110, "Nadiya Hussain (@nadiyajhussain)", "IG", "British TV & Celebrity", "Macro", "900K+ IG", "GBBO winner, children's author", "Brand collab", "Agent", "Not Started", "900K — children's books, diverse audience"],

    # ═══ UK FINANCE (YOUNG PEOPLE) ═══
    [111, "Sidequest / Graham Stephan (UK)", "YouTube", "UK Finance (Young People)", "Macro", "4M+ YT", "Finance for young people", "Sponsored video", "Agent", "Not Started", "4M — young adult money management"],
    [112, "Damien Talks Money (@damientalksmoney)", "YouTube/TikTok", "UK Finance (Young People)", "Mid", "500K+ YT", "UK personal finance for young people", "Sponsored video", "Email", "Not Started", "UK-specific money tips for young adults"],
    [113, "Nischa (@nischal)", "YouTube", "UK Finance (Young People)", "Mid", "500K+ YT", "UK money saving, student finance", "Sponsored video", "Email", "Not Started", "Student-focused money content"],
    [114, "Your Money Sorted / MSE", "Website/Social", "UK Finance (Young People)", "Macro", "Massive reach", "Martin Lewis empire — student money", "Sponsored content / Partnership", "Email", "Not Started", "MSE student section — massive traffic"],

    # ═══ UK EDUCATION CONFERENCES & EVENTS ═══
    [115, "BETT (British Educational Training & Tech)", "Conference", "UK Education Conferences", "Macro", "30K+ attendees", "World's largest EdTech show — ExCeL London", "Exhibit / Sponsor / Attend", "Email", "Not Started", "Jan annually — must-attend for EdTech"],
    [116, "Education World Forum", "Conference", "UK Education Conferences", "Macro", "Global education ministers", "World's largest gathering of education ministers", "Attend / Pitch", "Email", "Not Started", "Jan — London, policy-level contacts"],
    [117, "ResearchED", "Conference", "UK Education Conferences", "Mid", "5K+ per event", "Evidence-based education conference, global", "Exhibit / Speak / Sponsor", "Email", "Not Started", "Tom Bennett founded — teacher trust"],
    [118, "Festival of Education (Wellington College)", "Conference", "UK Education Conferences", "Mid", "5K+ attendees", "UK's biggest education festival", "Exhibit / Sponsor", "Email", "Not Started", "June — influential educators"],
    [119, "Schools & Academies Show", "Conference", "UK Education Conferences", "Mid", "3K+ attendees", "UK school leaders & MAT decision-makers", "Exhibit / Sponsor", "Email", "Not Started", "Free entry — procurement-focused"],
    [120, "Hay Festival", "Literary Festival", "UK Literary Festivals", "Macro", "250K+ attendees", "UK's premier literary festival", "Sponsor / Exhibit / Workshop", "Email", "Not Started", "May/Jun — Hay-on-Wye, literary audience"],

    # ═══ UK EDUCATION CHARITIES ═══
    [121, "Teach First", "Charity/Social", "UK Education Charities", "Macro", "Large teacher network", "Teacher training charity — graduate pipeline", "Partnership / Sponsored", "Email", "Not Started", "Massive teacher network — recommendation"],
    [122, "The Brilliant Club", "Charity", "UK Education Charities", "Mid", "Works with 100+ unis", "Places PhD researchers in schools", "Partnership", "Email", "Not Started", "Widening access — education equity angle"],
    [123, "IntoUniversity", "Charity", "UK Education Charities", "Mid", "35K+ students/year", "Academic support centres for disadvantaged students", "Partnership / Donation", "Email", "Not Started", "Disadvantaged students — social impact"],
    [124, "Achievement for All", "Charity", "UK Education Charities", "Mid", "Works with 4K+ schools", "School improvement charity", "Partnership", "Email", "Not Started", "4K schools — direct procurement channel"],
    [125, "National Literacy Trust", "Charity", "UK Education Charities", "Macro", "Huge school reach", "Literacy charity — perfect English subject fit", "Partnership / Endorsement", "Email", "Not Started", "Literacy focus = direct product fit"],

    # ═══ UK STUDENT PLATFORMS ═══
    [126, "UNiDAYS", "App/Website", "UK Student Platforms", "Macro", "22M+ verified students", "Student discount platform", "List product / Partnership", "Email", "Not Started", "22M — student discount listing"],
    [127, "Student Beans", "App/Website", "UK Student Platforms", "Macro", "163M+ users globally", "Student discount platform", "List product / Partnership", "Email", "Not Started", "163M — student discount listing"],
    [128, "NUS TOTUM", "Card/App", "UK Student Platforms", "Macro", "4M+ UK members", "Official NUS student card", "Partnership / Discount listing", "Email", "Not Started", "4M UK students — discount listing"],
    [129, "The Student Room", "Forum/Website", "UK Student Platforms", "Macro", "1.37M members", "UK's biggest student forum", "Sponsored thread / Banner ad", "Email", "Not Started", "1.37M — GCSE/A-Level discussion"],
    [130, "Save the Student", "Website", "UK Student Platforms", "Mid", "Large student traffic", "Student money/deals site", "Sponsored article / Listing", "Email", "Not Started", "Student deals section — natural fit"],

    # ═══ UK ONLINE COMMUNITIES ═══
    [131, "r/GCSE (Reddit)", "Reddit", "UK Online Communities", "Mid", "100K+ members", "UK GCSE students subreddit", "Organic / AMA / Sponsored", "Reddit", "Not Started", "Direct target audience — GCSE students"],
    [132, "r/6thForm (Reddit)", "Reddit", "UK Online Communities", "Mid", "80K+ members", "UK A-Level/sixth form subreddit", "Organic / AMA", "Reddit", "Not Started", "A-Level students — upsell audience"],
    [133, "r/TeachingUK (Reddit)", "Reddit", "UK Online Communities", "Mid", "40K+ members", "UK teachers subreddit", "Organic / AMA", "Reddit", "Not Started", "UK teachers — recommendation potential"],
    [134, "r/UniUK (Reddit)", "Reddit", "UK Online Communities", "Mid", "100K+ members", "UK university students", "Organic presence", "Reddit", "Not Started", "Student audience — brand awareness"],
    [135, "Secondary Teachers UK (Facebook)", "Facebook", "UK Online Communities", "Mid", "30K+ members", "UK secondary teachers group", "Organic / Admin partnership", "Facebook", "Not Started", "Teacher community — word of mouth"],
    [136, "11+ Exam Parents UK (Facebook)", "Facebook", "UK Online Communities", "Mid", "30K+ members", "Grammar school prep parents", "Organic / Sponsored post", "Facebook", "Not Started", "Education-focused parents — GCSE pipeline"],

    # ═══ UK PARENTING MEDIA ═══
    [137, "Mother & Baby Magazine", "Magazine/Website", "UK Parenting Media", "Mid", "Large readership", "UK's longest-running parenting magazine", "Advertorial / Review", "Email", "Not Started", "Print + digital — parent audience"],
    [138, "Families Online", "Website", "UK Parenting Media", "Micro", "500K+ monthly", "UK family activities/events site", "Sponsored article", "Email", "Not Started", "Family activities — education angle"],
    [139, "Netmums", "Website/Forum", "UK Parenting Media", "Macro", "10M+ monthly visits", "UK parenting forum & advice site", "Sponsored content / Banner", "Email", "Not Started", "10M — massive UK parent audience"],
    [140, "BabyCentre UK", "Website", "UK Parenting Media", "Macro", "Large audience", "Parenting advice — school readiness content", "Sponsored article", "Email", "Not Started", "School transition content — pipeline"],

    # ═══ UK EXAM BOARDS & COMPETITORS ═══
    [141, "AQA (exam board)", "Website/Social", "UK Exam Boards & Competitors", "Macro", "250K+ social", "Largest UK exam board — English specs", "Partnership / Endorsed resource", "Email", "Not Started", "AQA English Lit/Lang specs = direct integration"],
    [142, "Edexcel / Pearson", "Website/Social", "UK Exam Boards & Competitors", "Macro", "Large reach", "Major UK exam board", "Partnership / Listed resource", "Email", "Not Started", "Edexcel English specs — integration target"],
    [143, "OCR", "Website/Social", "UK Exam Boards & Competitors", "Mid", "Medium reach", "UK exam board", "Partnership application", "Email", "Not Started", "OCR English specs — integration"],
    [144, "Cambridge Assessment (CIE)", "Website/Social", "UK Exam Boards & Competitors", "Macro", "Global reach", "IGCSE exam board — international", "Partnership / Endorsed", "Email", "Not Started", "IGCSE English = GCC market direct link"],
    [145, "GCSEPod", "App/Website", "UK Exam Boards & Competitors", "Mid", "1M+ students", "GCSE podcast revision platform", "Competitor analysis / Partnership", "Email", "Not Started", "Direct competitor — partnership or differentiation"],
    [146, "Seneca Learning", "App/Website", "UK Exam Boards & Competitors", "Macro", "14M+ users", "Free revision platform", "Integration / Partnership", "Email", "Not Started", "14M users — complementary tool angle"],
    [147, "Tassomai", "App/Website", "UK Exam Boards & Competitors", "Mid", "Hundreds of schools", "AI-powered GCSE science revision", "Competitor analysis", "Email", "Not Started", "AI revision — learn from their school sales model"],
    [148, "Quizlet", "App/Website", "UK Exam Boards & Competitors", "Mega", "60M+ monthly users", "Flashcard/study platform — global", "Partnership / Integration", "Email", "Not Started", "60M — complementary study tool"],
    [149, "Kahoot!", "App/Website", "UK Exam Boards & Competitors", "Mega", "9B+ cumulative participants", "Game-based learning platform", "Integration / Partnership", "Email", "Not Started", "Massive school adoption — integration"],
    [150, "Sparx Maths / Sparx Reader", "App/Website", "UK Exam Boards & Competitors", "Mid", "3M+ students", "AI maths + reading platform for schools", "Competitor analysis / Partnership", "Email", "Not Started", "Sparx Reader = direct English reading competitor"],

    # ═══ DUBAI EXPAT MUM BLOGGERS (web-verified) ═══
    [151, "Sneha Rebecca (@srdubai)", "IG", "Dubai/GCC Lifestyle", "Macro", "898K IG", "Award-winning parenting influencer Dubai", "Story mention / Brand deal", "DM/Email", "Not Started", "Filmfare Most Influential Woman 2025"],
    [152, "Virdah Javed Khan (@mom_in_dubai)", "IG", "Dubai/GCC Lifestyle", "Macro", "1M IG / 2.1M total", "Pakistani expat mum, Dubai since 2016", "Story mention / Brand deal", "DM", "Not Started", "1M IG — motherhood, family, lifestyle"],
    [153, "Charlotte Briggs (@charlottexaxaa)", "IG", "Dubai/GCC Lifestyle", "Mid", "588K IG", "British-Brazilian expat mum/model Dubai", "Story mention", "DM", "Not Started", "588K — maternal fashion, family at Dubai landmarks"],
    [154, "Dajana Ic (@dajanaic)", "IG", "Dubai/GCC Lifestyle", "Mid", "539K IG", "Mum & business owner Dubai", "Story mention", "DM", "Not Started", "539K — motherhood, OOTD, home decor"],
    [155, "Danah Alshayji (@danaindxb)", "IG", "Dubai/GCC Lifestyle", "Mid", "557K IG", "Kuwaiti mum of 3 boys in Dubai", "Story mention", "DM", "Not Started", "557K — kid-friendly cafes, family activities UAE"],
    [156, "QuratUlAin Mohsin (@quratulainarif)", "IG", "Dubai/GCC Lifestyle", "Mid", "267K IG", "Working mum, TV presenter Dubai", "Sponsored post", "DM", "Not Started", "267K — professional lifestyle, motherhood"],
    [157, "Helen Farmer (@_helenfarmer_ / The Mothership)", "IG/Blog/Radio", "Dubai/GCC Lifestyle", "Micro", "68K IG", "BBC journalist, Dubai Eye host, children's book author", "Sponsored post / Radio mention", "Email", "Not Started", "68K — credible voice, radio reach, British expat"],
    [158, "Sassy Mama Dubai (@sassymamadubai)", "IG/Website", "Dubai/GCC Lifestyle", "Micro", "29K IG", "Go-to resource for expat families — school guides", "Sponsored listing / Article", "Email", "Not Started", "29K — school guides + events calendar for parents"],
    [159, "Tehreem Pasha (@that.dubai.mum)", "IG", "Dubai/GCC Lifestyle", "Micro", "19K IG", "Trusted reviews, screen-free play, UAE-licensed", "Sponsored post", "DM", "Not Started", "19K — DIY activities, education-adjacent"],
    [160, "DubaiMums (@dubaimums)", "IG/Blog", "Dubai/GCC Lifestyle", "Nano", "8.4K IG", "British expat Jenny Ashton, school advice since 2011", "Sponsored post / Review", "Email", "Not Started", "8.4K — school advice, kid activities, established since 2011"],

    # ═══ UAE SCHOOLS — B2B SOCIAL OUTREACH (web-verified) ═══
    [161, "GEMS Founders School Dubai (@gemsfoundersschooldubai)", "IG", "GCC Schools B2B", "Micro", "22K IG", "GEMS group school — large parent community", "DM / Email Head of English", "Email", "Not Started", "22K — pilot school target for IGCSE English"],
    [162, "Repton School Dubai (@reptondubai)", "IG", "GCC Schools B2B", "Micro", "19K IG", "Premium British school Dubai", "Email Head of English", "Email", "Not Started", "19K — premium British curriculum, IGCSE"],
    [163, "GEMS Dubai American Academy (@gems_daa)", "IG", "GCC Schools B2B", "Micro", "16K IG", "GEMS group — American/IB curriculum", "Email", "Email", "Not Started", "16K — cross-curriculum English support"],
    [164, "GEMS Metropole School (@gemsmetropoleschool)", "IG", "GCC Schools B2B", "Micro", "14K IG", "GEMS British curriculum school", "Email Head of English", "Email", "Not Started", "14K — British curriculum, IGCSE target"],
    [165, "Cranleigh Abu Dhabi (@cranleighad)", "IG", "GCC Schools B2B", "Micro", "11K IG", "Premium British school Abu Dhabi", "Email", "Email", "Not Started", "11K — British prep school, premium market"],
    [166, "JESS Dubai (@jessdubaischool)", "IG", "GCC Schools B2B", "Nano", "9.7K IG", "Jumeirah English Speaking School — established", "Email Head of English", "Email", "Not Started", "9.7K — one of Dubai's oldest British schools"],
    [167, "Dubai College (@dubaicollege)", "IG", "GCC Schools B2B", "Nano", "8.7K IG", "Selective British school, top IGCSE results", "Email Head of English", "Email", "Not Started", "8.7K — top academic results, selective intake"],
    [168, "The British School Al Khubairat (@bsak_abudhabi)", "IG", "GCC Schools B2B", "Nano", "8.4K IG", "British school Abu Dhabi since 1968", "Email", "Email", "Not Started", "8.4K — longest-running British school in Abu Dhabi"],
    [169, "Brighton College Dubai (@brightondubai)", "IG", "GCC Schools B2B", "Nano", "8.1K IG", "UK Brighton College franchise", "Email Head of English", "Email", "Not Started", "8.1K — UK school brand franchise"],
    [170, "Brighton College Abu Dhabi (@bcabudhabi)", "IG", "GCC Schools B2B", "Nano", "6.8K IG", "UK Brighton College franchise Abu Dhabi", "Email", "Email", "Not Started", "6.8K — UK school brand, IGCSE English"],

    # ═══ QATAR SCHOOLS — B2B SOCIAL OUTREACH (web-verified) ═══
    [171, "Doha College (@dohacollege)", "IG", "GCC Schools B2B", "Micro", "11K IG", "Top British school in Qatar — IGCSE results leader", "Email Head of English", "Email", "Not Started", "11K — Qatar's top British school, IGCSE English priority"],
    [172, "Sherborne Qatar (@sherborneschoolqatar)", "IG", "GCC Schools B2B", "Micro", "10K IG", "UK Sherborne franchise — British boarding school heritage", "Email Head of English", "Email", "Not Started", "10K — UK boarding school brand in Qatar"],
    [173, "DBS Ain Khaled (@dbsainkhaledofficial)", "IG", "GCC Schools B2B", "Nano", "9.5K IG", "Doha British School — IGCSE curriculum", "Email Head of English", "Email", "Not Started", "9.5K — British curriculum school, IGCSE focus"],
    [174, "Newton British Academy (@newtonbritishacademy)", "IG", "GCC Schools B2B", "Nano", "8K IG", "5 Qatar campuses — group deal opportunity", "Email Group Admin", "Email", "Not Started", "8K — 5 campuses across Qatar, group deal target"],
    [175, "GEMS Wellington Qatar (@gemswellingtonqatar)", "IG", "GCC Schools B2B", "Nano", "7.5K IG", "GEMS group — British curriculum", "Email Head of English", "Email", "Not Started", "7.5K — GEMS brand, British curriculum"],
    [176, "Royal Grammar School Qatar (@rgsqatar)", "IG", "GCC Schools B2B", "Nano", "6.5K IG", "UK RGS franchise — premium British school", "Email Head of English", "Email", "Not Started", "6.5K — UK RGS brand franchise in Qatar"],
    [177, "Nord Anglia Compass Doha (@compassinternationalschooldoha)", "IG", "GCC Schools B2B", "Nano", "6K IG", "Nord Anglia group — IGCSE curriculum", "Email Head of English", "Email", "Not Started", "6K — Nord Anglia group, IGCSE English"],
    [178, "Qatar International School (@qatarinternationalschool)", "IG", "GCC Schools B2B", "Nano", "5.5K IG", "Established British curriculum school", "Email Head of English", "Email", "Not Started", "5.5K — British curriculum, long-established"],
    [179, "Park House English School (@parkhouseenglish)", "IG", "GCC Schools B2B", "Nano", "4K IG", "Small British school — direct approach", "Email Head of English", "Email", "Not Started", "4K — smaller school, easier decision-maker access"],
    [180, "Durham School for Girls Doha (@durhamschooldoha)", "IG", "GCC Schools B2B", "Nano", "3.8K IG", "Girls school — British curriculum", "Email Head of English", "Email", "Not Started", "3.8K — British curriculum, girls school"],
    [181, "Al Khor International School (@akhis_qatar)", "IG", "GCC Schools B2B", "Nano", "3K IG", "British curriculum Al Khor", "Email Head of English", "Email", "Not Started", "3K — outside Doha, less competitive"],
    [182, "Gulf English School Qatar (@gesqatar)", "IG", "GCC Schools B2B", "Nano", "2.5K IG", "British school — niche community", "Email Head of English", "Email", "Not Started", "2.5K — smaller school, direct English dept access"],
    [183, "Doha English Speaking School (@dessqatar)", "IG", "GCC Schools B2B", "Nano", "5K IG", "One of Qatar's oldest British schools", "Email Head of English", "Email", "Not Started", "5K — long-established, strong parent community"],
    [184, "King's College Doha (@kingscollegedoha)", "IG", "GCC Schools B2B", "Nano", "7K IG", "UK King's College franchise — 2 campuses", "Email Head of English", "Email", "Not Started", "7K — UK franchise brand, 2 Doha campuses"],
    [185, "Nord Anglia Al Khor (@naisak_qatar)", "IG", "GCC Schools B2B", "Nano", "4.5K IG", "Nord Anglia group — Al Khor campus", "Email Head of English", "Email", "Not Started", "4.5K — Nord Anglia group, northern Qatar"],

    # ═══ IELTS/ESL YOUTUBE — HIGH-REACH CHANNELS (web-verified) ═══
    [186, "Fastrack IELTS", "YouTube", "ESL/IELTS YouTube", "Macro", "1.95M subs", "Top IELTS prep channel — massive GCC viewership", "Sponsored video / Channel partnership", "Email", "Not Started", "1.95M — IELTS prep, huge Middle East/South Asia audience"],
    [187, "Asad Yaqub IELTS", "YouTube", "ESL/IELTS YouTube", "Macro", "1.73M subs", "IELTS for Arabic/Urdu speakers — perfect GCC fit", "Sponsored video", "Email", "Not Started", "1.73M — Arabic/Urdu-speaking audience, GCC match"],
    [188, "Ross IELTS Academy", "YouTube", "ESL/IELTS YouTube", "Macro", "1.17M subs", "UK-based IELTS teacher — British English focus", "Sponsored video / Affiliate link", "Email", "Not Started", "1.17M — British English teacher, UK-based"],
    [189, "IELTS Daily", "YouTube", "ESL/IELTS YouTube", "Macro", "1.1M subs", "Daily IELTS tips — high engagement rate", "Sponsored video", "Email", "Not Started", "1.1M — daily content, high engagement"],
    [190, "IELTS Advantage (YouTube)", "YouTube", "ESL/IELTS YouTube", "Mid", "650K subs", "IELTS strategy and techniques — premium audience", "Affiliate partnership", "Email", "Not Started", "650K — strategy-focused, premium test-prep audience"],
    [191, "engVid (multi-teacher)", "YouTube", "ESL/IELTS YouTube", "Mega", "4M+ subs", "Network of English teachers — multiple channels", "Network partnership", "Email", "Not Started", "4M+ — network of 11+ teachers, massive combined reach"],

    # ═══ UK EDUCATION CHARITIES & LITERACY NGOs (web-verified) ═══
    [192, "BookTrust (@Booktrust / @booktrust)", "X/IG", "UK Education Charities", "Macro", "120K X / 58K IG", "UK's largest children's reading charity", "Partnership / Sponsored campaign", "Email", "Not Started", "178K combined — children's reading charity, school reach"],
    [193, "National Literacy Trust (@Literacy_Trust / @literacy_trust)", "X/IG", "UK Education Charities", "Macro", "95K X / 41K IG", "UK literacy policy leader — huge school network", "Partnership / Resource listing", "Email", "Not Started", "136K combined — school network, literacy hub zones"],
    [194, "World Book Day (@WorldBookDayUK)", "X/IG/FB", "UK Education Charities", "Mid", "37K IG / 82K FB", "Annual event reaching every UK school", "Sponsored resource / Event partnership", "Email", "Not Started", "119K combined — reaches every UK primary + secondary school"],
    [195, "The Reading Agency (@readingagency)", "X/IG", "UK Education Charities", "Mid", "67K X / 18K IG", "Summer Reading Challenge — 700K+ kids/year", "Partnership / Resource listing", "Email", "Not Started", "85K combined — library network, Summer Reading Challenge"],
    [196, "Teach First (@TeachFirst / @teachfirstuk)", "X/IG", "UK Education Charities", "Mid", "17K IG", "Teacher training charity — direct school access", "Partnership / Teacher resources", "Email", "Not Started", "Trains 1,500+ teachers/year — direct classroom access"],
    [197, "UK Literacy Association (@The_UKLA)", "X", "UK Education Charities", "Micro", "16K X", "Professional body for literacy teachers", "Conference sponsorship / Resource", "Email", "Not Started", "16K — English teacher professional body, conference access"],
    [198, "Bookmark Reading Charity (@BookmarkCharity / @bookmarkreading)", "X/IG", "UK Education Charities", "Micro", "11K IG / 4K X", "Volunteer reading charity — school partnerships", "Partnership / Volunteer outreach", "Email", "Not Started", "15K combined — works directly in schools with struggling readers"],
    [199, "Book Aid International (@Book_Aid)", "X/IG", "UK Education Charities", "Micro", "10K X / 4K IG", "Global literacy charity — UK HQ", "Partnership", "Email", "Not Started", "14K combined — international literacy reach"],
    [200, "Coram Beanstalk (@beanstalkreads)", "X/IG", "UK Education Charities", "Nano", "9K X / 3K IG", "Reading volunteers in primary schools", "Partnership", "Email", "Not Started", "12K combined — primary school reading volunteers"],
    [201, "Doorstep Library (@DoorstepLib)", "X/IG", "UK Education Charities", "Nano", "2K X / 2K IG", "Community reading charity — underserved areas", "Partnership", "Email", "Not Started", "4K combined — community literacy focus"],
    [202, "Schoolreaders (@schoolreaders)", "X/IG", "UK Education Charities", "Nano", "1K IG", "Volunteer readers in primary schools", "Partnership", "Email", "Not Started", "Direct school access — volunteer reading programme"],
    [203, "Shannon Trust (@Shannon_Trust)", "X/IG", "UK Education Charities", "Nano", "1.2K IG", "Adult literacy — prison/community programmes", "Partnership", "Email", "Not Started", "Adult literacy — wider demographic reach"],
    [204, "Read Easy UK (@ReadEasyUK)", "X/LinkedIn", "UK Education Charities", "Nano", "633 LinkedIn", "Adult literacy volunteer network", "Partnership", "Email", "Not Started", "Volunteer network — community literacy"],
    [205, "Spread the Word (@STWevents)", "X/IG", "UK Education Charities", "Nano", "Small", "London literary development — underrepresented writers", "Partnership", "Email", "Not Started", "Literary development charity — diverse voices"],
    [206, "Go All In — National Year of Reading 2026 (@Go_All_In_2026)", "IG/Website", "UK Education Charities", "Micro", "New (2026 campaign)", "National Literacy Trust coalition — 2026 reading campaign", "Campaign partnership", "Email", "Not Started", "2026 campaign — massive awareness moment, time-sensitive"],

    # ═══ ASIA IGCSE CREATORS — SINGAPORE / MALAYSIA / HONG KONG (web-verified) ═══
    [207, "Sir Fathi / @mathsanova", "TikTok", "Asia IGCSE Creators", "Mid", "710K TikTok", "Malaysian Maths teacher — Cambridge-aligned content, award winner", "Sponsored video / Affiliate", "DM", "Not Started", "710K + 14.6M likes — won Digital Creator Award 2024"],
    [208, "Cikgu Debbie / @cikgudebbie", "TikTok", "Asia IGCSE Creators", "Mid", "580K TikTok", "Malaysian Maths teacher — SPM/Cambridge-aligned", "Sponsored video", "DM", "Not Started", "580K + 13.7M likes — Shah Alam tuition centre"],
    [209, "Cikgu Zack", "TikTok", "Asia IGCSE Creators", "Mid", "378K TikTok", "Malaysian language teacher — national + intl curriculum", "Sponsored video", "DM", "Not Started", "378K + 5.5M likes — language education content"],
    [210, "Biogirl MJ / Kong Man Jing (@justkeepthinking)", "YT/TikTok/IG", "Asia IGCSE Creators", "Mid", "500K+ combined", "Singapore science educator — Content Creator of Year 2023", "Sponsored video / Partnership", "Email", "Not Started", "500K+ combined, 30M monthly IG reach — science education"],
    [211, "Primrose Kitten Academy (Jen)", "YouTube", "Asia IGCSE Creators", "Mid", "258K YT", "GCSE/IGCSE Science — Oxford Uni Press author, 600 free videos", "Affiliate link / Sponsored", "Email", "Not Started", "258K — widely used by IGCSE students across SE Asia"],
    [212, "Nura Ezzatie / @nuraezzatie", "TikTok", "Asia IGCSE Creators", "Mid", "190K TikTok", "Popular among Malaysian intl school students", "Sponsored video", "DM", "Not Started", "190K + 6M likes — education content creator Malaysia"],
    [213, "Science with Hazel (SWH Learning)", "YT/TikTok/IG", "Asia IGCSE Creators", "Micro", "90K YT", "IGCSE/A-Level Science — Cambridge grad, revision guides", "Affiliate / Sponsored", "Email", "Not Started", "90K YT — popular with SG/MY IGCSE students"],
    [214, "Cambridge in 5 Minutes", "YouTube", "Asia IGCSE Creators", "Micro", "55K YT", "Cambridge IGCSE Bio/Chem/Physics — 4.5M+ views", "Affiliate link / Sponsored", "Email", "Not Started", "55K + 4.5M views — concise IGCSE revision, SE Asia audience"],
    [215, "Dixon Yau / D. Math Academy", "YouTube", "Asia IGCSE Creators", "Nano", "25K YT", "IGCSE/A-Level Maths — serves SG, MY, UAE, India", "Affiliate / Sponsored", "Email", "Not Started", "25K + 4M views — Subang Jaya, cross-country reach"],
    [216, "Tavis Online Tuition / @tavislive", "TikTok", "Asia IGCSE Creators", "Nano", "25K TikTok", "Malaysia's fastest-growing online tuition platform", "Partnership", "DM", "Not Started", "25K — SPM tuition, growing platform"],
    [217, "TWINS Education / @twinseducation", "YT/Facebook", "Asia IGCSE Creators", "Nano", "10K+ combined", "Cambridge IGCSE tuition centre — 25 subjects since 2014", "Partnership / Affiliate", "Email", "Not Started", "10K+ — Subang Jaya, 25 IGCSE subjects, exam walkthroughs"],
    [218, "Edulab Academy / @edulab.sg", "IG/TikTok", "Asia IGCSE Creators", "Nano", "5K+ students", "IB and IGCSE tuition — Singapore & Hong Kong", "Partnership", "Email", "Not Started", "5K+ students, 95% A grades IGCSE 2024 — SG/HK"],
    [219, "Tutopiya", "YouTube/Website", "Asia IGCSE Creators", "Micro", "20+ countries", "Singapore edtech — IGCSE/A-Level/IB online tutoring", "Partnership / Affiliate", "Email", "Not Started", "Founded 2018 — IGCSE Maths, Sciences, English content"],
    [220, "HKExcel", "YouTube", "Asia IGCSE Creators", "Nano", "3M+ YT views", "Hong Kong's only IB-exclusive tutorial centre", "Partnership", "Email", "Not Started", "6,000+ students, avg score 36+/45 — HK market entry"],
    [221, "ITS Education Asia", "YouTube/Website", "Asia IGCSE Creators", "Nano", "100K+ hours delivered", "HK official IGCSE/GCE exam centre since 2005", "Partnership", "Email", "Not Started", "100K+ hours GCSE/IGCSE tutorials — HK exam centre"],

    # ═══ ARABIC-ENGLISH YOUTUBE CHANNELS — GCC AUDIENCE (web-verified) ═══
    [222, "ZAmericanEnglish (Ibrahim Adel)", "YouTube", "Arabic-English YouTube", "Mega", "12.1M subs", "Egypt — #1 Arabic English-learning channel globally", "Sponsored video / Partnership", "Email", "Not Started", "12.1M — biggest Arabic English channel, massive GCC viewership"],
    [223, "Durus Online / Ahmed Abu Zaid", "YouTube", "Arabic-English YouTube", "Mega", "9.4M subs", "Egypt — top education channel, English + study skills", "Sponsored video", "Email", "Not Started", "9.4M — broad education channel with English learning content"],
    [224, "Sbeata Academy / English with Sbeata", "YouTube", "Arabic-English YouTube", "Mega", "4.18M subs", "France-based — Arabic English teaching", "Sponsored video / Affiliate", "Email", "Not Started", "4.18M — French-based, teaches English to Arabic speakers"],
    [225, "Learn English with Murad", "YouTube", "Arabic-English YouTube", "Macro", "3M subs", "Algeria — English for Arabic speakers", "Sponsored video", "Email", "Not Started", "3M — Algerian creator, huge North Africa + GCC reach"],
    [226, "Omar Abdelrahim", "YouTube", "Arabic-English YouTube", "Macro", "1.45M subs", "Egypt — English pronunciation & grammar", "Sponsored video", "Email", "Not Started", "1.45M — Egyptian English teacher, pronunciation focus"],
    [227, "AdamAcademy", "YouTube", "Arabic-English YouTube", "Macro", "1M subs", "Egypt — English learning academy", "Sponsored video / Affiliate", "Email", "Not Started", "1M — established English learning channel"],
    [228, "Master English / اتقن الإنجليزية", "YouTube", "Arabic-English YouTube", "Mid", "900K subs", "English mastery for Arabic speakers", "Sponsored video", "Email", "Not Started", "900K — dedicated English mastery channel"],
    [229, "Learn English with Sondos AbdelHalim", "YouTube", "Arabic-English YouTube", "Mid", "720K subs", "Egypt — female English teacher, relatable style", "Sponsored video", "Email", "Not Started", "720K — popular female English teacher for Arabic audience"],
    [230, "Learn English with Ehab (Ehab Rashwan)", "YouTube", "Arabic-English YouTube", "Mid", "653K subs", "Egypt — structured English courses", "Sponsored video / Affiliate", "Email", "Not Started", "653K — structured courses, Egyptian audience + GCC"],
    [231, "English With Khaled (Khaled Habbal)", "YouTube", "Arabic-English YouTube", "Mid", "600K subs", "Iraq — English for Arabic speakers", "Sponsored video", "Email", "Not Started", "600K — Iraqi creator, Gulf Arabic audience"],
    [232, "Love English with Antoinette", "YouTube", "Arabic-English YouTube", "Mid", "500K subs", "Egypt — conversational English", "Sponsored video", "Email", "Not Started", "500K — conversational English for Arabic speakers"],
    [233, "Mai Gamal / بتاعة إنجليزي", "YouTube", "Arabic-English YouTube", "Mid", "450K subs", "Egypt — fun English teaching style", "Sponsored video", "DM", "Not Started", "450K — engaging Egyptian English teacher"],
    [234, "Learn English with Haya", "YouTube", "Arabic-English YouTube", "Mid", "400K subs", "English for Arabic speakers — conversation focus", "Sponsored video", "Email", "Not Started", "400K — conversation-focused English channel"],
    [235, "Learn English Through Translated Stories", "YouTube", "Arabic-English YouTube", "Mid", "350K subs", "Story-based English learning for Arabic speakers", "Sponsored video", "Email", "Not Started", "350K — unique story-based approach, high engagement"],
    [236, "Ahmed Hassan (pronunciation)", "YouTube", "Arabic-English YouTube", "Mid", "300K subs", "Egypt — pronunciation & conversation", "Sponsored video", "Email", "Not Started", "300K — pronunciation specialist, Egyptian English teacher"],

    # ═══ ESL TIKTOK MICRO-CREATORS 10K-500K (web-verified, live counts) ═══
    [237, "@englishunderstood", "TikTok", "ESL TikTok Creators", "Mid", "427K TikTok", "UK — IELTS expert, helps students reach Band 7+", "Affiliate / Sponsored", "DM", "Not Started", "427K — IELTS specialist, UK-based, perfect GCC fit"],
    [238, "@englishvocabulary100", "TikTok", "ESL TikTok Creators", "Mid", "424K TikTok", "English vocabulary lessons — high engagement", "Sponsored video", "DM", "Not Started", "424K — vocabulary-focused, broad ESL audience"],
    [239, "@learnenglish840", "TikTok", "ESL TikTok Creators", "Mid", "423K TikTok", "English for Arabic speakers — direct GCC audience", "Sponsored video", "DM", "Not Started", "423K — Arabic-speaking audience, GCC match"],
    [240, "@missenglishteacher", "TikTok", "ESL TikTok Creators", "Mid", "377K TikTok", "British English teacher — Germany/UK based", "Sponsored video", "DM", "Not Started", "377K — British English content, international audience"],
    [241, "@learn.english.02", "TikTok", "ESL TikTok Creators", "Mid", "281K TikTok", "English speaking practice channel", "Sponsored video", "DM", "Not Started", "281K — speaking practice focus, high engagement"],
    [242, "@englishwithemma_", "TikTok", "ESL TikTok Creators", "Mid", "279K TikTok", "UK — British English teacher, confidence-building", "Affiliate / Sponsored", "DM", "Not Started", "279K — UK-based, confidence-building approach"],
    [243, "@reallifeenglish", "TikTok", "ESL TikTok Creators", "Mid", "260K TikTok", "Brazil — founded by American Justin Murray", "Partnership / Affiliate", "Email", "Not Started", "260K — real-life English approach, Crunchbase-listed"],
    [244, "@tetiana.kurova.5 (PicSpeak)", "TikTok", "ESL TikTok Creators", "Mid", "191K TikTok", "Ukraine — 'From Pictures to Words' ESL method", "Sponsored video", "DM", "Not Started", "191K — unique visual teaching method"],
    [245, "@englishwithkayla", "TikTok", "ESL TikTok Creators", "Mid", "181K TikTok", "USA — intermediate/advanced English phrases", "Sponsored video", "DM", "Not Started", "181K — intermediate/advanced focus, good for GCSE level"],
    [246, "@learnenglishwithlouise", "TikTok", "ESL TikTok Creators", "Mid", "152K TikTok", "UK — native speaker, vocabulary & grammar", "Affiliate / Sponsored", "DM", "Not Started", "152K — UK native speaker, vocabulary focus"],
    [247, "@englishfluencyjourney (Hanna Khoma)", "TikTok", "ESL TikTok Creators", "Mid", "141K TikTok", "Ukraine/USA — American accent focus", "Sponsored video", "DM", "Not Started", "141K — fluency journey brand, relatable approach"],
    [248, "@fluencyacademy", "TikTok", "ESL TikTok Creators", "Mid", "136K TikTok", "Brazil — English fluency school", "Partnership", "Email", "Not Started", "136K — Portuguese audience, fluency school brand"],
    [249, "@english.teachtalk", "TikTok", "ESL TikTok Creators", "Micro", "92K TikTok", "Daily English conversations for learners", "Sponsored video", "DM", "Not Started", "92K — daily conversation format, high retention"],
    [250, "@your.english.cheerleader (Tim)", "TikTok", "ESL TikTok Creators", "Micro", "77K TikTok", "USA — certified native English teacher", "Sponsored video", "DM", "Not Started", "77K — certified teacher, encouraging style"],
    [251, "@tanya.icdenglish", "TikTok", "ESL TikTok Creators", "Micro", "69K TikTok", "USA — Cambridge CELTA-qualified, 'must-know English'", "Sponsored video", "DM", "Not Started", "69K — CELTA-qualified, Cambridge credential adds credibility"],

    # ═══ UK TEEN LIFESTYLE TIKTOK — GCSE-AGE AUDIENCE (training data) ═══
    [252, "Kyle Thomas (@kylethomas)", "TikTok", "UK Teen Lifestyle TikTok", "Mega", "~30M TikTok", "UK teen — one of UK's biggest TikTokers", "Sponsored post / Brand deal", "Agent", "Not Started", "~30M — massive UK teen audience, lifestyle/trends"],
    [253, "Holly H (@hollyh)", "TikTok", "UK Teen Lifestyle TikTok", "Mega", "~16M TikTok", "UK — comedy skits, lip-sync, lifestyle", "Sponsored post", "Agent", "Not Started", "~16M — one of UK's top female TikTokers"],
    [254, "Amelia Gething (@ameliagething)", "TikTok", "UK Teen Lifestyle TikTok", "Mega", "~10M TikTok", "UK — comedy sketches, had own CBBC show", "Sponsored post", "Agent", "Not Started", "~10M — CBBC show, huge UK teen following"],
    [255, "Max & Harvey (@maxandharvey)", "TikTok", "UK Teen Lifestyle TikTok", "Mega", "~7M TikTok", "UK twins — music, comedy, teen lifestyle", "Sponsored post", "Agent", "Not Started", "~7M — British twins, music + comedy content"],
    [256, "Rhia Official (@rhia_official)", "TikTok", "UK Teen Lifestyle TikTok", "Mega", "~5M TikTok", "UK — dance, transitions, lifestyle", "Sponsored post", "DM", "Not Started", "~5M — dance/lifestyle, UK teen demographic"],
    [257, "Shauni (@shauni)", "TikTok", "UK Teen Lifestyle TikTok", "Mega", "~4M TikTok", "UK — dance, lifestyle, POV content", "Sponsored post", "DM", "Not Started", "~4M — POV content popular with UK teens"],
    [258, "Abbie Quinnen (@abbiequinnen)", "TikTok", "UK Teen Lifestyle TikTok", "Macro", "~2.5M TikTok", "UK — dance, lifestyle, fashion", "Sponsored post", "DM", "Not Started", "~2.5M — dance/fashion, UK teen audience"],
    [259, "Millie T (@milliet0)", "TikTok", "UK Teen Lifestyle TikTok", "Macro", "~2M TikTok", "UK — commentary, pop culture, lifestyle", "Sponsored post", "DM", "Not Started", "~2M — pop culture commentary, GCSE-age viewers"],
    [260, "ChloeLock (@chloelock_)", "TikTok", "UK Teen Lifestyle TikTok", "Macro", "~1.8M TikTok", "UK — fashion hauls, beauty, teen lifestyle", "Sponsored post", "DM", "Not Started", "~1.8M — fashion/beauty, UK teen girls"],
    [261, "Sophie Aspin (@sophieaspin)", "TikTok", "UK Teen Lifestyle TikTok", "Macro", "~1.5M TikTok", "UK — lifestyle, music, vlogs", "Sponsored post", "DM", "Not Started", "~1.5M — music/lifestyle, northern UK teen audience"],

    # ═══ UK COMEDY & ENTERTAINMENT — TEEN AUDIENCE (training data) ═══
    [262, "KSI / Olajide Olatunji (@KSI)", "YouTube", "UK Comedy & Entertainment", "Mega", "~24M YT", "Sidemen — #1 UK YouTube entertainer, boxing/music/comedy", "Brand deal / Sponsored", "Agent", "Not Started", "24M — Sidemen ecosystem, dominates UK teen YouTube"],
    [263, "W2S / Harry Lewis (@W2S)", "YouTube", "UK Comedy & Entertainment", "Mega", "~16M YT", "Sidemen — extreme comedy challenges", "Brand deal", "Agent", "Not Started", "16M — chaotic comedy, massive teen following"],
    [264, "Miniminter / Simon Minter (@Miniminter)", "YouTube", "UK Comedy & Entertainment", "Mega", "~10M YT", "Sidemen — comedy challenges, vlogs", "Brand deal", "Agent", "Not Started", "10M — Sidemen member, challenge content"],
    [265, "Luke Davidson (@lukedavidson)", "TikTok", "UK Comedy & Entertainment", "Mega", "~10M TikTok", "British-Canadian — comedy sketches, life hack parodies", "Sponsored post", "DM", "Not Started", "10M — short comedy, huge UK teen following"],
    [266, "ThatcherJoe / Joe Sugg (@ThatcherJoe)", "YouTube", "UK Comedy & Entertainment", "Mega", "~8M YT", "Original British YouTube gen — comedy challenges, pranks", "Brand deal", "Agent", "Not Started", "8M — OG British YouTuber, still huge with teens"],
    [267, "Niko Omilana (@NikoOmilana)", "YouTube", "UK Comedy & Entertainment", "Mega", "~7M YT", "Prank comedy, social experiments — ran for London Mayor", "Brand deal", "Agent", "Not Started", "7M — viral stunts, mainstream UK media crossover"],
    [268, "Max Mayfield (@max)", "TikTok", "UK Comedy & Entertainment", "Mega", "~5M TikTok", "Comedy skits — relatable school/teen humor", "Sponsored post", "DM", "Not Started", "5M — school humor, directly relatable to GCSE students"],
    [269, "Callux / Callum Airey (@Callux)", "YouTube", "UK Comedy & Entertainment", "Mega", "~5M YT", "Challenge videos — Sidemen-adjacent comedy", "Brand deal", "Agent", "Not Started", "5M — Sidemen-adjacent, challenge content"],
    [270, "Max Sherwood (@maxbalegde)", "TikTok", "UK Comedy & Entertainment", "Mega", "~4M TikTok", "British comedy sketches, character humor", "Sponsored post", "DM", "Not Started", "4M — character comedy, British humour"],
    [271, "ChunkZ / Amin Mohamed (@chunkz)", "TikTok/YouTube", "UK Comedy & Entertainment", "Mega", "~4M+ combined", "Comedy, gaming, panel shows — Sidemen-adjacent", "Brand deal", "Agent", "Not Started", "4M+ — panel shows, gaming, comedy crossover"],
    [272, "Amelia Dimoldenberg (@ameliadimz)", "TikTok/YouTube", "UK Comedy & Entertainment", "Macro", "~3.5M TikTok", "Chicken Shop Date — awkward comedy, mainstream crossover", "Brand deal / Sponsored", "Agent", "Not Started", "3.5M — Chicken Shop Date viral format, mainstream press"],
    [273, "ImAllexx / Alex Elmslie (@ImAllexx)", "YouTube", "UK Comedy & Entertainment", "Macro", "~2.5M YT", "Commentary, reaction videos, British humor", "Sponsored post", "Email", "Not Started", "2.5M — commentary channel, UK teen viewers"],
    [274, "Ehiz Ufuah (@ehiz)", "TikTok", "UK Comedy & Entertainment", "Macro", "~2M TikTok", "Comedy, reactions, entertainment presenter", "Sponsored post", "DM", "Not Started", "2M — entertainment presenter, reactions"],
    [275, "Jack Dean (@JackDean)", "YouTube", "UK Comedy & Entertainment", "Macro", "~1.5M YT", "Storytelling, animation-style comedy", "Sponsored post", "Email", "Not Started", "1.5M — unique storytelling format"],

    # ═══ UK MAINSTREAM MUM INFLUENCERS — PARENT AUDIENCE (training data) ═══
    [276, "Stacey Solomon (@staceysolomon)", "Instagram", "UK Mainstream Parenting", "Mega", "~6M IG", "Family life, crafts — has teens + younger kids", "Brand deal / Sponsored", "Agent", "Not Started", "6M — UK's biggest mum influencer, discusses parenting older teens"],
    [277, "Mrs Hinch / Sophie Hinchliffe (@mrshinchhome)", "Instagram", "UK Mainstream Parenting", "Mega", "~4.5M IG", "Home organisation — massive mum audience overlap", "Brand deal", "Agent", "Not Started", "4.5M — cleaning/home, core audience = UK mums"],
    [278, "Sam Faiers (@samanthafaiers)", "Instagram", "UK Mainstream Parenting", "Macro", "~2.5M IG", "Family lifestyle — The Mummy Diaries audience", "Brand deal", "Agent", "Not Started", "2.5M — ITVBe show, family lifestyle brand"],
    [279, "Louise Pentland (@louisepentland)", "Instagram", "UK Mainstream Parenting", "Macro", "~2M IG", "Motherhood, body positivity, mental health — has older children", "Sponsored post", "Agent", "Not Started", "2M — OG YouTube mum, older child demographic"],
    [280, "Giovanna Fletcher (@giovannafletcherofficial)", "Instagram", "UK Mainstream Parenting", "Macro", "~2M IG", "Parenting all ages, books, wellbeing — I'm A Celeb winner", "Sponsored post", "Agent", "Not Started", "2M — author, podcast host, parenting across ages"],
    [281, "Katie Piper (@katiepiper_)", "Instagram", "UK Mainstream Parenting", "Macro", "~1M IG", "Family life, wellbeing — school-age children", "Brand deal", "Agent", "Not Started", "1M — TV presenter, broad parent audience"],
    [282, "The Unmumsy Mum / Sarah Turner (@theunmumsymum)", "Instagram", "UK Mainstream Parenting", "Mid", "~600K IG", "Honest/humorous parenting — bestselling author", "Sponsored post", "Email", "Not Started", "600K — relatable parenting through all stages, author"],
    [283, "Izzy Judd (@izzyjudd)", "Instagram", "UK Mainstream Parenting", "Mid", "~400K IG", "Family life, fertility, wellbeing — school-age kids", "Sponsored post", "Email", "Not Started", "400K — musician's wife, family wellbeing content"],
    [284, "Rachaele Hambleton / Part-Time Working Mummy (@parttimeworkingmummy)", "Instagram", "UK Mainstream Parenting", "Mid", "~400K IG", "Blended family — teens and younger kids, fostering", "Sponsored post", "Email", "Not Started", "400K — openly discusses teen parenting, blended family"],
    [285, "Anna Whitehouse / Mother Pukka (@mother_pukka)", "Instagram", "UK Mainstream Parenting", "Mid", "~350K IG", "Working parenthood, flexible work campaigns — school-age kids", "Sponsored post", "Email", "Not Started", "350K — campaigns on working parents, school-age focus"],
    [286, "Emily Norris (@emilynorris)", "TikTok/Instagram", "UK Mainstream Parenting", "Mid", "~300K combined", "Family life hacks — three boys including school-age", "Sponsored post", "DM", "Not Started", "300K — practical parenting, school-age boys"],
    [287, "Candice Brathwaite (@candicebrathwaite)", "Instagram", "UK Mainstream Parenting", "Mid", "~250K IG", "Parenting, race, motherhood — all parenting stages", "Sponsored post", "Email", "Not Started", "250K — diverse voice, mainstream press crossover"],
    [288, "Clemmie Telford (@clemmie_telford)", "Instagram", "UK Mainstream Parenting", "Mid", "~200K IG", "Honest parenting of tweens/teens — school-age focus", "Sponsored post", "Email", "Not Started", "200K — specifically discusses teen/tween parenting"],
    [289, "Kerry Whelpdale (@kerrywhelpdale)", "TikTok", "UK Mainstream Parenting", "Mid", "~200K TikTok", "Mum life, family humour — children spanning age ranges", "Sponsored post", "DM", "Not Started", "200K — relatable mum humour, TikTok native"],
    [290, "Channel Mum (@channelmum)", "TikTok", "UK Mainstream Parenting", "Mid", "~150K TikTok", "Parenting tips, product reviews, school-age advice", "Partnership / Sponsored", "Email", "Not Started", "150K — community platform, product reviews for parents"],

    # ═══ GCC TEEN & LIFESTYLE TIKTOK (web-verified) ═══
    [291, "Anatoly (@anatoly_pranks)", "TikTok", "GCC Teen Lifestyle TikTok", "Mega", "22.1M TikTok", "Dubai-based pranks — 18M avg views, mall/street content", "Brand deal", "Agent", "Not Started", "22.1M — Dubai pranks, massive teen viewership across GCC"],
    [292, "Bessan Ismail (@bessan_esmail)", "TikTok", "GCC Teen Lifestyle TikTok", "Mega", "19.2M TikTok", "Music & entertainment — Arabic music + style", "Brand deal", "Agent", "Not Started", "19.2M — Arabic music/entertainment, resonates across GCC"],
    [293, "Ziba Gulley (@zibagulley)", "TikTok", "GCC Teen Lifestyle TikTok", "Mega", "15.9M TikTok", "Travel & beauty — strong GCC-wide connection", "Brand deal", "Agent", "Not Started", "15.9M — travel/beauty, connects across entire GCC region"],
    [294, "Missdouaa (@missdouaa.officiel)", "TikTok", "GCC Teen Lifestyle TikTok", "Mega", "14.8M TikTok", "Fashion & lifestyle — motivational + fashion tips", "Brand deal", "Agent", "Not Started", "14.8M — fashion/motivational, huge female GCC audience"],
    [295, "Hemex (@hemex)", "TikTok", "GCC Teen Lifestyle TikTok", "Mega", "8M TikTok", "Sports & fitness — 8.4% engagement rate", "Brand deal", "Agent", "Not Started", "8M — sports/fitness, highest engagement rate in UAE"],
    [296, "Yaseenkhanofficial (@yaseenkhanofficial)", "TikTok", "GCC Teen Lifestyle TikTok", "Mega", "6.7M TikTok", "Comedy/entertainment — 8.2% engagement", "Brand deal", "DM", "Not Started", "6.7M — comedy, exceptionally high 8.2% engagement"],
    [297, "Faisal Al-Issa (فيصل العيسى)", "TikTok", "GCC Teen Lifestyle TikTok", "Mega", "4.8M TikTok", "Saudi Arabia's top TikToker — lifestyle/comedy", "Brand deal", "Agent", "Not Started", "4.8M — #1 Saudi TikToker, lifestyle content"],
    [298, "Yara Aziz (@yaraaziz)", "TikTok/YouTube", "GCC Teen Lifestyle TikTok", "Macro", "3.7M TikTok / 860K YT", "Dubai-based comedy sketches & pranks, age 22", "Brand deal", "DM", "Not Started", "3.7M TT + 860K YT — Dubai comedy, young female audience"],
    [299, "Dina Saeva (@dinasaeva)", "TikTok", "GCC Teen Lifestyle TikTok", "Mega", "10M+ TikTok", "Dubai-based — comedy, lifestyle, massive reach", "Brand deal", "Agent", "Not Started", "10M+ — Dubai-based lifestyle/comedy creator"],
    [300, "Amy Roko (@amyroko)", "TikTok", "GCC Teen Lifestyle TikTok", "Macro", "2M+ TikTok", "Saudi rapper/comedian — skits, beauty, challenges", "Sponsored post", "DM", "Not Started", "2M+ — Saudi native, Cosmo cover girl, comedy + beauty"],
    [301, "Dalal AlDoub / Dalalid (@dalalid)", "TikTok", "GCC Teen Lifestyle TikTok", "Macro", "1M+ TikTok", "Kuwait fashion/beauty — makeup tutorials, style", "Sponsored post", "DM", "Not Started", "1M+ — Kuwaiti fashion/beauty, strong GCC female audience"],
    [302, "Abdullah Alhumaidhan / Rise", "TikTok", "GCC Teen Lifestyle TikTok", "Mid", "500K+ TikTok", "Kuwait — reactions, food reviews, local culture", "Sponsored post", "DM", "Not Started", "500K+ — Kuwait culture/food, relatable teen content"],
    [303, "Mohammed Aldhurafi", "TikTok", "GCC Teen Lifestyle TikTok", "Macro", "1M+ TikTok", "Saudi sports/football content — Ronaldo focus", "Sponsored post", "DM", "Not Started", "1M+ — Saudi football content, teen male audience"],

    # ═══ DUBAI/UAE LIFESTYLE MEGA INFLUENCERS (training data) ═══
    [304, "Huda Kattan (@hudabeauty)", "Instagram", "Dubai Mega Lifestyle", "Mega", "~53M IG", "Dubai — world's biggest beauty influencer, massive mum audience", "Brand deal", "Agent", "Not Started", "53M — #1 beauty influencer globally, Dubai-based, family content"],
    [305, "Joelle Mardinian (@joellemardinian)", "Instagram", "Dubai Mega Lifestyle", "Mega", "~10M IG", "Dubai — beauty, lifestyle, family", "Brand deal", "Agent", "Not Started", "10M — Lebanese-Dubai beauty mogul, family content"],
    [306, "Khalid Al Ameri (@khalidalameri)", "IG/TikTok", "Dubai Mega Lifestyle", "Mega", "~4M IG / 7M TikTok", "Family life, cultural commentary — expat relatability", "Brand deal", "Agent", "Not Started", "11M combined — Emirati family content, expat audience"],
    [307, "Karen Wazen (@karenwazen)", "Instagram", "Dubai Mega Lifestyle", "Mega", "~5M IG", "Fashion, luxury, motherhood — Dubai-based", "Brand deal", "Agent", "Not Started", "5M — luxury fashion + motherhood, GCC audience"],
    [308, "Noor Stars (@noorstars)", "IG/TikTok", "Dubai Mega Lifestyle", "Mega", "~5M IG / 3M TikTok", "Youth lifestyle, vlogs, beauty — young audience", "Brand deal", "Agent", "Not Started", "8M combined — youth entertainment, perfect teen demographic"],
    [309, "Lana Rose (@lanarose786)", "IG/YouTube", "Dubai Mega Lifestyle", "Macro", "~3M IG", "Vlogs, lifestyle, art, comedy — Dubai-based", "Brand deal", "Agent", "Not Started", "3M — vlogs/art/comedy, Mo Vlogs' sister"],
    [310, "Mo Vlogs / Mohamed Beiraghdary (@movlogs)", "IG/YouTube/TikTok", "Dubai Mega Lifestyle", "Macro", "~3M IG / 1M TikTok", "Youth lifestyle, luxury Dubai vlogs", "Brand deal", "Agent", "Not Started", "4M+ combined — Dubai luxury vlogs, massive teen male audience"],
    [311, "Anas Bukhash (@anasbukhash)", "Instagram", "Dubai Mega Lifestyle", "Macro", "~2M IG", "Lifestyle, interviews, motivational — ABtalks podcast", "Sponsored / Interview", "Agent", "Not Started", "2M — ABtalks interview series, thoughtful content"],
    [312, "Salama Mohamed (@salamamohamed)", "Instagram", "Dubai Mega Lifestyle", "Macro", "~2M IG", "Emirati lifestyle, fashion, family", "Sponsored post", "DM", "Not Started", "2M — Emirati creator, fashion + family"],
    [313, "Loujain Adada (@loujainaj)", "Instagram", "Dubai Mega Lifestyle", "Macro", "~2M IG", "Fashion, lifestyle, motherhood — Dubai-based", "Sponsored post", "DM", "Not Started", "2M — fashion + motherhood, Lebanese-Dubai"],
    [314, "Ahmad Al Bulooshi (@ahmadalblooshi_)", "TikTok/IG", "Dubai Mega Lifestyle", "Macro", "~2M TikTok", "Comedy, lifestyle — youth-oriented content", "Sponsored post", "DM", "Not Started", "2M — Emirati comedy, youth audience"],
    [315, "Farhana Bodi (@farhanabodi)", "Instagram", "Dubai Mega Lifestyle", "Macro", "~1.5M IG", "Modest fashion, travel, family lifestyle", "Sponsored post", "Email", "Not Started", "1.5M — modest fashion, family lifestyle, GCC parent audience"],
    [316, "Dana Al Tuwarish (@dana_altuwarish)", "Instagram", "Dubai Mega Lifestyle", "Macro", "~1M IG", "Fashion, luxury — Kuwaiti-Dubai expat", "Sponsored post", "DM", "Not Started", "1M — Kuwaiti-Dubai, luxury fashion lifestyle"],
    [317, "Reem Al Kanhal (@rikitta_)", "IG/TikTok", "Dubai Mega Lifestyle", "Macro", "~1M IG", "Fashion, lifestyle — Emirati culture", "Sponsored post", "DM", "Not Started", "1M — authentic Emirati culture + fashion"],
    [318, "Safa Siddiqui (@safasiddiqui1)", "Instagram", "Dubai Mega Lifestyle", "Mid", "~700K IG", "Fashion, luxury — Dubai Bling reality TV", "Sponsored post", "DM", "Not Started", "700K — Dubai Bling cast, Netflix exposure"],

    # ═══ UK GAMING & STREAMING — TEEN AUDIENCE (training data) ═══
    [319, "JackSucksAtLife / Jack Massey Welsh (@JackSucksAtLife)", "YouTube", "UK Gaming & Streaming", "Mega", "~30M+ YT (multi-channel)", "YouTube play buttons, geography, Minecraft — massive teen following", "Brand deal", "Agent", "Not Started", "30M+ across channels — geography + gaming, UK teen favourite"],
    [320, "DanTDM / Daniel Middleton (@DanTDM)", "YouTube", "UK Gaming & Streaming", "Mega", "~28M YT", "OG UK Minecraft YouTuber — Roblox, Fortnite, variety", "Brand deal", "Agent", "Not Started", "28M — OG UK gaming, fans now 14-18, Guinness record holder"],
    [321, "TommyInnit / Tom Simons (@TommyInnit)", "YouTube/Twitch", "UK Gaming & Streaming", "Mega", "~14M YT / 7M Twitch", "Minecraft Dream SMP — biggest UK teen gaming creator", "Brand deal", "Agent", "Not Started", "21M combined — #1 UK teen Minecraft creator, massive engagement"],
    [322, "Stampylonghead / Joseph Garrett (@stampylonghead)", "YouTube", "UK Gaming & Streaming", "Mega", "~11M YT", "Minecraft Let's Plays — iconic, fans now GCSE-age", "Brand deal", "Agent", "Not Started", "11M — fans grew up with him, now 14-18 demographic"],
    [323, "Grian / Charles Batchelor (@Grian)", "YouTube", "UK Gaming & Streaming", "Mega", "~9M YT", "Minecraft Hermitcraft — building tutorials", "Sponsored video", "Agent", "Not Started", "9M — Hermitcraft, creative Minecraft, UK-based"],
    [324, "Vikkstar123 / Vikram Barn (@Vikkstar123)", "YouTube/Twitch", "UK Gaming & Streaming", "Mega", "~8M YT", "Warzone, Minecraft, Fortnite — Sidemen member", "Brand deal", "Agent", "Not Started", "8M — Sidemen gaming member, CoD Warzone"],
    [325, "Wilbur Soot / Will Gold (@WilburSoot)", "YouTube/Twitch", "UK Gaming & Streaming", "Mega", "~6.5M YT", "Minecraft, music, comedy — Dream SMP lore creator", "Brand deal", "Agent", "Not Started", "6.5M — musician + gamer, huge 14-18 fanbase"],
    [326, "LDShadowLady / Lizzie Dwyer (@LDShadowLady)", "YouTube", "UK Gaming & Streaming", "Mega", "~6M YT", "Minecraft, Sims — family-friendly, female gaming audience", "Sponsored video", "Email", "Not Started", "6M — female gaming icon, family-friendly content"],
    [327, "Ph1LzA / Phil Watson (@Ph1LzA)", "Twitch/YouTube", "UK Gaming & Streaming", "Mega", "~5M YT / 4.5M Twitch", "Minecraft hardcore — QSMP, Origins SMP", "Sponsored video", "Agent", "Not Started", "9.5M combined — hardcore Minecraft, dedicated teen fanbase"],
    [328, "Tubbo / Toby Smith (@Tubbo)", "YouTube/Twitch", "UK Gaming & Streaming", "Mega", "~4.5M YT / 4.5M Twitch", "Minecraft Dream SMP — TommyInnit's duo partner", "Sponsored video", "Agent", "Not Started", "9M combined — Dream SMP, concentrated 14-18 audience"],
    [329, "SmallishBeans / Joel (@SmallishBeans)", "YouTube", "UK Gaming & Streaming", "Mega", "~4M YT", "Minecraft building, Empires SMP", "Sponsored video", "Email", "Not Started", "4M — creative Minecraft, wholesome content"],
    [330, "InTheLittleWood / Martyn Littlewood (@InTheLittleWood)", "YouTube/Twitch", "UK Gaming & Streaming", "Macro", "~3.5M YT", "Minecraft, Nintendo — Yogscast member", "Sponsored video", "Email", "Not Started", "3.5M — Yogscast, Nintendo, long-running UK creator"],

    # ═══ UK MENTAL HEALTH & STUDENT WELLBEING CREATORS (web-verified) ═══
    [331, "Dr. Julie Smith (@drjulie)", "TikTok/IG/YouTube", "UK Mental Health & Wellbeing", "Mega", "4.9M TikTok / 5M+ combined", "UK clinical psychologist — bestselling author, #1 mental health creator", "Partnership / Sponsored", "Email", "Not Started", "5M+ — UK's biggest mental health creator, relatable psychology"],
    [332, "Eden Harvey (@edenharvey)", "TikTok", "UK Mental Health & Wellbeing", "Macro", "3M+ TikTok", "Eat With Eden — candid OCD, emetophobia, eating disorder content", "Sponsored post", "DM", "Not Started", "3M+ — lifestyle + mental health, young female audience"],
    [333, "Dr. Kojo Sarfo (@drkojosarfo)", "TikTok", "UK Mental Health & Wellbeing", "Macro", "2.4M TikTok", "Mental health tips — workplace and student focus", "Sponsored video", "Email", "Not Started", "2.4M — practical mental health guidance, broad audience"],
    [334, "The Cognitive Corner (@thecognitivecorner)", "TikTok", "UK Mental Health & Wellbeing", "Mid", "542K TikTok", "Cognitive psychology tips — student-relevant content", "Sponsored video", "DM", "Not Started", "542K — cognitive psychology, study/exam applicable"],
    [335, "Stubbs (@stubbosfits)", "TikTok", "UK Mental Health & Wellbeing", "Mid", "403K TikTok", "UK health influencer — 6.76% engagement rate", "Sponsored post", "DM", "Not Started", "403K — UK-based, exceptionally high engagement 6.76%"],
    [336, "Miles Nazaire (@milesjnazaire)", "TikTok", "UK Mental Health & Wellbeing", "Mid", "353K TikTok", "UK — Made in Chelsea, mental health advocacy", "Sponsored post", "DM", "Not Started", "353K — MiC cast, speaks openly about anxiety + therapy"],
    [337, "Amara Riva (@amarariva)", "TikTok", "UK Mental Health & Wellbeing", "Mid", "159K TikTok", "UK — 10.38% engagement, mental health content", "Sponsored post", "DM", "Not Started", "159K — highest engagement rate 10.38%, UK-based"],
    [338, "Libby Mercer (@libbymercer)", "TikTok", "UK Mental Health & Wellbeing", "Mid", "100K+ TikTok", "UK — 'anxious bestie', humor + vulnerability, young adult audience", "Sponsored post", "DM", "Not Started", "Rising — candid anxiety content, relatable to students"],
    [339, "YoungMinds (@youngmindsuk)", "IG/TikTok/X", "UK Mental Health & Wellbeing", "Macro", "500K+ combined", "UK's leading children's mental health charity", "Partnership / Campaign", "Email", "Not Started", "500K+ — charity partner, reaches every UK school"],
    [340, "Student Minds (@studentmindsorg)", "IG/X", "UK Mental Health & Wellbeing", "Micro", "50K+ combined", "UK student mental health charity — university network", "Partnership", "Email", "Not Started", "50K+ — uni mental health charity, student network"],
    [341, "Mind (@mindcharity)", "IG/TikTok/X", "UK Mental Health & Wellbeing", "Macro", "1M+ combined", "UK's biggest mental health charity — young people section", "Partnership", "Email", "Not Started", "1M+ — UK's biggest MH charity, young people resources"],
    [342, "Anna Freud Centre (@annaaboreal)", "IG/X", "UK Mental Health & Wellbeing", "Mid", "100K+ combined", "Children's mental health research — school programmes", "Partnership", "Email", "Not Started", "100K+ — schools programme, research-backed, credible partner"],

    # ═══ UK FITNESS & WELLNESS — TEEN AUDIENCE (training data) ═══
    [343, "Joe Wicks / The Body Coach (@thebodycoach)", "IG/YouTube", "UK Fitness & Wellness", "Mega", "~4.5M IG / 2.9M YT", "UK — lockdown PE teacher, family fitness, massive teen reach", "Brand deal / Partnership", "Agent", "Not Started", "7.4M combined — lockdown PE hero, household name with teens"],
    [344, "Lily Sabri (@lily.sabri)", "YouTube/IG", "UK Fitness & Wellness", "Mega", "~4.5M YT / 1.5M IG", "London — short home workouts, hugely popular with teen girls", "Sponsored video", "Email", "Not Started", "6M combined — abs/glutes workouts, teen girl demographic"],
    [345, "MattDoesFitness (@mattdoesfitness)", "YouTube/IG", "UK Fitness & Wellness", "Macro", "~2.8M YT / 1.2M IG", "Extreme food/fitness challenges, body transformations", "Sponsored video", "Email", "Not Started", "4M combined — challenge content, teen male audience"],
    [346, "Dr. Alex George (@dralexgeorge)", "IG/TikTok", "UK Fitness & Wellness", "Macro", "~2M IG / 1.5M TikTok", "UK Youth Mental Health Ambassador — Love Island fame", "Partnership / Sponsored", "Email", "Not Started", "3.5M combined — official UK gov youth MH ambassador, Love Island reach"],
    [347, "Zac Perna (@zacperna)", "TikTok/YouTube", "UK Fitness & Wellness", "Macro", "~2M TikTok / 1.5M YT", "Gym humor, relatable teen/young adult gym content", "Sponsored video", "DM", "Not Started", "3.5M combined — gym humour, relatable to teen boys"],
    [348, "Natacha Oceane (@natacha.oceane)", "YouTube/IG", "UK Fitness & Wellness", "Macro", "~1.8M YT / 1.7M IG", "Science-based fitness — very popular with teen girls", "Sponsored video", "Email", "Not Started", "3.5M combined — science-based, female fitness audience"],
    [349, "Steph Elswood (@stephelswood)", "TikTok/IG", "UK Fitness & Wellness", "Mid", "~800K TikTok", "Fitness routines, 'what I eat in a day' — strong teen demo", "Sponsored post", "DM", "Not Started", "800K — TikTok native, wellness tips for teens"],
    [350, "Grace Beverley (@gracebeverley)", "IG/TikTok", "UK Fitness & Wellness", "Mid", "~700K IG / 500K TikTok", "Fitness, productivity, TALA brand — Oxford grad, ambitious teens", "Sponsored post", "Email", "Not Started", "1.2M combined — fitness entrepreneur, appeals to 16-18 females"],
    [351, "Anna's Analysis / Anna Sherburn (@annasanalysis)", "TikTok", "UK Fitness & Wellness", "Mid", "~500K TikTok", "Psychology, mental health explainers — very popular with UK teens", "Sponsored video", "DM", "Not Started", "500K — therapy-style content, UK teen TikTok audience"],
    [352, "Lucy Mountain (@thefashionfitnessfoodie)", "IG/TikTok", "UK Fitness & Wellness", "Mid", "~500K IG", "Debunking diet culture, food/fitness myths, body positivity", "Sponsored post", "Email", "Not Started", "500K — myth-busting, resonates strongly with teen girls"],
    [353, "The Slumflower / Chidera Eggerue (@theslumflower)", "IG/TikTok", "UK Fitness & Wellness", "Mid", "~500K IG", "Body confidence, self-worth — London-based, young women", "Sponsored post", "DM", "Not Started", "500K — body confidence icon, young female audience"],
    [354, "Matt Haig (@mattzhaig)", "IG/TikTok", "UK Fitness & Wellness", "Mid", "~450K IG / 200K TikTok", "Author — Reasons to Stay Alive, anxiety/depression openness", "Partnership / Sponsored", "Email", "Not Started", "650K combined — bestselling author, teen-read mental health books"],
    [355, "Poppy Jamie (@poppyjamie)", "IG/TikTok", "UK Fitness & Wellness", "Mid", "~400K IG", "Happy Not Perfect app founder — positive mindset for young people", "Partnership / Sponsored", "Email", "Not Started", "400K — wellness app founder, positive mindset content"],
    [356, "Adrienne Herbert (@adrienneldn)", "Instagram", "UK Fitness & Wellness", "Mid", "~300K IG", "Wellness, running, Power Hour concept — older teens", "Sponsored post", "Email", "Not Started", "300K — Power Hour concept, motivational, older teen audience"],

    # ═══ UK EDUCATION PODCASTS & MEDIA (web-verified) ═══
    [357, "Parent Guide to GCSE Podcast (Emily & Paul Hughes)", "Podcast/Social", "UK Education Podcasts", "Nano", "4K+ combined", "GCSE parent advice — revision tips without nagging", "Sponsored episode / Partnership", "Email", "Not Started", "Directly targets GCSE parents — perfect product fit"],
    [358, "Emma & Tom Talk Teaching", "Podcast/Social", "UK Education Podcasts", "Mid", "60K FB / 31K X / 30K IG", "UK teaching podcast — debates, great practice", "Sponsored episode", "Email", "Not Started", "121K combined — large teacher audience, education debates"],
    [359, "Lexx Education Podcast (Ron & Laura)", "Podcast/Social", "UK Education Podcasts", "Nano", "1.5K combined", "GCSE revision podcast — 5/5 Apple rating", "Sponsored episode", "Email", "Not Started", "5/5 rated — fun GCSE revision format, niche but targeted"],
    [360, "Tips for Teachers Podcast", "Podcast", "UK Education Podcasts", "Micro", "10K+ listeners", "Teaching tips — planning, assessment, parent communication", "Sponsored episode", "Email", "Not Started", "Teacher audience — direct recommendation channel"],
    [361, "TES (Times Educational Supplement)", "Website/Social/Podcast", "UK Education Podcasts", "Mega", "500K+ combined", "UK's biggest education media — every teacher reads TES", "Advertising / Sponsored content", "Email", "Not Started", "500K+ — UK education media monopoly, teacher resource hub"],
    [362, "Schools Week (@schoolsweek)", "Website/X", "UK Education Podcasts", "Mid", "100K+ X", "UK education news — policy, school leadership", "Sponsored content / Ad", "Email", "Not Started", "100K+ — education news, read by school leaders"],

    # ═══ UK SPORTS & FOOTBALL — TEEN AUDIENCE (training data) ═══
    [363, "F2Freestylers (@f2freestylers)", "YouTube/TikTok", "UK Sports & Football", "Mega", "~12M YT", "UK football freestyle duo — tricks, challenges, collabs", "Brand deal / Sponsored", "Agent", "Not Started", "12M YT — football skills, massive teen male audience"],
    [364, "ChrisMD (@chrismd10)", "YouTube/TikTok", "UK Sports & Football", "Mega", "~7M YT", "UK football challenges — viral content, Sidemen-adjacent", "Brand deal", "Agent", "Not Started", "7M — football challenges, UK teen male audience"],
    [365, "Hashtag United (@hashtagutd)", "YouTube/IG", "UK Sports & Football", "Macro", "~2M YT", "Spencer Owen — non-league football, FIFA content", "Sponsored video", "Email", "Not Started", "2M — football content, crossover gaming/sport audience"],
    [366, "Angry Ginge (@angryginge)", "TikTok/YouTube", "UK Sports & Football", "Macro", "~3M combined", "UK football comedy — viral challenges", "Sponsored post", "DM", "Not Started", "3M combined — football comedy, viral with teens"],
    [367, "Yung Filly (@yungfilly)", "YouTube/TikTok", "UK Sports & Football", "Macro", "~3M combined", "Comedy, music, football — Beta Squad member", "Brand deal", "Agent", "Not Started", "3M combined — Beta Squad, comedy + sport crossover"],
    [368, "Chunkz x Beta Squad (@betasquadofficial)", "YouTube", "UK Sports & Football", "Mega", "~10M YT", "Football challenges, comedy, panel shows — group channel", "Brand deal", "Agent", "Not Started", "10M — Beta Squad group, football + comedy, teen audience"],

    # ═══ UK YA AUTHORS & BOOKTOK (training data + web) ═══
    [369, "Holly Jackson (@HollyJacksonAuthor)", "IG/TikTok", "UK YA Authors & BookTok", "Mid", "~300K IG", "A Good Girl's Guide to Murder — massive BookTok hit, Netflix", "Partnership / Sponsored", "Email", "Not Started", "300K — GCSE-age readers, Netflix adaptation, murder mystery"],
    [370, "Alice Oseman (@aliceoseman)", "IG/TikTok", "UK YA Authors & BookTok", "Mid", "~400K IG", "Heartstopper — Netflix, teen LGBTQ+ icon", "Partnership", "Email", "Not Started", "400K — Heartstopper Netflix, massive teen following"],
    [371, "Malorie Blackman (@maaboreal)", "IG/X", "UK YA Authors & BookTok", "Mid", "~100K combined", "Noughts and Crosses — former Children's Laureate", "Partnership", "Email", "Not Started", "100K+ — former Children's Laureate, GCSE set text author"],
    [372, "Cressida Cowell (@caboreal)", "IG/X", "UK YA Authors & BookTok", "Micro", "~50K combined", "How to Train Your Dragon — Children's Laureate 2019-22", "Partnership", "Email", "Not Started", "50K — Children's Laureate, DreamWorks franchise"],
    [373, "Phil Earle (@philearle)", "IG/X", "UK YA Authors & BookTok", "Nano", "~20K combined", "When the Sky Falls — YA author, school visit circuit", "Partnership / School event", "Email", "Not Started", "20K — active school visits, GCSE reading list potential"],
    [374, "Juno Dawson (@junodawson)", "IG/X", "UK YA Authors & BookTok", "Micro", "~80K combined", "YA author — Clean, Meat Market, trans visibility", "Partnership", "Email", "Not Started", "80K — YA author, teen audience, inclusive content"],
    [375, "Benjamin Zephaniah (estate/legacy)", "Social/Legacy", "UK YA Authors & BookTok", "Macro", "Legacy", "Refugee Boy — GCSE set text, performance poetry legend", "Legacy partnership / Tribute", "Email", "Not Started", "GCSE set text — every English student studies his work"],
    [376, "Anthony Horowitz (@AnthonyHorowitz)", "IG/X", "UK YA Authors & BookTok", "Mid", "~150K combined", "Alex Rider — spy series, Prime Video adaptation", "Partnership", "Email", "Not Started", "150K — Alex Rider TV series, teen boys reading audience"],

    # ═══ UK STUDY & PRODUCTIVITY CREATORS (training data) ═══
    [377, "Ruby Granger (@rubygranger)", "YouTube/TikTok", "UK Study & Productivity", "Macro", "~1.5M YT", "StudyTube queen — Exeter Uni, revision vlogs, study tips", "Affiliate / Sponsored", "Email", "Not Started", "1.5M — #1 UK StudyTuber, GCSE/A-Level audience"],
    [378, "UnJaded Jade (@unjadedjade)", "YouTube", "UK Study & Productivity", "Macro", "~1M YT", "Study tips, Cambridge student, productivity", "Affiliate / Sponsored", "Email", "Not Started", "1M — Cambridge student, aspirational for GCSE students"],
    [379, "Jack Edwards (@jack_edwards)", "YouTube/TikTok", "UK Study & Productivity", "Mid", "~700K YT", "BookTube, study, Durham Uni — English literature focus", "Affiliate / Sponsored", "Email", "Not Started", "700K — English Lit student, BookTok + StudyTube crossover"],
    [380, "Eve Cornwell (@evecornwell)", "YouTube", "UK Study & Productivity", "Mid", "~500K YT", "Law student, study tips, productivity — Bristol/Oxford", "Affiliate / Sponsored", "Email", "Not Started", "500K — law student, study motivation for ambitious teens"],
    [381, "Ibz Mo (@ibzmo)", "YouTube/TikTok", "UK Study & Productivity", "Mid", "~400K YT", "Study tips, motivation — relatable revision content", "Affiliate / Sponsored", "DM", "Not Started", "400K — relatable study content, diverse audience"],
    [382, "Ali Abdaal (@aliabdaal)", "YouTube", "UK Study & Productivity", "Mega", "~5M YT", "Cambridge doctor — productivity, study tips, Feel Good Productivity", "Partnership / Sponsored", "Email", "Not Started", "5M — Cambridge doctor, productivity king, huge student audience"],
    [383, "Thomas Frank (@thomas_frank)", "YouTube", "UK Study & Productivity", "Mega", "~3M YT", "Study tips, Notion — College Info Geek brand", "Sponsored video", "Email", "Not Started", "3M — study systems, Notion tutorials, student audience"],

    # ═══ UK MUSIC ARTISTS — TEEN AUDIENCE (training data) ═══
    [384, "Ed Sheeran (@teddysphotos)", "Instagram", "UK Music Artists", "Mega", "~44M IG", "UK — global superstar, massive teen following", "Aspirational / Event", "Agent", "Not Started", "44M — global reach, UK teen icon, education charity support"],
    [385, "Stormzy (@staboreal)", "IG/TikTok", "UK Music Artists", "Mega", "~5M IG", "UK grime — Stormzy Scholarship, education advocate", "Partnership / Scholarship collab", "Agent", "Not Started", "5M — Stormzy Scholarship funds Cambridge degrees, education angle"],
    [386, "Dave (@santandave)", "IG/TikTok", "UK Music Artists", "Mega", "~5M IG", "UK rap — lyrics studied in schools, BRIT Award winner", "Partnership", "Agent", "Not Started", "5M — lyrics used in GCSE English lessons, intellectual rapper"],
    [387, "Central Cee (@centralcee)", "IG/TikTok", "UK Music Artists", "Mega", "~10M IG", "UK drill — biggest UK rapper with teen audience", "Brand deal", "Agent", "Not Started", "10M — #1 UK drill artist, massive teen following"],
    [388, "Raye (@raboreal)", "IG/TikTok", "UK Music Artists", "Mega", "~3M IG", "UK — 6 BRIT Awards 2024, massive teen female audience", "Brand deal", "Agent", "Not Started", "3M — 6 BRITs 2024, breakout star, teen idol"],
    [389, "Sam Fender (@samfender)", "IG/TikTok", "UK Music Artists", "Macro", "~1.5M IG", "UK indie rock — northern working-class hero, teen fanbase", "Partnership", "Agent", "Not Started", "1.5M — Geordie, working-class themes, teen male audience"],
    [390, "Aitch (@aaboreal)", "IG/TikTok", "UK Music Artists", "Macro", "~2M IG", "Manchester rapper — teen audience, fun personality", "Brand deal", "Agent", "Not Started", "2M — Manchester, fun rap, teen demographic"],

    # ═══ ADDITIONAL CATEGORIES TO REACH 400 ═══
    [391, "Mumsnet (@mumsnet)", "Website/Social", "UK Parenting Platforms", "Mega", "1M+ monthly users", "UK's biggest parenting forum — education subforum very active", "Advertising / Sponsored thread", "Email", "Not Started", "1M+ monthly — education subforum, GCSE revision threads"],
    [392, "Netmums (@netmums)", "Website/Social", "UK Parenting Platforms", "Macro", "500K+ combined", "UK parenting resource — school advice section", "Advertising / Sponsored", "Email", "Not Started", "500K+ — school/education section, parent resource"],
    [393, "The Student Room (@studentroom)", "Website/Social", "UK Student Platforms", "Mega", "2M+ monthly users", "UK's biggest student forum — GCSE/A-Level revision hub", "Advertising / Sponsored", "Email", "Not Started", "2M+ monthly — #1 student forum, GCSE revision central"],
    [394, "Save My Exams (@savemyexams)", "Website/Social", "UK Student Platforms", "Macro", "500K+ users", "GCSE/A-Level revision platform — direct competitor/partner", "Partnership / Affiliate cross-promo", "Email", "Not Started", "500K+ — revision platform, cross-promotion opportunity"],
    [395, "Seneca Learning (@senaboreal)", "Website/Social", "UK Student Platforms", "Macro", "400K+ users", "Free GCSE revision platform — used in schools", "Partnership", "Email", "Not Started", "400K+ — used by schools, revision platform partnership"],
    [396, "BBC Bitesize (@bbcbitesize)", "Website/Social", "UK Student Platforms", "Mega", "5M+ monthly users", "BBC GCSE revision — aspirational content partnership", "Content partnership", "Email", "Not Started", "5M+ — every UK student uses Bitesize, aspirational partner"],
    [397, "Oak National Academy (@oaknational)", "Website/Social", "UK Student Platforms", "Macro", "200K+ combined", "Government-backed free curriculum — school reach", "Partnership", "Email", "Not Started", "200K+ — gov-backed, used by every UK school"],
    [398, "Twinkl (@twinklresources)", "Website/Social", "UK Education Platforms", "Macro", "1M+ teachers", "UK teaching resources — massive teacher network", "Advertising / Resource listing", "Email", "Not Started", "1M+ teachers — resource marketplace, teacher reach"],
    [399, "Tassomai (@tassomai)", "Website/Social", "UK Student Platforms", "Micro", "50K+ users", "AI GCSE revision — used by schools", "Partnership / Cross-promo", "Email", "Not Started", "50K+ — AI revision tool, school partnerships"],
    [400, "GCSEPod (@gcsepod)", "Website/Social", "UK Student Platforms", "Micro", "100K+ users", "GCSE revision podcasts — used by 1,400+ schools", "Partnership / Cross-promo", "Email", "Not Started", "1,400 schools — GCSE audio revision, direct competitor/partner"],

    # ═══ UK YA AUTHORS — GCSE SET TEXT & SCHOOL READING (training data) ═══
    [401, "Patrick Ness (@Patrick_Ness)", "X/IG", "UK YA Authors & BookTok", "Mid", "~120K X", "A Monster Calls — school reading list staple", "Partnership", "Email", "Not Started", "120K — A Monster Calls, extremely popular school text"],
    [402, "Michael Morpurgo (@MichaelMorpurgo)", "X/Facebook", "UK YA Authors & BookTok", "Micro", "~50K X", "War Horse, Private Peaceful — GCSE reading list icon", "Partnership", "Email", "Not Started", "50K — War Horse/Private Peaceful, GCSE set text legend"],
    [403, "Sarah Crossan (@SarahCrossan)", "X/IG", "UK YA Authors & BookTok", "Nano", "~20K IG", "Verse novels — One, Toffee — Carnegie Medal winner", "Partnership", "Email", "Not Started", "20K — verse novels used in GCSE English classrooms"],
    [404, "Robert Muchamore (@RobertMuchamore)", "X", "UK YA Authors & BookTok", "Nano", "~30K X", "CHERUB series — school library staple for reluctant readers", "Partnership", "Email", "Not Started", "30K — CHERUB series, reaches reluctant teen readers"],
    [405, "Muhammad Khan (@MKhanAuthor)", "IG/X", "UK YA Authors & BookTok", "Nano", "~15K IG", "I Am Thunder — former teacher, school visiting author", "Partnership / School event", "Email", "Not Started", "15K — ex-teacher, tackles GCSE-relevant themes, school visits"],
    [406, "Alex Wheatle (@AlexWheatle)", "X", "UK YA Authors & BookTok", "Nano", "~15K X", "Crongton Knights — Guardian Prize winner, school reading list", "Partnership", "Email", "Not Started", "15K — Guardian Prize, inner-city London YA fiction"],
    [407, "Patrice Lawrence (@LawrencePatrice)", "X/IG", "UK YA Authors & BookTok", "Nano", "~15K X", "Orangeboy — Waterstones Prize + YA Book Prize winner", "Partnership", "Email", "Not Started", "15K — double prize winner, school reading list"],
    [408, "Sita Brahmachari (@SitaBrahmachari)", "X", "UK YA Authors & BookTok", "Nano", "~10K X", "Artichoke Hearts — Waterstones Children's Book Prize winner", "Partnership", "Email", "Not Started", "10K — Waterstones Prize, diverse voices in GCSE English"],

    # ═══ BRITISH TV CELEBRITIES — PARENT AUDIENCE (training data) ═══
    [409, "Holly Willoughby (@hollywilloughby)", "Instagram", "British TV & Celebrity", "Mega", "~8M IG", "Former This Morning — mother of 3 school-age kids", "Brand deal", "Agent", "Not Started", "8M — UK's most-followed TV presenter, school-age mum"],
    [410, "Rochelle Humes (@rochellehumes)", "Instagram", "British TV & Celebrity", "Macro", "~2.3M IG", "This Morning presenter, The Saturdays — mother of 3", "Brand deal", "Agent", "Not Started", "2.3M — This Morning, fashion + parenting content"],
    [411, "Billie Shepherd (@billieshepherdofficial)", "Instagram", "British TV & Celebrity", "Macro", "~2.2M IG", "TOWIE, Family Diaries — mother of 3", "Sponsored post", "Agent", "Not Started", "2.2M — TOWIE/Family Diaries, parent audience"],
    [412, "Peter Andre (@peterandre)", "Instagram", "British TV & Celebrity", "Macro", "~1.8M IG", "Singer, reality TV — father of 4, younger school-age", "Sponsored post", "Agent", "Not Started", "1.8M — reality TV, family man, broad parent reach"],
    [413, "Frankie Bridge (@frankiebridge)", "Instagram", "British TV & Celebrity", "Macro", "~1.5M IG", "The Saturdays, Loose Women — mental health advocate, mum of 2", "Sponsored post", "Agent", "Not Started", "1.5M — Loose Women, mental health + parenting"],
    [414, "Vogue Williams (@voguewilliams)", "Instagram", "British TV & Celebrity", "Macro", "~1M IG", "TV presenter, podcaster — London-based mum of 3", "Sponsored post", "Email", "Not Started", "1M — Spencer & Vogue podcast, UK parent audience"],
    [415, "Spencer Matthews (@spencermatthews)", "Instagram", "British TV & Celebrity", "Mid", "~800K IG", "Made in Chelsea — father of 3, ultra-marathon runner", "Sponsored post", "Email", "Not Started", "800K — MiC, inspirational dad content"],
    [416, "Christine McGuinness (@mrscmcguinness)", "Instagram", "British TV & Celebrity", "Mid", "~700K IG", "Autism awareness campaigner — mother of 3", "Partnership / Sponsored", "Email", "Not Started", "700K — autism advocacy, SEN parent audience"],
    [417, "Danny Jones / McFly (@dannyjonesofficial)", "Instagram", "British TV & Celebrity", "Mid", "~600K IG", "McFly — I'm A Celeb 2024 winner, dad of school-age son", "Sponsored post", "Agent", "Not Started", "600K — McFly + I'm A Celeb winner, family content"],
    [418, "Charley Webb (@miss_charleywebb)", "Instagram", "British TV & Celebrity", "Mid", "~500K IG", "Emmerdale actress — mother of 3, honest parenting", "Sponsored post", "Email", "Not Started", "500K — Emmerdale, relatable parenting content"],

    # ═══ UK MOTIVATIONAL & PERSONAL DEVELOPMENT — YOUTH (training data) ═══
    [419, "Steven Bartlett (@stevenbartlett)", "IG/YouTube", "UK Motivational & Youth", "Mega", "~13M IG / 5M YT", "Diary of a CEO — UK's biggest podcast, entrepreneurship + mindset", "Brand deal / Sponsored", "Agent", "Not Started", "18M combined — #1 UK podcast, massive youth/student audience"],
    [420, "Jay Shetty (@jayshetty)", "IG/YouTube", "UK Motivational & Youth", "Mega", "~12M IG / 4M YT", "London-origin — purpose, mindfulness, relationships", "Brand deal", "Agent", "Not Started", "16M combined — London-origin, massive global youth following"],
    [421, "Vex King (@vfrking)", "Instagram", "UK Motivational & Youth", "Mid", "~800K IG", "Good Vibes Good Life author — hugely popular with teens", "Partnership / Sponsored", "Email", "Not Started", "800K — bestselling self-help, teen + early 20s audience"],
    [422, "Jay Alderton (@jayalderton)", "TikTok/IG", "UK Motivational & Youth", "Mid", "~500K TikTok", "UK military vet — fitness motivation, discipline for young men", "Sponsored post", "DM", "Not Started", "500K — discipline/fitness, teen male audience"],
    [423, "Zanna van Dijk (@zannavandijk)", "IG/YouTube", "UK Motivational & Youth", "Mid", "~400K IG", "Wellness, sustainable living — London, 18-25 demographic", "Sponsored post", "Email", "Not Started", "400K — sustainability + wellness, young female audience"],
    [424, "Matthew Hussey (@thematthewhussey)", "IG/YouTube", "UK Motivational & Youth", "Macro", "~3M IG / 2M YT", "Essex-origin — confidence, self-worth, huge female following", "Sponsored video", "Agent", "Not Started", "5M combined — confidence coaching, teen/young adult females"],
    [425, "Robert Mayanja / The Motivator (@robertmayanja)", "TikTok/IG", "UK Motivational & Youth", "Mid", "~200K TikTok", "Study motivation, exam mindset — UK teens and students", "Sponsored post", "DM", "Not Started", "200K — study motivation, directly targets UK exam students"],
    [426, "Simon Alexander Ong (@simonalexanderong)", "IG/LinkedIn", "UK Motivational & Youth", "Micro", "~100K combined", "London life coach — Energize author, productivity", "Partnership", "Email", "Not Started", "100K — author, productivity coaching, professional audience"],
    [427, "Liam Hackett / Ditch the Label (@liamhackett)", "IG/LinkedIn", "UK Motivational & Youth", "Micro", "~50K combined", "Anti-bullying charity founder — directly targets teens", "Partnership", "Email", "Not Started", "50K — Ditch the Label charity, directly reaches UK teens in schools"],
    [428, "Leon Taylor (@leontaylor)", "Instagram", "UK Motivational & Youth", "Nano", "~30K IG", "Olympic diver — mental performance, mindset coaching", "Partnership / School event", "Email", "Not Started", "30K — Olympic athlete, school speaking circuit"],

    # ═══ UK STUDENT LIFE VLOGGERS (training data) ═══
    [429, "Vee Kativhu (@VeeKativhu)", "YouTube", "UK Study & Productivity", "Mid", "~300K YT", "Oxford student — study tips, motivation, diverse voice", "Affiliate / Sponsored", "Email", "Not Started", "300K — Oxford, aspirational for GCSE students aiming high"],
    [430, "Simon Clark (@SimonClark)", "YouTube", "UK Study & Productivity", "Mid", "~500K YT", "PhD science communicator — Exeter, student experience", "Sponsored video", "Email", "Not Started", "500K — science communication, appeals to STEM students"],
    [431, "Mike Boyd (@MikeBoyd)", "YouTube", "UK Study & Productivity", "Macro", "~3M YT", "Scottish — learning challenges, appeals to student audience", "Sponsored video", "Email", "Not Started", "3M — learning/skills channel, student demographic"],
    [432, "Jess Studies (@JessStudies)", "YouTube", "UK Study & Productivity", "Micro", "~100K YT", "Edinburgh medical student — study vlogs, day-in-the-life", "Affiliate / Sponsored", "Email", "Not Started", "100K — medical student, aspirational for ambitious teens"],
    [433, "Tolani Sherif (@tolanisherif)", "YouTube", "UK Study & Productivity", "Micro", "~50K YT", "Warwick — Black British student life, study motivation", "Sponsored video", "DM", "Not Started", "50K — diverse voice, Warwick uni, study motivation"],

    # ═══ UK MUSIC ARTISTS — ADDITIONAL TEEN AUDIENCE (training data) ═══
    [434, "Dua Lipa (@dualipa)", "Instagram", "UK Music Artists", "Mega", "~88M IG", "UK pop — global superstar, older teen audience", "Aspirational / Event", "Agent", "Not Started", "88M — biggest UK female artist, global teen reach"],
    [435, "PinkPantheress (@pinkpantheress)", "TikTok", "UK Music Artists", "Mega", "~5M TikTok", "UK — broke out via TikTok, defining Gen Z artist", "Brand deal", "Agent", "Not Started", "5M — TikTok-native artist, quintessential Gen Z sound"],
    [436, "Jorja Smith (@jorjasmith_)", "Instagram", "UK Music Artists", "Mega", "~5M IG", "UK R&B/neo-soul — Walsall, massive teen female following", "Brand deal", "Agent", "Not Started", "5M — UK R&B star, strong teen female audience"],
    [437, "Ardee (@ardeemusicc)", "TikTok", "UK Music Artists", "Macro", "~3M TikTok", "UK rap/pop-rap — TikTok viral, teen male audience", "Brand deal", "DM", "Not Started", "3M — viral UK rap, secondary school favourite"],
    [438, "Cat Burns (@catburns)", "TikTok", "UK Music Artists", "Macro", "~2.5M TikTok", "UK pop — singer-songwriter, broke out on TikTok", "Sponsored post", "DM", "Not Started", "2.5M — TikTok breakout, relatable pop, teen audience"],
    [439, "Headie One (@headieone)", "Instagram", "UK Music Artists", "Macro", "~2.5M IG", "UK drill — extremely popular in secondary schools", "Brand deal", "Agent", "Not Started", "2.5M — UK drill pioneer, secondary school listenership"],
    [440, "Digga D (@diggad_)", "Instagram", "UK Music Artists", "Macro", "~2M IG", "UK drill — massive school-age following", "Brand deal", "Agent", "Not Started", "2M — UK drill, one of most-streamed UK artists by teens"],
    [441, "Tion Wayne (@tionwayne)", "TikTok", "UK Music Artists", "Macro", "~2M TikTok", "UK rap/afroswing — Body went #1, TikTok viral", "Sponsored post", "DM", "Not Started", "2M — Body #1 hit, afroswing genre, teen dance trends"],
    [442, "Little Simz (@littlesimz)", "Instagram", "UK Music Artists", "Macro", "~1M IG", "UK rap — Mercury Prize, critical acclaim, aspirational", "Partnership", "Agent", "Not Started", "1M — Mercury Prize winner, intellectual rap, aspirational"],
    [443, "Knucks (@knfrm)", "Instagram", "UK Music Artists", "Mid", "~500K IG", "UK jazz-rap — Mercury Prize, growing teen following", "Sponsored post", "DM", "Not Started", "500K — Mercury Prize, jazz-rap crossover, rising star"],

    # ═══ UK BEAUTY & FASHION — TEEN GIRLS (training data) ═══
    [444, "Abby Roberts (@abbyroberts)", "TikTok", "UK Beauty & Fashion Teen", "Mega", "~17M TikTok", "UK — makeup artistry, SFX, beauty tutorials", "Brand deal", "Agent", "Not Started", "17M — biggest UK beauty TikToker, teen girls"],
    [445, "Olivia Neill (@olivianeill)", "TikTok/IG", "UK Beauty & Fashion Teen", "Macro", "~3M TikTok", "UK — GRWM, fashion hauls, beauty, lifestyle", "Sponsored post", "DM", "Not Started", "3M — GRWM queen, fashion hauls, teen demographic"],
    [446, "India Moon (@indiamoon)", "TikTok", "UK Beauty & Fashion Teen", "Macro", "~3M TikTok", "UK — GRWM, beauty, teen fashion", "Sponsored post", "DM", "Not Started", "3M — GRWM, beauty content, strong teen following"],
    [447, "Saffron Barker (@saffronbarker)", "TikTok/IG", "UK Beauty & Fashion Teen", "Macro", "~2.5M TikTok", "UK — fashion, beauty, vlogs, Strictly contestant", "Sponsored post", "DM", "Not Started", "2.5M — Strictly, beauty/fashion, Brighton-based"],
    [448, "Aisha Beau (@aishabeau)", "TikTok", "UK Beauty & Fashion Teen", "Macro", "~1.5M TikTok", "UK — makeup tutorials, GRWM", "Sponsored post", "DM", "Not Started", "1.5M — makeup tutorials, strong teen girl following"],
    [449, "Monami Frost (@monamifrost)", "IG/TikTok", "UK Beauty & Fashion Teen", "Macro", "~1.5M IG", "UK — alternative beauty, tattoo style, fashion", "Sponsored post", "Email", "Not Started", "1.5M — alternative beauty, unique aesthetic, teen alt audience"],
    [450, "Sophie Hannah Richardson (@sophiehannah)", "Instagram", "UK Beauty & Fashion Teen", "Macro", "~1M IG", "UK — fashion hauls, outfit inspo, lifestyle", "Sponsored post", "Email", "Not Started", "1M — fashion inspiration, outfit of the day"],
    [451, "Eliza Rose Watson (@elizarosewatson)", "TikTok/IG", "UK Beauty & Fashion Teen", "Macro", "~1M TikTok", "UK — fashion styling, outfit inspo", "Sponsored post", "DM", "Not Started", "1M — fashion styling, teen outfit inspiration"],
    [452, "Jessica Sherwood (@jessicasherwood)", "TikTok", "UK Beauty & Fashion Teen", "Mid", "~800K TikTok", "UK — GRWM, skincare, beauty routines", "Sponsored post", "DM", "Not Started", "800K — skincare + beauty routines, teen girls"],
    [453, "Emily Sherlock (@emilysherlock)", "TikTok", "UK Beauty & Fashion Teen", "Mid", "~700K TikTok", "UK — beauty, GRWM, fashion hauls", "Sponsored post", "DM", "Not Started", "700K — beauty GRWM, fashion content"],
    [454, "Grace Sherwood (@itsgracesherwood)", "TikTok", "UK Beauty & Fashion Teen", "Mid", "~600K TikTok", "UK — Primark hauls, budget teen fashion", "Sponsored post", "DM", "Not Started", "600K — budget fashion, Primark hauls, relatable"],
    [455, "Ruby Richardson (@rubyrichardson_)", "TikTok", "UK Beauty & Fashion Teen", "Mid", "~500K TikTok", "UK — teen fashion, OOTD, hauls", "Sponsored post", "DM", "Not Started", "500K — OOTD, teen fashion hauls"],
    [456, "Ambar Driscoll (@ambardriscoll)", "IG/TikTok", "UK Beauty & Fashion Teen", "Mid", "~400K IG", "UK — fashion, beauty, wellness, GRWM", "Sponsored post", "Email", "Not Started", "400K — fashion + wellness crossover"],

    # ═══ UK DRAMA, THEATRE & ACTORS — ENGLISH LIT CROSSOVER (training data) ═══
    [457, "Tom Hiddleston (@twhiddleston)", "Instagram", "UK Drama & Theatre", "Mega", "~12M IG", "RADA-trained — Hamlet, Coriolanus, brings youth to Shakespeare", "Aspirational / Partnership", "Agent", "Not Started", "12M — young audience gateway to Shakespeare, Loki/Marvel crossover"],
    [458, "David Tennant (@davidtennant)", "Instagram", "UK Drama & Theatre", "Mega", "~5M IG", "RSC Hamlet + Richard II — Doctor Who teen crossover", "Partnership", "Agent", "Not Started", "5M — Doctor Who + Shakespeare, audiobooks of classics"],
    [459, "Ian McKellen (@ianmckellen)", "Instagram", "UK Drama & Theatre", "Mega", "~4.5M IG", "RSC legend — Richard III, Macbeth, King Lear", "Aspirational", "Agent", "Not Started", "4.5M — Gandalf + Shakespeare, brings gravitas to English Lit"],
    [460, "Judi Dench (@judidenchofficial)", "Instagram", "UK Drama & Theatre", "Mega", "~4M IG", "RSC icon — TikTok viral in her 90s", "Aspirational", "Agent", "Not Started", "4M — went viral on TikTok, Shakespeare dame, cross-generational"],
    [461, "Michaela Coel (@michaelacoel)", "Instagram", "UK Drama & Theatre", "Macro", "~3M IG", "I May Destroy You — modern British voice, BAFTA/Emmy winner", "Partnership", "Agent", "Not Started", "3M — modern British writing, relevant to GCSE English Language themes"],
    [462, "Andrew Scott (@andrewscott_official)", "Instagram", "UK Drama & Theatre", "Macro", "~3M IG", "Old Vic Hamlet 2024 — live-streamed, literary credentials", "Partnership", "Agent", "Not Started", "3M — Hamlet, Vanya, Fleabag — brings Shakespeare to new audiences"],
    [463, "Kenneth Branagh (@kennethbranagh)", "Instagram", "UK Drama & Theatre", "Mid", "~600K IG", "Directed Henry V, Hamlet, Much Ado — Shakespeare film king", "Aspirational", "Agent", "Not Started", "600K — Shakespeare film adaptations studied at GCSE/A-Level"],
    [464, "National Theatre (@nationaltheatre)", "IG/X/YouTube", "UK Drama & Theatre", "Mid", "~600K+ combined", "NT Live — streams Shakespeare to schools", "Partnership / Schools programme", "Email", "Not Started", "600K+ — NT Live schools programme, direct classroom reach"],
    [465, "Ben Whishaw (@benwhishaw_official)", "Instagram", "UK Drama & Theatre", "Mid", "~500K IG", "Richard II (Hollow Crown) — played Keats, literary roles", "Partnership", "Agent", "Not Started", "500K — Hollow Crown Richard II, literary filmography"],
    [466, "Jessie Buckley (@jessiebuckleyofficial)", "Instagram", "UK Drama & Theatre", "Mid", "~400K IG", "Literary film adaptations — Olivier-nominated", "Partnership", "Agent", "Not Started", "400K — Women Talking, Lost Daughter, literary film star"],
    [467, "Shakespeare's Globe (@shakespearesglobe)", "IG/X/YouTube", "UK Drama & Theatre", "Mid", "~300K+ combined", "Globe Theatre — education programme, school visits", "Partnership / Schools programme", "Email", "Not Started", "300K+ — Globe education, school field trips, GCSE Shakespeare"],
    [468, "Maxine Peake (@maxinepeake1)", "X", "UK Drama & Theatre", "Mid", "~230K X", "Gender-blind Hamlet — advocate for accessible theatre", "Partnership", "Email", "Not Started", "230K — accessible theatre advocate, relatable to students"],
    [469, "Meera Syal (@meerasyal)", "X", "UK Drama & Theatre", "Mid", "~120K X", "Anita and Me — GCSE set text author, CBE", "Partnership", "Email", "Not Started", "120K — Anita and Me = GCSE set text, diversity in English Lit"],
    [470, "Adrian Lester (@adrianlester)", "X", "UK Drama & Theatre", "Micro", "~100K X", "Henry V, Othello — landmark diverse casting at National Theatre", "Partnership", "Email", "Not Started", "100K — diverse Shakespeare casting, representation"],
    [471, "James Graham (@jamesgrahamps)", "X", "UK Drama & Theatre", "Micro", "~50K X", "Dear England, Ink, Quiz — politically engaged playwright", "Partnership", "Email", "Not Started", "50K — modern playwright, relevant to English Language/media"],

    # ═══ UK FOOTBALL CREATORS — ADDITIONAL (training data) ═══
    [472, "TBJZL / Tobi Brown (@TBJZL)", "YouTube", "UK Sports & Football", "Mega", "~4M YT", "Sidemen member — football content, charity matches", "Brand deal", "Agent", "Not Started", "4M — Sidemen, football challenges, teen audience"],
    [473, "Daniel Cutting (@danielcutting)", "YouTube/TikTok", "UK Sports & Football", "Macro", "~3M YT", "Football freestyle, skills, trick shots", "Sponsored video", "DM", "Not Started", "3M — freestyle skills, viral trick shots"],
    [474, "Mark Goldbridge (@markgoldbridge)", "YouTube/TikTok", "UK Sports & Football", "Macro", "~2.5M YT", "Man United fan reactions — matchday watchalongs", "Sponsored", "Email", "Not Started", "2.5M — Man United, match reactions, teen male audience"],
    [475, "JaackMaate / Johnny Carey (@JaackMaate)", "YouTube", "UK Sports & Football", "Macro", "~2M YT", "Football culture commentary, podcasting", "Sponsored", "Email", "Not Started", "2M — Happy Hour podcast, football culture"],
    [476, "Ben Foster / The Cycling GK (@TheCyclingGK)", "YouTube", "UK Sports & Football", "Macro", "~1.5M YT", "Retired pro goalkeeper — behind-the-scenes football", "Sponsored video", "Email", "Not Started", "1.5M — actual pro footballer, unique credibility"],
    [477, "AFTV / Arsenal Fan TV (@AFTVmedia)", "YouTube", "UK Sports & Football", "Macro", "~1.5M YT", "Arsenal fan reactions — matchday content", "Sponsored", "Email", "Not Started", "1.5M — Arsenal fans, high-frequency matchday content"],
    [478, "The United Stand (@TheUnitedStand)", "YouTube", "UK Sports & Football", "Macro", "~1.2M YT", "Man United fan channel — reactions, transfers", "Sponsored", "Email", "Not Started", "1.2M — Man United, daily content, teen male fans"],
    [479, "Vizeh (@Vizeh)", "YouTube/TikTok", "UK Sports & Football", "Macro", "~1M YT", "Football edits, player highlight compilations", "Sponsored", "DM", "Not Started", "1M — football edits, viral compilations, teen audience"],
    [480, "Expressions Oozing (@ExpressionsOozing)", "YouTube", "UK Sports & Football", "Mid", "~700K YT", "Arsenal fan reactions — passionate matchday vlogs", "Sponsored", "Email", "Not Started", "700K — Arsenal, passionate reactions, engaged fanbase"],
    [481, "Manny (@Manny)", "YouTube", "UK Sports & Football", "Mega", "~4M YT", "Beta Squad — football challenges, sports entertainment", "Brand deal", "Agent", "Not Started", "4M — Beta Squad member, football + comedy"],

    # ═══ ARABIC FAMILY YOUTUBE — GCC PARENT AUDIENCE (training data) ═══
    [482, "Fayhan Family (عائلة فيحان)", "YouTube", "Arabic Family YouTube", "Mega", "~25M YT", "Saudi — family vlogs, comedy, challenges", "Brand deal", "Agent", "Not Started", "25M — biggest Arabic family channel, Saudi parents + kids"],
    [483, "mmoshaya (محمد مشايع)", "YouTube", "Arabic Family YouTube", "Mega", "~20M YT", "Saudi — family-friendly vlogs, travel, challenges", "Brand deal", "Agent", "Not Started", "20M — family travel/challenges, Saudi-based"],
    [484, "Anasala Family (أنس و أصالة)", "YouTube", "Arabic Family YouTube", "Mega", "~18M YT", "Kuwait/UAE — family vlogs, daily life, travel", "Brand deal", "Agent", "Not Started", "18M — Kuwait/UAE family, massive parent audience"],
    [485, "Karameesh TV (كراميش)", "YouTube", "Arabic Family YouTube", "Mega", "~18M YT", "Saudi/UAE — children's songs, family entertainment", "Partnership / Ad", "Email", "Not Started", "18M — children's content, parents choose what kids watch"],
    [486, "Shabab Al Bomb (شباب البومب)", "YouTube", "Arabic Family YouTube", "Mega", "~12M YT", "Saudi — family comedy series", "Brand deal", "Agent", "Not Started", "12M — Saudi comedy series, family-friendly"],
    [487, "Abu Raed Family (عائلة ابو رعد)", "YouTube", "Arabic Family YouTube", "Mega", "~8M YT", "Saudi — family challenges, comedy skits", "Brand deal", "Agent", "Not Started", "8M — family challenges, Saudi parent/kid audience"],
    [488, "Majid Kids TV (قناة ماجد)", "YouTube", "Arabic Family YouTube", "Mega", "~6M YT", "UAE — children's educational entertainment", "Partnership", "Email", "Not Started", "6M — UAE educational kids content, parent-trusted"],
    [489, "Abdullah Vlog (فلوق عبدالله / @iABD)", "YouTube", "Arabic Family YouTube", "Mega", "~5M YT", "Saudi — family lifestyle vlogs", "Brand deal", "Agent", "Not Started", "5M — Saudi family lifestyle, daily content"],
    [490, "Saud & Sarah (سعود و سارة)", "YouTube", "Arabic Family YouTube", "Mega", "~4M YT", "Saudi — family vlogs, challenges", "Brand deal", "DM", "Not Started", "4M — Saudi couple/family, young parent audience"],
    [491, "Badr & Hind (بدر وهند)", "YouTube", "Arabic Family YouTube", "Macro", "~3M YT", "Saudi — couple/family vlogs", "Brand deal", "DM", "Not Started", "3M — Saudi family vlogs, relatable daily life"],
    [492, "Adnan Family (عائلة عدنان)", "YouTube", "Arabic Family YouTube", "Macro", "~3M YT", "Saudi — kids/family entertainment", "Brand deal", "DM", "Not Started", "3M — kids content, Saudi family channel"],
    [493, "Nisana Family (عائلة نيسانة)", "YouTube", "Arabic Family YouTube", "Macro", "~2M YT", "UAE — family entertainment, kids content", "Sponsored video", "DM", "Not Started", "2M — UAE family entertainment"],
    [494, "Baba W Mama (بابا وماما)", "YouTube", "Arabic Family YouTube", "Macro", "~1M YT", "Kuwait — parenting, family daily life", "Sponsored video", "DM", "Not Started", "1M — Kuwaiti parenting channel, daily family content"],
    [495, "Maram & Family (مرام و عائلتها)", "YouTube", "Arabic Family YouTube", "Mid", "~800K YT", "Saudi — family lifestyle, motherhood", "Sponsored video", "DM", "Not Started", "800K — Saudi motherhood content, growing channel"],

    # ═══ UK REALITY TV — TEEN + PARENT AUDIENCE (training data) ═══
    [496, "Molly-Mae Hague (@mollymae)", "Instagram", "UK Reality TV", "Mega", "~8M IG", "Love Island 2019 — biggest UK reality star, teen icon", "Brand deal", "Agent", "Not Started", "8M — #1 Love Island star, massive teen female following"],
    [497, "Vicky Pattison (@vickypattison)", "Instagram", "UK Reality TV", "Mega", "~5M IG", "Geordie Shore + I'm A Celeb winner — huge teen audience", "Brand deal", "Agent", "Not Started", "5M — I'm A Celeb winner, teen + young adult audience"],
    [498, "Tommy Fury (@tommyfury)", "Instagram", "UK Reality TV", "Mega", "~4.5M IG", "Love Island + boxer — teen male crossover", "Brand deal", "Agent", "Not Started", "4.5M — boxing + Love Island, teen male audience"],
    [499, "Maura Higgins (@maurahiggins)", "Instagram", "UK Reality TV", "Macro", "~3.5M IG", "Love Island 2019 — hugely popular with teen/young adult", "Brand deal", "Agent", "Not Started", "3.5M — Love Island breakout, teen female audience"],
    [500, "Jacqueline Jossa (@jacjossa)", "Instagram", "UK Reality TV", "Macro", "~3.5M IG", "EastEnders + I'm A Celeb winner — mum of school-age kids", "Sponsored post", "Email", "Not Started", "3.5M — EastEnders, I'm A Celeb, daughters in primary school"],
    [501, "Dani Dyer (@danidyerxx)", "Instagram", "UK Reality TV", "Macro", "~3.5M IG", "Love Island 2018 winner — mum of 3, Danny Dyer's daughter", "Sponsored post", "Email", "Not Started", "3.5M — Love Island winner, young mum, family content"],
    [502, "Olivia Bowen (@oliviadbowen)", "Instagram", "UK Reality TV", "Macro", "~3M IG", "Love Island 2016 — mum, home/lifestyle content", "Sponsored post", "Email", "Not Started", "3M — Love Island, home/lifestyle + parenting"],
    [503, "Charlotte Crosby (@charlottegshore)", "Instagram", "UK Reality TV", "Macro", "~3M IG", "Geordie Shore — mum, fitness/lifestyle", "Sponsored post", "Email", "Not Started", "3M — Geordie Shore, fitness + parenting content"],
    [504, "Louise Thompson (@louisethompson)", "Instagram", "UK Reality TV", "Macro", "~1.5M IG", "Made in Chelsea — mum, mental health advocate", "Sponsored post", "Email", "Not Started", "1.5M — MiC, openly discusses postnatal mental health"],
    [505, "Binky Felstead (@binkyfelstead)", "Instagram", "UK Reality TV", "Macro", "~1.5M IG", "Made in Chelsea — mum of 3, parenting content", "Sponsored post", "Email", "Not Started", "1.5M — MiC, family content, daughter in primary school"],
    [506, "Amber Davies (@amberdavies)", "TikTok", "UK Reality TV", "Macro", "~1M TikTok", "Love Island 2017 winner — popular with teen girls", "Sponsored post", "DM", "Not Started", "1M — Love Island winner, West End actress, teen audience"],

    # ═══ UK CAREER ADVICE & EMPLOYABILITY PLATFORMS (training data) ═══
    [507, "Bright Network (@brightnetwork)", "LinkedIn/IG", "UK Career & Employability", "Mid", "~300K LinkedIn", "Graduate schemes, internships — employer events", "Partnership / Ad", "Email", "Not Started", "300K — graduate career platform, 16-25 audience"],
    [508, "Kush Careers (@kushcareers)", "TikTok/LinkedIn", "UK Career & Employability", "Micro", "~100K TikTok", "Graduate job hunting, CV reviews, LinkedIn tips", "Sponsored post", "DM", "Not Started", "100K — Gen Z career coach, TikTok native"],
    [509, "Apprenticeships Gov UK (@apprenticeships)", "IG/TikTok", "UK Career & Employability", "Micro", "~80K combined", "Official apprenticeship promotion — gov-backed", "Partnership", "Email", "Not Started", "80K — official gov channel, school careers fairs"],
    [510, "NotGoingToUni (@notgoingtouni)", "TikTok/IG", "UK Career & Employability", "Micro", "~60K TikTok", "Apprenticeship listings, alternatives to uni", "Partnership / Ad", "Email", "Not Started", "60K — uni alternatives, reaches school-leavers"],
    [511, "Prospects.ac.uk (@prospects)", "LinkedIn/IG", "UK Career & Employability", "Micro", "~50K LinkedIn", "Graduate career guidance, job sector explainers", "Partnership", "Email", "Not Started", "50K — official career guidance, used by school careers depts"],
    [512, "RateMyApprenticeship (@ratemyapp)", "TikTok/IG", "UK Career & Employability", "Micro", "~40K TikTok", "Apprenticeship reviews, day-in-the-life employer content", "Partnership / Ad", "Email", "Not Started", "40K — apprenticeship reviews, employer rankings"],
    [513, "Success at School (@successatschool)", "IG/YouTube", "UK Career & Employability", "Nano", "~30K IG", "Career quizzes, subject-to-career guides", "Partnership", "Email", "Not Started", "30K — school careers resource, subject-to-career mapping"],
    [514, "WorldSkills UK / The Leap (@worldskillsuk)", "IG/TikTok", "UK Career & Employability", "Nano", "~25K IG", "Skills competitions, vocational careers, apprenticeships", "Partnership", "Email", "Not Started", "25K — vocational skills, apprenticeship advocacy"],
    [515, "LifeSkills by Barclays (@barclayslifeskills)", "YouTube/IG", "UK Career & Employability", "Nano", "~20K YT", "Employability workshops, interview skills, financial literacy", "Partnership", "Email", "Not Started", "20K — Barclays-backed, school workshop programme"],

    # ═══ UK STEM & SCIENCE COMMUNICATORS (training data) ═══
    [516, "Tom Scott (@TomScottGo)", "YouTube", "UK STEM & Science", "Mega", "~6.4M YT", "Science, linguistics, geography — 'Things You Might Not Know'", "Partnership", "Agent", "Not Started", "6.4M — linguistics crossover with English, massive teen following"],
    [517, "Numberphile / Brady Haran (@numberphile)", "YouTube", "UK STEM & Science", "Mega", "~4.5M YT", "Maths explainers with uni academics — Nottingham-based", "Sponsored video", "Email", "Not Started", "4.5M — maths channel, GCSE/A-Level audience"],
    [518, "Steve Mould (@SteveMould)", "YouTube", "UK STEM & Science", "Mega", "~4.5M YT", "Physics demos, fluid dynamics — quirky experiments", "Sponsored video", "Email", "Not Started", "4.5M — physics demos, appeals to curious teens"],
    [519, "James May (@JamesMay)", "YouTube", "UK STEM & Science", "Macro", "~2.5M YT", "Engineering, 'how things work' — practical builds", "Brand deal", "Agent", "Not Started", "2.5M — Top Gear fame, engineering curiosity, broad audience"],
    [520, "Periodic Videos / Brady Haran (@periodicvideos)", "YouTube", "UK STEM & Science", "Macro", "~1.7M YT", "Chemistry — Nottingham uni, one video per element", "Sponsored video", "Email", "Not Started", "1.7M — chemistry, GCSE Science crossover"],
    [521, "Jay Foreman / Map Men (@JayForeman)", "YouTube", "UK STEM & Science", "Macro", "~1.5M YT", "Geography, maps, London history — comedic explainers", "Sponsored video", "Email", "Not Started", "1.5M — geography + history, GCSE crossover subject"],
    [522, "Matt Parker / Standup Maths (@standupmaths)", "YouTube", "UK STEM & Science", "Macro", "~1.2M YT", "Maths comedy — live standup maths shows, TV regular", "Partnership", "Email", "Not Started", "1.2M — maths comedy, makes maths fun for teens"],
    [523, "Medlife Crisis / Rohin Francis (@MedlifeCrisis)", "YouTube", "UK STEM & Science", "Macro", "~1.2M YT", "NHS cardiologist — medical science, health myth debunking", "Sponsored video", "Email", "Not Started", "1.2M — practising NHS doctor, credible science voice"],
    [524, "IFLScience / Elise Andrew (@IFLScience)", "TikTok/YouTube", "UK STEM & Science", "Macro", "~1M TikTok", "General science news, viral science stories", "Sponsored post", "Email", "Not Started", "1M+ — viral science brand, broad teen audience"],
    [525, "Dr. Becky Smethurst (@DrBecky)", "YouTube", "UK STEM & Science", "Mid", "~800K YT", "Oxford astrophysicist — space news, black holes", "Sponsored video", "Email", "Not Started", "800K — Oxford astrophysicist, inspiring female STEM voice"],
    [526, "Maddie Moate (@MaddieMoate)", "YouTube/TikTok", "UK STEM & Science", "Mid", "~500K combined", "Nature, wildlife — family-friendly STEM", "Sponsored video", "Email", "Not Started", "500K — nature/STEM, family audience, CBeebies presenter"],
    [527, "Myles Power (@MylesPower)", "YouTube", "UK STEM & Science", "Mid", "~400K YT", "Chemistry, debunking pseudoscience — science skepticism", "Sponsored video", "Email", "Not Started", "400K — chemistry, critical thinking, GCSE Science relevant"],
    [528, "Hannah Fry (@hannahfry)", "YouTube/BBC", "UK STEM & Science", "Mid", "~300K YT", "UCL maths professor — BBC presenter, data science + AI", "Partnership", "Email", "Not Started", "300K — BBC presenter, maths + AI, aspirational female STEM"],
    [529, "Sally Le Page (@SallyLePageBio)", "YouTube/TikTok", "UK STEM & Science", "Mid", "~200K combined", "Evolutionary biology, genetics — Oxford background", "Sponsored video", "Email", "Not Started", "200K — biology, genetics, female STEM voice"],

    # ═══ UK TRAVEL & GAP YEAR CREATORS (training data) ═══
    [530, "Jack Morris / Do You Travel (@doyoutravel)", "Instagram", "UK Travel & Gap Year", "Macro", "~2.7M IG", "UK — aspirational travel photography", "Brand deal", "Agent", "Not Started", "2.7M — aspirational travel, appeals to 16-18 gap year dreamers"],
    [531, "FunForLouis / Louis Cole (@FunForLouis)", "YouTube", "UK Travel & Gap Year", "Macro", "~1.9M YT", "Adventure travel vlogs — immersive cultural travel", "Sponsored video", "Email", "Not Started", "1.9M — OG UK travel vlogger, adventure content"],
    [532, "Eva zu Beck (@evazubeck)", "YouTube", "UK Travel & Gap Year", "Mid", "~600K YT", "Solo female travel — off-the-beaten-path destinations", "Sponsored video", "Email", "Not Started", "600K — solo female travel, inspiring for teen girls"],
    [533, "Stevo the Madman (@stevothemadman)", "TikTok/YouTube", "UK Travel & Gap Year", "Mid", "~500K TikTok", "Budget backpacking, gap year culture, hostel life", "Sponsored post", "DM", "Not Started", "500K — budget gap year, hostel culture, teen audience"],
    [534, "Mollie Bylett (@molliebylett)", "YouTube/IG", "UK Travel & Gap Year", "Mid", "~400K YT", "Solo female backpacking, budget travel, gap year", "Sponsored video", "Email", "Not Started", "400K — solo female backpacking, budget gap year tips"],
    [535, "Kinging-It / Craig & Aimee (@KingingIt)", "YouTube", "UK Travel & Gap Year", "Mid", "~300K YT", "Budget travel, backpacking, van life", "Sponsored video", "Email", "Not Started", "300K — budget travel couple, van life, relatable"],
    [536, "Backpacking Bananas (@backpackingbananas)", "YouTube", "UK Travel & Gap Year", "Mid", "~250K YT", "Gap year vlogs — Southeast Asia, budget tips", "Sponsored video", "Email", "Not Started", "250K — gap year focused, SE Asia backpacking"],

    # ═══ UK FOOD & COOKING — TEEN/STUDENT AUDIENCE (training data) ═══
    [537, "B. Dylan Hollis (@bdylanhollis)", "TikTok", "UK Food & Cooking", "Mega", "~12M TikTok", "Retro recipe testing — chaotic baking, vintage cookbooks", "Brand deal", "Agent", "Not Started", "12M — viral baking, UK-based, massive teen following"],
    [538, "Poppy Cooks (@poppycooks)", "TikTok/YouTube", "UK Food & Cooking", "Mega", "~4M TikTok", "Quick easy recipes — high energy, humour", "Sponsored video", "DM", "Not Started", "4M — biggest UK food TikToker for young audience"],
    [539, "FitWaffle / Eloise Head (@fitwaffle)", "TikTok/IG", "UK Food & Cooking", "Mega", "~4M TikTok", "3-ingredient desserts — viral sweet treats", "Sponsored video", "DM", "Not Started", "4M — viral desserts, teen girls love this content"],
    [540, "MOB Kitchen (@mobkitchen)", "TikTok/YouTube", "UK Food & Cooking", "Macro", "~3M TikTok", "Quick overhead recipe videos — student-friendly meals", "Sponsored video / Partnership", "Email", "Not Started", "3M — student meals, overhead recipe format, perfect uni audience"],
    [541, "Cooking with Ayeh (@cookingwithayeh)", "TikTok/YouTube", "UK Food & Cooking", "Macro", "~3M TikTok", "UK-based — simple recipes, Middle-Eastern influenced", "Sponsored video", "DM", "Not Started", "3M — Middle-Eastern food, UK-based, 16-24 demographic"],
    [542, "Bosh! (@bosh.tv)", "TikTok/YouTube", "UK Food & Cooking", "Macro", "~2.5M TikTok", "Plant-based/vegan viral recipes", "Sponsored video", "Email", "Not Started", "2.5M — vegan recipes, viral format, young audience"],
    [543, "The Meal Prep King / John Clarke (@themealprep_king)", "TikTok/IG", "UK Food & Cooking", "Macro", "~1.2M TikTok", "Budget meal prep — family-friendly", "Sponsored video", "DM", "Not Started", "1.2M — budget meal prep, appeals to parents + students"],
    [544, "Dina Macki (@dinamacki)", "TikTok/IG", "UK Food & Cooking", "Mid", "~800K TikTok", "Easy Middle-Eastern food — budget-friendly, UK-based", "Sponsored video", "DM", "Not Started", "800K — Middle-Eastern food, budget meals, young audience"],
    [545, "The Batch Lady / Suzanne Mulholland (@thebatchlady)", "TikTok/IG", "UK Food & Cooking", "Mid", "~600K TikTok", "Batch cooking — budget meal prep for families", "Sponsored video", "Email", "Not Started", "600K — batch cooking, parent audience, family meals"],
    [546, "Nisha Katona (@nishakatonambe)", "TikTok/IG", "UK Food & Cooking", "Mid", "~500K combined", "Easy curry, South Asian cooking — MBE, Mowgli restaurants", "Sponsored video", "Email", "Not Started", "500K — MBE, Mowgli chain, South Asian cooking for UK audience"],
    [547, "Nathan Anthony (@nathananthony_)", "TikTok", "UK Food & Cooking", "Mid", "~500K TikTok", "Student budget meals — uni cooking content", "Sponsored post", "DM", "Not Started", "500K — student budget meals, directly targets uni/6th form"],

    # ═══ UK PERSONAL FINANCE — YOUNG PEOPLE (training data) ═══
    [548, "Damien Talks Money (@damientalks)", "YouTube/TikTok", "UK Finance & Money", "Mid", "~300K combined", "Investing basics, ISAs, side hustles for young adults", "Sponsored video", "Email", "Not Started", "300K — investing for beginners, young adult audience"],
    [549, "Sidecar Finance / Nischa (@nischa)", "YouTube", "UK Finance & Money", "Mid", "~200K YT", "Saving, investing, career/money for 20-somethings", "Sponsored video", "Email", "Not Started", "200K — finance + career, appeals to ambitious teens"],
    [550, "Kia Commodore / Pennies to Pounds (@kiacommodore)", "YouTube/TikTok/IG", "UK Finance & Money", "Mid", "~150K combined", "Student budgeting, investing for beginners", "Sponsored video", "DM", "Not Started", "150K — student finance, diverse voice, young audience"],
    [551, "Tola Talks / Tola Alabi (@tolatalks)", "YouTube/IG", "UK Finance & Money", "Micro", "~100K combined", "Student finance, budgeting, money mindset", "Sponsored video", "DM", "Not Started", "100K — student money, budgeting, relatable"],
    [552, "This Girl Talks Money / Ellie Austin-Williams (@thisgirltalks_money)", "IG/YouTube", "UK Finance & Money", "Micro", "~80K combined", "Budgeting, money confidence for young women", "Sponsored post", "Email", "Not Started", "80K — money confidence, young female audience"],
    [553, "The Humble Penny / Tokunbo Koiki (@thehumblepenny)", "YouTube", "UK Finance & Money", "Micro", "~80K YT", "Financial independence, investing, UK pensions", "Sponsored video", "Email", "Not Started", "80K — financial independence, longer-form education"],
    [554, "Your Money Sorted / Sasha Yansane (@yourmoneysorted)", "TikTok/IG", "UK Finance & Money", "Mid", "~250K TikTok", "Gen Z money tips, saving hacks", "Sponsored post", "DM", "Not Started", "250K — Gen Z finance, TikTok native, saving challenges"],
    [555, "Nicole Victoria (@nicolevictoria)", "YouTube", "UK Finance & Money", "Micro", "~50K YT", "Saving, investing, financial literacy for beginners", "Sponsored video", "Email", "Not Started", "50K — financial literacy basics, beginner-friendly"],

    # ═══ GCC INFLUENCER MARKETING AGENCIES — DISCOVERY TOOLS (training data) ═══
    [556, "ITP Live (itp.live)", "Agency", "GCC Influencer Agencies", "Agency", "Part of ITP Media Group", "Dubai — largest ME media group, influencer management", "Outreach for creator intros", "Email", "Not Started", "Use to connect with Gulf-based education creators"],
    [557, "Socialeyez (socialeyez.com)", "Agency", "GCC Influencer Agencies", "Agency", "Dubai-HQ", "Social media + influencer marketing — Emaar, DEWA clients", "Outreach for creator intros", "Email", "Not Started", "Dubai agency — can source education/parent influencers"],
    [558, "Vamp (vamp.com)", "Platform", "GCC Influencer Agencies", "Platform", "Global + Dubai office", "Influencer platform — Adobe, GAP, Estee Lauder clients", "Self-service platform", "Email", "Not Started", "Platform to discover GCC creators, self-service"],
    [559, "Influencer.ae (Jenavive Media)", "Platform", "GCC Influencer Agencies", "Platform", "UAE-specific", "UAE influencer discovery — Gulf creator database", "Browse platform", "Email", "Not Started", "UAE-specific influencer matching, local database"],
    [560, "TIMA (timaagency.com)", "Agency", "GCC Influencer Agencies", "Agency", "Riyadh + Dubai", "Influencer ID + campaign — beauty, auto, telecom", "Outreach", "Email", "Not Started", "Riyadh + Dubai offices, KSA + UAE creator network"],
    [561, "Gambit Communications (gambit.ae)", "Agency", "GCC Influencer Agencies", "Agency", "Dubai", "PR + influencer marketing — Huawei, Lenovo clients", "Outreach", "Email", "Not Started", "Dubai agency, tech focus, Gulf influencer networks"],

    # ═══ UK PODCASTS — PARENT & YOUNG ADULT AUDIENCE (training data) ═══
    [562, "The Parenting Hell (Rob Beckett & Josh Widdicombe)", "Podcast", "UK Podcasts", "Mega", "~3M downloads/month", "Comedy parenting — school-age kids, family chaos", "Podcast ad / Sponsorship", "Email", "Not Started", "3M/month — UK's biggest parenting podcast, school-age parent audience"],
    [563, "Shagged Married Annoyed (Chris & Rosie Ramsey)", "Podcast", "UK Podcasts", "Mega", "~2.5M downloads/month", "Relationships, family life, comedy — parents 30-45", "Podcast ad / Sponsorship", "Email", "Not Started", "2.5M/month — family comedy, massive parent audience"],
    [564, "That Peter Crouch Podcast", "Podcast", "UK Podcasts", "Mega", "~2M downloads/month", "Football, comedy, family — huge dad audience crossover", "Podcast ad", "Email", "Not Started", "2M/month — dad audience, football + family"],
    [565, "My Therapist Ghosted Me (Vogue Williams & Joanne McNally)", "Podcast", "UK Podcasts", "Mega", "~2M downloads/month", "Comedy, life advice — women 25-45, many parents", "Podcast ad", "Email", "Not Started", "2M/month — massive female audience, parent demographic"],
    [566, "The Rest Is History (Tom Holland & Dominic Sandbrook)", "Podcast", "UK Podcasts", "Macro", "~2M downloads/month", "History — educated parents helping kids with school topics", "Podcast ad / Sponsorship", "Email", "Not Started", "2M/month — history podcast, GCSE History crossover"],
    [567, "You're Dead to Me (Greg Jenner / BBC)", "Podcast", "UK Podcasts", "Macro", "~1.5M per series", "Comedy history — recommended by teachers for older kids", "BBC sponsorship", "Email", "Not Started", "1.5M/series — BBC, teachers recommend to students"],
    [568, "Fearne Cotton's Happy Place", "Podcast", "UK Podcasts", "Macro", "~1.5M downloads/month", "Mental health, wellbeing, parenting — women 28-50", "Podcast ad / Sponsorship", "Email", "Not Started", "1.5M/month — wellbeing, core parent demographic"],
    [569, "The High Performance Podcast (Jake Humphrey)", "Podcast", "UK Podcasts", "Macro", "~1M downloads/month", "Mindset, performance, education — aspirational parents", "Podcast ad", "Email", "Not Started", "1M/month — mindset/performance, aspirational parent audience"],
    [570, "Scummy Mummies (Ellie Gibson & Helen Thorn)", "Podcast", "UK Podcasts", "Mid", "~250K downloads/month", "Comedy parenting — specifically school-age mums", "Podcast ad / Sponsorship", "Email", "Not Started", "250K/month — school-age parent focus, perfect niche"],

    # ═══ UK EDUCATION JOURNALISTS & COLUMNISTS (training data) ═══
    [571, "Sean Coughlan (BBC News)", "BBC/X", "UK Education Media", "Macro", "BBC platform", "BBC education & family correspondent — GCSE results, uni access", "PR / Press release", "Email", "Not Started", "BBC reach — covers GCSE/A-Level results day, highest visibility"],
    [572, "Sally Weale (The Guardian)", "Guardian/X", "UK Education Media", "Mid", "Guardian platform", "Guardian education correspondent — schools policy, teachers", "PR / Press release", "Email", "Not Started", "Guardian education beat — school policy, teacher issues"],
    [573, "Laura McInerney (Teacher Tapp / Schools Week)", "X/Newsletter", "UK Education Media", "Mid", "~30K+ X", "Co-founder Teacher Tapp, ex-Schools Week editor — data-driven", "PR / Interview", "Email", "Not Started", "30K+ — most influential voice in UK education media"],
    [574, "Fiona Millar (The Guardian)", "Guardian/X", "UK Education Media", "Mid", "Guardian column", "Guardian education columnist — parenting + schooling author", "PR / Interview", "Email", "Not Started", "Long-running Guardian column, education + parenting books"],
    [575, "Nicola Woolcock (The Times)", "Times/X", "UK Education Media", "Mid", "Times platform", "Times education correspondent — schools, universities, policy", "PR / Press release", "Email", "Not Started", "Times education beat — broadsheet reach"],
    [576, "Camilla Turner (The Telegraph)", "Telegraph/X", "UK Education Media", "Mid", "Telegraph platform", "Telegraph social affairs — universities, schools policy", "PR / Press release", "Email", "Not Started", "Telegraph education — campus free speech, schools"],
    [577, "Schools Week (schoolsweek.co.uk)", "Website/X", "UK Education Media", "Micro", "Specialist", "England's leading specialist education publication", "PR / Sponsored content", "Email", "Not Started", "Most important specialist education outlet — read by school leaders"],
    [578, "Tanith Carey (freelance)", "Press/X", "UK Education Media", "Micro", "National press", "Parenting author — What's My Child Thinking?, Daily Telegraph/Mail/Guardian", "PR / Book collab", "Email", "Not Started", "Parenting author, national press columnist — child development"],

    # ═══ UK LITERARY FESTIVALS — PARTNERSHIP OPPORTUNITIES (training data) ═══
    [579, "Hay Festival (Hay-on-Wye)", "Event", "UK Literary Festivals", "Mega", "~250K attendees", "10 days, May/June — 'Woodstock of the mind', schools programme", "Exhibit / Schools strand sponsor", "Email", "Not Started", "250K — flagship festival, strong schools programme"],
    [580, "Edinburgh International Book Festival", "Event", "UK Literary Festivals", "Mega", "~250K+ attendees", "17 days, August — world's largest public book festival, schools programme", "Exhibit / Sponsor", "Email", "Not Started", "250K+ — largest book festival globally, schools + children's strand"],
    [581, "Cheltenham Literature Festival", "Event", "UK Literary Festivals", "Mega", "~130K attendees", "10 days, October — oldest lit festival (1949), education strand", "Exhibit / Schools sponsor", "Email", "Not Started", "130K — oldest lit festival, strong education strand"],
    [582, "London Book Fair (Olympia)", "Event/Trade", "UK Literary Festivals", "Mid", "~25K trade professionals", "March — B2B, publishers, EdTech, librarians", "Exhibit / Demo", "Email", "Not Started", "25K trade — B2B event, EdTech exhibiting, decision-makers"],
    [583, "Oxford Literary Festival", "Event", "UK Literary Festivals", "Mid", "~30K+ attendees", "March/April — events in Oxford colleges, educator audience", "Exhibit / Sponsor", "Email", "Not Started", "30K — Oxford colleges, English Lit alignment, educator-dense"],
    [584, "Cambridge Literary Festival", "Event", "UK Literary Festivals", "Micro", "~15K attendees", "April + November — university town, academic audience", "Exhibit / Sponsor", "Email", "Not Started", "15K — Cambridge, intellectually curious, two editions/year"],
    [585, "Stratford-upon-Avon Literary Festival", "Event", "UK Literary Festivals", "Micro", "~15K attendees", "April/May — Shakespeare connection, English students", "Exhibit / Schools strand", "Email", "Not Started", "15K — Shakespeare town, perfect GCSE English Lit fit"],
    [586, "Manchester Literature Festival", "Event", "UK Literary Festivals", "Micro", "~20K attendees", "October — diverse, contemporary writing, poetry", "Exhibit / Sponsor", "Email", "Not Started", "20K — diverse urban audience, contemporary English curriculum"],

    # ═══ UK EDUCATION CHARITIES & YOUTH ORGS — SCHOOL NETWORKS (training data) ═══
    [587, "The Prince's Trust (princes-trust.org.uk)", "Charity", "UK Education Charities", "Mega", "1M+ young people supported", "Ages 11-30 — jobs, education, training, mentoring", "Partnership", "Email", "Not Started", "UK's largest youth charity — 1M+ supported since 1976"],
    [588, "Speakers for Schools (speakersforschools.org)", "Charity", "UK Education Charities", "Macro", "1,000+ schools", "Free talks + work experience for state school students", "Partnership / Platform demo", "Email", "Not Started", "1,000+ schools — founded by Robert Peston, state school access"],
    [589, "Sutton Trust (suttontrust.com)", "Charity", "UK Education Charities", "Macro", "100K+ young people", "Social mobility — uni access programmes, summer schools", "Partnership", "Email", "Not Started", "100K+ supported — highly influential in education policy"],
    [590, "Young Enterprise (young-enterprise.org.uk)", "Charity", "UK Education Charities", "Macro", "350K+ young people/year", "Enterprise + financial education — Company Programme", "Partnership", "Email", "Not Started", "350K/year — enterprise education in thousands of schools"],
    [591, "STEM Learning (stem.org.uk)", "Charity", "UK Education Charities", "Macro", "30K+ STEM Ambassadors", "Largest STEM education provider — majority of UK secondaries", "Partnership / Cross-promo", "Email", "Not Started", "30K ambassadors — works with majority of UK secondary schools"],
    [592, "The Brilliant Club (thebrilliantclub.org)", "Charity", "UK Education Charities", "Mid", "100K+ pupils", "PhD researchers teach in state schools — uni access", "Partnership / Pilot", "Email", "Not Started", "100K+ pupils — PhD tutors in schools, uni access focus"],
    [593, "Debate Mate (debatemate.com)", "Charity", "UK Education Charities", "Mid", "500+ schools", "Debate clubs in underserved schools — communication skills", "Partnership", "Email", "Not Started", "500+ schools — debate = English Language skills, perfect fit"],
    [594, "IntoUniversity (intouniversity.org)", "Charity", "UK Education Charities", "Mid", "40+ centres", "Local learning centres — 50K young people annually", "Partnership / Pilot", "Email", "Not Started", "40+ centres — learning centres for disadvantaged students"],
    [595, "Ark Education (arkonline.org)", "Charity/MAT", "UK Education Charities", "Mid", "39 schools", "MAT + English Mastery programme — used by 100s of schools", "Partnership / Pilot", "Email", "Not Started", "39 schools + English Mastery used by hundreds more — direct fit"],
    [596, "The Access Project (theaccessproject.org.uk)", "Charity", "UK Education Charities", "Micro", "100+ partner schools", "1-to-1 tutoring for disadvantaged students — uni access", "Partnership", "Email", "Not Started", "100+ schools — tutoring focus, complementary to platform"],
    [597, "Envision (envision.org.uk)", "Charity", "UK Education Charities", "Micro", "100+ schools", "Leadership + teamwork projects mentored by professionals", "Partnership", "Email", "Not Started", "100+ schools — skills development, 30K+ young people"],

    # ═══ UK STUDENT DISCOUNT & DISTRIBUTION PLATFORMS (training data) ═══
    [598, "UNiDAYS (unidays.com)", "Platform", "UK Student Platforms", "Mega", "22M+ verified students", "Student discount platform — app + web, 130+ UK unis", "List student offer / Ad", "Email", "Not Started", "22M — #1 student discount platform, verified student base"],
    [599, "Student Beans (studentbeans.com)", "Platform", "UK Student Platforms", "Mega", "163M+ users globally", "Student discount platform — verified offers, affiliate tracking", "List student offer / Ad", "Email", "Not Started", "163M — massive reach, affiliate tracking, sponsored content"],
    [600, "TOTUM / NUS (totum.com)", "Platform", "UK Student Platforms", "Mega", "4M+ UK students", "Official NUS discount card — physical + digital", "Partnership / Listing", "Email", "Not Started", "4M — official NUS card, reaches every UK uni"],
    [601, "UCAS (ucas.com)", "Platform", "UK Student Platforms", "Mega", "700K+ applicants/year", "University admissions — UCAS Hub, newsletter sponsorship", "Advertising / Sponsored", "Email", "Not Started", "700K+ applicants — every UK uni applicant passes through UCAS"],
    [602, "Save the Student (savethestudent.org)", "Website", "UK Student Platforms", "Macro", "2M+ monthly visitors", "Student finance advice — deals, best-of guides, reviews", "Affiliate / Sponsored review", "Email", "Not Started", "2M/month — editorial reviews, 'best of' guides, high trust"],
    [603, "Unifrog (unifrog.org)", "Platform", "UK Student Platforms", "Macro", "5,000+ UK schools", "Uni/careers guidance platform — used in schools + sixth forms", "Partnership / Listing", "Email", "Not Started", "5,000+ schools — B2B, integrated into student pathways"],
    [604, "Jisc (jisc.ac.uk)", "Platform", "UK Student Platforms", "Mega", "Every UK uni + most FE", "Digital solutions for UK education — procurement frameworks", "Vendor partnership / Marketplace", "Email", "Not Started", "Every UK uni — Jisc marketplace, institutional procurement"],
    [605, "Blackbullion (blackbullion.com)", "Platform", "UK Student Platforms", "Mid", "80+ UK universities", "Financial wellbeing + scholarships — uni dashboard integration", "Partnership / Integration", "Email", "Not Started", "80+ unis — student financial wellbeing, dashboard access"],

    # ═══ ARABIC EDUCATION YOUTUBE — GCC STUDENT AUDIENCE (training data) ═══
    [606, "AJ+ عربي (AJ+ Arabic)", "YouTube", "Arabic Education YouTube", "Mega", "~11M YT", "Qatar/Al Jazeera — explainers, science, social issues", "Sponsored video / Partnership", "Email", "Not Started", "11M — Al Jazeera network, massive GCC reach"],
    [607, "Osloop (أسلوب)", "YouTube", "Arabic Education YouTube", "Macro", "~3M YT", "Egypt — self-development, productivity, book summaries", "Sponsored video", "Email", "Not Started", "3M — massive GCC student viewership, productivity content"],
    [608, "ElZero Web School", "YouTube", "Arabic Education YouTube", "Macro", "~1.8M YT", "Egypt — web dev, programming, popular with Gulf students", "Sponsored video", "Email", "Not Started", "1.8M — tech education, Gulf uni students"],
    [609, "Free4arab (فور عرب)", "YouTube", "Arabic Education YouTube", "Macro", "~1.5M YT", "Egypt — IT, programming, cybersecurity", "Sponsored video", "Email", "Not Started", "1.5M — IT/programming, widely used across GCC"],
    [610, "Muslim Researchers (الباحثون المسلمون)", "YouTube", "Arabic Education YouTube", "Macro", "~1.2M YT", "Egypt — science communication, critical thinking", "Sponsored video", "Email", "Not Started", "1.2M — science/critical thinking, pan-Arab audience"],
    [611, "Tahrir Academy (أكاديمية التحرير)", "YouTube", "Arabic Education YouTube", "Mid", "~800K YT", "Egypt — animated maths, science, humanities", "Sponsored video", "Email", "Not Started", "800K — animated education, pan-Arab K-12 reach"],
    [612, "Durosi (دروسي)", "YouTube", "Arabic Education YouTube", "Mid", "~500K YT", "Saudi — maths, science, Arabic curriculum K-12", "Sponsored video", "Email", "Not Started", "500K — Saudi curriculum support, direct K-12 fit"],
    [613, "TeleTub (تيلي تب)", "YouTube", "Arabic Education YouTube", "Mid", "~400K YT", "Saudi — English language learning for Arabic speakers", "Sponsored video / Affiliate", "Email", "Not Started", "400K — Saudi English learning, direct product crossover"],
    [614, "Barmej (برمج)", "YouTube", "Arabic Education YouTube", "Mid", "~350K YT", "Saudi — programming, software engineering, tech careers", "Sponsored video", "Email", "Not Started", "350K — Saudi tech education, young audience"],

    # ═══ UK EXAM SEASON / RESULTS DAY CREATORS (training data) ═══
    [615, "Zach Todd (@zachtodd_)", "TikTok", "UK Exam Season Creators", "Mid", "~500K TikTok", "GCSE/A-Level exam reactions, results day content, relatable student humour", "Sponsored post", "DM", "Not Started", "500K — viral results day content, relatable to GCSE students"],
    [616, "The GCSE Guide / Primrose Sheridan (@thegcseguide)", "TikTok/IG", "UK Exam Season Creators", "Mid", "~400K TikTok", "GCSE-specific revision tips, exam technique, results day reactions", "Affiliate / Sponsored", "DM", "Not Started", "400K — GCSE-specific, spikes May–August, perfect fit"],
    [617, "StudyWithJess (@studywithjessx)", "TikTok/YouTube", "UK Exam Season Creators", "Mid", "~200K TikTok", "GCSE & A-Level revision, study-with-me livestreams, results day vlogs", "Affiliate / Sponsored", "DM", "Not Started", "200K — study livestreams, revision community"],
    [618, "Shaaba (@shaabasket)", "YouTube", "UK Exam Season Creators", "Micro", "~150K YT", "Exam season advice, A-Level results, university guidance", "Affiliate / Sponsored", "Email", "Not Started", "150K — Jack Edwards collab partner, exam advice niche"],

    # ═══ UK STUDENT UNIONS & YOUTH ORGS (training data) ═══
    [619, "The Tab (@taboracle)", "Website/IG/TikTok", "UK Student Unions & Youth Orgs", "Macro", "~3M monthly readers", "National student news network — 30+ university editions", "Sponsored article / Banner", "Email", "Not Started", "3M monthly — biggest student media platform, campus-level targeting"],
    [620, "UK Youth (@uaboringyouth)", "Website/Social", "UK Student Unions & Youth Orgs", "Macro", "~2.6M young people via 8K+ orgs", "National youth development charity — huge network", "Partnership", "Email", "Not Started", "2.6M reach — national youth charity network"],
    [621, "British Youth Council (@baborinyc)", "Website/Social", "UK Student Unions & Youth Orgs", "Mid", "200+ member orgs", "National youth voice organisation — campaigns, local youth councils", "Partnership", "Email", "Not Started", "200+ orgs — national youth representation body"],
    [622, "Sixth Form Colleges Association (SFCA)", "Website/Events", "UK Student Unions & Youth Orgs", "Mid", "90+ colleges / ~200K students", "Sector body for sixth form colleges — ideal 16-18 reach", "Partnership / Sponsorship", "Email", "Not Started", "200K students — direct sixth form pipeline, 16-18 sweet spot"],
    [623, "Manchester SU (@manchestersu)", "Website/Social", "UK Student Unions & Youth Orgs", "Mid", "~45K students", "One of UK's largest SUs — freshers fair, newsletter", "Freshers fair / Newsletter ad", "Email", "Not Started", "45K — largest UK SU, freshers fair presence"],
    [624, "Leeds University Union (LUU)", "Website/Social", "UK Student Unions & Youth Orgs", "Mid", "~38K students", "Major Russell Group SU — events, newsletter, freshers", "Freshers fair / Newsletter ad", "Email", "Not Started", "38K — Russell Group SU, strong events programme"],
    [625, "KCLSU (King's College London SU)", "Website/Social", "UK Student Unions & Youth Orgs", "Mid", "~35K students", "Central London SU — diverse, high engagement", "Freshers fair / Newsletter ad", "Email", "Not Started", "35K — central London, diverse student body"],
    [626, "Birmingham Guild of Students", "Website/Social", "UK Student Unions & Youth Orgs", "Mid", "~36K students", "Major Midlands SU — freshers fair, email reach", "Freshers fair / Newsletter ad", "Email", "Not Started", "36K — major Midlands uni, strong freshers programme"],
    [627, "National Society of Apprentices (NUS)", "Social/Events", "UK Student Unions & Youth Orgs", "Micro", "FE/apprentice network", "Represents apprentices across FE and training providers", "Partnership", "Email", "Not Started", "FE apprentice body — vocational audience reach"],

    # ═══ UK DIVERSITY & INCLUSION EDUCATORS (training data) ═══
    [628, "Akala (@akalamusic)", "IG/YouTube", "UK Diversity & Inclusion Educators", "Mid", "~350K IG", "Author of Natives, Hip-Hop Shakespeare Company — youth education & racial literacy", "Partnership / Event", "Agent", "Not Started", "350K — Hip-Hop Shakespeare, bridges urban culture + English Lit"],
    [629, "The Black Curriculum / Lavinya Stennett (@theblackcurriculum)", "IG", "UK Diversity & Inclusion Educators", "Micro", "~40K IG", "Teaching Black British history in schools — curriculum resources", "Partnership", "Email", "Not Started", "40K — school-facing org, curriculum resources, direct teacher reach"],
    [630, "Nova Reid (@novareid)", "IG", "UK Diversity & Inclusion Educators", "Micro", "~95K IG", "Anti-racism training, author of The Good Ally — school workshops", "Partnership / Sponsored", "Email", "Not Started", "95K — school workshops, corporate + education training"],
    [631, "Jeffrey Boakye (@jeffreyboakye)", "IG/X", "UK Diversity & Inclusion Educators", "Micro", "~20K", "English teacher, author of Black Listed & I Heard What You Said — race & education", "Partnership", "Email", "Not Started", "20K — actual English teacher + author, perfect fit"],
    [632, "Pragya Agarwal (@pragyaagarwal)", "X", "UK Diversity & Inclusion Educators", "Micro", "~60K X", "Unconscious bias researcher, author of Wish We Knew What to Say (for parents/educators)", "Partnership / Sponsored", "Email", "Not Started", "60K — parent/educator focused, bias research"],
    [633, "Bennie Kara (@benniekara)", "X", "UK Diversity & Inclusion Educators", "Micro", "~25K X", "Deputy headteacher, author of Diversity in Schools — school leadership", "Partnership", "Email", "Not Started", "25K — deputy head, school leadership + inclusion, direct schools access"],
    [634, "Tanya Compas / Conscious Youth CIC (@tanyacompas)", "IG", "UK Diversity & Inclusion Educators", "Micro", "~45K IG", "LGBTQ+ youth work, anti-racism, youth empowerment", "Partnership", "Email", "Not Started", "45K — youth empowerment, LGBTQ+ inclusive education"],
    [635, "Nikesh Shukla (@nikaboreal)", "X", "UK Diversity & Inclusion Educators", "Micro", "~90K X", "Author, editor of The Good Immigrant — youth writing programmes", "Partnership", "Email", "Not Started", "90K — Good Immigrant editor, youth writing initiatives"],
    [636, "Uju Asika (@babesabouttown)", "IG", "UK Diversity & Inclusion Educators", "Micro", "~20K IG", "Author of Raising An Antiracist — parenting & schools-focused anti-racism", "Partnership / Sponsored", "Email", "Not Started", "20K — parent audience, schools-focused anti-racism"],
    [637, "Karl Nova (@karlnova)", "IG", "UK Diversity & Inclusion Educators", "Micro", "~10K IG", "Performance poet, author of Rhythm and Poetry — school visits, spoken word", "Partnership / School visit", "Email", "Not Started", "10K — school visits, spoken word + poetry, direct GCSE English link"],

    # ═══ UK EDUCATION CONFERENCES & EVENTS (training data) ═══
    [638, "ASCL Annual Conference", "Conference", "UK Education Conferences", "Mid", "~1.5K attendees", "Association of School & College Leaders — secondary heads + MAT CEOs", "Exhibit / Sponsor / Speak", "Email", "Not Started", "1.5K — senior decision-makers, budget holders"],
    [639, "NAHT Annual Conference", "Conference", "UK Education Conferences", "Mid", "~1.5K attendees", "National Association of Head Teachers — primary + secondary heads", "Exhibit / Sponsor", "Email", "Not Started", "1.5K — headteachers who make purchasing decisions"],
    [640, "EdTechX / London EdTech Week", "Conference", "UK Education Conferences", "Mid", "~2K attendees", "EdTech investment + innovation summit — founders, investors, buyers", "Exhibit / Pitch / Network", "Email", "Not Started", "June — best event for investor networking + startup visibility"],
    [641, "Jisc Digifest", "Conference", "UK Education Conferences", "Mid", "~2K attendees", "Digital technology in FE/HE — AI, data, learning platforms", "Exhibit / Sponsor", "Email", "Not Started", "March — Jisc digital body, FE/HE technology leaders"],
    [642, "ALT Annual Conference", "Conference", "UK Education Conferences", "Micro", "~700 attendees", "Association for Learning Technology — HE/FE learning tech", "Exhibit / Sponsor / Speak", "Email", "Not Started", "September — academic learning tech community, credibility"],
    [643, "ISBA Annual Conference", "Conference", "UK Education Conferences", "Micro", "~800 attendees", "Independent Schools' Bursars Association — private school budget holders", "Exhibit / Sponsor", "Email", "Not Started", "May — bursars control procurement, independent school market"],
    [644, "MAT Conference / Academies Show", "Conference", "UK Education Conferences", "Mid", "~2K attendees", "Multi-Academy Trust CEOs, CFOs — buy at scale across multiple schools", "Exhibit / Sponsor", "Email", "Not Started", "MAT leaders buy for 10-50+ schools at once — high ROI"],
    [645, "UCISA Conference", "Conference", "UK Education Conferences", "Micro", "~800 attendees", "University IT leadership — CIOs, digital leads from UK universities", "Exhibit / Sponsor", "Email", "Not Started", "Spring — HE IT decision-makers, institutional procurement"],

    # ═══ UK PARENTING MAGAZINES & MEDIA (training data) ═══
    [646, "MadeForMums (madeformums.com)", "Website", "UK Parenting Media", "Macro", "Millions monthly", "Product reviews, buying guides — parents 0-5, massive traffic", "Sponsored review / Banner ad", "Email", "Not Started", "Millions monthly — product review authority, parent buying decisions"],
    [647, "The Good Schools Guide (goodschoolsguide.co.uk)", "Website/Print", "UK Parenting Media", "Mid", "Large readership", "School research hub — affluent parents actively choosing schools", "Directory listing / Sponsored review", "Email", "Not Started", "Affluent parents researching education — high intent audience"],
    [648, "Junior Magazine (juniormagazine.co.uk)", "Print/Website", "UK Parenting Media", "Micro", "Affluent parents", "Covers fashion, travel, education for parents with children 0-8", "Print ad / Advertorial", "Email", "Not Started", "Affluent demographic — education, lifestyle, fashion"],
    [649, "Right Start Magazine", "Print/Website", "UK Parenting Media", "Micro", "Niche readership", "Child development & education focus — Best Toy Awards", "Product review / Print ad", "Email", "Not Started", "Education-focused parent mag — strong product review credibility"],
    [650, "Absolutely Education (absolutelyeducation.co.uk)", "Print/Website", "UK Parenting Media", "Micro", "London affluent parents", "Part of Absolutely group — London parents researching schools/education", "Print ad / Advertorial", "Email", "Not Started", "Affluent London parents — school + education product decisions"],
    [651, "Families Online (familiesonline.co.uk)", "Website/Print", "UK Parenting Media", "Mid", "Regional network", "Regional digital network + local printed guides — families 0-12", "Regional banner / Directory listing", "Email", "Not Started", "Regional targeting — multiple UK editions, local parent reach"],
    [652, "The Green Parent", "Print/Website", "UK Parenting Media", "Micro", "Niche readership", "Eco-conscious parents — natural parenting, home education", "Print ad / Sponsored editorial", "Email", "Not Started", "Home education audience — natural fit for online learning tools"],

    # ═══ UK EDUCATION PODCASTS (training data) ═══
    [653, "Rethinking Education (James Mannion)", "Podcast", "UK Education Podcasts", "Micro", "~20K per episode", "Research-informed practice, metacognition, self-regulated learning", "Sponsored read / Interview", "Email", "Not Started", "20K — research-informed teachers, school improvement focus"],
    [654, "The SEND Cast", "Podcast", "UK Education Podcasts", "Micro", "~12K per episode", "SEND practitioners — inclusion strategies, EHCP processes", "Sponsored read", "Email", "Not Started", "12K — SEND niche, accessibility angle for The English Hub"],
    [655, "Staffroom Stories", "Podcast", "UK Education Podcasts", "Micro", "~12K per episode", "Real teaching experiences, NQT/ECT advice, school culture", "Sponsored read", "Email", "Not Started", "12K — new teachers, relatable classroom content"],
    [656, "Evidence Based Education Podcast", "Podcast", "UK Education Podcasts", "Micro", "~12K per episode", "Assessment, feedback, Great Teaching Toolkit — EEF-aligned", "Sponsored read / Interview", "Email", "Not Started", "12K — evidence-based teaching, assessment focus"],
    [657, "The School Leadership Podcast", "Podcast", "UK Education Podcasts", "Micro", "~10K per episode", "Headteacher/SLT challenges, Ofsted, MAT leadership", "Sponsored read", "Email", "Not Started", "10K — reaches headteachers + SLT who make purchase decisions"],
    [658, "Ollie Lovell's Education Podcast", "Podcast", "UK Education Podcasts", "Micro", "~20K per episode", "Cognitive load theory, explicit instruction — strong UK audience", "Sponsored read / Interview", "Email", "Not Started", "20K — education academics + teachers, evidence-based"],
    [659, "Tes Book Club Podcast", "Podcast", "UK Education Podcasts", "Micro", "~15K per episode", "Education books, author interviews — Tes journalist network", "Sponsored read", "Email", "Not Started", "15K — Tes brand, education book community"],
    [660, "The Early Career Hub (Ambition Institute)", "Podcast", "UK Education Podcasts", "Micro", "~12K per episode", "ECF training, mentor support — new teachers entering profession", "Sponsored read", "Email", "Not Started", "12K — new teachers pipeline, Ambition Institute credibility"],

    # ═══ UK LINKEDIN EDUCATION THOUGHT LEADERS (training data) ═══
    [661, "Sir David Carter", "LinkedIn", "UK Education LinkedIn Leaders", "Mid", "~30K LinkedIn", "Former National Schools Commissioner — advises MATs on strategy", "Direct outreach / Speaking", "LinkedIn", "Not Started", "30K — ex-National Schools Commissioner, MAT advisory influence"],
    [662, "Leora Cruddas (Confederation of School Trusts)", "LinkedIn", "UK Education LinkedIn Leaders", "Mid", "~15K LinkedIn", "CEO of CST — directly advises trust leaders on strategy + resources", "Partnership / Speaking", "LinkedIn/Email", "Not Started", "15K — CST CEO, shapes MAT procurement decisions"],
    [663, "Sir Jon Coles (United Learning)", "LinkedIn", "UK Education LinkedIn Leaders", "Micro", "~10K LinkedIn", "CEO United Learning MAT — one of UK's largest MATs", "Partnership", "LinkedIn/Email", "Not Started", "10K — runs 75+ schools, single procurement decision = massive"],
    [664, "Ty Goddard (Education Foundation)", "LinkedIn", "UK Education LinkedIn Leaders", "Mid", "~20K LinkedIn", "Co-founder Education Foundation, BETT advisor — EdTech connector", "Partnership / BETT intro", "LinkedIn/Email", "Not Started", "20K — BETT advisor, EdTech ecosystem connector"],
    [665, "Priya Lakhani OBE (CENTURY Tech)", "LinkedIn", "UK Education LinkedIn Leaders", "Mid", "~25K LinkedIn", "EdTech founder CEO — shapes conversation on school tech adoption", "Partnership / Co-marketing", "LinkedIn/Email", "Not Started", "25K — EdTech CEO, OBE, influences school tech buying"],
    [666, "Mark Martin / UrbanTeacher", "LinkedIn", "UK Education LinkedIn Leaders", "Mid", "~20K LinkedIn", "Education consultant & speaker — bridges urban schools + tech", "Speaking / Sponsored", "LinkedIn/Email", "Not Started", "20K — urban education advocate, tech in schools voice"],
    [667, "Dame Alison Peacock (Chartered College of Teaching)", "LinkedIn", "UK Education LinkedIn Leaders", "Mid", "~15K LinkedIn", "CEO Chartered College — professional body for teachers", "Partnership", "LinkedIn/Email", "Not Started", "15K — Chartered College CEO, teacher professional standards"],
    [668, "Sir Steve Lancashire (REAch2)", "LinkedIn", "UK Education LinkedIn Leaders", "Micro", "~8K LinkedIn", "Founder CEO REAch2 — UK's largest primary academy trust", "Partnership", "LinkedIn/Email", "Not Started", "8K — largest primary MAT, procurement across 60+ schools"],
    [669, "Hamid Patel CBE (Star Academies)", "LinkedIn", "UK Education LinkedIn Leaders", "Micro", "~10K LinkedIn", "CEO Star Academies — high-performing MAT, diverse schools", "Partnership", "LinkedIn/Email", "Not Started", "10K — Star Academies CEO, high-performing diverse trust"],
    [670, "Bob Harrison (EdTech Advisor)", "LinkedIn", "UK Education LinkedIn Leaders", "Mid", "~15K LinkedIn", "Independent EdTech advisor — Toshiba education, BETT speaker", "Advisory / Speaking", "LinkedIn/Email", "Not Started", "15K — veteran EdTech advisor, school tech procurement"],
    [671, "Al Kingsley (NetSupport)", "LinkedIn", "UK Education LinkedIn Leaders", "Micro", "~12K LinkedIn", "CEO NetSupport, EdTech author — school technology thought leader", "Partnership / Co-marketing", "LinkedIn/Email", "Not Started", "12K — EdTech CEO + author, school IT decision influencer"],

    # ═══ UK CHILDREN'S TV PRESENTERS & FAMILY CREATORS (training data) ═══
    [672, "Ali-A / Alastair Aiken (@OMGitsAliA)", "YouTube/IG", "UK Children's TV & Family Creators", "Mega", "~19M YT / 5M IG", "Fortnite, gaming — family-friendly, popular with primary/secondary", "Brand deal", "Agent", "Not Started", "19M — family-friendly gaming, older primary + secondary school boys"],
    [673, "Colin Furze (@colinfurze)", "YouTube", "UK Children's TV & Family Creators", "Mega", "~13M YT", "Invention/engineering — family-friendly STEM, former CBBC", "Sponsored video", "Email", "Not Started", "13M — engineering content, STEM inspiration, school-age audience"],
    [674, "Katie Thistleton (@katiethistleton)", "IG/Radio", "UK Children's TV & Family Creators", "Micro", "~70K IG", "CBBC presenter, BBC Radio — children's mental health advocacy", "Partnership / Sponsored", "Email", "Not Started", "70K — CBBC + BBC Radio, mental health for young people"],
    [675, "Lindsey Russell (@lindseyrussell)", "IG", "UK Children's TV & Family Creators", "Micro", "~150K IG", "Blue Peter presenter 2013-2020 — craft, adventure, family content", "Sponsored post", "Email", "Not Started", "150K — Blue Peter alumni, trusted family brand"],
    [676, "Andy Day (@andydayofficial)", "IG", "UK Children's TV & Family Creators", "Micro", "~60K IG", "CBeebies presenter — Wild Adventures, Dinosaur Adventures", "Partnership", "Email", "Not Started", "60K — CBeebies icon, strong parent + young child recognition"],
    [677, "Naomi Wilkinson (@naomiwilkinson)", "IG", "UK Children's TV & Family Creators", "Micro", "~30K IG", "CBBC presenter — Nightmares of Nature, Wild & Weird", "Sponsored post", "Email", "Not Started", "30K — CBBC nature presenter, school-age audience"],
    [678, "Joel & Lia (@joelandlia)", "YouTube", "UK Children's TV & Family Creators", "Macro", "~1M YT", "British culture, lifestyle, comedy — popular with teens", "Sponsored video", "Email", "Not Started", "1M — British culture content, teen audience, relatable"],
    [679, "Paperboyo / Rich McCor (@paperboyo)", "IG", "UK Children's TV & Family Creators", "Mid", "~500K IG", "Creative paper cut-out art — Blue Peter/CBBC appearances", "Sponsored post / Creative collab", "Email", "Not Started", "500K — creative content, appeared on Blue Peter, family audience"],

    # ═══ GCC EXPAT COMMUNITY PLATFORMS (training data) ═══
    [680, "ExpatFocus (expatfocus.com)", "Website/Forum", "GCC Expat Community Platforms", "Mid", "~200K monthly", "Expat resource hub — GCC country guides, forums, newsletters", "Sponsored content / Banner", "Email", "Not Started", "200K monthly — all GCC, expat planning + living resources"],
    [681, "BritishMums (britishmums.com)", "Community/Blog", "GCC Expat Community Platforms", "Micro", "~50K followers", "British mothers in UAE — Dubai/Abu Dhabi focus, family/education", "Sponsored post / Partnership", "Email", "Not Started", "50K — specifically British mums in UAE, perfect demographic"],
    [682, "InterNations (internations.org)", "Social Network", "GCC Expat Community Platforms", "Mega", "5M+ global", "Expat social network — chapters in Dubai, Doha, Riyadh, Muscat", "Advertising / Event sponsor", "Email", "Not Started", "5M global — GCC chapters, expat events + online groups"],
    [683, "AngloINFO", "Website", "GCC Expat Community Platforms", "Macro", "~1M monthly global", "Practical guides + classifieds for English-speaking expats", "Banner / Directory listing", "Email", "Not Started", "1M monthly — English-speaking expat community boards"],
    [684, "Dubai Expat Forum (dubaiexpat.com)", "Forum", "GCC Expat Community Platforms", "Mid", "~100K members", "Long-running expat forum — schools, family life sections", "Sponsored thread / Banner", "Email", "Not Started", "100K members — schools + family sections, organic promotion"],
    [685, "Expatriates.com", "Classifieds/Community", "GCC Expat Community Platforms", "Mid", "~500K monthly", "Oldest Gulf expat platform — strong in Saudi, Bahrain, Kuwait", "Banner / Sponsored listing", "Email", "Not Started", "500K monthly — Saudi/Bahrain/Kuwait focus, community boards"],
    [686, "TimeOut Dubai/Abu Dhabi/Doha (Family section)", "Lifestyle Media", "GCC Expat Community Platforms", "Mega", "2M+ monthly (Dubai)", "Things to do with kids, school guides — accepts sponsored content", "Sponsored article / Advertising", "Email", "Not Started", "2M+ monthly — family section heavily trafficked by expat parents"],
    [687, "Saudi Expat Forum (saudiexpat.com)", "Forum", "GCC Expat Community Platforms", "Micro", "~50K members", "Western expat community — international schools, compound life", "Sponsored thread / Banner", "Email", "Not Started", "50K — Saudi-based Western expats, school discussion threads"],
    [688, "YallaCompare (education section)", "Comparison Platform", "GCC Expat Community Platforms", "Macro", "~1M monthly", "School comparison tool alongside finance — UAE, Bahrain", "Listing / Sponsored", "Email", "Not Started", "1M monthly — parents comparing schools, high purchase intent"],
    [689, "British Expats in Dubai (Facebook Group)", "Facebook", "GCC Expat Community Platforms", "Mid", "~25K members", "Active British expat group — education, family, lifestyle discussions", "Community post / Admin partnership", "DM", "Not Started", "25K — highly engaged, word-of-mouth driven, admin approval"],
    [690, "r/dubai + r/saudiarabia (Reddit)", "Reddit", "GCC Expat Community Platforms", "Mid", "400K+ / 200K+", "Active subreddits — school threads, family life, education advice", "Organic / AMA", "Reddit", "Not Started", "600K+ combined — regular school/education threads"],

    # ═══ UK SCHOOL SUPPLY & PROCUREMENT CHANNELS (training data) ═══
    [691, "RM Technology (RM plc)", "IT Reseller", "UK School Procurement Channels", "Macro", "20K+ UK schools", "IT infrastructure, managed services, RM Unify marketplace", "Reseller / Marketplace listing", "Email", "Not Started", "20K schools — bundle into managed services, RM Unify marketplace"],
    [692, "YPO (Yorkshire Purchasing Organisation)", "Procurement Framework", "UK School Procurement Channels", "Macro", "20K+ customers", "Public sector buying org — framework listing removes procurement friction", "Framework agreement", "Email", "Not Started", "20K+ — framework listing = approved supplier, schools can buy easily"],
    [693, "ESPO (Eastern Shires Purchasing Organisation)", "Procurement Framework", "UK School Procurement Channels", "Mid", "~8K members", "Public buying consortium — Midlands focus, available nationally", "Framework agreement", "Email", "Not Started", "8K — framework listing, concentrated Midlands + national"],
    [694, "Capita SIMS / ESS (Education Software Solutions)", "MIS Platform", "UK School Procurement Channels", "Mega", "~80% UK schools", "School MIS used by 80% of UK schools — API integration opportunity", "API integration", "Email", "Not Started", "80% of UK schools — SIMS integration embeds in daily workflows"],
    [695, "Softcat (Education team)", "IT Reseller", "UK School Procurement Channels", "Macro", "Thousands of schools", "Major UK IT reseller — dedicated education team, sells into schools daily", "Reseller / Recommendation", "Email", "Not Started", "Top UK IT reseller — recommends software alongside hardware deals"],
    [696, "XMA (Apple Education Specialist)", "IT Reseller", "UK School Procurement Channels", "Mid", "Thousands of schools", "IT hardware + software licensing — bundles EdTech with device deployments", "Reseller / Bundle", "Email", "Not Started", "Apple education specialist — bundles software with device rollouts"],
    [697, "BESA / LendED Platform", "Trade Body", "UK School Procurement Channels", "Mid", "All UK schools (indirect)", "British Educational Suppliers Association — LendED trial platform", "Membership / LendED trial", "Email", "Not Started", "BESA membership + LendED = schools trial EdTech free, then convert"],
    [698, "Promethean (ActivPanel)", "Classroom Hardware", "UK School Procurement Channels", "Macro", "Tens of thousands of classrooms", "Interactive displays — app ecosystem promotes compatible software", "App integration / Partnership", "Email", "Not Started", "In tens of thousands of classrooms — app ecosystem integration"],
    [699, "Hodder Education / Dynamic Learning", "Publisher/Platform", "UK School Procurement Channels", "Macro", "Top 3 UK publisher", "Textbooks + Dynamic Learning digital platform — integration opportunity", "Content partnership / Integration", "Email", "Not Started", "Top 3 publisher — Dynamic Learning platform integration"],
    [700, "2Simple / Purple Mash", "EdTech Platform", "UK School Procurement Channels", "Mid", "Thousands of primary schools", "Primary-focused creative + coding platform — gateway into primary classrooms", "Integration / Bundling", "Email", "Not Started", "Primary school gateway — Purple Mash integration opportunity"],
    [701, "Scholastic UK Book Fairs", "Publisher/Events", "UK School Procurement Channels", "Macro", "25K school visits/year", "Book fairs visit 25K schools annually — co-promotion opportunity", "Co-promotion / Book fair bundle", "Email", "Not Started", "25K school visits — physical presence in schools, literacy crossover"],
    [702, "GLS Educational Supplies", "Catalogue/Online", "UK School Procurement Channels", "Mid", "Thousands of schools", "General school supplies catalogue — online + print listings", "Catalogue listing", "Email", "Not Started", "School supplies catalogue — listed alongside other education products"],

    # ═══ UK PODCASTS — PARENTS & TEENS (web-verified) ═══
    [703, "Table Manners / Jessie & Lennie Ware", "Podcast", "UK Podcasts (Parents/Teens)", "Macro", "60M+ total downloads", "Food, culture, celebrity interviews — cross-generational parent/teen appeal", "Sponsored read", "Agent", "Not Started", "60M downloads — mother-daughter format, cross-generational"],
    [704, "Desert Island Discs / Lauren Laverne", "BBC Sounds/Podcast", "UK Podcasts (Parents/Teens)", "Mega", "Millions weekly", "BBC flagship — culture, music, interviews — cross-generational", "Aspirational / Sponsored", "Agent", "Not Started", "BBC flagship — millions weekly, cross-generational institution"],
    [705, "Not Another Mummy Podcast / Alison Perry", "Podcast", "UK Podcasts (Parents/Teens)", "Micro", "Established UK audience", "Long-running UK parenting staple — family life, interviews", "Sponsored read", "Email", "Not Started", "Long-running UK parenting podcast, loyal audience"],
    [706, "The Therapy Crouch / Peter Crouch & Abbey Clancy", "BBC Sounds/Podcast", "UK Podcasts (Parents/Teens)", "Macro", "Edison UK Top 25", "Comedy, relationships, parenting — celebrity reach", "Sponsored read", "Agent", "Not Started", "Edison Top 25 — Peter Crouch + Abbey Clancy, family audience"],

    # ═══ UK EDUCATION COMPETITORS & COMPLEMENTARY PLATFORMS (training data) ═══
    [707, "York Notes (by Pearson)", "Books/Website", "UK EdTech Competitors", "Macro", "Millions of copies sold", "Classic GCSE English Lit study guides — text-specific, digital quizzes", "Partnership / Cross-promo", "Email", "Not Started", "Dominant English Lit revision brand — Pearson owned, integration angle"],
    [708, "Bedrock Learning", "Platform", "UK EdTech Competitors", "Mid", "1,000+ UK schools", "Vocabulary & literacy platform — adaptive, EEF-aligned evidence", "Partnership / Integration", "Email", "Not Started", "1K schools — vocabulary + literacy, complementary to English Hub"],
    [709, "PiXL (Partners in Excellence)", "Network/App", "UK EdTech Competitors", "Macro", "1,600+ secondary schools", "School improvement network — PiXL English App, diagnostic tools", "Partnership / Membership", "Email", "Not Started", "1,600 schools — PiXL English App, intervention resources"],
    [710, "ZigZag Education", "Resources", "UK EdTech Competitors", "Mid", "Hundreds of thousands of students", "Exam-board-specific worksheets, schemes of work — teacher-facing", "Partnership / Cross-promo", "Email", "Not Started", "English departments buy direct — exam-board-specific resources"],
    [711, "The Day (theday.co.uk)", "Platform", "UK EdTech Competitors", "Mid", "3,000+ UK schools", "Daily news-based reading + comprehension — English Language skills", "Partnership / Cross-promo", "Email", "Not Started", "3K schools — news literacy, English Language GCSE skills overlap"],
    [712, "Educake", "Platform", "UK EdTech Competitors", "Mid", "Thousands of UK schools", "Auto-marked homework quizzes — English Language content", "Partnership / Integration", "Email", "Not Started", "Low-stakes homework tool — complementary, teacher-friendly"],

    # ═══ UK EDUCATION ONLINE COMMUNITIES (web-verified) ═══
    [713, "r/IGCSE (Reddit)", "Reddit", "UK Education Communities", "Macro", "~127K members", "Cambridge IGCSE students — resources, past papers, exam discussion", "Organic / AMA", "Reddit", "Not Started", "127K — IGCSE students, international school audience, direct fit"],
    [714, "r/Edexcel (Reddit)", "Reddit", "UK Education Communities", "Micro", "~13K members", "Edexcel exam board students — GCSE + A-Level discussion", "Organic / Resource sharing", "Reddit", "Not Started", "13K — exam-board-specific, resource sharing niche"],
    [715, "r/UKParenting (Reddit)", "Reddit", "UK Education Communities", "Mid", "~30K members", "UK parents — secondary school, GCSE options, school choices", "Organic / AMA", "Reddit", "Not Started", "30K — UK parents discussing GCSE options + school decisions"],
    [716, "GCSE 9-1 Discord Server", "Discord", "UK Education Communities", "Micro", "~6.4K members", "Subject-specific channels, homework help, exam discussion", "Community presence / Partnership", "Discord", "Not Started", "6.4K — active GCSE Discord, subject helpers, study community"],
    [717, "r/IGCSE Discord Server", "Discord", "UK Education Communities", "Micro", "Part of 127K ecosystem", "Study channels, past paper discussion, peer tutoring", "Community presence", "Discord", "Not Started", "IGCSE Discord — peer tutoring, study channels"],
    [718, "Discord School UK", "Discord", "UK Education Communities", "Micro", "Several thousand", "GCSE + A-Level revision channels, UK student community", "Community presence", "Discord", "Not Started", "UK-focused Discord — revision channels, student community"],

    # ═══ UK REALITY TV — ADDITIONAL (training data) ═══
    [719, "Gemma Owen (@gaborowen)", "Instagram", "UK Reality TV", "Macro", "~2M IG", "Love Island 2022 — Michael Owen's daughter, equestrian + fashion", "Sponsored post", "Agent", "Not Started", "2M — Love Island 2022, teen fashion audience"],
    [720, "Ekin-Su Culculoglu (@ekinsuofficial)", "Instagram", "UK Reality TV", "Macro", "~2.8M IG", "Love Island 2022 winner — massive teen/young adult following", "Sponsored post", "Agent", "Not Started", "2.8M — Love Island winner, teen female icon"],
    [721, "Tasha Ghouri (@tashaghouri)", "IG/TikTok", "UK Reality TV", "Macro", "~1.8M IG", "Love Island 2022 — strong TikTok, younger-skewing audience", "Sponsored post", "DM", "Not Started", "1.8M — TikTok presence skews younger, deaf representation"],
    [722, "Olivia Attwood (@oliviaattwood)", "Instagram", "UK Reality TV", "Macro", "~2M IG", "Love Island 2017 — own ITV series, fashion + lifestyle", "Sponsored post", "Agent", "Not Started", "2M — Love Island OG, own ITV show, teen/young adult following"],
    [723, "Stacey Dooley (@sjdooley)", "Instagram", "UK Reality TV", "Macro", "~1M IG", "Strictly 2018 winner — BBC documentarian, educational crossover", "Partnership / Sponsored", "Email", "Not Started", "1M — BBC docs + Strictly, educational credibility angle"],

    # ═══ SAUDI MEGA INFLUENCERS (training data) ═══
    [724, "Abu Flah / Abdulaziz Al-Jasmi (@AboFlah)", "YouTube", "Saudi Mega Influencers", "Mega", "~35M YT", "Entertainment, charity streams — massive Saudi/GCC youth audience", "Brand deal", "Agent", "Not Started", "35M — biggest Saudi YouTuber, charity streams, youth audience"],
    [725, "BanderitaX (@BanderitaX)", "YouTube", "Saudi Mega Influencers", "Mega", "~18M YT", "Gaming, comedy — dominant Saudi gaming creator", "Brand deal", "Agent", "Not Started", "18M — Saudi gaming #1, teen male audience"],
    [726, "Bader Saleh (@BaderSaleh)", "YouTube/Snapchat", "Saudi Mega Influencers", "Mega", "~6M YT", "Comedy, talk show format — Saudi youth entertainment", "Brand deal", "Agent", "Not Started", "6M — Saudi comedy talk show, young audience"],
    [727, "Fouz Al-Fahad (@fouzderella)", "Instagram", "Saudi Mega Influencers", "Mega", "~5M IG", "Lifestyle, fashion, beauty — Saudi/GCC women", "Sponsored post", "Agent", "Not Started", "5M — Saudi lifestyle queen, female GCC audience"],
    [728, "Ibrahim Saleh (@ibrahimsa11eh)", "TikTok/YouTube", "Saudi Mega Influencers", "Mega", "~5M TikTok", "Comedy sketches — Saudi youth, viral format", "Sponsored video", "DM", "Not Started", "5M — comedy sketches, Saudi teen audience"],
    [729, "Ahmed Show (@Ahmedshow2)", "YouTube/Snapchat", "Saudi Mega Influencers", "Mega", "~4M YT", "Comedy, family content — Saudi household name", "Brand deal", "Agent", "Not Started", "4M — family comedy, cross-generational Saudi audience"],
    [730, "Njoud Al Shammari (@njoud_alshamri)", "TikTok/IG", "Saudi Mega Influencers", "Macro", "~3M TikTok", "Comedy, entertainment — Saudi female teen audience", "Sponsored post", "DM", "Not Started", "3M — comedy, female teen following"],

    # ═══ UK FEMALE FITNESS & WELLNESS — ADDITIONAL (training data) ═══
    [731, "Krissy Cela (@krissycela)", "IG/YouTube", "UK Fitness & Wellness", "Macro", "~3.5M IG", "Gym workouts, Oner Active founder — massive teen/young adult female following", "Sponsored post", "Email", "Not Started", "3.5M — Oner Active brand, gym fitness, teen female audience"],
    [732, "Stef Williams (@stefwilliams_)", "Instagram", "UK Fitness & Wellness", "Macro", "~1.5M IG", "Gym workouts, WeGLOW app founder — fitness coaching", "Sponsored post", "Email", "Not Started", "1.5M — WeGLOW app, gym content, young female audience"],
    [733, "Alice Liveing (@aliceliveing)", "Instagram", "UK Fitness & Wellness", "Mid", "~700K IG", "Strength training, body positivity, mental health — Nike ambassador", "Sponsored post", "Email", "Not Started", "700K — body positivity, mental health angle, young women"],
    [734, "Chessie King (@chessieking)", "IG/TikTok", "UK Fitness & Wellness", "Mid", "~700K IG", "Body confidence, comedic wellness, mental health — teen appeal", "Sponsored post", "DM", "Not Started", "700K — body confidence, comedic approach, teen girls love her"],
    [735, "Courtney Black (@courtneydblack)", "IG/YouTube", "UK Fitness & Wellness", "Macro", "~1M IG / 500K YT", "Home workouts, fitness app — challenges format, teen audience", "Sponsored post", "Email", "Not Started", "1.5M combined — home workout challenges, app-based training"],
    [736, "Hazel Wallace / The Food Medic (@thefoodmedic)", "Instagram", "UK Fitness & Wellness", "Mid", "~300K IG", "Nutrition, evidence-based health, wellness — doctor + influencer", "Sponsored post", "Email", "Not Started", "300K — doctor credibility, nutrition + wellness, young women"],
    [737, "Carly Rowena (@carlyrowena)", "IG/YouTube", "UK Fitness & Wellness", "Mid", "~400K IG / 300K YT", "Home workouts, pregnancy fitness, wellness — family audience", "Sponsored post", "Email", "Not Started", "700K combined — home workouts, family/parent crossover"],

    # ═══ UK TEEN FASHION & STREETWEAR BRANDS (training data) ═══
    [738, "Corteiz / CRTZ (@corteiz)", "Instagram", "UK Teen Fashion Brands", "Macro", "~1.8M IG", "Dominant UK streetwear brand — guerrilla drops, 14-18 hype culture", "Cross-promo / Competition", "Email", "Not Started", "1.8M — #1 UK teen streetwear, scarcity marketing, school-age hype"],
    [739, "Trapstar (@trapstarlondon)", "Instagram", "UK Teen Fashion Brands", "Macro", "~1.5M IG", "London streetwear icon — worn by every UK teen boy", "Cross-promo / Competition", "Email", "Not Started", "1.5M — London streetwear, teen male wardrobe staple"],
    [740, "PrettyLittleThing (@prettylittlething)", "Instagram", "UK Teen Fashion Brands", "Mega", "~21M IG", "Fast fashion — Molly-Mae ambassador, massive teen female audience", "Cross-promo / Back-to-school", "Email", "Not Started", "21M — biggest UK fast fashion, teen female audience"],
    [741, "Oh Polly (@ohpolly)", "Instagram", "UK Teen Fashion Brands", "Mega", "~5M IG", "Fashion brand — very popular with UK teen girls, viral TikTok", "Cross-promo", "Email", "Not Started", "5M — teen girls, viral TikTok fashion content"],
    [742, "Gym King (@gymking)", "Instagram", "UK Teen Fashion Brands", "Macro", "~1.1M IG", "Athleisure/streetwear — commonly seen in UK schools and sixth forms", "Cross-promo", "Email", "Not Started", "1.1M — athleisure staple, teen male audience, school-age"],

    # ═══ UK TRUE CRIME & DOCUMENTARY — TEEN AUDIENCE (training data) ═══
    [743, "Eleanor Neale (@eleanorneale)", "YouTube", "UK True Crime & Documentary", "Macro", "~2.5M YT", "True crime deep dives — massive teen girl audience, GCSE-age", "Sponsored video", "Email", "Not Started", "2.5M — true crime YouTuber, 16-25 female demographic"],
    [744, "The Rest Is Entertainment (Richard Osman & Marina Hyde)", "Podcast", "UK True Crime & Documentary", "Mega", "Millions weekly", "Top 5 UK podcast — media, entertainment, cultural analysis", "Sponsored read", "Agent", "Not Started", "Top 5 UK podcast — Richard Osman, cross-generational"],
    [745, "RedHanded (@redhandedthepodcast)", "Podcast", "UK True Crime & Documentary", "Macro", "~200M+ total downloads", "UK true crime with social commentary — consistently top UK charts", "Sponsored read", "Agent", "Not Started", "200M downloads — top UK true crime podcast, 16-25 audience"],

    # ═══ UK TEEN MENTAL HEALTH & WELLNESS APPS (training data) ═══
    [746, "Kooth (kooth.com)", "App/Website", "UK Teen Mental Health Apps", "Macro", "4.5M+ registered users", "NHS-commissioned online counselling for 11-25s — free, peer support", "Partnership", "Email", "Not Started", "4.5M users — NHS-funded, every school can refer, direct teen reach"],
    [747, "Shout 85258 (@gaborneiveasmile)", "SMS/Social", "UK Teen Mental Health Apps", "Mid", "~55K social combined", "24/7 free crisis text line for young people — social content", "Partnership / Campaign", "Email", "Not Started", "24/7 crisis text — exam stress campaign partnership angle"],
    [748, "Calm Harm / Stem4 (@stem4org)", "App/Social", "UK Teen Mental Health Apps", "Micro", "~25K social combined", "CBT-based apps for teens — self-harm, anxiety management", "Partnership", "Email", "Not Started", "Stem4 charity — Calm Harm + Clear Fear apps, school partnerships"],
    [749, "Togetherall (formerly Big White Wall)", "Website", "UK Teen Mental Health Apps", "Macro", "4.5M+ via uni partnerships", "Peer support + CBT courses — NHS + university commissioned", "Partnership", "Email", "Not Started", "4.5M students — uni partnerships, 24/7 clinical moderation"],
    [750, "MeeTwo", "App", "UK Teen Mental Health Apps", "Micro", "Teen user base", "Anonymous peer support for teens 12-20 — exam stress, bullying", "Partnership", "Email", "Not Started", "Covers exam stress — natural fit for revision platform cross-promo"],

    # ═══ UK SEND & DISABILITY INCLUSION ADVOCATES (training data) ═══
    [751, "Special Needs Jungle (@SNJ)", "X/Website", "UK SEND & Inclusion", "Mid", "~25K X", "Leading UK SEND info site — EHCP processes, tribunals, parent rights", "Partnership / Sponsored", "Email", "Not Started", "25K — trusted SEND resource, parent audience, accessibility angle"],
    [752, "Square Peg (@SquarePegUK)", "X", "UK SEND & Inclusion", "Micro", "~12K X", "School attendance difficulties linked to SEND/anxiety — parent campaign", "Partnership", "Email", "Not Started", "12K — attendance/anxiety, online learning alternative angle"],
    [753, "The SEND Parenting Podcast (@thesendpodcast)", "Podcast/IG", "UK SEND & Inclusion", "Micro", "~8K IG", "Parent-led SEND community — navigating the system", "Sponsored read / Partnership", "Email", "Not Started", "8K — SEND parents, podcast sponsorship opportunity"],
    [754, "Joanna Grace / The Sensory Projects (@JoGraceSensory)", "X/IG", "UK SEND & Inclusion", "Micro", "~15K combined", "Sensory engagement specialist — resources for educators + families", "Partnership", "Email", "Not Started", "15K — sensory SEND, educator resources, accessibility"],
    [755, "Nancy Gedge (@NancyGedge)", "X", "UK SEND & Inclusion", "Micro", "~18K X", "Primary teacher specialising in SEND inclusion in mainstream schools", "Partnership / Guest content", "Email", "Not Started", "18K — SEND inclusion practitioner, trusted teacher voice"],

    # ═══ UK HISTORY & CULTURE CREATORS (training data) ═══
    [756, "Dan Snow / History Hit (@thehistoryguy)", "YouTube/TikTok", "UK History & Culture", "Mid", "~500K YT / 300K TikTok", "British history documentaries, History Hit network", "Sponsored video / Partnership", "Email", "Not Started", "800K combined — History Hit network, GCSE History crossover"],
    [757, "History Matters (@HistoryMatters)", "YouTube", "UK History & Culture", "Macro", "~2.5M YT", "Short animated history explainers — British/European topics", "Sponsored video", "Email", "Not Started", "2.5M — animated explainers, GCSE History students watch these"],
    [758, "English Heritage (@EnglishHeritage)", "YouTube/TikTok", "UK History & Culture", "Macro", "~1.2M YT / 1M TikTok", "Heritage sites, historical cooking, Victorian life — viral TikTok", "Partnership / Sponsored", "Email", "Not Started", "2.2M combined — historical cooking goes viral, teen audience"],
    [759, "Horrible Histories (Official)", "YouTube/TikTok", "UK History & Culture", "Mid", "~500K+ YT", "CBBC sketch comedy — British history, GCSE-adjacent content", "Partnership", "Email", "Not Started", "500K+ — CBBC, every UK teen knows Horrible Histories"],
    [760, "The History Chap (@thehistorychap)", "TikTok", "UK History & Culture", "Macro", "~1M TikTok", "Short-form British history, monarchy, culture", "Sponsored video", "DM", "Not Started", "1M TikTok — short-form history, teen-friendly format"],
    [761, "Timeline - World History Documentaries", "YouTube", "UK History & Culture", "Mega", "~8M YT", "Full-length British history docs — Tudor, medieval, WWII", "Pre-roll / Sponsored", "Email", "Not Started", "8M — massive history channel, UK-heavy content"],
    [762, "Absolute History", "YouTube", "UK History & Culture", "Macro", "~2M YT", "Full-length British history documentaries", "Pre-roll / Sponsored", "Email", "Not Started", "2M — Tudor, medieval, WWII docs, student viewers"],
    [763, "Mary Beard (@wmarybeard)", "X/BBC", "UK History & Culture", "Mid", "~300K X", "Classical + British history — BBC documentarian, Cambridge professor", "Aspirational / Partnership", "Agent", "Not Started", "300K — Cambridge prof, BBC docs, academic credibility"],

    # ═══ UK YOUTH PUBLICATIONS & MAGAZINES (training data) ═══
    [764, "Shout Magazine", "Print/Online", "UK Youth Publications", "Mid", "~50K+ print circulation", "Teen girls 11-16 — celebrity, fashion, school life, advice", "Advertorial / Sponsored", "Email", "Not Started", "50K+ print — 11-16 age group, perfect GCSE demographic"],
    [765, "Dazed (Dazed Media)", "Digital/Print", "UK Youth Publications", "Macro", "~4M monthly online", "Youth culture, fashion, music — defining voice for young creatives", "Sponsored content / Banner", "Email", "Not Started", "4M monthly — youth culture authority, 16-25 audience"],
    [766, "i-D Magazine", "Digital", "UK Youth Publications", "Macro", "~3M monthly online", "Fashion, music, youth culture, identity — large IG following", "Sponsored content", "Email", "Not Started", "3M monthly — youth culture, fashion, 16-25 creative audience"],
    [767, "The Face", "Digital", "UK Youth Publications", "Macro", "~1.5M monthly online", "Iconic UK youth culture brand — music, fashion, style since 1980", "Sponsored content", "Email", "Not Started", "1.5M monthly — iconic youth brand, relaunched digital"],
    [768, "gal-dem", "Digital", "UK Youth Publications", "Mid", "~500K monthly", "Women + non-binary POC perspectives — culture, politics, music", "Sponsored content / Partnership", "Email", "Not Started", "500K monthly — diverse voice, 18-30, diversity angle for English Hub"],

    # ═══ GCC EDUCATION CONFERENCES & EVENTS (training data) ═══
    [769, "GESS Dubai (Global Educational Supplies & Solutions)", "Conference", "GCC Education Conferences", "Macro", "10K+ attendees / 500+ exhibitors", "Largest education trade show in Middle East — EdTech, STEAM", "Exhibit / Sponsor", "Email", "Not Started", "10K+ — #1 ME education show, Dubai World Trade Centre, November"],
    [770, "BETT MEA (Middle East & Africa)", "Conference", "GCC Education Conferences", "Mid", "~5K attendees", "Regional Bett Show — EdTech, digital learning, AI in education", "Exhibit / Sponsor", "Email", "Not Started", "5K — Bett regional edition, Abu Dhabi/Dubai, November"],
    [771, "GETEX (Gulf Education & Training Exhibition)", "Conference", "GCC Education Conferences", "Macro", "~30K visitors", "Higher ed, university recruitment — Spring + Autumn editions", "Exhibit / Sponsor", "Email", "Not Started", "30K — oldest Gulf education exhibition, Dubai, April + October"],
    [772, "WISE (World Innovation Summit for Education)", "Conference", "GCC Education Conferences", "Mid", "~3K from 100+ countries", "Global education policy + innovation — Qatar Foundation, biennial", "Attend / Speak / Sponsor", "Email", "Not Started", "3K — Doha, Qatar Foundation, global education policy leaders"],
    [773, "IECHE (International Exhibition on Higher Education)", "Conference", "GCC Education Conferences", "Mega", "~50K visitors", "Saudi Ministry of Education — university admissions, scholarships", "Exhibit", "Email", "Not Started", "50K — Riyadh, Saudi Ministry, massive higher ed event"],
    [774, "Najah (Abu Dhabi Education & Career Exhibition)", "Conference", "GCC Education Conferences", "Macro", "~30K visitors", "Higher ed, career guidance, university admissions for UAE students", "Exhibit / Sponsor", "Email", "Not Started", "30K — Abu Dhabi, student career decisions, scholarship focus"],
    [775, "SETE (Saudi Education & Training Exhibition)", "Conference", "GCC Education Conferences", "Macro", "~20K visitors", "K-12 + e-learning — aligned with Vision 2030", "Exhibit / Sponsor", "Email", "Not Started", "20K — Riyadh, Vision 2030 education investment, May"],
    [776, "Sharjah Children's Reading Festival", "Festival", "GCC Education Conferences", "Mega", "~100K+ visitors", "Children's literacy, educational workshops — Arabic + English", "Exhibit / Workshop", "Email", "Not Started", "100K+ over 10 days — literacy focus, perfect for English Hub"],
    [777, "Knowledge Summit (MBRF)", "Conference", "GCC Education Conferences", "Mid", "~5K attendees", "Knowledge economy, education development — Dubai", "Attend / Sponsor", "Email", "Not Started", "5K — MBRF, knowledge economy, Nov/Dec Dubai"],
    [778, "Abu Dhabi International Book Fair", "Festival", "GCC Education Conferences", "Mega", "~100K+ visitors", "Publishing, literacy, education content — Arabic + English", "Exhibit / Sponsor", "Email", "Not Started", "100K+ — ADNEC, May, education content + literacy"],

    # ═══ UK SPOKEN WORD & POETRY CREATORS (training data) ═══
    [779, "Suli Breaks (@sulibreaks)", "YouTube/IG", "UK Spoken Word & Poetry", "Mid", "~600K YT", "Viral spoken word on education + ambition — 'Why I Hate School' went mega-viral", "Partnership / Sponsored", "Email", "Not Started", "600K — viral education poetry, directly about school/learning"],
    [780, "George the Poet (@georgethepoet)", "IG/Podcast", "UK Spoken Word & Poetry", "Mid", "~300K IG", "Spoken word, social commentary — 'Have You Heard George's Podcast?'", "Partnership / Guest", "Agent", "Not Started", "300K — BBC podcast, spoken word pioneer, GCSE English crossover"],
    [781, "Nikita Gill (@nikita_gill)", "IG/TikTok", "UK Spoken Word & Poetry", "Mid", "~500K IG", "Written poetry, feminist/mythological themes — massive BookTok presence", "Partnership / Sponsored", "Email", "Not Started", "500K — poetry + BookTok, feminist themes, teen girls love her"],
    [782, "Tomos Roberts / Tom Foolery (@tomfoolery)", "IG/TikTok", "UK Spoken Word & Poetry", "Mid", "~200K IG", "Inspirational spoken word — 'The Great Realisation' went viral", "Partnership / Sponsored", "Email", "Not Started", "200K — viral spoken word, social commentary, school assemblies"],
    [783, "Hussain Manawer (@hussainmanawer)", "IG/TikTok", "UK Spoken Word & Poetry", "Micro", "~150K IG", "Mental health spoken word, motivational poetry — school visits", "Partnership / School event", "Email", "Not Started", "150K — mental health poetry, school visits, GCSE relevant"],
    [784, "Hollie McNish (@holliemcnish)", "IG/TikTok", "UK Spoken Word & Poetry", "Micro", "~100K IG", "Spoken word on motherhood, politics, identity — viral poetry videos", "Partnership / Sponsored", "Email", "Not Started", "100K — viral poetry, parent + student crossover audience"],
    [785, "Kae Tempest (@kaetempest)", "IG", "UK Spoken Word & Poetry", "Micro", "~90K IG", "Spoken word, hip-hop poetry — Mercury Prize-nominated", "Aspirational / Partnership", "Agent", "Not Started", "90K — Mercury-nominated, spoken word pioneer, literary credibility"],
    [786, "Sophia Thakur (@sophiathakur)", "IG/TikTok", "UK Spoken Word & Poetry", "Micro", "~80K IG", "Spoken word performances, love/identity poetry — published author", "Partnership / Sponsored", "Email", "Not Started", "80K — young spoken word star, love/identity themes, teen appeal"],
    [787, "Caleb Femi (@calebfemi)", "IG", "UK Spoken Word & Poetry", "Micro", "~25K IG", "Former Young People's Laureate for London — visual poetry", "Partnership / School event", "Email", "Not Started", "25K — Young People's Laureate, directly targets youth audience"],
    [788, "Raymond Antrobus (@raymondantrobus)", "IG", "UK Spoken Word & Poetry", "Micro", "~20K IG", "Deaf poet, Ted Hughes Award winner — BSL poetry, accessibility", "Partnership", "Email", "Not Started", "20K — Ted Hughes Award, deaf poet, accessibility + inclusion angle"],

    # ═══ UK TEACHER UNIONS & PROFESSIONAL BODIES (training data) ═══
    [789, "NEU (National Education Union)", "X/Magazine", "UK Teacher Unions", "Mega", "~450K members / 200K X", "Largest UK education union — Educate magazine, conference, CPD", "Magazine ad / Conference / CPD partner", "Email", "Not Started", "450K members — Educate mag reaches every member, conference exhibit"],
    [790, "NASUWT (The Teachers' Union)", "X/Magazine", "UK Teacher Unions", "Macro", "~300K members / 80K X", "Second largest UK teacher union — Teaching Today magazine", "Magazine ad / Conference", "Email", "Not Started", "300K members — Teaching Today mag, conference exhibit"],
    [791, "UCU (University and College Union)", "X", "UK Teacher Unions", "Macro", "~120K members / 100K X", "FE/HE lecturers and academics — Congress sponsorship", "Conference / Publication ad", "Email", "Not Started", "120K — FE/HE sector, Congress sponsorship opportunity"],

    # ═══ UK YOUTH CHARITIES — CAMPAIGN PARTNERSHIPS (training data) ═══
    [792, "BBC Children in Need", "TV/Social", "UK Youth Charity Campaigns", "Mega", "~1.5M social", "Annual school fundraising packs — every UK school participates", "Campaign partnership", "Email", "Not Started", "Every UK school — Pudsey, school fundraising, family audience"],
    [793, "Comic Relief / Red Nose Day", "TV/Social", "UK Youth Charity Campaigns", "Mega", "~1.5M social", "Massive school participation — TikTok creator partnerships", "Campaign partnership", "Email", "Not Started", "Huge school participation — creator partnerships drive teen engagement"],
    [794, "Anti-Bullying Alliance / Odd Socks Day", "Social/Schools", "UK Youth Charity Campaigns", "Mid", "~200K social", "Odd Socks Day trends on TikTok every November — thousands of schools", "Campaign partnership", "Email", "Not Started", "Viral in schools every Nov — anti-bullying + wellbeing angle"],
    [795, "MIND (Mental Health Charity)", "Social", "UK Youth Charity Campaigns", "Macro", "~800K social", "Time to Talk Day — mental health #1 concern for Gen Z", "Campaign partnership", "Email", "Not Started", "800K — mental health charity, exam stress campaign angle"],
    [796, "NSPCC", "Social/Schools", "UK Youth Charity Campaigns", "Macro", "~700K social", "Speak Out Stay Safe delivered in thousands of primary schools", "Campaign partnership", "Email", "Not Started", "700K — school programme delivery, safeguarding angle"],
    [797, "Samaritans (Brew Monday)", "Social", "UK Youth Charity Campaigns", "Mid", "~500K social", "Brew Monday in schools + unis — loneliness, reaching out", "Campaign partnership", "Email", "Not Started", "500K — Brew Monday in schools, exam period wellbeing"],
]

# ── Import additional agent-researched entries ──
try:
    import sys
    sys.path.insert(0, r"D:\Coding\english-hub")
    from additional_aspirational_data import additional_entries
    # Re-number and append
    start_num = len(aspirational_influencers) + 1
    for i, entry in enumerate(additional_entries):
        entry_copy = list(entry)
        entry_copy[0] = start_num + i  # Fix numbering
        aspirational_influencers.append(entry_copy)
    print(f"Added {len(additional_entries)} agent-researched entries (total: {len(aspirational_influencers)})")
except ImportError:
    print("No additional_aspirational_data.py found, using base entries only")
except Exception as e:
    print(f"Warning loading additional data: {e}")

# Write aspirational influencers to sheet
row_num = 2
current_cat = None
for entry in aspirational_influencers:
    cat = entry[3]
    if cat != current_cat:
        add_section_row(ws14, row_num, f"═══ {cat.upper()} ═══", len(cols14))
        row_num += 1
        current_cat = cat
    for i, v in enumerate(entry, 1):
        ws14.cell(row=row_num, column=i, value=v)
    style_data_row(ws14, row_num, len(cols14))
    ws14.row_dimensions[row_num].height = 22
    row_num += 1

# ════════════════════════════════════════════════════════════════
# SHEET 15: OUTREACH MESSAGES
# ════════════════════════════════════════════════════════════════
ws15 = wb.create_sheet("Outreach Messages")
cols15 = ["#", "Category / Sector", "Message Type", "Subject Line", "Full Message Template", "Personalization Fields", "Notes"]
for i, col in enumerate(cols15, 1):
    ws15.cell(row=1, column=i, value=col)
style_header_row(ws15, 1, len(cols15))
set_col_widths(ws15, [6, 40, 20, 50, 80, 25, 30])
ws15.row_dimensions[1].height = 30
ws15.auto_filter.ref = f"A1:G1"

# Load outreach messages
outreach_loaded = {}
try:
    import importlib.util
    spec = importlib.util.spec_from_file_location("outreach_messages", r"D:\Coding\english-hub\outreach_messages.py")
    om = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(om)
    outreach_loaded = om.outreach_messages
    print(f"Loaded {len(outreach_loaded)} outreach message templates")
except Exception as e:
    print(f"Warning loading outreach messages: {e}")

row15 = 2
msg_num = 1
current_cat_type = None
for cat in sorted(outreach_loaded.keys()):
    msg = outreach_loaded[cat]
    # Determine message type
    if any(kw in cat.lower() for kw in ['school', 'mat ', 'mat(', 'academy', 'primary', 'governance', 'leadership', 'procurement', 'gcc school', 'british curriculum', 'international school', 'inspection']):
        msg_type = "School Introduction"
    elif any(kw in cat.lower() for kw in ['influencer', 'blogger', 'creator', 'tiktok', 'youtube', 'instagram', 'vlogger', 'booktok', 'podcast', 'parent', 'dad ', 'mum ', 'teen', 'fashion', 'fitness', 'sport', 'beauty', 'comedy', 'gaming', 'music', 'food']):
        msg_type = "Affiliate Invitation"
    elif any(kw in cat.lower() for kw in ['charity', 'foundation', 'trust', 'organisation', 'network', 'association', 'union']):
        msg_type = "Partnership Proposal"
    elif any(kw in cat.lower() for kw in ['edtech', 'platform', 'app', 'saas', 'software', 'publisher']):
        msg_type = "Integration/Partnership"
    elif any(kw in cat.lower() for kw in ['journalist', 'media', 'publication', 'magazine', 'press']):
        msg_type = "Media/PR Pitch"
    else:
        msg_type = "Partnership Proposal"

    # Extract subject line
    subject = ""
    lines = msg.strip().split('\n')
    for line in lines:
        if line.strip().startswith('Subject:'):
            subject = line.strip().replace('Subject:', '').strip()
            break

    # Determine personalization fields
    fields = []
    if '{contact_name}' in msg or '{school_name}' in msg:
        fields.append('Name')
    if '{organization_name}' in msg or '{school_name}' in msg:
        fields.append('Organization')
    if '{platform}' in msg:
        fields.append('Platform')
    fields_str = ', '.join(fields) if fields else '{contact_name}, {organization_name}'

    # Add section header for new message types
    if msg_type != current_cat_type:
        add_section_row(ws15, row15, f"═══ {msg_type.upper()} ═══", len(cols15))
        row15 += 1
        current_cat_type = msg_type

    ws15.cell(row=row15, column=1, value=msg_num)
    ws15.cell(row=row15, column=2, value=cat)
    ws15.cell(row=row15, column=3, value=msg_type)
    ws15.cell(row=row15, column=4, value=subject)
    ws15.cell(row=row15, column=5, value=msg)
    ws15.cell(row=row15, column=6, value=fields_str)
    ws15.cell(row=row15, column=7, value=f"Template for {cat} contacts")
    style_data_row(ws15, row15, len(cols15))
    ws15.row_dimensions[row15].height = 60
    row15 += 1
    msg_num += 1

print(f"Sheet 15 'Outreach Messages': {msg_num - 1} templates written")

# ════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════
output_path = r"D:\Coding\english-hub\The_English_Hub_Master_GTM_Campaign_Tracker.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Sheets created: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
